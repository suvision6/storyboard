#!/usr/bin/env python3
"""Build and validate stable 7-column su-fenjingskill-zh deliveries."""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import math
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - reported at runtime.
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None

VERSION = "2.4.2"
RULE_REVISION = "2.4.2-source-lock-entry-guard-2026-07-07"
RULE_REVISIONS = {
    "2.3.4": "2.3.4-overcompression-guard-2026-06-30",
    "2.3.5": "2.3.5-adjacent-motion-guard-2026-06-30",
    "2.3.6": "2.3.6-tri-source-audit-guard-2026-06-30",
    "2.4.0": "2.4.0-human-gate-stability-2026-07-06",
    "2.4.1": "2.4.1-source-lock-2026-07-07",
    "2.4.2": RULE_REVISION,
}
SHEET_NAME = "分镜表"
SUMMARY_SHEET_NAME = "校验摘要"
HEADERS = [
    "镜号",
    "场景",
    "原剧本段落",
    "镜头时长(秒)",
    "运镜+主画面描述(含台词)",
    "备注",
    "Prompt",
]
PROMPT_FIELDS = ["时间：", "景别：", "构图：", "运镜手法：", "画面内容："]
DURATION_FIELDS = [
    "sync_action_seconds",
    "sync_dialogue_seconds",
    "non_sync_action_seconds",
    "emotional_pause_seconds",
]
FACT_TYPES = {
    "character",
    "action",
    "dialogue",
    "prop",
    "space",
    "position",
    "emotion",
    "sound",
    "reality",
}
SHOT_TYPES = {
    "master",
    "action",
    "dialogue",
    "reaction",
    "insert",
    "transition",
    "vfx_anchor",
    "safety",
}
SPLIT_REASONS = {
    "spatial_anchor",
    "performance_continuity",
    "new_information",
    "prop_state_change",
    "new_vfx_state",
    "new_sound_source",
    "reality_layer_shift",
    "causal_reveal",
    "emotional_turn",
    "continuity_migration",
}
INSERT_PRIORITIES = {"none", "recommended", "must_have"}
CUT_PRIORITIES = {"normal", "recommended", "must_isolate"}
CUT_CATEGORIES = {"space", "prop", "action", "emotion", "sound", "reality", "dialogue", "vfx", "character"}
CRITICAL_SPLIT_REASONS = {
    "causal_reveal",
    "reality_layer_shift",
    "new_vfx_state",
    "new_sound_source",
    "emotional_turn",
    "prop_state_change",
}
MUST_ISOLATE_REASONS = {
    "causal_reveal",
    "reality_layer_shift",
    "new_vfx_state",
    "new_sound_source",
    "emotional_turn",
}
LONG_TAKE_SUPPORTS = {
    "shot_size_change",
    "character_blocking",
    "foreground_background_change",
    "sound_source_change",
    "vfx_state_change",
    "emotional_turn",
    "spatial_progression",
}
INSERT_LIKE_TYPES = {"insert", "reaction", "vfx_anchor", "safety"}
HYBRID_VERSIONS = {"2.3.4", "2.3.5", "2.3.6", "2.4.0", "2.4.1", "2.4.2"}
ADJACENT_MOTION_GUARD_VERSIONS = {"2.3.5", "2.3.6", "2.4.0", "2.4.1", "2.4.2"}
STRUCTURED_AUDIT_VERSIONS = {"2.3.6", "2.4.0", "2.4.1", "2.4.2"}
HUMAN_GATE_VERSIONS = {"2.4.0", "2.4.1", "2.4.2"}
SOURCE_LOCK_VERSIONS = {"2.4.1", "2.4.2"}
APPROVED_SCRIPT_PATH_VERSIONS = {"2.4.2"}
MUST_ISOLATE_DENSITY_THRESHOLD = 0.5
VALID_STATUSES = {"PASS", "WARN", "FAIL", "NOT_RUN"}
ANCHOR_TYPES = {"space", "multi_character", "both", "single_continuation", "subjective"}
REQUIRED_REFERENCE_KEYS = ("continuity-shot-data", "hybrid-shot-audit", "camera-language", "seedance-prompt-rules")
REQUIRED_SCRIPT_KEYS = ("storyboard_delivery.py", "validate_storyboard.js")
GATES = {"GATE_0", "GATE_A", "GATE_B", "GATE_C"}
LONG_TAKE_CLASSIFICATIONS = {"not_applicable", "dialogue_long", "action_long", "spatial_long", "emotional_long"}
NOTE_MARKERS = [
    "[合理补足]",
    "[时长估算]",
    "[长镜头]",
    "[保留理由]",
    "[不可拆说明]",
    "[景别跨度]",
    "[非连续Beat]",
    "[必拆相邻]",
    "[反转动机]",
    "[安全镜]",
    "[人工批准]",
    "[声音]",
    "[reference missing]",
]
PANORAMIC_SHOT_SIZE_KEYWORDS = ("大远景", "大全景", "中全景", "全景")
PANORAMIC_REQUIRED_ANCHORS = {"space", "multi_character", "both"}
AERIAL_ALLOWED_SHOT_SIZE_KEYWORDS = ("大远景", "大全景", "全景")
AERIAL_ALLOWED_ANGLE_KEYWORDS = ("俯拍", "高角度", "正俯拍", "微俯视")
WIDE_SHOT_SIZE_KEYWORDS = ("大远景", "大全景", "远景", "全景", "中全景")
CLOSE_SHOT_SIZE_KEYWORDS = ("中近景", "近景", "中特写", "特写", "微距特写")
SHOT_SIZE_RANKS = [
    (0, ("大远景", "大全景", "远景")),
    (1, ("全景", "中全景")),
    (2, ("中景", "七分身", "半身")),
    (3, ("中近景", "近景", "中特写")),
    (4, ("微距特写", "特写")),
]
STATIC_OR_SOFT_MOVEMENT_KEYWORDS = (
    "固定",
    "轻微",
    "缓慢",
    "微晃",
    "呼吸式",
    "下摇",
    "纵摇",
    "横摇",
)
PUSH_IN_MOVEMENT_KEYWORDS = (
    "推轨推进",
    "缓慢推进",
    "航拍推进",
    "急推",
    "推进",
)
PULL_OUT_MOVEMENT_KEYWORDS = (
    "推轨拉出",
    "缓慢拉出",
    "拉远",
    "拉出",
)
SPAN_MOVEMENT_KEYWORDS = (
    "光学变焦",
    "快速推拉",
    "急推",
    "急拉",
    "推轨推进",
    "推轨拉出",
    "缓慢推进",
    "缓慢拉出",
    "伸缩摇臂",
)
INCOMPLETE_DIALOGUE_ENDINGS = ("，", "、", "：", "；", ",", ":", ";")
EXPRESSION_ONLY_KEYWORDS = (
    "笑",
    "苦笑",
    "嘴角",
    "眼神",
    "抬眼",
    "闭眼",
    "呼吸",
    "停顿",
    "沉默",
    "表情",
    "眉",
    "瞳孔",
    "下颌",
    "嘴唇",
)
NEW_FACT_TYPES_FOR_CUT = {"position", "prop", "space", "sound", "reality"}
STATEFUL_CONTINUITY_FIELDS = {"position", "facing", "state", "owner", "value", "presence"}
POSITION_UPDATE_FIELDS = {"position", "facing", "presence"}
PROMPT_COMPOSITION_FORBIDDEN_PATTERNS = [
    re.compile(r"[“”\"]"),
    re.compile(r"[\u4e00-\u9fa5A-Za-z0-9·]{1,8}(?:说|喊|问|答|念)"),
    re.compile(r"画外声|对白|台词|传来|响起|震动|发出"),
    re.compile(r"走到|冲向|拿起|放下|击中|钻入|劈下|吞没|笑|哭|喊|抬手|跪地"),
]
FACT_TEXT_ALLOWLIST = {
    "画面",
    "镜头",
    "机位",
    "朝向",
    "轴线",
    "左侧",
    "右侧",
    "前景",
    "背景",
    "可见",
}
INTERNAL_LABELS = [
    "【镜内变化】",
    "【人物关系】",
    "【人物关系补充】",
    "【站位位移】",
    "【机位逻辑】",
    "【场景首镜站位】",
]
PROMPT_FORBIDDEN = [
    "【镜内变化】",
    "【人物关系】",
    "【人物关系补充】",
    "【站位位移】",
    "【机位逻辑】",
    "【场景首镜站位】",
    "运镜修正",
    "关键帧",
    "生图",
    "AI生图提示词",
    "风格：",
    "禁止：",
] + NOTE_MARKERS
UNSUPPORTED_MARKERS = [
    "承上镜动作",
    "承上镜",
    "补充镜头",
    "新增镜头",
    "新增过渡",
    "无原文依据",
    "非原文",
    "原文未写",
    "自行添加",
    "自行补充",
    "为了过渡",
    "建议保留此镜",
]
FORBIDDEN_VISUAL_WORDS = [
    "令人叹为观止",
    "令人惊叹",
    "令人着迷",
    "精心打造",
    "匠心独运",
    "独具匠心",
    "视觉盛宴",
    "光影交响",
    "完美呈现",
    "极致体验",
    "引人入胜",
    "震撼人心",
    "巧妙融合",
    "仿佛",
    "犹如",
    "宛如",
    "好似",
]

REQUIRED_CUT_POINT_CATEGORIES = [
    (
        "发言权转移",
        [re.compile(r"[\u4e00-\u9fa5A-Za-z0-9·]{1,8}[：:]")],
        True,
    ),
    ("问答关系变化", [re.compile(r"问|回答|回应|反问|是不是|为什么|什么|吗|？|\?")], False),
    (
        "角色明显反应",
        [re.compile(r"一怔|脸色|眼眶|嘴唇|冷笑|大笑|苦笑|痛苦|不可思议|回头|低头|抬头|摇头|点头|踉跄|绷|僵|停住")],
        False,
    ),
    (
        "道具状态变化",
        [re.compile(r"手环|短棍|长棍|磁卡|魂钉|裂痕|项链|绿灯|显示|亮起|暗下|恢复|变长|延伸|变成|熄灭|铜镯")],
        False,
    ),
    (
        "攻击/命中/结果",
        [re.compile(r"抬手|挥|扑|冲|拽|拦|扫|斩|震|炸|击|挡|断裂|碎裂|灰飞烟灭|跳入|钻进|坠落|吞没")],
        False,
    ),
    (
        "空间方向改变",
        [re.compile(r"东侧|西侧|南侧|中央|高台|祭池边|走向|走到|冲向|转身|回到|两侧|不同方向|分头|裂缝|地底|神殿|人间")],
        False,
    ),
    ("层级切换", [re.compile(r"切回|切——|梦境|幻境|现实|闪回|脑海|灵魂|亡魂|残魂|画外声|VO|回忆")], False),
    (
        "阵法/VFX状态变化",
        [re.compile(r"阵眼|破阵|续阵|阵破|符文|祭池|锁魂柱|柱身|黑纹|雾气|黑雾|龙卷|漩涡|回收|膨胀|光柱|封印|赤光|蓝光")],
        False,
    ),
]


@dataclass
class ValidationResult:
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def status(self) -> str:
        if self.errors:
            return "FAIL"
        if self.warnings:
            return "WARN"
        return "PASS"

    def error(self, message: str) -> None:
        self.errors.append(message)

    def warn(self, message: str) -> None:
        if message not in self.warnings:
            self.warnings.append(message)


def warning_id(message: str) -> str:
    return "W-" + hashlib.sha1(clean_text(message).encode("utf-8")).hexdigest()[:12]


def warning_items(warnings: list[str]) -> list[dict[str, str]]:
    return [{"warn_id": warning_id(message), "message": message} for message in warnings]


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig") as handle:
        value = json.load(handle)
    if not isinstance(value, dict):
        raise ValueError("shot_data 顶层必须是 JSON 对象。")
    return value


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def data_without_validation_report(data: dict[str, Any]) -> dict[str, Any]:
    value = copy.deepcopy(data)
    value.pop("validation_report", None)
    return value


def canonical_data_hash(data: dict[str, Any]) -> str:
    payload = json.dumps(data_without_validation_report(data), ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def report_summary_items(data: dict[str, Any]) -> list[tuple[str, str]]:
    metadata = as_dict(data.get("metadata"))
    report = as_dict(data.get("validation_report"))
    reference_status = as_dict(metadata.get("reference_status"))
    items = [
        ("version", clean_text(metadata.get("version"))),
        ("rule_revision", clean_text(metadata.get("rule_revision"))),
        ("source_json_hash", clean_text(report.get("source_json_hash"))),
        ("status", clean_text(report.get("status"))),
        ("warnings", str(len(as_list(report.get("warnings"))))),
        ("errors", str(len(as_list(report.get("errors"))))),
        ("reference_status", json.dumps(reference_status, ensure_ascii=False, sort_keys=True)),
    ]
    if clean_text(metadata.get("version")) in HUMAN_GATE_VERSIONS:
        items.extend(
            [
                ("human_gate_audit", json.dumps(as_dict(report.get("human_gate_audit")), ensure_ascii=False, sort_keys=True)),
                ("warn_resolution_audit", json.dumps(as_dict(report.get("warn_resolution_audit")), ensure_ascii=False, sort_keys=True)),
                ("scene_id_audit", json.dumps(as_dict(report.get("scene_id_audit")), ensure_ascii=False, sort_keys=True)),
            ]
        )
    return items


def resolved_warning_ids(data: dict[str, Any]) -> set[str]:
    ids: set[str] = set()
    for item in as_list(data.get("warn_resolutions")):
        if isinstance(item, dict) and clean_text(item.get("warn_id")):
            ids.add(clean_text(item.get("warn_id")))
    return ids


def human_gate_audit(data: dict[str, Any], *, final_signoff: bool = False) -> dict[str, Any]:
    required = ["GATE_A", "GATE_B"]
    if data.get("batch_plan"):
        required.insert(0, "GATE_0")
    if final_signoff:
        required.append("GATE_C")
    approved = [gate for gate in sorted(GATES) if approved_review(data, gate)]
    missing = [gate for gate in required if gate not in approved]
    return {"required": required, "approved": approved, "missing": missing}


def warn_resolution_audit(data: dict[str, Any], warnings: list[str]) -> dict[str, Any]:
    required = {warning_id(message) for message in warnings}
    resolved = resolved_warning_ids(data)
    return {"required": sorted(required), "resolved": sorted(resolved), "missing": sorted(required - resolved)}


def scene_id_audit(data: dict[str, Any]) -> dict[str, Any]:
    logs = [log for log in as_list(data.get("continuity_logs")) if isinstance(log, dict)]
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    scene_ids = [clean_text(log.get("scene_id")) for log in logs if clean_text(log.get("scene_id"))]
    shot_scene_ids = [clean_text(shot.get("scene_id")) for shot in shots if clean_text(shot.get("scene_id"))]
    return {
        "scene_ids": scene_ids,
        "shot_scene_ids": shot_scene_ids,
        "missing_in_logs": sorted(set(shot_scene_ids) - set(scene_ids)),
    }


def as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def clean_text(value: Any) -> str:
    return str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()


def one_line(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def fmt_number(value: Any) -> str:
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:g}"


def numeric_beat_id(beat_id: str) -> int | None:
    match = re.fullmatch(r"B(\d{3})", str(beat_id))
    return int(match.group(1)) if match else None


def format_beat_ids(beat_ids: list[Any]) -> str:
    ids = [str(item) for item in beat_ids if item]
    if not ids:
        return "B000"
    numbers = [numeric_beat_id(item) for item in ids]
    if all(number is not None for number in numbers):
        first = numbers[0]
        expected = list(range(first, first + len(numbers)))
        if numbers == expected:
            if len(ids) == 1:
                return ids[0]
            return f"{ids[0]}-{ids[-1]}"
    return "+".join(ids)


def source_with_beats(shot: dict[str, Any]) -> str:
    return f"{format_beat_ids(as_list(shot.get('beat_ids')))}～{one_line(shot.get('source_paragraph'))}"


def escape_cell(value: Any) -> str:
    return clean_text(value).replace("|", "｜").replace("\n", "<br>")


def parse_triad(camera: str) -> tuple[str, str, str]:
    first = clean_text(camera).split("\n", 1)[0].strip()
    match = re.fullmatch(r"\[([^,\]]+),\s*([^,\]]+),\s*([^\]]+)\]", first)
    if not match:
        return "平视", "中景", "固定镜头"
    return tuple(part.strip() for part in match.groups())


def is_panoramic_shot_size(shot_size: str) -> bool:
    value = clean_text(shot_size)
    return any(keyword in value for keyword in PANORAMIC_SHOT_SIZE_KEYWORDS)


def shot_size_rank(shot_size: str) -> int | None:
    value = clean_text(shot_size)
    for rank, keywords in SHOT_SIZE_RANKS:
        if any(keyword in value for keyword in keywords):
            return rank
    return None


def shot_size_has_span(shot_size: str) -> bool:
    value = clean_text(shot_size)
    return has_any_keyword(value, WIDE_SHOT_SIZE_KEYWORDS) and has_any_keyword(value, CLOSE_SHOT_SIZE_KEYWORDS)


def angle_family(angle: str) -> str:
    value = clean_text(angle)
    if "主观" in value:
        return "subjective"
    if "过肩" in value:
        return "over_shoulder"
    if "仰" in value:
        return "low"
    if "俯" in value or "高角度" in value:
        return "high"
    if "侧" in value:
        return "side"
    return "level"


def is_static_or_soft_movement(movement: str) -> bool:
    return has_any_keyword(movement, STATIC_OR_SOFT_MOVEMENT_KEYWORDS)


def motion_axis_direction(movement: str) -> str | None:
    value = clean_text(movement)
    if has_any_keyword(value, PUSH_IN_MOVEMENT_KEYWORDS):
        return "push_in"
    if has_any_keyword(value, PULL_OUT_MOVEMENT_KEYWORDS):
        return "pull_out"
    return None


def has_continuity_position_change(shot: dict[str, Any]) -> bool:
    if "【站位位移】" in clean_text(shot.get("camera_main_image")):
        return True
    for update in as_list(shot.get("continuity_updates")):
        if isinstance(update, dict) and update.get("field") in {"position", "facing"}:
            return True
    return False


def normalized_visible_characters(shot: dict[str, Any]) -> tuple[str, ...]:
    return tuple(sorted(str(item) for item in as_list(shot.get("visible_characters")) if str(item)))


def fact_types_for_shot(shot: dict[str, Any], fact_types: dict[str, str]) -> set[str]:
    return {fact_types[fact_id] for fact_id in map(str, as_list(shot.get("covered_fact_ids"))) if fact_id in fact_types}


def quoted_dialogue_text(shot: dict[str, Any]) -> str:
    camera = clean_text(shot.get("camera_main_image"))
    quoted = "".join(re.findall(r"[“\"]([^”\"]+)[”\"]", camera))
    if quoted:
        return quoted.strip()
    source = one_line(shot.get("source_paragraph"))
    match = re.search(r"[：:](.+)$", source)
    return match.group(1).strip() if match else ""


def dialogue_speaker(shot: dict[str, Any]) -> str:
    source = one_line(shot.get("source_paragraph"))
    match = re.match(r"([^：:]{1,20})[：:]", source)
    return match.group(1).strip() if match else ""


def is_expression_only_response(shot: dict[str, Any], fact_types: dict[str, str]) -> bool:
    types = fact_types_for_shot(shot, fact_types)
    if types & NEW_FACT_TYPES_FOR_CUT:
        return False
    if not (types & {"action", "emotion", "character"}):
        return False
    text = one_line(shot.get("source_paragraph")) + " " + one_line(shot.get("camera_main_image"))
    return has_any_keyword(text, EXPRESSION_ONLY_KEYWORDS)


def has_any_keyword(value: str, keywords: tuple[str, ...]) -> bool:
    text = clean_text(value)
    return any(keyword in text for keyword in keywords)


def metadata_version(data: dict[str, Any]) -> str:
    return clean_text(as_dict(data.get("metadata")).get("version"))


def requires_hybrid_fields(data: dict[str, Any]) -> bool:
    return metadata_version(data) in HYBRID_VERSIONS


def requires_adjacent_motion_guard(data: dict[str, Any]) -> bool:
    return metadata_version(data) in ADJACENT_MOTION_GUARD_VERSIONS


def requires_structured_audit(data: dict[str, Any]) -> bool:
    return metadata_version(data) in STRUCTURED_AUDIT_VERSIONS


def requires_human_gate(data: dict[str, Any]) -> bool:
    return metadata_version(data) in HUMAN_GATE_VERSIONS


def requires_source_lock(data: dict[str, Any]) -> bool:
    return metadata_version(data) in SOURCE_LOCK_VERSIONS


def current_rule_revision(data: dict[str, Any]) -> str:
    return RULE_REVISIONS.get(metadata_version(data), RULE_REVISION)


def normalized_list(value: Any) -> list[str]:
    return [str(item) for item in as_list(value) if str(item)]


def scene_key_for(data: dict[str, Any], item: dict[str, Any]) -> str:
    if requires_human_gate(data):
        return clean_text(item.get("scene_id"))
    return clean_text(item.get("scene"))


def approved_review(data: dict[str, Any], gate: str, *, note_contains: str | None = None) -> bool:
    for review in as_list(data.get("human_reviews")):
        if not isinstance(review, dict):
            continue
        if clean_text(review.get("gate")) != gate or clean_text(review.get("status")) != "approved":
            continue
        if note_contains and note_contains not in clean_text(review.get("notes")):
            continue
        return True
    return False


def approved_safety_exception(data: dict[str, Any], shot: dict[str, Any]) -> bool:
    notes = clean_text(shot.get("notes"))
    return (
        requires_human_gate(data)
        and clean_text(shot.get("shot_type")) == "safety"
        and "[安全镜]" in notes
        and "[人工批准]" in notes
        and (approved_review(data, "GATE_B", note_contains="安全镜") or approved_review(data, "GATE_B", note_contains="safety"))
    )


def fact_lookup(data: dict[str, Any]) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        for fact in as_list(beat.get("facts")):
            if isinstance(fact, dict) and fact.get("fact_id"):
                lookup[str(fact.get("fact_id"))] = fact
    return lookup


def fact_cut_reasons(fact: dict[str, Any]) -> set[str]:
    return set(normalized_list(fact.get("cut_reasons")))


def is_fact_critical(fact: dict[str, Any]) -> bool:
    return clean_text(fact.get("cut_priority")) == "must_isolate" or bool(fact_cut_reasons(fact) & CRITICAL_SPLIT_REASONS)


def find_required_cut_point_categories(text_value: str) -> list[str]:
    found: list[str] = []
    for name, patterns, count_matches in REQUIRED_CUT_POINT_CATEGORIES:
        if count_matches:
            count = sum(len(pattern.findall(text_value)) for pattern in patterns)
            if count >= 2:
                found.append(name)
            continue
        if any(pattern.search(text_value) for pattern in patterns):
            found.append(name)
    return found


def extract_source_speakers(text_value: str) -> list[str]:
    forbidden_labels = {"时间", "景别", "构图", "运镜手法", "画面内容", "场景", "主体", "画面", "风格", "禁止"}
    speakers: list[str] = []
    for match in re.finditer(r"(?:^|[\n/。；;，,])\s*([\u4e00-\u9fa5A-Za-z0-9·]{1,8})(?:（VO）|\(VO\))?[：:]", clean_text(text_value)):
        speaker = match.group(1).strip()
        if speaker and speaker not in forbidden_labels and speaker not in speakers:
            speakers.append(speaker)
    return speakers


def source_marks_offscreen_voice(text_value: str, speaker: str) -> bool:
    source = clean_text(text_value)
    return bool(re.search(rf"{re.escape(speaker)}\s*(?:（VO）|\(VO\)|VO|画外声)", source))


def shot_fact_objects(shot: dict[str, Any], facts_by_id: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    facts: list[dict[str, Any]] = []
    for fact_id in normalized_list(shot.get("covered_fact_ids")):
        fact = facts_by_id.get(fact_id)
        if fact is not None:
            facts.append(fact)
    return facts


def shot_cut_groups(facts: list[dict[str, Any]], *, must_only: bool) -> dict[str, list[str]]:
    groups: dict[str, list[str]] = {}
    for fact in facts:
        if must_only and clean_text(fact.get("cut_priority")) != "must_isolate":
            continue
        group = clean_text(fact.get("cut_group"))
        if not group:
            continue
        groups.setdefault(group, []).append(str(fact.get("fact_id")))
    return groups


def fact_beat_id(fact: dict[str, Any]) -> str:
    match = re.match(r"(B\d{3})-F\d{2}$", str(fact.get("fact_id", "")))
    return match.group(1) if match else ""


def fact_structural_cut_key(fact: dict[str, Any], *, include_category: bool) -> tuple[str, ...]:
    if include_category:
        return (
            fact_beat_id(fact),
            clean_text(fact.get("cut_category")),
            clean_text(fact.get("cut_moment_id")),
        )
    return (
        fact_beat_id(fact),
        clean_text(fact.get("cut_moment_id")),
    )


def structural_cut_groups(
    facts: list[dict[str, Any]],
    *,
    must_only: bool,
    include_category: bool,
) -> dict[tuple[str, ...], list[str]]:
    groups: dict[tuple[str, ...], list[str]] = {}
    for fact in facts:
        if must_only and clean_text(fact.get("cut_priority")) != "must_isolate":
            continue
        key = fact_structural_cut_key(fact, include_category=include_category)
        if not all(key):
            continue
        groups.setdefault(key, []).append(str(fact.get("fact_id")))
    return groups


def structured_key_label(key: tuple[str, ...]) -> str:
    return "/".join(key)


def strip_internal_prefix(line: str) -> str:
    value = line.strip()
    for label in INTERNAL_LABELS:
        if value.startswith(label):
            return value[len(label):].lstrip(" ：:").strip()
    return value


def build_overcompression_audit(data: dict[str, Any]) -> dict[str, Any]:
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    facts_by_id = fact_lookup(data)
    at_risk_shots: list[dict[str, Any]] = []
    total_duration = 0.0
    short_3 = 0
    multi_beat_critical = 0
    for shot in shots:
        duration = float(shot.get("duration_seconds") or 0)
        total_duration += duration
        short_3 += int(duration <= 3)
        facts = shot_fact_objects(shot, facts_by_id)
        critical_facts = [fact for fact in facts if is_fact_critical(fact)]
        must_groups = shot_cut_groups(facts, must_only=True)
        cut_points = find_required_cut_point_categories(
            clean_text(shot.get("source_paragraph")) + "\n" + clean_text(shot.get("camera_main_image"))
        )
        beat_count = len(as_list(shot.get("beat_ids")))
        has_multi_beat_critical = beat_count >= 2 and bool(critical_facts)
        multi_beat_critical += int(has_multi_beat_critical)
        if critical_facts or len(must_groups) > 1 or cut_points or beat_count >= 3 or duration >= 9:
            at_risk_shots.append(
                {
                    "shot_no": shot.get("shot_no"),
                    "duration_seconds": int(duration) if duration.is_integer() else duration,
                    "beat_count": beat_count,
                    "shot_type": shot.get("shot_type"),
                    "critical_fact_count": len(critical_facts),
                    "must_isolate_groups": sorted(must_groups),
                    "required_cut_points": cut_points,
                    "multi_beat_critical": has_multi_beat_critical,
                }
            )
    cut_per_minute = round(((len(shots) - 1) / total_duration) * 60, 2) if total_duration > 0 and len(shots) > 1 else 0
    short_ratio = round(short_3 / len(shots), 4) if shots else 0
    return {
        "rule_revision": current_rule_revision(data),
        "at_risk_shots": at_risk_shots,
        "multi_beat_critical_shots": multi_beat_critical,
        "cut_per_minute": cut_per_minute,
        "short_shot_ratio": short_ratio,
    }


def prompt_visual_text(camera: str, source: str) -> str:
    lines = clean_text(camera).split("\n")
    body: list[str] = []
    for index, raw in enumerate(lines):
        line = raw.strip()
        if not line:
            continue
        if index == 0 and line.startswith("["):
            continue
        if line.startswith("【机位逻辑】") or line.startswith("【场景首镜站位】"):
            continue
        if line.startswith("【站位位移】"):
            continue
        stripped = strip_internal_prefix(line)
        if stripped:
            body.append(stripped)
    text = "；".join(body) if body else one_line(source)
    return sanitize_prompt_text(text)


def sanitize_prompt_text(text: str) -> str:
    value = clean_text(text)
    for label in PROMPT_FORBIDDEN:
        value = value.replace(label, "")
    value = re.sub(r"\s*；\s*", "；", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip(" ；")


def first_sentence(text: str) -> str:
    value = sanitize_prompt_text(text)
    if not value:
        return "当前镜头可见主体动作。"
    match = re.search(r"[。！？；]", value)
    if match:
        return value[: match.end()]
    return value[:80]


def prompt_composition_text(shot: dict[str, Any]) -> str:
    camera = clean_text(shot.get("camera_main_image"))
    first_position = next(
        (strip_internal_prefix(line) for line in camera.split("\n") if line.strip().startswith("【场景首镜站位】")),
        "",
    )
    if first_position:
        return sanitize_prompt_text(first_position)
    characters = normalized_list(shot.get("visible_characters"))
    props = normalized_list(shot.get("visible_props"))
    parts: list[str] = []
    if characters:
        parts.append("可见主体：" + "、".join(characters))
    if props:
        parts.append("可见道具：" + "、".join(props))
    if parts:
        return "；".join(parts) + "。"
    return "当前镜头可见主体关系。"


def prompt_fields(prompt: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    for line in clean_text(prompt).splitlines():
        for field_name in PROMPT_FIELDS:
            if line.startswith(field_name):
                fields[field_name] = line[len(field_name):].strip()
    return fields


def quotes_are_balanced(text: str) -> bool:
    value = clean_text(text)
    return value.count("“") == value.count("”") and value.count('"') % 2 == 0


def composition_has_field_leak(value: str) -> bool:
    text = clean_text(value)
    return any(pattern.search(text) for pattern in PROMPT_COMPOSITION_FORBIDDEN_PATTERNS)


def fact_terms(text: str) -> list[str]:
    value = re.sub(r"[，。！？；：、“”\"（）()\s]", " ", clean_text(text))
    terms = re.findall(r"[A-Za-z0-9_]+|[\u4e00-\u9fff]{2,}", value)
    return [term for term in terms if term and term not in FACT_TEXT_ALLOWLIST]


def unsupported_fact_terms(fact_text: str, source_text: str) -> list[str]:
    source = clean_text(source_text)
    missing: list[str] = []
    for term in fact_terms(fact_text):
        if term not in source and term not in missing:
            missing.append(term)
    return missing


def derive_prompt(shot: dict[str, Any]) -> str:
    duration = fmt_number(shot.get("duration_seconds", 0))
    _angle, shot_size, movement = parse_triad(clean_text(shot.get("camera_main_image")))
    prompt_shot_size = re.sub(r"\s*(?:->|→)\s*", "→", shot_size)
    visual = prompt_visual_text(
        clean_text(shot.get("camera_main_image")),
        clean_text(shot.get("source_paragraph")),
    )
    composition = prompt_composition_text(shot)
    return "\n".join(
        [
            f"时间：0秒-{duration}秒",
            f"景别：{prompt_shot_size}",
            f"构图：{composition}",
            f"运镜手法：{movement}",
            f"画面内容：{visual}",
        ]
    )


def derive_prompts(data: dict[str, Any]) -> None:
    for shot in as_list(data.get("shots")):
        if isinstance(shot, dict):
            shot["prompt"] = derive_prompt(shot)


def expected_rows(data: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for shot in as_list(data.get("shots")):
        if not isinstance(shot, dict):
            continue
        rows.append(
            [
                shot.get("shot_no", ""),
                shot.get("scene", ""),
                source_with_beats(shot),
                shot.get("duration_seconds", ""),
                clean_text(shot.get("camera_main_image")),
                clean_text(shot.get("notes")),
                clean_text(shot.get("prompt")),
            ]
        )
    return rows


def update_validation_report(data: dict[str, Any], result: ValidationResult) -> None:
    metadata = as_dict(data.get("metadata"))
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    scene_counts: dict[str, int] = {}
    scene_logs = {
        scene_key_for(data, log): log
        for log in as_list(data.get("continuity_logs"))
        if isinstance(log, dict) and scene_key_for(data, log)
    }
    first_shot_anchor_audit: list[dict[str, Any]] = []
    seen_scenes: set[str] = set()
    total_duration = 0.0
    over_6 = over_8 = over_10 = 0
    short_3 = 0
    long_6_or_more = 0
    shot_type_counts: dict[str, int] = {}
    insert_priority_counts: dict[str, int] = {}
    long_take_audit: list[dict[str, Any]] = []
    for shot in shots:
        scene = scene_key_for(data, shot) or str(shot.get("scene", "")).split(" ")[0] or str(shot.get("scene", ""))
        scene_counts[scene] = scene_counts.get(scene, 0) + 1
        full_scene = scene_key_for(data, shot) or str(shot.get("scene", ""))
        if full_scene not in seen_scenes:
            seen_scenes.add(full_scene)
            _angle, shot_size, _movement = parse_triad(clean_text(shot.get("camera_main_image")))
            anchor_type = as_dict(scene_logs.get(full_scene)).get("first_shot_anchor_type", "missing")
            visible_count = len(as_list(shot.get("visible_characters")))
            requires_panorama = anchor_type in PANORAMIC_REQUIRED_ANCHORS or (
                visible_count >= 2 and anchor_type != "single_continuation"
            )
            first_shot_anchor_audit.append(
                {
                    "scene": full_scene,
                    "shot_no": shot.get("shot_no"),
                    "anchor_type": anchor_type,
                    "shot_size": shot_size,
                    "visible_character_count": visible_count,
                    "requires_panoramic_first_shot": requires_panorama,
                    "passed": (not requires_panorama) or is_panoramic_shot_size(shot_size),
                }
            )
        duration = float(shot.get("duration_seconds") or 0)
        total_duration += duration
        over_6 += int(duration > 6)
        over_8 += int(duration > 8)
        over_10 += int(duration > 10)
        short_3 += int(duration <= 3)
        long_6_or_more += int(duration >= 6)
        shot_type = clean_text(shot.get("shot_type")) or "missing"
        insert_priority = clean_text(shot.get("insert_priority")) or "missing"
        shot_type_counts[shot_type] = shot_type_counts.get(shot_type, 0) + 1
        insert_priority_counts[insert_priority] = insert_priority_counts.get(insert_priority, 0) + 1
        if duration > 10:
            long_take_audit.append(
                {
                    "shot_no": shot.get("shot_no"),
                    "duration_seconds": int(duration) if duration.is_integer() else duration,
                    "long_take_support": normalized_list(shot.get("long_take_support")),
                }
            )
    cut_per_minute = round(((len(shots) - 1) / total_duration) * 60, 2) if total_duration > 0 and len(shots) > 1 else 0
    short_ratio = round(short_3 / len(shots), 4) if shots else 0
    long_ratio = round(long_6_or_more / len(shots), 4) if shots else 0
    hybrid_audit = {
        "short_shots_le_3_seconds": short_3,
        "short_shot_ratio": short_ratio,
        "long_shots_ge_6_seconds": long_6_or_more,
        "long_shot_ratio": long_ratio,
        "cut_per_minute": cut_per_minute,
        "shot_type_counts": shot_type_counts,
        "insert_priority_counts": insert_priority_counts,
        "long_take_audit": long_take_audit,
        "soft_targets": {
            "short_shot_ratio": "0.10-0.18 for full-episode references only",
            "long_shot_ratio": "0.40-0.55 for full-episode references only",
            "cut_per_minute": "10.5-12.5 for full-episode references only",
        },
    }
    data["validation_report"] = {
        "status": result.status,
        "warnings": result.warnings,
        "warning_items": warning_items(result.warnings),
        "errors": result.errors,
        "version": clean_text(metadata.get("version")),
        "rule_revision": clean_text(metadata.get("rule_revision")),
        "source_json_hash": canonical_data_hash(data),
        "reference_status": as_dict(metadata.get("reference_status")),
        "warning_count": len(result.warnings),
        "error_count": len(result.errors),
        "warning_budget": {
            "status": "WARN is deliverable only when every warning is explained or accepted by the user",
            "unresolved_warning_count": len(result.warnings),
        },
        "shot_count": len(shots),
        "scene_counts": scene_counts,
        "total_duration_seconds": int(total_duration) if total_duration.is_integer() else total_duration,
        "average_duration_seconds": round(total_duration / len(shots), 2) if shots else 0,
        "shots_over_6_seconds": over_6,
        "shots_over_8_seconds": over_8,
        "shots_over_10_seconds": over_10,
        "column_contract": "7-column stable storyboard table; keyframe column removed",
        "first_shot_anchor_audit": first_shot_anchor_audit,
        "hybrid_audit": hybrid_audit,
        "overcompression_audit": build_overcompression_audit(data),
    }
    if requires_human_gate(data):
        data["validation_report"]["human_gate_audit"] = human_gate_audit(data)
        data["validation_report"]["warn_resolution_audit"] = warn_resolution_audit(data, result.warnings)
        data["validation_report"]["scene_id_audit"] = scene_id_audit(data)


def validate_metadata(
    data: dict[str, Any],
    result: ValidationResult,
    *,
    strict_status: bool,
    final_signoff: bool,
) -> None:
    metadata = as_dict(data.get("metadata"))
    if metadata.get("skill_name") != "su-fenjingskill-zh":
        result.error("metadata.skill_name 必须为 su-fenjingskill-zh。")
    version = metadata_version(data)
    expected_rule_revision = RULE_REVISIONS.get(version)
    if expected_rule_revision and metadata.get("rule_revision") != expected_rule_revision:
        result.error(f"metadata.rule_revision 必须为 {expected_rule_revision}。")
    reference_status = as_dict(metadata.get("reference_status"))
    if "reference_status" not in metadata:
        if requires_human_gate(data):
            result.error("metadata.reference_status 缺失。")
        else:
            result.warn("metadata.reference_status 缺失，建议记录 reference loaded/missing 状态。")
    if requires_human_gate(data):
        for key in REQUIRED_REFERENCE_KEYS:
            status = clean_text(reference_status.get(key))
            if status not in {"loaded", "missing"}:
                result.error(f"metadata.reference_status.{key} 必须为 loaded / missing。")
        reference_proof = as_dict(metadata.get("reference_proof"))
        for key, status in reference_status.items():
            if status == "loaded" and not clean_text(reference_proof.get(key)):
                result.error(f"metadata.reference_proof.{key} 必须记录已读文件首个标题行。")
        script_status = as_dict(metadata.get("script_status"))
        for key in REQUIRED_SCRIPT_KEYS:
            status = clean_text(script_status.get(key))
            if status not in {"available", "missing"}:
                result.error(f"metadata.script_status.{key} 必须为 available / missing。")
        if not isinstance(metadata.get("revision_log"), list):
            result.error("metadata.revision_log 必须是数组。")
        title = clean_text(metadata.get("title"))
        if title and not re.fullmatch(r"[\w\u4e00-\u9fff-]+", title):
            result.error("metadata.title 只能包含中英文、数字、下划线和连字符。")
    report = as_dict(data.get("validation_report"))
    status = report.get("status")
    if strict_status and status is not None and status not in VALID_STATUSES:
        result.error("validation_report.status 只允许 PASS / WARN / FAIL / NOT_RUN。")
    if requires_human_gate(data) and clean_text(status) == "NOT_RUN":
        if clean_text(report.get("source_json_hash")):
            result.error("validation_report.status 为 NOT_RUN 时 source_json_hash 必须为空字符串。")
        if not approved_review(data, "GATE_C", note_contains="accepted_without_validation"):
            result.error("NOT_RUN 交付必须在 Gate C 记录 accepted_without_validation。")
    if strict_status and requires_structured_audit(data) and report.get("source_json_hash") and clean_text(status) != "NOT_RUN":
        expected_hash = canonical_data_hash(data)
        if report.get("source_json_hash") != expected_hash:
            result.error("validation_report.source_json_hash 与当前 shot_data 内容不一致。")


def validate_human_reviews(data: dict[str, Any], result: ValidationResult, *, final_signoff: bool) -> None:
    if not requires_human_gate(data):
        return
    reviews = as_list(data.get("human_reviews"))
    if not isinstance(data.get("human_reviews"), list):
        result.error("2.4.0 数据必须包含 human_reviews 数组。")
        reviews = []
    for review in reviews:
        if not isinstance(review, dict):
            result.error("human_reviews 每项必须是对象。")
            continue
        if clean_text(review.get("gate")) not in GATES:
            result.error("human_reviews.gate 必须为 GATE_0 / GATE_A / GATE_B / GATE_C。")
        if clean_text(review.get("status")) not in {"approved", "rejected"}:
            result.error("human_reviews.status 必须为 approved / rejected。")
        if not isinstance(review.get("round"), int) or int(review.get("round")) < 1:
            result.error("human_reviews.round 必须为从 1 开始的整数。")
        if not clean_text(review.get("reviewer")):
            result.error("human_reviews.reviewer 不能为空。")
    audit = human_gate_audit(data, final_signoff=final_signoff)
    for gate in audit["missing"]:
        result.error(f"2.4.0 缺少 {gate} approved 人工审核记录。")


def script_fragment_fingerprint(value: str) -> str:
    return re.sub(r"\s+", "", clean_text(value))


def script_text_hash(value: str) -> str:
    return hashlib.sha256(script_fragment_fingerprint(value).encode("utf-8")).hexdigest()


def raw_script_text(value: Any) -> str:
    return str(value or "").replace("\r\n", "\n").replace("\r", "\n")


def normalize_source_spans(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, dict):
        return [value]
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    return []


def extract_script_span_text(
    locked_text: str,
    spans: list[dict[str, Any]],
    result: ValidationResult,
    label: str,
) -> str | None:
    parts: list[str] = []
    for index, span in enumerate(spans, start=1):
        try:
            start = int(span.get("start"))
            end = int(span.get("end"))
        except (TypeError, ValueError):
            result.error(f"{label} source_span #{index} must contain integer start/end.")
            return None
        if start < 0 or end <= start or end > len(locked_text):
            result.error(f"{label} source_span #{index} is outside script_lock.locked_text.")
            return None
        extracted = locked_text[start:end]
        expected_hash = clean_text(span.get("text_hash"))
        if expected_hash and expected_hash != script_text_hash(extracted):
            result.error(f"{label} source_span #{index} text_hash does not match locked text.")
        parts.append(extracted)
    return "\n".join(parts)


def validate_source_text_against_lock(
    locked_text: str,
    item: dict[str, Any],
    text_key: str,
    result: ValidationResult,
    label: str,
) -> None:
    spans = normalize_source_spans(item.get("source_span")) or normalize_source_spans(item.get("source_spans"))
    if not spans:
        result.error(f"{label} must include source_span or source_spans for script-lock validation.")
        return
    extracted = extract_script_span_text(locked_text, spans, result, label)
    if extracted is None:
        return
    declared = clean_text(item.get(text_key))
    if not declared:
        result.error(f"{label} {text_key} must not be empty.")
        return
    if script_fragment_fingerprint(declared) != script_fragment_fingerprint(extracted):
        result.error(f"{label} {text_key} is not an exact extract from script_lock.locked_text.")


def validate_script_lock(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_source_lock(data):
        return
    lock = as_dict(data.get("script_lock"))
    if not lock:
        result.error("2.4.1+ data must include top-level script_lock.")
        return
    if clean_text(lock.get("status")) != "locked":
        result.error('script_lock.status must be "locked".')
    locked_text = raw_script_text(lock.get("locked_text"))
    if not locked_text:
        result.error("script_lock.locked_text must contain the full human-approved script text.")
        return
    expected_hash = script_text_hash(locked_text)
    supplied_hash = clean_text(lock.get("locked_text_hash"))
    if supplied_hash != expected_hash:
        result.error("script_lock.locked_text_hash does not match locked_text.")
    if not isinstance(lock.get("approved_corrections", []), list):
        result.error("script_lock.approved_corrections must be an array when present.")
    if metadata_version(data) in APPROVED_SCRIPT_PATH_VERSIONS and not clean_text(lock.get("approved_script_path")):
        result.error("2.4.2 data must include script_lock.approved_script_path.")
    for beat in as_list(data.get("beats")):
        if isinstance(beat, dict):
            label = clean_text(beat.get("beat_id")) or "beat"
            validate_source_text_against_lock(locked_text, beat, "source_text", result, label)
    for shot in as_list(data.get("shots")):
        if isinstance(shot, dict):
            label = f"shot {shot.get('shot_no')}"
            validate_source_text_against_lock(locked_text, shot, "source_paragraph", result, label)


def is_auto_whitelist_warning(message: str) -> bool:
    return any(token in message for token in ("reference missing", "[reference missing]", "[合理补足]", "节奏", "cut/min", "短镜比例"))


def validate_warn_resolutions(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_human_gate(data):
        return
    resolutions = as_list(data.get("warn_resolutions"))
    if not isinstance(data.get("warn_resolutions"), list):
        result.error("2.4.0 数据必须包含 warn_resolutions 数组。")
        resolutions = []
    by_id: dict[str, dict[str, Any]] = {}
    for item in resolutions:
        if not isinstance(item, dict):
            result.error("warn_resolutions 每项必须是对象。")
            continue
        warn_id_value = clean_text(item.get("warn_id"))
        if not warn_id_value:
            result.error("warn_resolutions.warn_id 不能为空。")
            continue
        by_id[warn_id_value] = item
        if clean_text(item.get("resolution")) not in {"keep", "revise", "accepted_without_change"}:
            result.error(f"{warn_id_value} resolution 必须为 keep / revise / accepted_without_change。")
        if clean_text(item.get("resolved_by")) not in {"human", "auto_whitelist"}:
            result.error(f"{warn_id_value} resolved_by 必须为 human / auto_whitelist。")
        if not clean_text(item.get("note")):
            result.error(f"{warn_id_value} note 不能为空。")
    for message in result.warnings:
        warn_id_value = warning_id(message)
        item = by_id.get(warn_id_value)
        if item is None:
            result.error(f"WARN 缺少处置记录：{warn_id_value}。")
            continue
        resolved_by = clean_text(item.get("resolved_by"))
        if resolved_by == "auto_whitelist" and not is_auto_whitelist_warning(message):
            result.error(f"{warn_id_value} 不是白名单 WARN，必须由 human 处置。")


def validate_continuity_logs(data: dict[str, Any], result: ValidationResult) -> dict[str, dict[str, Any]]:
    logs_by_scene: dict[str, dict[str, Any]] = {}
    logs: list[dict[str, Any]] = []
    use_scene_id = requires_human_gate(data)
    for log in as_list(data.get("continuity_logs")):
        if not isinstance(log, dict):
            result.error("continuity_logs 中每一项必须是对象。")
            continue
        scene = scene_key_for(data, log)
        if not scene:
            result.error("continuity_logs 每项必须包含 scene_id。" if use_scene_id else "continuity_logs 每项必须包含 scene。")
            continue
        if scene in logs_by_scene:
            result.error(f"continuity_logs 场景键重复：{scene}")
        logs_by_scene[scene] = log
        logs.append(log)
        if use_scene_id and not clean_text(log.get("scene")):
            result.error(f"{scene} continuity_logs 必须包含展示用 scene。")
        anchor_type = log.get("first_shot_anchor_type")
        if anchor_type not in ANCHOR_TYPES:
            result.error(
                f"{scene} first_shot_anchor_type 必须为 "
                "space / multi_character / both / single_continuation / subjective。"
            )
    if requires_structured_audit(data):
        for log in logs:
            scene = scene_key_for(data, log)
            parent_scene = clean_text(log.get("inherits_from"))
            if not parent_scene:
                continue
            parent = logs_by_scene.get(parent_scene)
            if parent is None:
                result.error(f"{scene} inherits_from 指向不存在的场景：{parent_scene}")
                continue
            inherited = set(normalized_list(log.get("inherited_states")))
            diverged = set(normalized_list(log.get("diverged_states")))
            overlap = inherited & diverged
            if overlap:
                result.error(f"{scene} inherited_states 与 diverged_states 重叠：{', '.join(sorted(overlap))}")
            for field_name in inherited:
                if field_name not in parent or field_name not in log:
                    result.error(f"{scene} 继承字段 {field_name} 必须同时存在于父场景和当前场景。")
                    continue
                if log.get(field_name) != parent.get(field_name):
                    result.error(f"{scene} 继承字段 {field_name} 与 {parent_scene} 不一致。")
    return logs_by_scene


def state_key(entity_type: Any, entity: Any, field_name: Any) -> tuple[str, str, str]:
    return clean_text(entity_type), clean_text(entity), clean_text(field_name)


def register_entity_state(states: dict[tuple[str, str, str], str], entity_type: str, entity: str, state: Any) -> None:
    if not entity:
        return
    if isinstance(state, dict):
        for field_name in STATEFUL_CONTINUITY_FIELDS:
            if field_name in state:
                states[state_key(entity_type, entity, field_name)] = clean_text(state.get(field_name))
        return
    if isinstance(state, str):
        text = clean_text(state)
        for field_name in STATEFUL_CONTINUITY_FIELDS:
            match = re.search(rf"{field_name}[：:]\s*([^，。；;]+)", text)
            if match:
                states[state_key(entity_type, entity, field_name)] = match.group(1).strip()


def initial_scene_states(scene_log: dict[str, Any]) -> dict[tuple[str, str, str], str]:
    states: dict[tuple[str, str, str], str] = {}
    for field_name, entity_type in [("characters", "character"), ("props", "prop"), ("fixed_objects", "fixed_object")]:
        collection = scene_log.get(field_name)
        if isinstance(collection, dict):
            for entity, state in collection.items():
                register_entity_state(states, entity_type, str(entity), state)
        for item in as_list(collection):
            if isinstance(item, dict):
                entity = clean_text(item.get("name") or item.get("entity") or item.get("character") or item.get("prop"))
                register_entity_state(states, entity_type, entity, item)
    reality = clean_text(scene_log.get("reality_layer"))
    if reality:
        states[state_key("reality_layer", "", "value")] = reality
    return states


def collect_facts(data: dict[str, Any], result: ValidationResult) -> tuple[dict[str, str], dict[str, str], set[str]]:
    beat_to_facts: dict[str, set[str]] = {}
    fact_to_beat: dict[str, str] = {}
    fact_types: dict[str, str] = {}
    seen_beats: set[str] = set()
    previous_beat_number: int | None = None
    scene_ids = {
        scene_key_for(data, log)
        for log in as_list(data.get("continuity_logs"))
        if isinstance(log, dict) and scene_key_for(data, log)
    }
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            result.error("beats 中每一项必须是对象。")
            continue
        beat_id = str(beat.get("beat_id", ""))
        beat_number = numeric_beat_id(beat_id)
        if beat_number is None:
            result.error(f"Beat ID 不合法：{beat_id}")
        elif requires_human_gate(data):
            if previous_beat_number is not None and beat_number <= previous_beat_number:
                result.error(f"2.4.0 Beat ID 必须唯一单调递增且禁止重编号：{beat_id}")
            previous_beat_number = beat_number
        if beat_id in seen_beats:
            result.error(f"Beat ID 重复：{beat_id}")
        seen_beats.add(beat_id)
        if requires_human_gate(data):
            scene_id = clean_text(beat.get("scene_id"))
            if not scene_id:
                result.error(f"{beat_id} 缺少 scene_id。")
            elif scene_ids and scene_id not in scene_ids:
                result.error(f"{beat_id} scene_id 未在 continuity_logs 登记：{scene_id}")
        facts = as_list(beat.get("facts"))
        if not facts:
            result.warn(f"{beat_id} 没有 facts。")
        beat_to_facts.setdefault(beat_id, set())
        for fact in facts:
            if not isinstance(fact, dict):
                result.error(f"{beat_id} facts 中存在非对象项。")
                continue
            fact_id = str(fact.get("fact_id", ""))
            if not re.fullmatch(rf"{beat_id}-F\d{{2}}", fact_id):
                result.error(f"事实 ID 必须绑定所属 Beat：{fact_id}")
            if fact.get("type") not in FACT_TYPES:
                result.error(f"{fact_id} 事实类型不合法：{fact.get('type')}")
            fact_to_beat[fact_id] = beat_id
            fact_types[fact_id] = str(fact.get("type", ""))
            beat_to_facts[beat_id].add(fact_id)
    return fact_to_beat, fact_types, set(fact_to_beat)


def validate_first_scene_shot(shot: dict[str, Any], scene_log: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    camera = clean_text(shot.get("camera_main_image"))
    _angle, shot_size, movement = parse_triad(camera)
    anchor_type = scene_log.get("first_shot_anchor_type")
    visible_count = len(as_list(shot.get("visible_characters")))
    has_first_position = "【场景首镜站位】" in camera
    if not has_first_position:
        result.error(f"镜号{shot_no} 是场景首镜，必须包含【场景首镜站位】。")
    requires_panorama = anchor_type in PANORAMIC_REQUIRED_ANCHORS or (
        visible_count >= 2 and anchor_type != "single_continuation"
    )
    if requires_panorama and not is_panoramic_shot_size(shot_size):
        result.error(
            f"镜号{shot_no} 为 {anchor_type or '未标注'} 锚定场景首镜，"
            f"且可见人物数 {visible_count}，景别必须为大远景/大全景/全景/中全景；当前为 {shot_size}。"
        )
    if anchor_type in {"space", "both"} and "斯坦尼康" in clean_text(movement):
        result.error(f"镜号{shot_no} 为 {anchor_type} 锚定场景首镜，不得使用斯坦尼康建立大空间。")
    if visible_count >= 2 and anchor_type != "single_continuation":
        first_position_line = next(
            (line for line in camera.split("\n") if line.strip().startswith("【场景首镜站位】")),
            "",
        )
        blocking_text = first_position_line + "\n" + clean_text(scene_log.get("spatial_axis"))
        if not any(keyword in blocking_text for keyword in ["左", "右", "轴线", "对视", "朝向", "面向"]):
            result.error(f"镜号{shot_no} 多人场景首镜站位必须写清左右关系、朝向或主要对视轴线。")


def validate_adjacent_shot_design(
    shots: list[Any],
    fact_types: dict[str, str],
    result: ValidationResult,
    *,
    guard_reverse_axis_motion: bool,
    allow_marked_exceptions: bool,
) -> None:
    previous: dict[str, Any] | None = None
    for item in shots:
        if not isinstance(item, dict):
            continue
        item_scene = clean_text(item.get("scene_id")) or clean_text(item.get("scene"))
        previous_scene = clean_text(previous.get("scene_id")) or clean_text(previous.get("scene")) if previous is not None else ""
        if previous is not None and item_scene == previous_scene:
            validate_adjacent_pair(
                previous,
                item,
                fact_types,
                result,
                guard_reverse_axis_motion=guard_reverse_axis_motion,
                allow_marked_exceptions=allow_marked_exceptions,
            )
        previous = item


def validate_adjacent_pair(
    prev: dict[str, Any],
    curr: dict[str, Any],
    fact_types: dict[str, str],
    result: ValidationResult,
    *,
    guard_reverse_axis_motion: bool,
    allow_marked_exceptions: bool,
) -> None:
    prev_no = prev.get("shot_no")
    curr_no = curr.get("shot_no")
    prev_angle, prev_size, prev_movement = parse_triad(clean_text(prev.get("camera_main_image")))
    curr_angle, curr_size, curr_movement = parse_triad(clean_text(curr.get("camera_main_image")))
    if guard_reverse_axis_motion:
        prev_axis = motion_axis_direction(prev_movement)
        curr_axis = motion_axis_direction(curr_movement)
        if {prev_axis, curr_axis} == {"push_in", "pull_out"}:
            if allow_marked_exceptions and "[反转动机]" in clean_text(curr.get("notes")):
                result.warn(
                    f"镜号{prev_no}-{curr_no} 同场相邻镜头使用 [反转动机] 保留轴向反转，需 Gate B 复核。"
                )
            else:
                result.error(
                    f"镜号{prev_no}-{curr_no} 同场相邻镜头出现无动机轴向反转："
                    f"{prev_movement} 后接 {curr_movement}。应改为固定、横移、环绕、垂直摇移、"
                    "切换主体或切换景别。"
                )
    prev_rank = shot_size_rank(prev_size)
    curr_rank = shot_size_rank(curr_size)
    same_characters = normalized_visible_characters(prev) == normalized_visible_characters(curr)
    if same_characters and normalized_visible_characters(prev):
        no_position_change = not has_continuity_position_change(prev) and not has_continuity_position_change(curr)
        close_size = prev_rank is not None and curr_rank is not None and abs(prev_rank - curr_rank) <= 1
        same_angle = angle_family(prev_angle) == angle_family(curr_angle)
        soft_movement = is_static_or_soft_movement(prev_movement) and is_static_or_soft_movement(curr_movement)
        if no_position_change and close_size and same_angle and soft_movement:
            if allow_marked_exceptions and ("[必拆相邻]" in clean_text(prev.get("notes")) or "[必拆相邻]" in clean_text(curr.get("notes"))):
                result.warn(f"镜号{prev_no}-{curr_no} 使用 [必拆相邻] 豁免低增量相邻镜组，需 Gate B 复核。")
            else:
                result.error(
                    f"镜号{prev_no}-{curr_no} 同场相邻镜头主体、视角、景别和运动过近，"
                    "且无站位/朝向迁移；必须合并或提供明确画面信息增量。"
                )

    prev_dialogue = quoted_dialogue_text(prev)
    curr_dialogue = quoted_dialogue_text(curr)
    if prev_dialogue and curr_dialogue and prev_dialogue.endswith(INCOMPLETE_DIALOGUE_ENDINGS):
        prev_speaker = dialogue_speaker(prev)
        curr_speaker = dialogue_speaker(curr)
        if not prev_speaker or not curr_speaker or prev_speaker == curr_speaker:
            result.error(
                f"镜号{prev_no}-{curr_no} 将同一说话人的未完成台词切成相邻两镜；"
                "应合并为一镜，用表演或机位运动承接。"
            )

    prev_types = fact_types_for_shot(prev, fact_types)
    curr_types = fact_types_for_shot(curr, fact_types)
    if "dialogue" in prev_types and is_expression_only_response(curr, fact_types):
        if same_characters or not normalized_visible_characters(curr):
            result.error(
                f"镜号{prev_no}-{curr_no} 是对白后同一人物短表情反应，且没有新的空间、道具、声音或位置事实；"
                "应合并进同一镜。"
            )
    try:
        curr_duration = float(curr.get("duration_seconds") or 0)
    except Exception:
        curr_duration = 0
    if (
        clean_text(curr.get("shot_type")) == "reaction"
        and clean_text(curr.get("insert_priority")) == "none"
        and curr_duration <= 3
        and not (curr_types & NEW_FACT_TYPES_FOR_CUT)
        and not has_continuity_position_change(prev)
        and not has_continuity_position_change(curr)
    ):
        result.warn(f"[可合并] 镜号{curr_no} 是无新增空间/道具/声音/位置/现实层事实的短 reaction，建议与相邻镜头合并。")


def validate_hybrid_fields(shot: dict[str, Any], result: ValidationResult, *, required: bool) -> None:
    shot_no = shot.get("shot_no")
    duration = float(shot.get("duration_seconds") or 0)
    if not required:
        return
    shot_type = clean_text(shot.get("shot_type"))
    if not shot_type:
        result.error(f"镜号{shot_no} 2.3.4+ 必须填写 shot_type。")
    elif shot_type not in SHOT_TYPES:
        result.error(f"镜号{shot_no} shot_type 不合法：{shot_type}")
    split_reasons = normalized_list(shot.get("split_reason"))
    if not split_reasons:
        result.error(f"镜号{shot_no} 2.3.4+ 必须填写 split_reason。")
    for reason in split_reasons:
        if reason not in SPLIT_REASONS:
            result.error(f"镜号{shot_no} split_reason 不合法：{reason}")
    insert_priority = clean_text(shot.get("insert_priority"))
    if not insert_priority:
        result.error(f"镜号{shot_no} 2.3.4+ 必须填写 insert_priority。")
    elif insert_priority not in INSERT_PRIORITIES:
        result.error(f"镜号{shot_no} insert_priority 不合法：{insert_priority}")
    supports = normalized_list(shot.get("long_take_support"))
    for support in supports:
        if support not in LONG_TAKE_SUPPORTS:
            result.error(f"镜号{shot_no} long_take_support 不合法：{support}")
    if duration > 10 and len(supports) < 2:
        result.error(f"镜号{shot_no} 超过10秒，2.3.4 要求至少填写两项 long_take_support。")
    if shot_type in INSERT_LIKE_TYPES and duration > 5:
        result.warn(f"镜号{shot_no} 是 {shot_type}，插镜建议 2-5 秒；当前 {duration:g} 秒，请确认是否承担镜内推进。")
    causal_reasons = {"new_vfx_state", "causal_reveal", "prop_state_change"}
    if causal_reasons & set(split_reasons) and shot_type not in {"insert", "vfx_anchor", "transition", "action"}:
        result.warn(f"镜号{shot_no} 承担真相/道具/VFX 因果落点，但 shot_type 为 {shot_type}，建议复核插镜意识。")


def validate_fact_cut_contract(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_hybrid_fields(data):
        return
    seen_groups: set[str] = set()
    structured_required = requires_structured_audit(data)
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        beat_id = str(beat.get("beat_id", ""))
        for fact in as_list(beat.get("facts")):
            if not isinstance(fact, dict):
                continue
            fact_id = str(fact.get("fact_id", ""))
            priority = clean_text(fact.get("cut_priority"))
            if priority not in CUT_PRIORITIES:
                result.error(f"{fact_id} cut_priority 必须为 normal / recommended / must_isolate。")
            if "cut_reasons" not in fact:
                result.error(f"{fact_id} 缺少 cut_reasons。")
                reasons: list[str] = []
            else:
                reasons = normalized_list(fact.get("cut_reasons"))
            for reason in reasons:
                if reason not in SPLIT_REASONS:
                    result.error(f"{fact_id} cut_reasons 不合法：{reason}")
            if priority in {"recommended", "must_isolate"} and not reasons:
                result.error(f"{fact_id} cut_priority 为 {priority} 时 cut_reasons 不能为空。")
            group = clean_text(fact.get("cut_group"))
            if not group:
                result.error(f"{fact_id} 缺少 cut_group。")
            else:
                seen_groups.add(group)
            if structured_required:
                category = clean_text(fact.get("cut_category"))
                moment_id = clean_text(fact.get("cut_moment_id"))
                if category not in CUT_CATEGORIES:
                    result.error(f"{fact_id} cut_category 必须为 {' / '.join(sorted(CUT_CATEGORIES))}。")
                if not moment_id:
                    result.error(f"{fact_id} 缺少 cut_moment_id。")
            reason_set = set(reasons)
            if reason_set & MUST_ISOLATE_REASONS and priority != "must_isolate":
                result.error(f"{fact_id} 包含关键切点 {sorted(reason_set & MUST_ISOLATE_REASONS)}，cut_priority 必须为 must_isolate。")
            if reason_set & CRITICAL_SPLIT_REASONS and priority == "normal":
                result.error(f"{fact_id} 包含关键切点 {sorted(reason_set & CRITICAL_SPLIT_REASONS)}，cut_priority 不得为 normal。")
            if group and not group.startswith(beat_id):
                result.warn(f"{fact_id} cut_group 建议以所属 Beat 开头，当前为 {group}。")
    if not seen_groups:
        result.error("2.3.4 数据必须登记事实级 cut_group。")


def validate_fact_text_preservation(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_structured_audit(data):
        return
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        source = clean_text(beat.get("source_text"))
        if not source:
            continue
        for fact in as_list(beat.get("facts")):
            if not isinstance(fact, dict):
                continue
            missing = unsupported_fact_terms(clean_text(fact.get("text")), source)
            if missing:
                result.warn(
                    f"{fact.get('fact_id')} Fact 文本疑似引入 source_text 未出现词组："
                    f"{', '.join(missing[:5])}。请确认不是编造事实。"
                )


def validate_must_isolate_density(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_structured_audit(data):
        return
    scene_stats: dict[str, dict[str, Any]] = {}
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        scene = clean_text(beat.get("scene"))
        stats = scene_stats.setdefault(scene, {"beats": 0, "must_beats": 0, "facts": []})
        stats["beats"] += 1
        beat_must = False
        for fact in as_list(beat.get("facts")):
            if isinstance(fact, dict) and clean_text(fact.get("cut_priority")) == "must_isolate":
                beat_must = True
                stats["facts"].append(str(fact.get("fact_id")))
        if beat_must:
            stats["must_beats"] += 1
    for scene, stats in scene_stats.items():
        beats = int(stats["beats"])
        if beats <= 0:
            continue
        ratio = int(stats["must_beats"]) / beats
        if ratio > MUST_ISOLATE_DENSITY_THRESHOLD:
            result.warn(
                f"场景 {scene} must_isolate Beat 占比 {ratio:.0%}（{stats['must_beats']}/{beats} Beats），"
                f"可能存在过度标记；请复核 {', '.join(stats['facts'][:8])} 是否应降级为 recommended。"
            )


def validate_overcompression(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_hybrid_fields(data):
        return
    structured_required = requires_structured_audit(data)
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    facts_by_id = fact_lookup(data)
    total_duration = 0.0
    short_3 = 0
    has_multi_beat_critical = False
    for shot in shots:
        shot_no = shot.get("shot_no")
        duration = float(shot.get("duration_seconds") or 0)
        total_duration += duration
        short_3 += int(duration <= 3)
        notes = clean_text(shot.get("notes"))
        shot_type = clean_text(shot.get("shot_type"))
        beat_count = len(as_list(shot.get("beat_ids")))
        facts = shot_fact_objects(shot, facts_by_id)
        critical_facts = [fact for fact in facts if is_fact_critical(fact)]
        must_groups = shot_cut_groups(facts, must_only=True)
        must_facts = [fact for fact in facts if clean_text(fact.get("cut_priority")) == "must_isolate"]
        structural_groups = structural_cut_groups(
            must_facts,
            must_only=False,
            include_category=not requires_human_gate(data),
        )
        cut_points = find_required_cut_point_categories(
            clean_text(shot.get("source_paragraph")) + "\n" + clean_text(shot.get("camera_main_image"))
        )
        if beat_count >= 3 and critical_facts:
            has_multi_beat_critical = True
            result.error(f"镜号{shot_no} 覆盖 {beat_count} 个 Beat 且包含关键切点，属于过度压缩。")
        if beat_count >= 2 and critical_facts:
            has_multi_beat_critical = True
        if structured_required:
            if len(structural_groups) > 1:
                result.error(
                    f"镜号{shot_no} 同时覆盖多个 must_isolate 结构切点："
                    f"{', '.join(structured_key_label(key) for key in sorted(structural_groups))}；"
                    "必须拆镜，不能只靠 cut_group 或 cut_category 合并。"
                )
            if len(must_facts) > 1 and len(structural_groups) == 1 and "[不可拆说明]" not in notes:
                result.error(f"镜号{shot_no} 合并同一结构切点的多个 must_isolate 事实时，备注必须写 [不可拆说明]。")
        elif len(must_groups) > 1:
            result.error(
                f"镜号{shot_no} 同时覆盖多个 must_isolate cut_group：{', '.join(sorted(must_groups))}；"
                "除同一不可拆瞬间外必须拆镜。"
            )
        if shot_type == "master" and critical_facts:
            result.error(f"镜号{shot_no} master 镜不得承载关键切点，应拆为 transition / vfx_anchor / insert / action。")
        if shot_type == "dialogue" and critical_facts:
            if len(must_groups) > 1 or duration >= 8 or len(cut_points) >= 2:
                result.error(f"镜号{shot_no} dialogue 镜承载多个/过长关键切点，必须拆分真相、声源或反应落点。")
        if shot_type in INSERT_LIKE_TYPES and duration > 5:
            supports = normalized_list(shot.get("long_take_support"))
            if "[保留理由]" not in notes or len(supports) < 2:
                result.error(f"镜号{shot_no} {shot_type} 超过5秒，必须写 [保留理由] 且至少两项 long_take_support。")
        if 9 <= duration <= 11:
            if "[长镜头]" not in notes or "[保留理由]" not in notes:
                result.error(f"镜号{shot_no} 9-11秒高风险长镜头必须写 [长镜头] 和 [保留理由]。")
        if duration >= 12:
            missing = [label for label in ["[长镜头]", "[保留理由]", "[不可拆说明]"] if label not in notes]
            if missing:
                result.error(f"镜号{shot_no} 12秒及以上镜头缺少 {'、'.join(missing)}。")
        if duration >= 9 and len(cut_points) >= 2:
            result.error(f"镜号{shot_no} {duration:g}秒镜头包含多个必拆切点：{', '.join(cut_points)}。")
        source = clean_text(shot.get("source_paragraph"))
        camera = clean_text(shot.get("camera_main_image"))
        visible = set(normalized_list(shot.get("visible_characters")))
        for speaker in extract_source_speakers(source):
            if speaker not in visible and re.search(rf"{re.escape(speaker)}[^。；;\n]{{0,8}}画外声", camera):
                if not source_marks_offscreen_voice(source, speaker):
                    result.error(f"镜号{shot_no} 将原文现场对白角色 {speaker} 改为画外声，疑似绕过多人关系镜头。")
    cut_per_minute = round(((len(shots) - 1) / total_duration) * 60, 2) if total_duration > 0 and len(shots) > 1 else 0
    short_ratio = round(short_3 / len(shots), 4) if shots else 0
    if cut_per_minute and cut_per_minute < 10.5:
        message = f"cut/min {cut_per_minute} 低于 10.5，疑似整体压缩。"
        if has_multi_beat_critical:
            result.error(message + " 且存在多 Beat 关键镜头，必须重拆。")
        else:
            result.warn(message)
    if shots and short_ratio < 0.10:
        message = f"短镜比例 {short_ratio:.2%} 低于 10%，疑似缺少必要插镜。"
        if has_multi_beat_critical:
            result.error(message + " 且存在多 Beat 关键镜头，必须重拆。")
        else:
            result.warn(message)


def validate_rows(
    data: dict[str, Any],
    result: ValidationResult,
    fact_to_beat: dict[str, str],
    fact_types: dict[str, str],
    all_fact_ids: set[str],
    continuity_logs: dict[str, dict[str, Any]],
) -> None:
    shots = as_list(data.get("shots"))
    if not shots:
        result.error("shots 不能为空。")
        return
    expected_no = 1
    covered: set[str] = set()
    scenes = set(continuity_logs)
    seen_scenes: set[str] = set()
    has_update = False
    hybrid_required = requires_hybrid_fields(data)
    structured_required = requires_structured_audit(data)
    use_scene_id = requires_human_gate(data)
    scene_states = {
        scene: initial_scene_states(log)
        for scene, log in continuity_logs.items()
    } if structured_required else {}
    for shot in shots:
        if not isinstance(shot, dict):
            result.error("shots 中每一项必须是对象。")
            continue
        shot_no = shot.get("shot_no")
        if shot_no != expected_no:
            result.error(f"镜号必须从1连续递增：期望 {expected_no}，实际 {shot_no}")
        expected_no += 1
        for key in ["scene", "source_paragraph", "duration_seconds", "duration_breakdown", "camera_main_image", "notes", "prompt", "beat_ids", "covered_fact_ids", "continuity_updates"]:
            if key not in shot:
                result.error(f"镜号{shot_no} 缺少字段：{key}")
        if "keyframe" in shot:
            result.error(f"镜号{shot_no} 不得包含 keyframe 字段；关键帧列已从主表删除。")
        if shot.get("old_shot_no") is not None:
            result.error(f"镜号{shot_no} 不得保留 old_shot_no；禁止后处理合并痕迹进入交付数据。")
        scene = scene_key_for(data, shot)
        if use_scene_id:
            if not scene:
                result.error(f"镜号{shot_no} 缺少 scene_id。")
            if not clean_text(shot.get("scene")):
                result.error(f"镜号{shot_no} 缺少展示用 scene。")
        if scenes and scene not in scenes:
            result.warn(f"镜号{shot_no} 场景键未在 continuity_logs 中登记：{scene}")
        is_first_scene_shot = str(scene) not in seen_scenes
        if is_first_scene_shot:
            seen_scenes.add(str(scene))
            validate_first_scene_shot(shot, as_dict(continuity_logs.get(str(scene))), result)
        beat_ids = [str(item) for item in as_list(shot.get("beat_ids"))]
        fact_ids = [str(item) for item in as_list(shot.get("covered_fact_ids"))]
        if not beat_ids:
            result.error(f"镜号{shot_no} beat_ids 不能为空。")
        safety_exception = approved_safety_exception(data, shot)
        if not fact_ids and not safety_exception:
            result.error(f"镜号{shot_no} covered_fact_ids 不能为空。")
        for fact_id in fact_ids:
            covered.add(fact_id)
            owner = fact_to_beat.get(fact_id)
            if owner is None:
                result.error(f"镜号{shot_no} 覆盖了不存在的事实 ID：{fact_id}")
            elif owner not in beat_ids:
                result.error(f"镜号{shot_no} 覆盖事实 {fact_id}，但 beat_ids 未包含 {owner}。")
        if use_scene_id and clean_text(shot.get("shot_type")) in {"transition", "safety"} and fact_ids:
            bound_types = fact_types_for_shot(shot, fact_types)
            if not (bound_types & {"space", "sound", "reality"}):
                result.error(f"镜号{shot_no} {shot.get('shot_type')} 镜必须绑定空间、声音或现实层事实。")
        validate_duration(shot, result, version_24=use_scene_id)
        validate_camera(shot, result)
        validate_prompt(shot, result)
        validate_hybrid_fields(shot, result, required=hybrid_required)
        updates = as_list(shot.get("continuity_updates"))
        if updates:
            has_update = True
        has_station_move = "【站位位移】" in clean_text(shot.get("camera_main_image"))
        if has_station_move and not updates:
            result.error(f"镜号{shot_no} 写了【站位位移】，但 continuity_updates 为空。")
        for update in updates:
            current_states = scene_states.get(str(scene)) if structured_required else None
            validate_update(shot, update, result, has_station_move, current_states=current_states)
    missing = sorted(all_fact_ids - covered)
    if missing:
        result.error("存在未覆盖事实 ID：" + ", ".join(missing[:20]))
    if any("【站位位移】" in clean_text(shot.get("camera_main_image")) for shot in shots if isinstance(shot, dict)) and not has_update:
        result.error("分镜出现站位位移，但没有任何 continuity_updates。")
    validate_adjacent_shot_design(
        shots,
        fact_types,
        result,
        guard_reverse_axis_motion=requires_adjacent_motion_guard(data),
        allow_marked_exceptions=use_scene_id,
    )


def validate_duration(shot: dict[str, Any], result: ValidationResult, *, version_24: bool = False) -> None:
    shot_no = shot.get("shot_no")
    try:
        duration = float(shot.get("duration_seconds"))
    except Exception:
        result.error(f"镜号{shot_no} 镜头时长必须是数字。")
        return
    if duration <= 0:
        result.error(f"镜号{shot_no} 镜头时长必须大于0。")
    breakdown = as_dict(shot.get("duration_breakdown"))
    for field_name in DURATION_FIELDS:
        if field_name not in breakdown:
            result.error(f"镜号{shot_no} duration_breakdown 缺少 {field_name}。")
            return
    try:
        values = [float(breakdown[field_name]) for field_name in DURATION_FIELDS]
        if version_24:
            for field_name, value in zip(DURATION_FIELDS, values):
                if value < 0 or not value.is_integer():
                    result.error(f"镜号{shot_no} duration_breakdown.{field_name} 必须是非负整数。")
        expected = math.ceil(max(values[0], values[1]) + values[2] + values[3])
    except Exception:
        result.error(f"镜号{shot_no} duration_breakdown 必须全为数字。")
        return
    if int(duration) != expected:
        result.error(f"镜号{shot_no} 时长不符合公式：表格 {duration:g} 秒，公式 {expected} 秒。")
    quoted = "".join(re.findall(r"[“\"]([^”\"]+)[”\"]", clean_text(shot.get("camera_main_image"))))
    spoken_chars = len(re.sub(r"[，。！？……、\s]", "", quoted))
    if spoken_chars:
        min_dialogue_seconds = math.ceil(spoken_chars / 4)
        if float(breakdown["sync_dialogue_seconds"]) < min_dialogue_seconds:
            result.error(f"镜号{shot_no} 对白时长过短：{spoken_chars}字至少约 {min_dialogue_seconds} 秒。")
    notes = clean_text(shot.get("notes"))
    if duration > 6 and "[时长估算]" not in notes:
        result.error(f"镜号{shot_no} 超过6秒，备注必须写 [时长估算]。")
    if duration > 10:
        long_take = as_dict(shot.get("long_take"))
        classification = clean_text(long_take.get("classification"))
        if version_24:
            if classification not in LONG_TAKE_CLASSIFICATIONS or classification == "not_applicable":
                result.error(f"镜号{shot_no} 超过10秒，long_take.classification 必须为有效长镜分类。")
        elif long_take.get("classification") in (None, "not_applicable") and "[长镜头]" not in notes:
            result.error(f"镜号{shot_no} 超过10秒，必须标注长镜头分类或人工复核理由。")


def validate_camera(shot: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    camera = clean_text(shot.get("camera_main_image"))
    lines = [line.strip() for line in camera.split("\n") if line.strip()]
    if len(lines) < 2:
        result.error(f"镜号{shot_no} 运镜列至少需要三联首行和【机位逻辑】。")
        return
    if not re.fullmatch(r"\[[^,\]]+,\s*[^,\]]+,\s*[^\]]+\]", lines[0]):
        result.error(f"镜号{shot_no} 运镜首行必须是三要素：[视角/高度, 景别/特殊视角, 运镜方式]。")
    angle, shot_size, movement = parse_triad(camera)
    validate_camera_movement_compatibility(shot_no, angle, shot_size, movement, result, clean_text(shot.get("notes")))
    if not lines[1].startswith("【机位逻辑】"):
        result.error(f"镜号{shot_no} 运镜第二行必须以【机位逻辑】开头。")
    for word in FORBIDDEN_VISUAL_WORDS:
        if word in camera:
            result.error(f"镜号{shot_no} 运镜列包含禁用表达：{word}")
    for marker in UNSUPPORTED_MARKERS:
        if marker in camera or marker in clean_text(shot.get("notes")):
            result.error(f"镜号{shot_no} 包含后处理/无依据标记：{marker}")
    if "【镜内变化】" in camera:
        result.error(f"镜号{shot_no} 不得在最终主表出现【镜内变化】。")
    if "运镜修正" in clean_text(shot.get("notes")):
        result.error(f"镜号{shot_no} 备注不得保留运镜修正痕迹。")


def validate_camera_movement_compatibility(
    shot_no: Any,
    angle: str,
    shot_size: str,
    movement: str,
    result: ValidationResult,
    notes: str = "",
) -> None:
    if "斯坦尼康" in clean_text(movement) and has_any_keyword(shot_size, ("大远景", "大全景")):
        result.error(f"镜号{shot_no} 大远景/大全景不得使用斯坦尼康；应改用摇臂、伸缩摇臂或航拍。")
    if "航拍" in clean_text(movement):
        if not has_any_keyword(shot_size, AERIAL_ALLOWED_SHOT_SIZE_KEYWORDS):
            result.error(f"镜号{shot_no} 航拍必须服务于大范围空间，景别应为大远景/大全景/全景。")
        if not has_any_keyword(angle, AERIAL_ALLOWED_ANGLE_KEYWORDS):
            result.error(f"镜号{shot_no} 航拍必须匹配俯拍、高角度、正俯拍或微俯视。")
    size_text = f"{angle} {shot_size}"
    if shot_size_has_span(size_text):
        has_span_movement = has_any_keyword(movement, SPAN_MOVEMENT_KEYWORDS)
        has_span_note = "[景别跨度]" in clean_text(notes)
        if not has_span_movement or not has_span_note:
            result.error(
                f"镜号{shot_no} 景别同时包含远景类和近景/特写类，"
                "必须使用光学变焦、快速推拉、急推/急拉、推轨推进/拉出或伸缩摇臂等跨度运镜，"
                "并在备注写 [景别跨度] 理由。"
            )


def validate_prompt(shot: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    prompt = clean_text(shot.get("prompt"))
    expected = derive_prompt(shot)
    if prompt != expected:
        result.error(f"镜号{shot_no} Prompt 必须由锁定主表派生，且与标准五字段模板一致。")
    if len(prompt) > 800:
        result.warn(f"镜号{shot_no} Prompt 总长超过800字符，需申请回流 Gate B 裁决拆镜或保留。")
    lines = [line.strip() for line in prompt.split("\n") if line.strip()]
    if len(lines) != 5:
        result.error(f"镜号{shot_no} Prompt 必须且只能包含五行。")
    for field_name in PROMPT_FIELDS:
        if not any(line.startswith(field_name) for line in lines):
            result.error(f"镜号{shot_no} Prompt 缺少字段：{field_name}")
    for item in PROMPT_FORBIDDEN:
        if item in prompt:
            result.error(f"镜号{shot_no} Prompt 包含禁止进入派生列的内部/关键帧内容：{item}")
    if not quotes_are_balanced(prompt):
        result.error(f"镜号{shot_no} Prompt 引号不闭合。")
    fields = prompt_fields(prompt)
    composition = fields.get("构图：", "")
    if composition and composition_has_field_leak(composition):
        result.error(f"镜号{shot_no} Prompt 构图字段混入对白、声音或动作链；构图只能写主体/位置/道具关系。")


def validate_update(
    shot: dict[str, Any],
    update: Any,
    result: ValidationResult,
    has_station_move: bool,
    *,
    current_states: dict[tuple[str, str, str], str] | None = None,
) -> None:
    shot_no = shot.get("shot_no")
    if not isinstance(update, dict):
        result.error(f"镜号{shot_no} continuity_updates 每项必须是对象。")
        return
    required = ["entity_type", "entity", "field", "from", "to", "evidence_fact_ids"]
    for key in required:
        if key not in update:
            result.error(f"镜号{shot_no} continuity_update 缺少 {key}。")
    if update.get("from") == update.get("to"):
        result.error(f"镜号{shot_no} continuity_update 的 from/to 不得相同。")
    evidence = [str(item) for item in as_list(update.get("evidence_fact_ids"))]
    covered = {str(item) for item in as_list(shot.get("covered_fact_ids"))}
    if not evidence:
        result.error(f"镜号{shot_no} continuity_update 必须有 evidence_fact_ids。")
    for fact_id in evidence:
        if fact_id not in covered:
            result.error(f"镜号{shot_no} continuity_update 证据 {fact_id} 未被当前镜头覆盖。")
    if update.get("field") in POSITION_UPDATE_FIELDS and update.get("entity_type") == "character" and not has_station_move:
        result.error(f"镜号{shot_no} 登记了位置/朝向迁移，但主画面缺少【站位位移】。")
    if current_states is not None and clean_text(update.get("field")) in STATEFUL_CONTINUITY_FIELDS:
        key = state_key(update.get("entity_type"), update.get("entity"), update.get("field"))
        current = current_states.get(key)
        from_value = clean_text(update.get("from"))
        to_value = clean_text(update.get("to"))
        if current is None:
            result.warn(f"镜号{shot_no} continuity_update 缺少可校验的上一状态：{key[0]}/{key[1]}/{key[2]}。")
        elif current != from_value:
            result.error(
                f"镜号{shot_no} continuity_update.from 与上一状态不一致："
                f"{key[0]}/{key[1]}/{key[2]} 当前为 {current}，from 写为 {from_value}。"
            )
        if to_value:
            current_states[key] = to_value


def validate_data(data: dict[str, Any], *, strict_status: bool, final_signoff: bool = False) -> ValidationResult:
    result = ValidationResult()
    validate_metadata(data, result, strict_status=strict_status, final_signoff=final_signoff)
    validate_human_reviews(data, result, final_signoff=final_signoff)
    validate_script_lock(data, result)
    continuity_logs = validate_continuity_logs(data, result)
    fact_to_beat, fact_types, all_fact_ids = collect_facts(data, result)
    validate_fact_cut_contract(data, result)
    validate_fact_text_preservation(data, result)
    validate_rows(data, result, fact_to_beat, fact_types, all_fact_ids, continuity_logs)
    validate_overcompression(data, result)
    validate_must_isolate_density(data, result)
    validate_warn_resolutions(data, result)
    return result


def markdown_text(data: dict[str, Any]) -> str:
    lines = ["# 分镜表", ""]
    lines.append("| " + " | ".join(HEADERS) + " |")
    lines.append("| " + " | ".join(["---"] * len(HEADERS)) + " |")
    for row in expected_rows(data):
        lines.append("| " + " | ".join(escape_cell(item) for item in row) + " |")
    report = as_dict(data.get("validation_report"))
    if report:
        lines.extend(
            [
                "",
                "## 校验摘要",
                "",
                f"- status: {report.get('status', '')}",
                f"- shot_count: {report.get('shot_count', '')}",
                f"- column_contract: {report.get('column_contract', '')}",
            ]
        )
        if requires_structured_audit(data):
            for key, value in report_summary_items(data):
                lines.append(f"- {key}: {value}")
    return "\n".join(lines) + "\n"


def build_markdown(data: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(markdown_text(data), encoding="utf-8")


def build_excel(data: dict[str, Any], path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl 不可用，无法生成 Excel。")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    for row in expected_rows(data):
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, width in enumerate([8, 24, 44, 14, 72, 36, 56], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 36 if row_index == 1 else 108
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if requires_structured_audit(data):
        summary = workbook.create_sheet(SUMMARY_SHEET_NAME)
        summary.append(["字段", "值"])
        for key, value in report_summary_items(data):
            summary.append([key, value])
        for cell in summary[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        summary.column_dimensions["A"].width = 24
        summary.column_dimensions["B"].width = 96
        for row in summary.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def build_report(data: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    report = as_dict(data.get("validation_report"))
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_markdown_table(path: Path) -> list[list[str]]:
    text = path.read_text(encoding="utf-8-sig")
    rows: list[list[str]] = []
    for line in text.splitlines():
        if not line.startswith("|"):
            continue
        cells = [cell.strip().replace("<br>", "\n").replace("｜", "|") for cell in line.strip().strip("|").split("|")]
        if cells and all(re.fullmatch(r":?-{3,}:?", cell) for cell in cells):
            continue
        rows.append(cells)
    return rows


def compare_markdown(data: dict[str, Any], path: Path, result: ValidationResult) -> None:
    if not path.exists():
        result.error(f"Markdown 文件不存在：{path}")
        return
    rows = parse_markdown_table(path)
    if not rows:
        result.error("Markdown 未找到分镜表。")
        return
    if rows[0] != HEADERS:
        result.error("Markdown 表头必须是稳定 7 列，且不得包含关键帧列。")
    expected = [[str(item) for item in row] for row in expected_rows(data)]
    actual = [[str(item) for item in row] for row in rows[1:]]
    if actual != expected:
        result.error("Markdown 表格内容与 shot_data 派生结果不一致。")
    if requires_structured_audit(data):
        text = path.read_text(encoding="utf-8-sig")
        for key, value in report_summary_items(data):
            expected_line = f"- {key}: {value}"
            if expected_line not in text:
                result.error(f"Markdown 校验摘要缺少或不一致：{key}")


def compare_excel(data: dict[str, Any], path: Path, result: ValidationResult) -> None:
    if load_workbook is None:
        result.error("openpyxl 不可用，无法校验 Excel。")
        return
    if not path.exists():
        result.error(f"Excel 文件不存在：{path}")
        return
    workbook = load_workbook(path, read_only=True)
    try:
        expected_sheets = [SHEET_NAME, SUMMARY_SHEET_NAME] if requires_structured_audit(data) else [SHEET_NAME]
        if workbook.sheetnames != expected_sheets:
            result.error(f"Excel Sheet 必须为：{', '.join(expected_sheets)}。")
            return
        sheet = workbook[SHEET_NAME]
        rows = list(sheet.iter_rows(values_only=True))
        summary_rows = []
        if requires_structured_audit(data):
            summary_rows = list(workbook[SUMMARY_SHEET_NAME].iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        result.error("Excel 分镜表为空。")
        return
    header = [str(item or "") for item in rows[0]]
    if header != HEADERS:
        result.error("Excel 表头必须是稳定 7 列，且不得包含关键帧列。")
    expected = [[str(item) for item in row] for row in expected_rows(data)]
    actual = [[str(item or "") for item in row] for row in rows[1:]]
    if actual != expected:
        result.error("Excel 内容与 shot_data 派生结果不一致。")
    if requires_structured_audit(data):
        expected_summary = [("字段", "值"), *report_summary_items(data)]
        actual_summary = [(str(row[0] or ""), str(row[1] or "")) for row in summary_rows]
        if actual_summary != [(str(key), str(value)) for key, value in expected_summary]:
            result.error("Excel 校验摘要与 shot_data.validation_report 不一致。")


def compare_report(data: dict[str, Any], path: Path, result: ValidationResult) -> None:
    if not path.exists():
        result.error(f"validation_report 文件不存在：{path}")
        return
    try:
        report = json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception as exc:
        result.error(f"validation_report 不是合法 JSON：{exc}")
        return
    if report != as_dict(data.get("validation_report")):
        result.error("validation_report 文件与 shot_data.validation_report 不一致。")


def print_result(result: ValidationResult) -> None:
    if result.errors:
        print("ERRORS:", file=sys.stderr)
        for item in result.errors:
            print(f"- {item}", file=sys.stderr)
    if result.warnings:
        print("WARNINGS:")
        for item in result.warnings:
            print(f"- {item}")
    print(f"{result.status}: storyboard validation {'passed' if not result.errors else 'failed'}.")


def command_build(args: argparse.Namespace) -> int:
    data_path = Path(args.data)
    data = load_json(data_path)
    derive_prompts(data)
    result = validate_data(data, strict_status=False, final_signoff=False)
    update_validation_report(data, result)
    if result.errors:
        write_json(data_path, data)
        if args.report:
            build_report(data, Path(args.report))
        print_result(result)
        return 1
    write_json(data_path, data)
    build_markdown(data, Path(args.markdown))
    build_excel(data, Path(args.excel))
    if args.report:
        build_report(data, Path(args.report))
    print_result(result)
    return 0


def command_validate(args: argparse.Namespace) -> int:
    data = load_json(Path(args.data))
    result = validate_data(data, strict_status=True, final_signoff=bool(args.final_signoff))
    compare_markdown(data, Path(args.markdown), result)
    compare_excel(data, Path(args.excel), result)
    if args.report:
        compare_report(data, Path(args.report), result)
    print_result(result)
    return 1 if result.errors else 0


def make_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    build = subparsers.add_parser("build", help="derive Prompt, update JSON, and write Markdown/Excel")
    build.add_argument("--data", required=True)
    build.add_argument("--markdown", required=True)
    build.add_argument("--excel", required=True)
    build.add_argument("--report")
    build.set_defaults(func=command_build)
    validate = subparsers.add_parser("validate", help="validate JSON/Markdown/Excel consistency")
    validate.add_argument("--data", required=True)
    validate.add_argument("--markdown", required=True)
    validate.add_argument("--excel", required=True)
    validate.add_argument("--report")
    validate.add_argument("--final-signoff", action="store_true", help="require Gate C approved for 2.4.x final deliveries")
    validate.set_defaults(func=command_validate)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = make_parser()
    args = parser.parse_args(argv)
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
