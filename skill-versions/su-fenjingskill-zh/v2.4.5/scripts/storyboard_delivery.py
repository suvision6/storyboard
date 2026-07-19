#!/usr/bin/env python3
"""Build and validate stable 7-column su-fenjingskill-zh deliveries."""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import math
import os
import re
import sys
import tempfile
import unicodedata
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - reported at runtime.
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None

VERSION = "2.4.5"
RULE_REVISION = "2.4.5-director-shot-integrity-guard-2026-07-18"

# Keep version behavior in one fail-closed table.  Historical entries are
# validation-only; only VERSION may be built by this release.
VERSION_PROFILES: dict[str, Any] = {
    "2.3.2": {},
    "2.3.3": {},
    "2.3.4": "2.3.4-overcompression-guard-2026-06-30",
    "2.3.5": "2.3.5-adjacent-motion-guard-2026-06-30",
    "2.3.6": "2.3.6-tri-source-audit-guard-2026-06-30",
    "2.4.0": "2.4.0-human-gate-stability-2026-07-06",
    "2.4.1": "2.4.1-source-lock-2026-07-07",
    "2.4.2": "2.4.2-source-lock-entry-guard-2026-07-07",
    "2.4.3": {
        "rule_revision": "2.4.3-contract-integrity-p2-2026-07-12",
        "contract_integrity": True,
    },
    "2.4.4": {
        "rule_revision": "2.4.4-emotion-performance-guard-2026-07-16",
        "contract_integrity": True,
        "emotion_performance_guard": True,
    },
    "2.4.5": {
        "rule_revision": RULE_REVISION,
        "contract_integrity": True,
        "emotion_performance_guard": True,
        "director_integrity_guard": True,
    },
}

# Expand the compact declarations above into explicit profiles without
# duplicating feature-version sets throughout the validator.
for _version, _profile in list(VERSION_PROFILES.items()):
    if isinstance(_profile, str):
        _profile = {"rule_revision": _profile}
    _profile.setdefault("rule_revision", None)
    _profile["hybrid"] = _version >= "2.3.4"
    _profile["adjacent_motion_guard"] = _version >= "2.3.5"
    _profile["structured_audit"] = _version >= "2.3.6"
    _profile["human_gate"] = _version >= "2.4.0"
    _profile["source_lock"] = _version >= "2.4.1"
    _profile["approved_script_path"] = _version >= "2.4.2"
    _profile.setdefault("contract_integrity", False)
    _profile.setdefault("emotion_performance_guard", False)
    _profile.setdefault("director_integrity_guard", False)
    VERSION_PROFILES[_version] = _profile

RULE_REVISIONS = {
    version: str(profile["rule_revision"])
    for version, profile in VERSION_PROFILES.items()
    if profile.get("rule_revision")
}
SHEET_NAME = "分镜表"
SUMMARY_SHEET_NAME = "校验摘要"
HEADERS = [
    "镜号",
    "场景",
    "原剧本段落",
    "镜头时长(秒)",
    "运镜+主画面描述(含台词)",
    "备注",
    "Prompt",
]
PROMPT_FIELDS = ["时间：", "景别：", "构图：", "运镜手法：", "画面内容："]
DURATION_FIELDS = [
    "sync_action_seconds",
    "sync_dialogue_seconds",
    "non_sync_action_seconds",
    "emotional_pause_seconds",
]
FACT_TYPES = {
    "character",
    "action",
    "dialogue",
    "prop",
    "space",
    "position",
    "emotion",
    "sound",
    "reality",
}
SHOT_TYPES = {
    "master",
    "action",
    "dialogue",
    "reaction",
    "insert",
    "transition",
    "vfx_anchor",
    "safety",
}
SPLIT_REASONS = {
    "spatial_anchor",
    "performance_continuity",
    "new_information",
    "prop_state_change",
    "new_vfx_state",
    "new_sound_source",
    "reality_layer_shift",
    "causal_reveal",
    "emotional_turn",
    "continuity_migration",
}
INSERT_PRIORITIES = {"none", "recommended", "must_have"}
CUT_PRIORITIES = {"normal", "recommended", "must_isolate"}
CUT_CATEGORIES = {"space", "prop", "action", "emotion", "sound", "reality", "dialogue", "vfx", "character"}
CRITICAL_SPLIT_REASONS = {
    "causal_reveal",
    "reality_layer_shift",
    "new_vfx_state",
    "new_sound_source",
    "emotional_turn",
    "prop_state_change",
}
MUST_ISOLATE_REASONS = {
    "causal_reveal",
    "reality_layer_shift",
    "new_vfx_state",
    "new_sound_source",
    "emotional_turn",
}
LONG_TAKE_SUPPORTS = {
    "shot_size_change",
    "character_blocking",
    "foreground_background_change",
    "sound_source_change",
    "vfx_state_change",
    "emotional_turn",
    "spatial_progression",
}
INSERT_LIKE_TYPES = {"insert", "reaction", "vfx_anchor", "safety"}
HYBRID_VERSIONS = {version for version, profile in VERSION_PROFILES.items() if profile["hybrid"]}
ADJACENT_MOTION_GUARD_VERSIONS = {
    version for version, profile in VERSION_PROFILES.items() if profile["adjacent_motion_guard"]
}
STRUCTURED_AUDIT_VERSIONS = {
    version for version, profile in VERSION_PROFILES.items() if profile["structured_audit"]
}
HUMAN_GATE_VERSIONS = {version for version, profile in VERSION_PROFILES.items() if profile["human_gate"]}
SOURCE_LOCK_VERSIONS = {version for version, profile in VERSION_PROFILES.items() if profile["source_lock"]}
APPROVED_SCRIPT_PATH_VERSIONS = {
    version for version, profile in VERSION_PROFILES.items() if profile["approved_script_path"]
}
MUST_ISOLATE_DENSITY_THRESHOLD = 0.5
VALID_STATUSES = {"PASS", "WARN", "FAIL", "NOT_RUN"}
ANCHOR_TYPES = {"space", "multi_character", "both", "single_continuation", "subjective"}
REQUIRED_REFERENCE_KEYS = ("continuity-shot-data", "hybrid-shot-audit", "camera-language", "seedance-prompt-rules")
REQUIRED_SCRIPT_KEYS = ("storyboard_delivery.py", "validate_storyboard.js")
GATES = {"GATE_0", "GATE_A", "GATE_B", "GATE_C"}
LONG_TAKE_CLASSIFICATIONS = {"not_applicable", "dialogue_long", "action_long", "spatial_long", "emotional_long"}
NOTE_MARKERS = [
    "[合理补足]",
    "[时长估算]",
    "[长镜头]",
    "[保留理由]",
    "[不可拆说明]",
    "[景别跨度]",
    "[非连续Beat]",
    "[必拆相邻]",
    "[反转动机]",
    "[安全镜]",
    "[人工批准]",
    "[声音]",
    "[reference missing]",
    "[跳切说明]",
    "[越轴说明]",
    "[动作链拆分]",
    "[原文复用]",
]
LEGACY_NOTE_MARKERS = NOTE_MARKERS[:-4]
PANORAMIC_SHOT_SIZE_KEYWORDS = ("大远景", "大全景", "中全景", "全景")
PANORAMIC_REQUIRED_ANCHORS = {"space", "multi_character", "both"}
AERIAL_ALLOWED_SHOT_SIZE_KEYWORDS = ("大远景", "大全景", "全景")
AERIAL_ALLOWED_ANGLE_KEYWORDS = ("俯拍", "高角度", "正俯拍", "微俯视")
WIDE_SHOT_SIZE_KEYWORDS = ("大远景", "大全景", "远景", "全景", "中全景")
CLOSE_SHOT_SIZE_KEYWORDS = ("中近景", "近景", "中特写", "特写", "微距特写")
SHOT_SIZE_RANKS = [
    (0, ("大远景", "大全景", "远景")),
    (1, ("全景", "中全景")),
    (2, ("中景", "七分身", "半身")),
    (3, ("中近景", "近景", "中特写")),
    (4, ("微距特写", "特写")),
]
STATIC_OR_SOFT_MOVEMENT_KEYWORDS = (
    "固定",
    "轻微",
    "缓慢",
    "微晃",
    "呼吸式",
    "下摇",
    "纵摇",
    "横摇",
)
PUSH_IN_MOVEMENT_KEYWORDS = (
    "推轨推进",
    "缓慢推进",
    "航拍推进",
    "急推",
    "推进",
)
PULL_OUT_MOVEMENT_KEYWORDS = (
    "推轨拉出",
    "缓慢拉出",
    "拉远",
    "拉出",
)
SPAN_MOVEMENT_KEYWORDS = (
    "光学变焦",
    "快速推拉",
    "急推",
    "急拉",
    "推轨推进",
    "推轨拉出",
    "缓慢推进",
    "缓慢拉出",
    "伸缩摇臂",
)
INCOMPLETE_DIALOGUE_ENDINGS = ("，", "、", "：", "；", ",", ":", ";")
EXPRESSION_ONLY_KEYWORDS = (
    "笑",
    "苦笑",
    "嘴角",
    "眼神",
    "抬眼",
    "闭眼",
    "呼吸",
    "停顿",
    "沉默",
    "表情",
    "眉",
    "瞳孔",
    "下颌",
    "嘴唇",
)
NEW_FACT_TYPES_FOR_CUT = {"position", "prop", "space", "sound", "reality"}
STATEFUL_CONTINUITY_FIELDS = {"position", "facing", "state", "owner", "value", "presence", "visibility"}
CONTINUITY_ENTITY_TYPES = {"character", "prop", "fixed_object", "sound_source", "reality_layer"}
POSITION_UPDATE_FIELDS = {"position", "facing", "presence"}
SOUND_SOURCE_UPDATE_FIELDS = {"position", "state", "visibility"}
SOUND_SOURCE_VISIBILITIES = {"onscreen", "offscreen"}
SPATIAL_AXIS_TYPES = {"eyeline", "movement", "action", "spatial"}
CAMERA_AXIS_SIDES = {"side_a", "side_b", "on_axis", "not_applicable"}
CAMERA_FACINGS = {
    "toward_end_a",
    "toward_end_b",
    "along_a_to_b",
    "along_b_to_a",
    "across_axis",
    "subjective",
    "overhead",
    "not_applicable",
}
ANGLE_DELTAS = {"scene_start", "same", "minor", "substantial", "axis_cross", "not_comparable"}
SCREEN_DIRECTION_KINDS = {"facing", "eyeline", "movement"}
SCREEN_DIRECTIONS = {"screen_left", "screen_right", "toward_camera", "away_camera", "neutral"}
PERFORMANCE_CHAIN_ROLES = {"action", "reaction", "dialogue"}
PERFORMANCE_CHAIN_BREAK_REASONS = {
    "must_isolate",
    "reality_shift",
    "time_jump",
    "spatial_discontinuity",
    "intentional_editorial_emphasis",
}
PERFORMANCE_CUE_CHANNELS = {"gaze_face", "breath_voice", "posture_hand", "stillness"}
ENVIRONMENT_CUE_KINDS = {"sound", "light", "contact"}
ENVIRONMENT_CUE_BASES = {"fact", "inevitable_action"}
DIRECTORIAL_CONSTRAINT_TYPES = {"same_shot", "same_beat", "forbid_literal", "direct_handoff"}
DIALOGUE_DELIVERIES = {"onscreen", "offscreen", "vo"}
PROJECT_LEXICON_KEYS = {
    "prop_terms",
    "space_terms",
    "vfx_terms",
    "reality_terms",
    "sound_terms",
}
CURRENT_BEAT_ID_PATTERN = r"B(?:00[1-9]|0[1-9][0-9]|[1-9][0-9]{2,})"
LEGACY_BEAT_ID_PATTERN = r"B[0-9]{3}"
SCENE_ID_PATTERN = r"S[0-9]{2,}"
BATCH_ID_PATTERN = r"BT[0-9]{2,}"
SOUND_SOURCE_ID_PATTERN = r"SS[0-9]{2,}"
SPATIAL_AXIS_ID_PATTERN = r"AX[0-9]{2,}"
PERFORMANCE_CHAIN_ID_PATTERN = r"PC[0-9]{3,}"
DIRECTORIAL_CONSTRAINT_ID_PATTERN = r"DC[0-9]{3,}"
CURRENT_FACT_SUFFIX_PATTERN = r"F[0-9]{2}"
PROMPT_COMPOSITION_FORBIDDEN_PATTERNS = [
    re.compile(r"[“”\"]"),
    re.compile(r"[\u4e00-\u9fa5A-Za-z0-9·]{1,8}(?:说|喊|问|答|念)"),
    re.compile(r"画外声|对白|台词|传来|响起|震动|发出"),
    re.compile(r"走到|冲向|拿起|放下|击中|钻入|劈下|吞没|笑|哭|喊|抬手|跪地"),
]
FACT_TEXT_ALLOWLIST = {
    "画面",
    "镜头",
    "机位",
    "朝向",
    "轴线",
    "左侧",
    "右侧",
    "前景",
    "背景",
    "可见",
}
INTERNAL_LABELS = [
    "【镜内变化】",
    "【人物关系】",
    "【人物关系补充】",
    "【站位位移】",
    "【机位逻辑】",
    "【场景首镜站位】",
]
PROMPT_FORBIDDEN = [
    "【镜内变化】",
    "【人物关系】",
    "【人物关系补充】",
    "【站位位移】",
    "【机位逻辑】",
    "【场景首镜站位】",
    "运镜修正",
    "关键帧",
    "生图",
    "AI生图提示词",
    "风格：",
    "禁止：",
] + NOTE_MARKERS
UNSUPPORTED_MARKERS = [
    "承上镜动作",
    "承上镜",
    "补充镜头",
    "新增镜头",
    "新增过渡",
    "无原文依据",
    "非原文",
    "原文未写",
    "自行添加",
    "自行补充",
    "为了过渡",
    "建议保留此镜",
]
FORBIDDEN_VISUAL_WORDS = [
    "令人叹为观止",
    "令人惊叹",
    "令人着迷",
    "精心打造",
    "匠心独运",
    "独具匠心",
    "视觉盛宴",
    "光影交响",
    "完美呈现",
    "极致体验",
    "引人入胜",
    "震撼人心",
    "巧妙融合",
    "仿佛",
    "犹如",
    "宛如",
    "好似",
]
EMOTION_PERFORMANCE_FORBIDDEN_WORDS = [
    "心如刀割",
    "怒火中烧",
    "万箭穿心",
    "悲痛欲绝",
    "杀气腾腾",
    "柔情似水",
    "绝望如深渊",
    "时间像河流般流淌",
    "城市在夜色中沉睡",
]
PRECISE_EXECUTION_VALUE_PATTERN = re.compile(
    r"(?:\d+(?:\.\d+)?|[零〇一二两三四五六七八九十百千万]+)\s*"
    r"(?:mm|cm|m(?![A-Za-z])|Hz|K|毫米|厘米|米|赫兹|开尔文|°|度)",
    re.IGNORECASE,
)

REQUIRED_CUT_POINT_CATEGORIES = [
    (
        "发言权转移",
        [re.compile(r"[\u4e00-\u9fa5A-Za-z0-9·]{1,8}[：:]")],
        True,
        None,
    ),
    ("问答关系变化", [re.compile(r"问|回答|回应|反问|是不是|为什么|什么|吗|？|\?")], False, None),
    (
        "角色明显反应",
        [re.compile(r"一怔|脸色|眼眶|嘴唇|冷笑|大笑|苦笑|痛苦|不可思议|回头|低头|抬头|摇头|点头|踉跄|绷|僵|停住")],
        False,
        None,
    ),
    (
        "道具状态变化",
        [re.compile(r"显示|亮起|暗下|恢复|变长|延伸|变成|熄灭")],
        False,
        "prop_terms",
    ),
    (
        "攻击/命中/结果",
        [re.compile(r"抬手|挥|扑|冲|拽|拦|扫|斩|震|炸|击|挡|断裂|碎裂|灰飞烟灭|跳入|钻进|坠落|吞没")],
        False,
        None,
    ),
    (
        "空间方向改变",
        [re.compile(r"中央|走向|走到|冲向|转身|回到|两侧|不同方向|分头")],
        False,
        "space_terms",
    ),
    (
        "层级切换",
        [re.compile(r"切回|切——|梦境|幻境|现实|闪回|脑海|画外声|VO|回忆")],
        False,
        ("reality_terms", "sound_terms"),
    ),
    (
        "视觉/VFX状态变化",
        [re.compile(r"雾气|膨胀")],
        False,
        "vfx_terms",
    ),
]


@dataclass
class ValidationResult:
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def status(self) -> str:
        if self.errors:
            return "FAIL"
        if self.warnings:
            return "WARN"
        return "PASS"

    def error(self, message: str) -> None:
        self.errors.append(message)

    def warn(self, message: str) -> None:
        if message not in self.warnings:
            self.warnings.append(message)


def warning_id(message: str) -> str:
    return "W-" + hashlib.sha1(clean_text(message).encode("utf-8")).hexdigest()[:12]


def warning_items(warnings: list[str]) -> list[dict[str, str]]:
    return [{"warn_id": warning_id(message), "message": message} for message in warnings]


def reject_json_constant(token: str) -> None:
    raise ValueError(f"JSON 不允许非标准数值：{token}")


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig") as handle:
        value = json.load(handle, parse_constant=reject_json_constant)
    if not isinstance(value, dict):
        raise ValueError("shot_data 顶层必须是 JSON 对象。")
    return value


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")


def data_without_validation_report(data: dict[str, Any]) -> dict[str, Any]:
    value = copy.deepcopy(data)
    value.pop("validation_report", None)
    # Audit records are procedurally append-only. Current-snapshot validation
    # cannot prove immutable history. They remain outside creative content
    # hashes so adding Gate C does not invalidate the content it approved.
    if is_current_contract(data):
        value.pop("human_reviews", None)
        metadata = as_dict(value.get("metadata"))
        if metadata:
            metadata.pop("revision_log", None)
    return value


def canonical_data_hash(data: dict[str, Any]) -> str:
    payload = json.dumps(
        data_without_validation_report(data),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def final_content_hash(data: dict[str, Any]) -> str:
    """Public name for the stable Gate-C/source-data content hash."""

    return canonical_data_hash(data)


def _batch_scene_ids(data: dict[str, Any], batch_id: str | None) -> set[str] | None:
    if not batch_id:
        return None
    for batch in as_list(as_dict(data.get("batch_plan")).get("batches")):
        if isinstance(batch, dict) and clean_text(batch.get("batch_id")) == batch_id:
            return set(normalized_list(batch.get("scene_ids")))
    return set()


def batch_id_for_scene(data: dict[str, Any], scene_id: Any) -> str | None:
    scene = clean_text(scene_id)
    for batch in as_list(as_dict(data.get("batch_plan")).get("batches")):
        if isinstance(batch, dict) and scene in set(normalized_list(batch.get("scene_ids"))):
            return clean_text(batch.get("batch_id")) or None
    return None


def _items_for_scenes(items: Any, scene_ids: set[str] | None) -> list[dict[str, Any]]:
    values = [copy.deepcopy(item) for item in as_list(items) if isinstance(item, dict)]
    if scene_ids is None:
        return values
    return [item for item in values if clean_text(item.get("scene_id")) in scene_ids]


def gate_review_payload(data: dict[str, Any], gate: str, batch_id: str | None = None) -> dict[str, Any]:
    """Return the deterministic payload reviewed at a human gate."""

    gate = clean_text(gate)
    metadata = as_dict(data.get("metadata"))
    scenes = _batch_scene_ids(data, clean_text(batch_id) or None)
    base: dict[str, Any] = {
        "version": metadata_version(data),
        "rule_revision": clean_text(metadata.get("rule_revision")),
        "title": clean_text(metadata.get("title")),
        "script_lock": copy.deepcopy(as_dict(data.get("script_lock"))),
        "project_lexicon": copy.deepcopy(metadata.get("project_lexicon")),
    }
    if gate == "GATE_0":
        base["batch_plan"] = copy.deepcopy(data.get("batch_plan"))
        return base
    if gate in {"GATE_A", "GATE_B"}:
        base["batch_id"] = clean_text(batch_id)
        base["continuity_logs"] = _items_for_scenes(data.get("continuity_logs"), scenes)
        base["beats"] = _items_for_scenes(data.get("beats"), scenes)
        if uses_director_integrity_guard(data):
            base["performance_chains"] = _items_for_scenes(data.get("performance_chains"), scenes)
            constraints = [
                copy.deepcopy(item)
                for item in as_list(data.get("directorial_constraints"))
                if isinstance(item, dict)
                and (gate == "GATE_B" or clean_text(item.get("freeze_gate")) == "GATE_A")
            ]
            base["directorial_constraints"] = constraints
    if gate == "GATE_B":
        protected_shots: list[dict[str, Any]] = []
        for shot in _items_for_scenes(data.get("shots"), scenes):
            shot.pop("prompt", None)
            shot.pop("keyframe", None)
            protected_shots.append(shot)
        base["shots"] = protected_shots
        return base
    if gate == "GATE_A":
        return base
    if gate == "GATE_C":
        return {"final_content_hash": final_content_hash(data)}
    raise ValueError(f"unknown gate: {gate}")


def gate_review_hash(data: dict[str, Any], gate: str, batch_id: str | None = None) -> str:
    try:
        if clean_text(gate) == "GATE_C":
            return final_content_hash(data)
        payload = json.dumps(
            gate_review_payload(data, gate, batch_id),
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )
    except (TypeError, ValueError):
        # Public review helpers stay total for diagnostic/test construction;
        # validate_json_compatible still makes the underlying data an
        # unconditional FAIL before this sentinel can authorize anything.
        return hashlib.sha256(b"invalid-json-review-payload").hexdigest()
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def report_summary_items(data: dict[str, Any]) -> list[tuple[str, str]]:
    metadata = as_dict(data.get("metadata"))
    report = as_dict(data.get("validation_report"))
    reference_status = as_dict(metadata.get("reference_status"))
    items = [
        ("version", clean_text(metadata.get("version"))),
        ("rule_revision", clean_text(metadata.get("rule_revision"))),
        ("source_json_hash", clean_text(report.get("source_json_hash"))),
        ("status", clean_text(report.get("status"))),
        ("warnings", str(len(as_list(report.get("warnings"))))),
        ("errors", str(len(as_list(report.get("errors"))))),
        ("reference_status", json.dumps(reference_status, ensure_ascii=False, sort_keys=True)),
    ]
    if clean_text(metadata.get("version")) in HUMAN_GATE_VERSIONS:
        items.extend(
            [
                ("human_gate_audit", json.dumps(as_dict(report.get("human_gate_audit")), ensure_ascii=False, sort_keys=True)),
                ("warn_resolution_audit", json.dumps(as_dict(report.get("warn_resolution_audit")), ensure_ascii=False, sort_keys=True)),
                ("scene_id_audit", json.dumps(as_dict(report.get("scene_id_audit")), ensure_ascii=False, sort_keys=True)),
            ]
        )
    return items


def resolved_warning_ids(data: dict[str, Any]) -> set[str]:
    ids: set[str] = set()
    for item in as_list(data.get("warn_resolutions")):
        if (
            isinstance(item, dict)
            and clean_text(item.get("warn_id"))
            and not (is_current_contract(data) and clean_text(item.get("resolution")) == "revise")
        ):
            ids.add(clean_text(item.get("warn_id")))
    return ids


def human_gate_audit(data: dict[str, Any], *, final_signoff: bool = False) -> dict[str, Any]:
    if not is_current_contract(data):
        required = ["GATE_A", "GATE_B"]
        if data.get("batch_plan"):
            required.insert(0, "GATE_0")
        if final_signoff:
            required.append("GATE_C")
        approved = [gate for gate in sorted(GATES) if approved_review(data, gate)]
        missing = [gate for gate in required if gate not in approved]
        return {"required": required, "approved": approved, "missing": missing}

    batches = [
        clean_text(batch.get("batch_id"))
        for batch in as_list(as_dict(data.get("batch_plan")).get("batches"))
        if isinstance(batch, dict) and clean_text(batch.get("batch_id"))
    ]
    required_pairs: list[tuple[str, str | None]] = []
    if batches:
        required_pairs.append(("GATE_0", None))
        required_pairs.extend((gate, batch_id) for batch_id in batches for gate in ("GATE_A", "GATE_B"))
    else:
        required_pairs.extend((("GATE_A", None), ("GATE_B", None)))
    if final_signoff:
        required_pairs.append(("GATE_C", None))
    required = [f"{gate}:{batch}" if batch else gate for gate, batch in required_pairs]
    approved = [
        f"{gate}:{batch}" if batch else gate
        for gate, batch in required_pairs
        if approved_review(data, gate, batch_id=batch)
    ]
    missing = [item for item in required if item not in approved]
    return {"required": required, "approved": approved, "missing": missing}


def warn_resolution_audit(data: dict[str, Any], warnings: list[str]) -> dict[str, Any]:
    required = {warning_id(message) for message in warnings}
    resolved = resolved_warning_ids(data)
    return {"required": sorted(required), "resolved": sorted(resolved), "missing": sorted(required - resolved)}


def scene_id_audit(data: dict[str, Any]) -> dict[str, Any]:
    logs = [log for log in as_list(data.get("continuity_logs")) if isinstance(log, dict)]
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    scene_ids = [clean_text(log.get("scene_id")) for log in logs if clean_text(log.get("scene_id"))]
    shot_scene_ids = [clean_text(shot.get("scene_id")) for shot in shots if clean_text(shot.get("scene_id"))]
    return {
        "scene_ids": scene_ids,
        "shot_scene_ids": shot_scene_ids,
        "missing_in_logs": sorted(set(shot_scene_ids) - set(scene_ids)),
    }


def as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def clean_text(value: Any) -> str:
    return str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()


def is_json_integer(value: Any, *, minimum: int | None = None) -> bool:
    if isinstance(value, bool) or not isinstance(value, int):
        return False
    return minimum is None or value >= minimum


def safe_finite_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError, OverflowError):
        return None
    return number if math.isfinite(number) else None


def contract_duration_number(value: Any, *, strict_contract: bool) -> int | float:
    if strict_contract and is_json_integer(value):
        return value
    return safe_finite_number(value) or 0.0


def duration_output_value(value: int | float) -> int | float:
    if isinstance(value, int):
        return value
    return int(value) if value.is_integer() else value


def cut_rate_per_minute(shot_count: int, total_duration: int | float) -> float:
    if total_duration <= 0 or shot_count <= 1:
        return 0
    if isinstance(total_duration, int):
        return float(round((Decimal(shot_count - 1) * Decimal(60)) / Decimal(total_duration), 2))
    return round(((shot_count - 1) / total_duration) * 60, 2)


def average_duration_value(total_duration: int | float, shot_count: int) -> int | float | str:
    if shot_count <= 0:
        return 0
    if not isinstance(total_duration, int) or abs(total_duration) <= 9_007_199_254_740_991:
        return round(total_duration / shot_count, 2)
    quotient, remainder = divmod(total_duration, shot_count)
    if remainder == 0:
        return quotient
    scaled = (total_duration * 100 + shot_count // 2) // shot_count
    whole, fraction = divmod(scaled, 100)
    return f"{whole}.{fraction:02d}".rstrip("0").rstrip(".")


def validate_json_compatible(
    value: Any,
    result: ValidationResult,
    *,
    path: str = "$",
    seen: set[int] | None = None,
) -> None:
    if seen is None:
        seen = set()
    if isinstance(value, float) and not math.isfinite(value):
        suffix = "；时长字段还必须是整数" if "duration" in path or path.endswith("_seconds") else ""
        result.error(f"{path} 不得包含 NaN 或 Infinity{suffix}。")
        return
    if isinstance(value, dict):
        identity = id(value)
        if identity in seen:
            result.error(f"{path} 包含循环引用，无法作为 JSON。")
            return
        seen.add(identity)
        for key, item in value.items():
            if not isinstance(key, str):
                result.error(f"{path} 的 JSON 对象键必须是字符串。")
                continue
            validate_json_compatible(item, result, path=f"{path}.{key}", seen=seen)
        seen.remove(identity)
        return
    if isinstance(value, list):
        identity = id(value)
        if identity in seen:
            result.error(f"{path} 包含循环引用，无法作为 JSON。")
            return
        seen.add(identity)
        for index, item in enumerate(value):
            validate_json_compatible(item, result, path=f"{path}[{index}]", seen=seen)
        seen.remove(identity)
        return
    if value is not None and not isinstance(value, (str, int, float, bool)):
        result.error(f"{path} 包含非 JSON 类型：{type(value).__name__}。")


def one_line(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def fmt_number(value: Any) -> str:
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    try:
        number = float(value)
    except (TypeError, ValueError, OverflowError):
        return clean_text(value)
    if not math.isfinite(number):
        return clean_text(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:g}"


def numeric_beat_id(beat_id: str) -> int | None:
    match = re.fullmatch(r"B([0-9]{3,})", str(beat_id))
    if match is None:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        # Python protects against pathologically long integer strings.  Such
        # IDs are invalid input, but validation must remain a structured FAIL.
        return None


def parse_beat_order(value: Any) -> Decimal:
    if not isinstance(value, str) or not re.fullmatch(r"(?:0|[1-9][0-9]*)(?:\.[0-9]*[1-9])?", value):
        raise ValueError("beat_order 必须是无前导零、无尾随小数零的正数 canonical decimal string。")
    try:
        order = Decimal(value)
    except InvalidOperation as exc:
        raise ValueError("beat_order 不是合法十进制定点数。") from exc
    if not order.is_finite() or order <= 0:
        raise ValueError("beat_order 必须大于0。")
    return order


def beat_order_index(data: dict[str, Any]) -> dict[str, int]:
    return {
        clean_text(beat.get("beat_id")): index
        for index, beat in enumerate(as_list(data.get("beats")))
        if isinstance(beat, dict) and clean_text(beat.get("beat_id"))
    }


def format_beat_ids(beat_ids: list[Any], beat_positions: dict[str, int] | None = None) -> str:
    ids = [str(item) for item in beat_ids if item]
    if not ids:
        return "B000"
    numbers = [numeric_beat_id(item) for item in ids]
    positions_contiguous = True
    if beat_positions is not None:
        positions = [beat_positions.get(item) for item in ids]
        positions_contiguous = all(position is not None for position in positions) and all(
            int(current) == int(previous) + 1
            for previous, current in zip(positions, positions[1:])
            if previous is not None and current is not None
        )
    if all(number is not None for number in numbers) and positions_contiguous:
        if len(ids) == 1:
            return ids[0]
        deltas = [int(current) - int(previous) for previous, current in zip(numbers, numbers[1:])]
        if deltas and len(set(deltas)) == 1 and deltas[0] == 1:
            return f"{ids[0]}-{ids[-1]}"
    return "+".join(ids)


def source_with_beats(
    shot: dict[str, Any],
    beat_positions: dict[str, int] | None = None,
    *,
    preserve_source: bool = False,
) -> str:
    source = (
        normalize_script_text(shot.get("source_paragraph"))
        if preserve_source
        else one_line(shot.get("source_paragraph"))
    )
    return (
        f"{format_beat_ids(as_list(shot.get('beat_ids')), beat_positions)}～"
        f"{source}"
    )


def escape_cell(value: Any) -> str:
    return (
        canonical_cell_text(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("|", "&#124;")
        .replace("\n", "<br>")
    )


def unescape_cell(value: str) -> str:
    return (
        str(value)
        .replace("<br>", "\n")
        .replace("&#124;", "|")
        .replace("&gt;", ">")
        .replace("&lt;", "<")
        .replace("&amp;", "&")
    )


def canonical_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return fmt_number(value)
    return clean_text(value)


def excel_cell_value(value: Any) -> Any:
    if isinstance(value, bool):
        return canonical_cell_text(value)
    if isinstance(value, int):
        # XLSX numeric cells are IEEE-754 doubles.  Preserve larger JSON
        # integers exactly by writing their canonical decimal text.
        return value if abs(value) <= 9_007_199_254_740_991 else str(value)
    if isinstance(value, float) and math.isfinite(value):
        return int(value) if value.is_integer() else value
    return canonical_cell_text(value)


def parse_triad(camera: str) -> tuple[str, str, str]:
    first = clean_text(camera).split("\n", 1)[0].strip()
    match = re.fullmatch(r"\[([^,\]]+),\s*([^,\]]+),\s*([^\]]+)\]", first)
    if not match:
        return "平视", "中景", "固定镜头"
    return tuple(part.strip() for part in match.groups())


def is_panoramic_shot_size(shot_size: str) -> bool:
    value = clean_text(shot_size)
    return any(keyword in value for keyword in PANORAMIC_SHOT_SIZE_KEYWORDS)


def shot_size_rank(shot_size: str) -> int | None:
    value = clean_text(shot_size)
    for rank, keywords in SHOT_SIZE_RANKS:
        if any(keyword in value for keyword in keywords):
            return rank
    return None


def shot_size_has_span(shot_size: str) -> bool:
    value = clean_text(shot_size)
    return has_any_keyword(value, WIDE_SHOT_SIZE_KEYWORDS) and has_any_keyword(value, CLOSE_SHOT_SIZE_KEYWORDS)


def angle_family(angle: str) -> str:
    value = clean_text(angle)
    if "主观" in value:
        return "subjective"
    if "过肩" in value:
        return "over_shoulder"
    if "仰" in value:
        return "low"
    if "俯" in value or "高角度" in value:
        return "high"
    if "侧" in value:
        return "side"
    return "level"


def is_static_or_soft_movement(movement: str) -> bool:
    return has_any_keyword(movement, STATIC_OR_SOFT_MOVEMENT_KEYWORDS)


def depth_motion_direction(movement: str) -> str | None:
    value = clean_text(movement)
    if has_any_keyword(value, PUSH_IN_MOVEMENT_KEYWORDS):
        return "push_in"
    if has_any_keyword(value, PULL_OUT_MOVEMENT_KEYWORDS):
        return "pull_out"
    return None


def has_continuity_position_change(shot: dict[str, Any]) -> bool:
    if "【站位位移】" in clean_text(shot.get("camera_main_image")):
        return True
    for update in as_list(shot.get("continuity_updates")):
        if isinstance(update, dict) and update.get("field") in {"position", "facing"}:
            return True
    return False


def normalized_visible_characters(shot: dict[str, Any]) -> tuple[str, ...]:
    return tuple(sorted(str(item) for item in as_list(shot.get("visible_characters")) if str(item)))


def fact_types_for_shot(shot: dict[str, Any], fact_types: dict[str, str]) -> set[str]:
    return {fact_types[fact_id] for fact_id in map(str, as_list(shot.get("covered_fact_ids"))) if fact_id in fact_types}


def quoted_dialogue_text(shot: dict[str, Any]) -> str:
    camera = clean_text(shot.get("camera_main_image"))
    quoted = "".join(re.findall(r"[“\"]([^”\"]+)[”\"]", camera))
    if quoted:
        return quoted.strip()
    source = one_line(shot.get("source_paragraph"))
    match = re.search(r"[：:](.+)$", source)
    return match.group(1).strip() if match else ""


def dialogue_speaker(shot: dict[str, Any]) -> str:
    source = one_line(shot.get("source_paragraph"))
    match = re.match(r"([^：:]{1,20})[：:]", source)
    return match.group(1).strip() if match else ""


def is_expression_only_response(shot: dict[str, Any], fact_types: dict[str, str]) -> bool:
    types = fact_types_for_shot(shot, fact_types)
    if types & NEW_FACT_TYPES_FOR_CUT:
        return False
    if not (types & {"action", "emotion", "character"}):
        return False
    text = one_line(shot.get("source_paragraph")) + " " + one_line(shot.get("camera_main_image"))
    return has_any_keyword(text, EXPRESSION_ONLY_KEYWORDS)


def has_any_keyword(value: str, keywords: tuple[str, ...]) -> bool:
    text = clean_text(value)
    return any(keyword in text for keyword in keywords)


def metadata_version(data: dict[str, Any]) -> str:
    return clean_text(as_dict(data.get("metadata")).get("version"))


def version_profile(data: dict[str, Any]) -> dict[str, Any] | None:
    return VERSION_PROFILES.get(metadata_version(data))


def is_current_contract(data: dict[str, Any]) -> bool:
    profile = version_profile(data)
    return bool(profile and profile.get("contract_integrity"))


def uses_emotion_performance_guard(data: dict[str, Any]) -> bool:
    profile = version_profile(data)
    return bool(profile and profile.get("emotion_performance_guard"))


def uses_director_integrity_guard(data: dict[str, Any]) -> bool:
    profile = version_profile(data)
    return bool(profile and profile.get("director_integrity_guard"))


def descriptive_text_without_dialogue(value: Any) -> str:
    text = clean_text(value)
    text = re.sub(r"“[^”]*”", "", text, flags=re.DOTALL)
    return re.sub(r'"[^"]*"', "", text, flags=re.DOTALL)


def requires_hybrid_fields(data: dict[str, Any]) -> bool:
    profile = version_profile(data)
    return bool(profile and profile.get("hybrid"))


def requires_adjacent_motion_guard(data: dict[str, Any]) -> bool:
    profile = version_profile(data)
    return bool(profile and profile.get("adjacent_motion_guard"))


def requires_structured_audit(data: dict[str, Any]) -> bool:
    profile = version_profile(data)
    return bool(profile and profile.get("structured_audit"))


def requires_human_gate(data: dict[str, Any]) -> bool:
    profile = version_profile(data)
    return bool(profile and profile.get("human_gate"))


def requires_source_lock(data: dict[str, Any]) -> bool:
    profile = version_profile(data)
    return bool(profile and profile.get("source_lock"))


def current_rule_revision(data: dict[str, Any]) -> str:
    profile = version_profile(data)
    return clean_text(profile.get("rule_revision")) if profile else ""


def normalized_list(value: Any) -> list[str]:
    return [str(item) for item in as_list(value) if str(item)]


def scene_key_for(data: dict[str, Any], item: dict[str, Any]) -> str:
    if requires_human_gate(data):
        return clean_text(item.get("scene_id"))
    return clean_text(item.get("scene"))


def latest_review(data: dict[str, Any], gate: str, batch_id: str | None = None) -> dict[str, Any] | None:
    candidates = [
        review
        for review in as_list(data.get("human_reviews"))
        if isinstance(review, dict)
        and clean_text(review.get("gate")) == gate
        and (clean_text(review.get("batch_id")) or None) == (clean_text(batch_id) or None)
        and isinstance(review.get("round"), int)
        and not isinstance(review.get("round"), bool)
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda review: int(review["round"]))


def approved_review(
    data: dict[str, Any],
    gate: str,
    *,
    batch_id: str | None = None,
    note_contains: str | None = None,
    approved_item: str | None = None,
) -> bool:
    if is_current_contract(data):
        review = latest_review(data, gate, batch_id)
        if review is None or clean_text(review.get("status")) != "approved":
            return False
        if clean_text(review.get("reviewed_hash")) != gate_review_hash(data, gate, batch_id):
            return False
        if note_contains and note_contains not in clean_text(review.get("notes")):
            return False
        if approved_item and approved_item not in normalized_list(review.get("approved_items")):
            return False
        return True
    for review in as_list(data.get("human_reviews")):
        if not isinstance(review, dict):
            continue
        if clean_text(review.get("gate")) != gate or clean_text(review.get("status")) != "approved":
            continue
        if note_contains and note_contains not in clean_text(review.get("notes")):
            continue
        return True
    return False


def approved_safety_exception(data: dict[str, Any], shot: dict[str, Any]) -> bool:
    notes = clean_text(shot.get("notes"))
    if is_current_contract(data):
        batch_id = batch_id_for_scene(data, shot.get("scene_id"))
        return (
            clean_text(shot.get("shot_type")) == "safety"
            and "[安全镜]" in notes
            and "[人工批准]" in notes
            and approved_review(
                data,
                "GATE_B",
                batch_id=batch_id,
                approved_item=f"shot:{shot.get('shot_no')}:safety",
            )
        )
    return (
        requires_human_gate(data)
        and clean_text(shot.get("shot_type")) == "safety"
        and "[安全镜]" in notes
        and "[人工批准]" in notes
        and (approved_review(data, "GATE_B", note_contains="安全镜") or approved_review(data, "GATE_B", note_contains="safety"))
    )


def fact_lookup(data: dict[str, Any]) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        for fact in as_list(beat.get("facts")):
            if isinstance(fact, dict) and fact.get("fact_id"):
                lookup[str(fact.get("fact_id"))] = fact
    return lookup


def fact_cut_reasons(fact: dict[str, Any]) -> set[str]:
    return set(normalized_list(fact.get("cut_reasons")))


def is_fact_critical(fact: dict[str, Any]) -> bool:
    return clean_text(fact.get("cut_priority")) == "must_isolate" or bool(fact_cut_reasons(fact) & CRITICAL_SPLIT_REASONS)


def validate_project_lexicon(data: dict[str, Any], result: ValidationResult) -> None:
    if not is_current_contract(data):
        return
    metadata = as_dict(data.get("metadata"))
    if "project_lexicon" not in metadata:
        return
    lexicon = metadata.get("project_lexicon")
    if not isinstance(lexicon, dict) or not lexicon:
        result.error("metadata.project_lexicon 存在时必须是非空对象。")
        return
    unknown = sorted(
        (key for key in lexicon if not isinstance(key, str) or key not in PROJECT_LEXICON_KEYS),
        key=lambda item: str(item),
    )
    if unknown:
        result.error("metadata.project_lexicon 包含未知键：" + ", ".join(str(item) for item in unknown))
    for key, terms in lexicon.items():
        if key not in PROJECT_LEXICON_KEYS:
            continue
        if not isinstance(terms, list) or not terms:
            result.error(f"metadata.project_lexicon.{key} 必须是非空字符串数组。")
            continue
        if any(
            not isinstance(term, str)
            or not term
            or term != clean_text(term)
            or any(unicodedata.category(char) in {"Cc", "Cf", "Cs"} for char in term)
            for term in terms
        ):
            result.error(
                f"metadata.project_lexicon.{key} 每项必须是无首尾空白、无 Unicode 控制/格式/代理字符的非空 canonical 字符串。"
            )
        if len(terms) != len(set(terms)):
            result.error(f"metadata.project_lexicon.{key} 不得包含重复词条。")


def project_lexicon_patterns(data: dict[str, Any]) -> dict[str, re.Pattern[str]]:
    if not is_current_contract(data):
        return {}
    lexicon = as_dict(as_dict(data.get("metadata")).get("project_lexicon"))
    patterns: dict[str, re.Pattern[str]] = {}
    for key in PROJECT_LEXICON_KEYS:
        terms = [term for term in as_list(lexicon.get(key)) if isinstance(term, str) and term]
        if terms:
            patterns[key] = re.compile("|".join(re.escape(term) for term in terms))
    return patterns


def find_required_cut_point_categories(
    text_value: str,
    project_patterns: dict[str, re.Pattern[str]] | None = None,
    *,
    current_contract: bool = False,
) -> list[str]:
    found: list[str] = []
    project_patterns = project_patterns or {}
    for name, base_patterns, count_matches, lexicon_key in REQUIRED_CUT_POINT_CATEGORIES:
        patterns = list(base_patterns)
        lexicon_keys = (lexicon_key,) if isinstance(lexicon_key, str) else tuple(lexicon_key or ())
        for key in lexicon_keys:
            if key in project_patterns:
                patterns.append(project_patterns[key])
        output_name = "层级/声音来源变化" if current_contract and name == "层级切换" else name
        if count_matches:
            count = sum(len(pattern.findall(text_value)) for pattern in patterns)
            if count >= 2:
                found.append(output_name)
            continue
        if any(pattern.search(text_value) for pattern in patterns):
            found.append(output_name)
    return found


def extract_source_speakers(text_value: str) -> list[str]:
    forbidden_labels = {"时间", "景别", "构图", "运镜手法", "画面内容", "场景", "主体", "画面", "风格", "禁止"}
    speakers: list[str] = []
    for match in re.finditer(r"(?:^|[\n/。；;，,])\s*([\u4e00-\u9fa5A-Za-z0-9·]{1,8})(?:（VO）|\(VO\))?[：:]", clean_text(text_value)):
        speaker = match.group(1).strip()
        if speaker and speaker not in forbidden_labels and speaker not in speakers:
            speakers.append(speaker)
    return speakers


def source_marks_offscreen_voice(text_value: str, speaker: str) -> bool:
    source = clean_text(text_value)
    return bool(re.search(rf"{re.escape(speaker)}\s*(?:（VO）|\(VO\)|VO|画外声)", source))


def shot_fact_objects(shot: dict[str, Any], facts_by_id: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    facts: list[dict[str, Any]] = []
    for fact_id in normalized_list(shot.get("covered_fact_ids")):
        fact = facts_by_id.get(fact_id)
        if fact is not None:
            facts.append(fact)
    return facts


def shot_cut_groups(facts: list[dict[str, Any]], *, must_only: bool) -> dict[str, list[str]]:
    groups: dict[str, list[str]] = {}
    for fact in facts:
        if must_only and clean_text(fact.get("cut_priority")) != "must_isolate":
            continue
        group = clean_text(fact.get("cut_group"))
        if not group:
            continue
        groups.setdefault(group, []).append(str(fact.get("fact_id")))
    return groups


def fact_beat_id(fact: dict[str, Any]) -> str:
    match = re.match(r"(B[0-9]{3,})-F[0-9]{2}$", str(fact.get("fact_id", "")))
    return match.group(1) if match else ""


def fact_structural_cut_key(fact: dict[str, Any], *, include_category: bool) -> tuple[str, ...]:
    if include_category:
        return (
            fact_beat_id(fact),
            clean_text(fact.get("cut_category")),
            clean_text(fact.get("cut_moment_id")),
        )
    return (
        fact_beat_id(fact),
        clean_text(fact.get("cut_moment_id")),
    )


def structural_cut_groups(
    facts: list[dict[str, Any]],
    *,
    must_only: bool,
    include_category: bool,
) -> dict[tuple[str, ...], list[str]]:
    groups: dict[tuple[str, ...], list[str]] = {}
    for fact in facts:
        if must_only and clean_text(fact.get("cut_priority")) != "must_isolate":
            continue
        key = fact_structural_cut_key(fact, include_category=include_category)
        if not all(key):
            continue
        groups.setdefault(key, []).append(str(fact.get("fact_id")))
    return groups


def structured_key_label(key: tuple[str, ...]) -> str:
    return "/".join(key)


def strip_internal_prefix(line: str) -> str:
    value = line.strip()
    for label in INTERNAL_LABELS:
        if value.startswith(label):
            return value[len(label):].lstrip(" ：:").strip()
    return value


def build_overcompression_audit(data: dict[str, Any]) -> dict[str, Any]:
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    facts_by_id = fact_lookup(data)
    at_risk_shots: list[dict[str, Any]] = []
    total_duration: int | float = 0 if is_current_contract(data) else 0.0
    short_3 = 0
    multi_beat_critical = 0
    lexicon_patterns = project_lexicon_patterns(data)
    for shot in shots:
        duration = contract_duration_number(
            shot.get("duration_seconds"), strict_contract=is_current_contract(data)
        )
        total_duration += duration
        short_3 += int(duration <= 3)
        facts = shot_fact_objects(shot, facts_by_id)
        critical_facts = [fact for fact in facts if is_fact_critical(fact)]
        must_groups = shot_cut_groups(facts, must_only=True)
        cut_points = find_required_cut_point_categories(
            clean_text(shot.get("source_paragraph")) + "\n" + clean_text(shot.get("camera_main_image")),
            lexicon_patterns,
            current_contract=is_current_contract(data),
        )
        beat_count = len(as_list(shot.get("beat_ids")))
        has_multi_beat_critical = beat_count >= 2 and bool(critical_facts)
        multi_beat_critical += int(has_multi_beat_critical)
        if critical_facts or len(must_groups) > 1 or cut_points or beat_count >= 3 or duration >= 9:
            at_risk_shots.append(
                {
                    "shot_no": shot.get("shot_no"),
                    "duration_seconds": duration_output_value(duration),
                    "beat_count": beat_count,
                    "shot_type": shot.get("shot_type"),
                    "critical_fact_count": len(critical_facts),
                    "must_isolate_groups": sorted(must_groups),
                    "required_cut_points": cut_points,
                    "multi_beat_critical": has_multi_beat_critical,
                }
            )
    cut_per_minute = cut_rate_per_minute(len(shots), total_duration)
    short_ratio = round(short_3 / len(shots), 4) if shots else 0
    return {
        "rule_revision": current_rule_revision(data),
        "at_risk_shots": at_risk_shots,
        "multi_beat_critical_shots": multi_beat_critical,
        "cut_per_minute": cut_per_minute,
        "short_shot_ratio": short_ratio,
    }


def prompt_visual_text(camera: str, source: str) -> str:
    lines = clean_text(camera).split("\n")
    body: list[str] = []
    for index, raw in enumerate(lines):
        line = raw.strip()
        if not line:
            continue
        if index == 0 and line.startswith("["):
            continue
        if line.startswith("【机位逻辑】") or line.startswith("【场景首镜站位】"):
            continue
        if line.startswith("【站位位移】"):
            continue
        stripped = strip_internal_prefix(line)
        if stripped:
            body.append(stripped)
    text = "；".join(body) if body else one_line(source)
    return sanitize_prompt_text(text)


def sanitize_prompt_text(text: str) -> str:
    value = clean_text(text)
    for label in PROMPT_FORBIDDEN:
        value = value.replace(label, "")
    value = re.sub(r"\s*；\s*", "；", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip(" ；")


def first_sentence(text: str) -> str:
    value = sanitize_prompt_text(text)
    if not value:
        return "当前镜头可见主体动作。"
    match = re.search(r"[。！？；]", value)
    if match:
        return value[: match.end()]
    return value[:80]


def prompt_composition_text(shot: dict[str, Any]) -> str:
    camera = clean_text(shot.get("camera_main_image"))
    first_position = next(
        (strip_internal_prefix(line) for line in camera.split("\n") if line.strip().startswith("【场景首镜站位】")),
        "",
    )
    if first_position:
        return sanitize_prompt_text(first_position)
    characters = normalized_list(shot.get("visible_characters"))
    props = normalized_list(shot.get("visible_props"))
    parts: list[str] = []
    if characters:
        parts.append("可见主体：" + "、".join(characters))
    if props:
        parts.append("可见道具：" + "、".join(props))
    if parts:
        return "；".join(parts) + "。"
    return "当前镜头可见主体关系。"


def prompt_fields(prompt: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    for line in clean_text(prompt).splitlines():
        for field_name in PROMPT_FIELDS:
            if line.startswith(field_name):
                fields[field_name] = line[len(field_name):].strip()
    return fields


def quotes_are_balanced(text: str) -> bool:
    value = clean_text(text)
    return value.count("“") == value.count("”") and value.count('"') % 2 == 0


def composition_has_field_leak(value: str) -> bool:
    text = clean_text(value)
    return any(pattern.search(text) for pattern in PROMPT_COMPOSITION_FORBIDDEN_PATTERNS)


def fact_terms(text: str) -> list[str]:
    value = re.sub(r"[，。！？；：、“”\"（）()\s]", " ", clean_text(text))
    terms = re.findall(r"[A-Za-z0-9_]+|[\u4e00-\u9fff]{2,}", value)
    return [term for term in terms if term and term not in FACT_TEXT_ALLOWLIST]


def unsupported_fact_terms(fact_text: str, source_text: str) -> list[str]:
    source = clean_text(source_text)
    missing: list[str] = []
    for term in fact_terms(fact_text):
        if term not in source and term not in missing:
            missing.append(term)
    return missing


def derive_prompt(shot: dict[str, Any]) -> str:
    duration = fmt_number(shot.get("duration_seconds", 0))
    _angle, shot_size, movement = parse_triad(clean_text(shot.get("camera_main_image")))
    prompt_shot_size = re.sub(r"\s*(?:->|→)\s*", "→", shot_size)
    visual = prompt_visual_text(
        clean_text(shot.get("camera_main_image")),
        clean_text(shot.get("source_paragraph")),
    )
    composition = prompt_composition_text(shot)
    return "\n".join(
        [
            f"时间：0秒-{duration}秒",
            f"景别：{prompt_shot_size}",
            f"构图：{composition}",
            f"运镜手法：{movement}",
            f"画面内容：{visual}",
        ]
    )


def derive_prompts(data: dict[str, Any]) -> None:
    for shot in as_list(data.get("shots")):
        if isinstance(shot, dict):
            shot["prompt"] = derive_prompt(shot)


def expected_rows(data: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    positions = beat_order_index(data) if is_current_contract(data) else None
    for shot in as_list(data.get("shots")):
        if not isinstance(shot, dict):
            continue
        rows.append(
            [
                shot.get("shot_no", ""),
                shot.get("scene", ""),
                source_with_beats(shot, positions, preserve_source=is_current_contract(data)),
                shot.get("duration_seconds", ""),
                clean_text(shot.get("camera_main_image")),
                clean_text(shot.get("notes")),
                clean_text(shot.get("prompt")),
            ]
        )
    return rows


def update_validation_report(
    data: dict[str, Any], result: ValidationResult, *, final_signoff: bool = False
) -> None:
    metadata = as_dict(data.get("metadata"))
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    scene_counts: dict[str, int] = {}
    scene_logs = {
        scene_key_for(data, log): log
        for log in as_list(data.get("continuity_logs"))
        if isinstance(log, dict) and scene_key_for(data, log)
    }
    first_shot_anchor_audit: list[dict[str, Any]] = []
    seen_scenes: set[str] = set()
    total_duration: int | float = 0 if is_current_contract(data) else 0.0
    over_6 = over_8 = over_10 = 0
    short_3 = 0
    long_6_or_more = 0
    shot_type_counts: dict[str, int] = {}
    insert_priority_counts: dict[str, int] = {}
    long_take_audit: list[dict[str, Any]] = []
    for shot in shots:
        scene = scene_key_for(data, shot) or str(shot.get("scene", "")).split(" ")[0] or str(shot.get("scene", ""))
        scene_counts[scene] = scene_counts.get(scene, 0) + 1
        full_scene = scene_key_for(data, shot) or str(shot.get("scene", ""))
        if full_scene not in seen_scenes:
            seen_scenes.add(full_scene)
            _angle, shot_size, _movement = parse_triad(clean_text(shot.get("camera_main_image")))
            anchor_type = as_dict(scene_logs.get(full_scene)).get("first_shot_anchor_type", "missing")
            visible_count = len(set(normalized_list(shot.get("visible_characters"))))
            requires_panorama = anchor_type in PANORAMIC_REQUIRED_ANCHORS or (
                not is_current_contract(data)
                and visible_count >= 2
                and anchor_type != "single_continuation"
            )
            first_shot_anchor_audit.append(
                {
                    "scene": full_scene,
                    "shot_no": shot.get("shot_no"),
                    "anchor_type": anchor_type,
                    "shot_size": shot_size,
                    "visible_character_count": visible_count,
                    "requires_panoramic_first_shot": requires_panorama,
                    "passed": (not requires_panorama) or is_panoramic_shot_size(shot_size),
                }
            )
        duration = contract_duration_number(
            shot.get("duration_seconds"), strict_contract=is_current_contract(data)
        )
        total_duration += duration
        over_6 += int(duration > 6)
        over_8 += int(duration > 8)
        over_10 += int(duration > 10)
        short_3 += int(duration <= 3)
        long_6_or_more += int(duration >= 6)
        shot_type = clean_text(shot.get("shot_type")) or "missing"
        insert_priority = clean_text(shot.get("insert_priority")) or "missing"
        shot_type_counts[shot_type] = shot_type_counts.get(shot_type, 0) + 1
        insert_priority_counts[insert_priority] = insert_priority_counts.get(insert_priority, 0) + 1
        if duration > 10:
            long_take_audit.append(
                {
                    "shot_no": shot.get("shot_no"),
                    "duration_seconds": duration_output_value(duration),
                    "long_take_support": normalized_list(shot.get("long_take_support")),
                }
            )
    cut_per_minute = cut_rate_per_minute(len(shots), total_duration)
    short_ratio = round(short_3 / len(shots), 4) if shots else 0
    long_ratio = round(long_6_or_more / len(shots), 4) if shots else 0
    hybrid_audit = {
        "short_shots_le_3_seconds": short_3,
        "short_shot_ratio": short_ratio,
        "long_shots_ge_6_seconds": long_6_or_more,
        "long_shot_ratio": long_ratio,
        "cut_per_minute": cut_per_minute,
        "shot_type_counts": shot_type_counts,
        "insert_priority_counts": insert_priority_counts,
        "long_take_audit": long_take_audit,
        "soft_targets": {
            "short_shot_ratio": "0.10-0.18 for full-episode references only",
            "long_shot_ratio": "0.40-0.55 for full-episode references only",
            "cut_per_minute": "10.5-12.5 for full-episode references only",
        },
    }
    director_integrity_audit = {
        "spatial_axis_count": sum(
            len(as_list(log.get("spatial_axes")))
            for log in as_list(data.get("continuity_logs"))
            if isinstance(log, dict)
        ),
        "performance_chain_count": len(as_list(data.get("performance_chains"))),
        "directorial_constraint_count": len(as_list(data.get("directorial_constraints"))),
        "shots_with_camera_geometry": sum(isinstance(shot.get("camera_geometry"), dict) for shot in shots),
        "shots_with_performance_cues": sum(bool(as_list(shot.get("performance_cues"))) for shot in shots),
        "shots_with_action_progressions": sum(bool(as_list(shot.get("action_progressions"))) for shot in shots),
        "shots_with_environment_cues": sum(bool(as_list(shot.get("environment_cues"))) for shot in shots),
        "source_reuse_count": sum(isinstance(shot.get("source_reuse"), dict) for shot in shots),
        "rhythm": {
            "cut_per_minute": cut_per_minute,
            "short_shot_ratio": short_ratio,
            "long_shot_ratio": long_ratio,
        },
    }
    data["validation_report"] = {
        "status": result.status,
        "warnings": result.warnings,
        "warning_items": warning_items(result.warnings),
        "errors": result.errors,
        "version": clean_text(metadata.get("version")),
        "rule_revision": clean_text(metadata.get("rule_revision")),
        "source_json_hash": canonical_data_hash(data),
        "reference_status": as_dict(metadata.get("reference_status")),
        "warning_count": len(result.warnings),
        "error_count": len(result.errors),
        "warning_budget": {
            "status": "WARN is deliverable only when every warning is explained or accepted by the user",
            "unresolved_warning_count": len(warn_resolution_audit(data, result.warnings)["missing"]),
        },
        "shot_count": len(shots),
        "scene_counts": scene_counts,
        "total_duration_seconds": duration_output_value(total_duration),
        "average_duration_seconds": average_duration_value(total_duration, len(shots)),
        "shots_over_6_seconds": over_6,
        "shots_over_8_seconds": over_8,
        "shots_over_10_seconds": over_10,
        "column_contract": "7-column stable storyboard table; keyframe column removed",
        "first_shot_anchor_audit": first_shot_anchor_audit,
        "hybrid_audit": hybrid_audit,
        "overcompression_audit": build_overcompression_audit(data),
    }
    if uses_director_integrity_guard(data):
        data["validation_report"]["director_integrity_audit"] = director_integrity_audit
    if requires_human_gate(data):
        data["validation_report"]["human_gate_audit"] = human_gate_audit(data, final_signoff=final_signoff)
        data["validation_report"]["warn_resolution_audit"] = warn_resolution_audit(data, result.warnings)
        data["validation_report"]["scene_id_audit"] = scene_id_audit(data)


def validate_metadata(
    data: dict[str, Any],
    result: ValidationResult,
    *,
    strict_status: bool,
    final_signoff: bool,
) -> None:
    metadata = as_dict(data.get("metadata"))
    if metadata.get("skill_name") != "su-fenjingskill-zh":
        result.error("metadata.skill_name 必须为 su-fenjingskill-zh。")
    version = metadata_version(data)
    profile = VERSION_PROFILES.get(version)
    if profile is None:
        result.error(f"metadata.version 缺失或不受支持：{version or '<empty>'}。")
    expected_rule_revision = clean_text(profile.get("rule_revision")) if profile else ""
    if expected_rule_revision and clean_text(metadata.get("rule_revision")) != expected_rule_revision:
        result.error(f"metadata.rule_revision 必须为 {expected_rule_revision}。")
    if profile is not None and not expected_rule_revision and clean_text(metadata.get("rule_revision")):
        result.error(f"metadata.version={version} 的历史合同不得填写未知 rule_revision。")
    reference_status = as_dict(metadata.get("reference_status"))
    if "reference_status" not in metadata:
        if requires_human_gate(data):
            result.error("metadata.reference_status 缺失。")
        else:
            result.warn("metadata.reference_status 缺失，建议记录 reference loaded/missing 状态。")
    if requires_human_gate(data):
        for key in REQUIRED_REFERENCE_KEYS:
            status = clean_text(reference_status.get(key))
            if status not in {"loaded", "missing"}:
                result.error(f"metadata.reference_status.{key} 必须为 loaded / missing。")
        reference_proof = as_dict(metadata.get("reference_proof"))
        for key, status in reference_status.items():
            if status == "loaded" and not clean_text(reference_proof.get(key)):
                result.error(f"metadata.reference_proof.{key} 必须记录已读文件首个标题行。")
        script_status = as_dict(metadata.get("script_status"))
        for key in REQUIRED_SCRIPT_KEYS:
            status = clean_text(script_status.get(key))
            if status not in {"available", "missing"}:
                result.error(f"metadata.script_status.{key} 必须为 available / missing。")
        if is_current_contract(data):
            script_dir = Path(__file__).resolve().parent
            actual_scripts = {
                "storyboard_delivery.py": Path(__file__).resolve().is_file(),
                "validate_storyboard.js": (script_dir / "validate_storyboard.js").is_file(),
            }
            for key, exists in actual_scripts.items():
                expected_status = "available" if exists else "missing"
                if clean_text(script_status.get(key)) != expected_status:
                    result.error(
                        f"metadata.script_status.{key} 与实际文件状态不一致；必须为 {expected_status}。"
                    )
            if not actual_scripts["storyboard_delivery.py"]:
                result.error("storyboard_delivery.py builder 缺失，禁止生成交付物。")
        if not isinstance(metadata.get("revision_log"), list):
            result.error("metadata.revision_log 必须是数组。")
        title = clean_text(metadata.get("title"))
        if title and not re.fullmatch(r"[\w\u4e00-\u9fff-]+", title):
            result.error("metadata.title 只能包含中英文、数字、下划线和连字符。")
    report = as_dict(data.get("validation_report"))
    status = report.get("status")
    if strict_status and status is not None and status not in VALID_STATUSES:
        result.error("validation_report.status 只允许 PASS / WARN / FAIL / NOT_RUN。")
    if requires_human_gate(data) and clean_text(status) == "NOT_RUN":
        if clean_text(report.get("source_json_hash")):
            result.error("validation_report.status 为 NOT_RUN 时 source_json_hash 必须为空字符串。")
        if not approved_review(data, "GATE_C", note_contains="accepted_without_validation"):
            result.error("NOT_RUN 交付必须在 Gate C 记录 accepted_without_validation。")
    if strict_status and requires_structured_audit(data) and clean_text(status) != "NOT_RUN":
        supplied_hash = clean_text(report.get("source_json_hash"))
        if is_current_contract(data) and not re.fullmatch(r"[0-9a-f]{64}", supplied_hash):
            result.error("validation_report.source_json_hash 必须是当前内容的 64 位小写 SHA-256。")
        elif supplied_hash and supplied_hash != canonical_data_hash(data):
            result.error("validation_report.source_json_hash 与当前 shot_data 内容不一致。")


def validate_batch_plan(data: dict[str, Any], result: ValidationResult) -> dict[str, set[str]]:
    """Validate strict-contract batch ownership and return batch -> scene IDs."""

    if not is_current_contract(data):
        return {}
    raw_plan = data.get("batch_plan")
    if raw_plan is None:
        return {}
    if not isinstance(raw_plan, dict):
        result.error("batch_plan 必须是对象或 null。")
        return {}
    raw_batches = raw_plan.get("batches")
    if not isinstance(raw_batches, list) or not raw_batches:
        result.error("batch_plan.batches 必须是非空数组；空计划不得触发 Gate 0。")
        return {}
    batches: dict[str, set[str]] = {}
    scene_owner: dict[str, str] = {}
    dependencies: dict[str, list[str]] = {}
    for index, batch in enumerate(raw_batches, start=1):
        if not isinstance(batch, dict):
            result.error(f"batch_plan.batches #{index} 必须是对象。")
            continue
        raw_batch_id = batch.get("batch_id")
        batch_id = clean_text(raw_batch_id)
        if (
            not isinstance(raw_batch_id, str)
            or raw_batch_id != batch_id
            or not re.fullmatch(BATCH_ID_PATTERN, batch_id)
        ):
            result.error(f"batch_plan.batches #{index} batch_id 必须符合 BT[0-9]{{2,}} 格式。")
            continue
        if batch_id in batches:
            result.error(f"batch_plan batch_id 重复：{batch_id}")
        raw_scenes = batch.get("scene_ids")
        if not isinstance(raw_scenes, list) or not raw_scenes:
            result.error(f"{batch_id} scene_ids 必须是非空数组。")
            scene_ids: list[str] = []
        else:
            scene_ids = [clean_text(scene) for scene in raw_scenes]
            if any(
                not isinstance(raw_scene, str)
                or raw_scene != scene_id
                or not re.fullmatch(SCENE_ID_PATTERN, scene_id)
                for raw_scene, scene_id in zip(raw_scenes, scene_ids)
            ):
                result.error(f"{batch_id} scene_ids 每项必须是 S[0-9]{{2,}} 格式字符串。")
            if len(scene_ids) != len(set(scene_ids)):
                result.error(f"{batch_id} scene_ids 不得重复。")
        batches[batch_id] = set(scene_ids)
        for scene_id in scene_ids:
            previous = scene_owner.get(scene_id)
            if previous and previous != batch_id:
                result.error(f"场景 {scene_id} 同时属于 {previous} 与 {batch_id}。")
            scene_owner[scene_id] = batch_id
        expected = batch.get("expected_shot_count")
        if not is_json_integer(expected, minimum=1):
            result.error(f"{batch_id} expected_shot_count 必须是大于0的整数。")
        raw_depends = batch.get("depends_on")
        if not isinstance(raw_depends, list):
            result.error(f"{batch_id} depends_on 必须是数组。")
            dependencies[batch_id] = []
        else:
            depends = [clean_text(item) for item in raw_depends]
            if any(
                not isinstance(raw_dependency, str)
                or raw_dependency != dependency
                or not re.fullmatch(BATCH_ID_PATTERN, dependency)
                for raw_dependency, dependency in zip(raw_depends, depends)
            ):
                result.error(f"{batch_id} depends_on 每项必须是 BT[0-9]{{2,}} 格式字符串。")
            if len(depends) != len(set(depends)):
                result.error(f"{batch_id} depends_on 不得重复。")
            dependencies[batch_id] = depends
    for batch_id, depends in dependencies.items():
        for dependency in depends:
            if dependency not in batches:
                result.error(f"{batch_id} depends_on 指向不存在批次：{dependency}")
            if dependency == batch_id:
                result.error(f"{batch_id} depends_on 不得自指。")
    visiting: set[str] = set()
    visited: set[str] = set()

    def visit(batch_id: str) -> None:
        if batch_id in visiting:
            result.error(f"batch_plan.depends_on 存在循环：{batch_id}")
            return
        if batch_id in visited:
            return
        visiting.add(batch_id)
        for dependency in dependencies.get(batch_id, []):
            if dependency in batches:
                visit(dependency)
        visiting.remove(batch_id)
        visited.add(batch_id)

    for batch_id in batches:
        visit(batch_id)
    registered_scenes = {
        clean_text(item.get("scene_id"))
        for collection in (data.get("continuity_logs"), data.get("beats"), data.get("shots"))
        for item in as_list(collection)
        if isinstance(item, dict) and clean_text(item.get("scene_id"))
    }
    missing = sorted(registered_scenes - set(scene_owner))
    if missing:
        result.error("batch_plan 未覆盖场景：" + ", ".join(missing))
    declared_scenes = set(scene_owner)
    log_scenes = {
        clean_text(item.get("scene_id"))
        for item in as_list(data.get("continuity_logs"))
        if isinstance(item, dict) and clean_text(item.get("scene_id"))
    }
    beat_scenes = {
        clean_text(item.get("scene_id"))
        for item in as_list(data.get("beats"))
        if isinstance(item, dict) and clean_text(item.get("scene_id"))
    }
    shot_scenes = {
        clean_text(item.get("scene_id"))
        for item in as_list(data.get("shots"))
        if isinstance(item, dict) and clean_text(item.get("scene_id"))
    }
    for label, present in (("continuity_logs", log_scenes), ("beats", beat_scenes), ("shots", shot_scenes)):
        absent = sorted(declared_scenes - present)
        if absent:
            result.error(f"batch_plan 声明的场景未被 {label} 覆盖：{', '.join(absent)}")
    return batches


def validate_human_reviews(data: dict[str, Any], result: ValidationResult, *, final_signoff: bool) -> None:
    if not requires_human_gate(data):
        return
    reviews = as_list(data.get("human_reviews"))
    if not isinstance(data.get("human_reviews"), list):
        result.error("2.4.0 数据必须包含 human_reviews 数组。")
        reviews = []
    current = is_current_contract(data)
    batch_ids = {
        clean_text(batch.get("batch_id"))
        for batch in as_list(as_dict(data.get("batch_plan")).get("batches"))
        if isinstance(batch, dict) and clean_text(batch.get("batch_id"))
    }
    last_round: dict[tuple[str, str | None], int] = {}
    seen_approved: set[tuple[str, str | None]] = set()
    latest_records: dict[tuple[str, str | None], tuple[int, dict[str, Any]]] = {}
    for review_index, review in enumerate(reviews):
        if not isinstance(review, dict):
            result.error("human_reviews 每项必须是对象。")
            continue
        raw_gate = review.get("gate")
        gate = clean_text(raw_gate)
        if gate not in GATES:
            result.error("human_reviews.gate 必须为 GATE_0 / GATE_A / GATE_B / GATE_C。")
        if current and (not isinstance(raw_gate, str) or raw_gate != gate):
            result.error("human_reviews.gate 必须是无首尾空白的 canonical 字符串。")
        raw_status = review.get("status")
        status = clean_text(raw_status)
        if status not in {"approved", "rejected"}:
            result.error("human_reviews.status 必须为 approved / rejected。")
        if current and (not isinstance(raw_status, str) or raw_status != status):
            result.error("human_reviews.status 必须是无首尾空白的 canonical 字符串。")
        round_value = review.get("round")
        if not is_json_integer(round_value, minimum=1):
            result.error("human_reviews.round 必须为从 1 开始的整数。")
        if not clean_text(review.get("reviewer")) or (current and not isinstance(review.get("reviewer"), str)):
            result.error("human_reviews.reviewer 不能为空。")
        if not isinstance(review.get("notes"), str) or not clean_text(review.get("notes")):
            result.error("human_reviews.notes 必须是非空字符串。")
        if not current or gate not in GATES or not is_json_integer(round_value, minimum=1):
            continue
        raw_batch_id = review.get("batch_id")
        batch_id = clean_text(raw_batch_id) or None
        if batch_id is not None and (
            not isinstance(raw_batch_id, str)
            or raw_batch_id != batch_id
            or not re.fullmatch(BATCH_ID_PATTERN, batch_id)
        ):
            result.error(f"{gate} batch_id 必须是 BT[0-9]{{2,}} 格式字符串。")
        if gate == "GATE_0" and not batch_ids:
            result.error("仅 batch_plan.batches 非空时允许 GATE_0。")
        if gate in {"GATE_0", "GATE_C"} and batch_id is not None:
            result.error(f"{gate} 必须是全局审核，不得填写 batch_id。")
        if gate in {"GATE_A", "GATE_B"}:
            if batch_ids and batch_id not in batch_ids:
                result.error(f"{gate} 必须绑定已登记 batch_id。")
            if not batch_ids and batch_id is not None:
                result.error(f"非分批任务的 {gate} 不得填写 batch_id。")
        key = (gate, batch_id)
        latest_records[key] = (review_index, review)
        previous = last_round.get(key, 0)
        if int(round_value) != previous + 1:
            result.error(
                f"human_reviews {gate}{':' + batch_id if batch_id else ''} round 必须按 1 开始连续追加；"
                f"期望 {previous + 1}，实际 {round_value}。"
            )
        last_round[key] = max(previous, int(round_value))
        supplied_hash = clean_text(review.get("reviewed_hash"))
        if not re.fullmatch(r"[0-9a-f]{64}", supplied_hash):
            result.error(f"{gate} round {round_value} reviewed_hash 必须是 64 位小写 SHA-256。")
        approved_items = review.get("approved_items", [])
        if "approved_items" not in review:
            result.error(f"{gate} round {round_value} 必须显式包含 approved_items 数组，可为空数组。")
        if not isinstance(approved_items, list) or any(
            not isinstance(item, str) or not clean_text(item) for item in as_list(approved_items)
        ):
            result.error(f"{gate} round {round_value} approved_items 必须是非空字符串数组或空数组。")
        elif any(
            not re.fullmatch(
                (
                    r"shot:[1-9][0-9]*:(?:safety|axis-reversal|required-adjacent|jump-cut|axis-cross|performance-chain-break|source-reuse)"
                    if uses_director_integrity_guard(data)
                    else r"shot:[1-9][0-9]*:(?:safety|axis-reversal|required-adjacent)"
                ),
                item,
            )
            for item in approved_items
        ):
            result.error(
                f"{gate} round {round_value} approved_items 含非法逐镜例外 token。"
            )
        elif len(approved_items) != len(set(approved_items)):
            result.error(f"{gate} round {round_value} approved_items 不得重复。")
        elif gate != "GATE_B" and approved_items:
            result.error(f"{gate} 不得填写 approved_items；只有 GATE_B 可批准逐镜例外。")
        if gate == "GATE_B" and status == "approved" and ("GATE_A", batch_id) not in seen_approved:
            result.error(f"{gate} 不得在对应 GATE_A approved 之前批准。")
        if gate == "GATE_A" and batch_ids and status == "approved" and ("GATE_0", None) not in seen_approved:
            result.error("分批任务的 GATE_A 不得在全局 GATE_0 approved 之前批准。")
        if gate == "GATE_C" and status == "approved":
            required_b = {("GATE_B", batch) for batch in batch_ids} if batch_ids else {("GATE_B", None)}
            if not required_b.issubset(seen_approved):
                result.error("GATE_C 不得在全部 GATE_B approved 之前批准。")
        if status == "approved":
            seen_approved.add(key)
        elif status == "rejected":
            seen_approved.discard(key)
    # Historical records retain the hash of the content seen at that time.
    # Only the latest record for each gate/batch must bind the current content.
    for (gate, batch_id), (_review_index, review) in latest_records.items():
        supplied_hash = clean_text(review.get("reviewed_hash"))
        if re.fullmatch(r"[0-9a-f]{64}", supplied_hash):
            if supplied_hash != gate_review_hash(data, gate, batch_id):
                result.error(
                    f"{gate}{':' + batch_id if batch_id else ''} 最新 round reviewed_hash 与当前审核范围不一致。"
                )

    def latest_approved(key: tuple[str, str | None]) -> bool:
        record = latest_records.get(key)
        return bool(record and clean_text(record[1].get("status")) == "approved")

    def require_latest_after(
        prerequisite: tuple[str, str | None],
        downstream: tuple[str, str | None],
    ) -> None:
        downstream_record = latest_records.get(downstream)
        if not downstream_record or clean_text(downstream_record[1].get("status")) != "approved":
            return
        prerequisite_record = latest_records.get(prerequisite)
        if (
            not latest_approved(prerequisite)
            or prerequisite_record is None
            or prerequisite_record[0] >= downstream_record[0]
        ):
            prerequisite_label = f"{prerequisite[0]}:{prerequisite[1]}" if prerequisite[1] else prerequisite[0]
            downstream_label = f"{downstream[0]}:{downstream[1]}" if downstream[1] else downstream[0]
            result.error(
                f"{downstream_label} 最新 approved 必须晚于 {prerequisite_label} 最新 approved；"
                "上游 Gate 新增任何轮次后，旧的下游批准失效。"
            )

    if batch_ids:
        for batch_id in sorted(batch_ids):
            require_latest_after(("GATE_0", None), ("GATE_A", batch_id))
            require_latest_after(("GATE_A", batch_id), ("GATE_B", batch_id))
    else:
        require_latest_after(("GATE_A", None), ("GATE_B", None))
    for batch_id in sorted(batch_ids) if batch_ids else [None]:
        require_latest_after(("GATE_B", batch_id), ("GATE_C", None))
    audit = human_gate_audit(data, final_signoff=final_signoff)
    for gate in audit["missing"]:
        result.error(f"{metadata_version(data)} 缺少有效的 {gate} approved 人工审核记录。")


def script_fragment_fingerprint(value: str) -> str:
    return re.sub(r"\s+", "", clean_text(value))


def legacy_script_text_hash(value: str) -> str:
    return hashlib.sha256(script_fragment_fingerprint(value).encode("utf-8")).hexdigest()


def script_text_hash(value: str) -> str:
    """Backward-compatible name for the 2.4.1/2.4.2 hash algorithm."""

    return legacy_script_text_hash(value)


def normalize_script_text(value: Any) -> str:
    text = str(value) if value is not None else ""
    if text.startswith("\ufeff"):
        text = text[1:]
    return text.replace("\r\n", "\n").replace("\r", "\n")


def normalized_script_text_hash(value: Any) -> str:
    return hashlib.sha256(normalize_script_text(value).encode("utf-8")).hexdigest()


def raw_script_text(value: Any) -> str:
    return normalize_script_text(value)


def _discover_workspace_root(data_path: Path | None) -> Path:
    starts = []
    if data_path is not None:
        starts.append(data_path.resolve(strict=False).parent)
    starts.append(Path.cwd().resolve(strict=False))
    for start in starts:
        for candidate in (start, *start.parents):
            if (candidate / ".git").exists() or (candidate / "AGENTS.md").is_file():
                return candidate
    return starts[0]


def resolve_approved_script_path(
    declared: Any,
    data_path: Path | None,
    workspace_root: Path | None,
) -> Path:
    text = str(declared or "")
    if not text:
        raise ValueError("script_lock.approved_script_path 不能为空。")
    relative = Path(text)
    if relative.is_absolute() or relative.drive:
        raise ValueError("script_lock.approved_script_path 必须是工作区内相对路径。")
    root = (workspace_root or _discover_workspace_root(data_path)).resolve(strict=False)
    candidate = (root / relative).resolve(strict=False)
    try:
        candidate.relative_to(root)
    except ValueError:
        raise ValueError("script_lock.approved_script_path 不得越出工作区。")
    if candidate.is_file():
        return candidate
    raise ValueError(f"script_lock.approved_script_path 文件不存在或不是普通文件：{text}")


def normalize_source_spans(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, dict):
        return [value]
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    return []


def extract_script_span_text(
    locked_text: str,
    spans: list[dict[str, Any]],
    result: ValidationResult,
    label: str,
    *,
    strict_contract: bool,
) -> str | None:
    parts: list[str] = []
    previous_end: int | None = None
    seen_ranges: set[tuple[int, int]] = set()
    for index, span in enumerate(spans, start=1):
        if strict_contract:
            if not is_json_integer(span.get("start")) or not is_json_integer(span.get("end")):
                result.error(f"{label} source_span #{index} must contain JSON integer start/end.")
                return None
            start = int(span["start"])
            end = int(span["end"])
        else:
            try:
                start = int(span.get("start"))
                end = int(span.get("end"))
            except (TypeError, ValueError):
                result.error(f"{label} source_span #{index} must contain integer start/end.")
                return None
        if (start, end) in seen_ranges:
            result.error(f"{label} source_span #{index} duplicates an earlier range.")
            return None
        seen_ranges.add((start, end))
        if strict_contract and previous_end is not None and start < previous_end:
            result.error(f"{label} source_spans must be ordered and non-overlapping.")
            return None
        previous_end = end
        if start < 0 or end <= start or end > len(locked_text):
            result.error(f"{label} source_span #{index} is outside script_lock.locked_text.")
            return None
        extracted = locked_text[start:end]
        expected_hash = clean_text(span.get("text_hash"))
        actual_hash = normalized_script_text_hash(extracted) if strict_contract else legacy_script_text_hash(extracted)
        if expected_hash and expected_hash != actual_hash:
            result.error(f"{label} source_span #{index} text_hash does not match locked text.")
        parts.append(extracted)
    return "\n".join(parts)


def validate_source_text_against_lock(
    locked_text: str,
    item: dict[str, Any],
    text_key: str,
    result: ValidationResult,
    label: str,
    *,
    strict_contract: bool,
) -> None:
    if strict_contract:
        has_single = "source_span" in item
        has_multiple = "source_spans" in item
        if has_single and has_multiple:
            result.error(f"{label} source_span 与 source_spans 互斥，不得同时存在。")
            return
        if has_single:
            if not isinstance(item.get("source_span"), dict):
                result.error(f"{label} source_span 必须是对象。")
                return
            spans = [item["source_span"]]
        elif has_multiple:
            raw_spans = item.get("source_spans")
            if not isinstance(raw_spans, list) or not raw_spans or any(not isinstance(span, dict) for span in raw_spans):
                result.error(f"{label} source_spans 必须是非空对象数组。")
                return
            spans = raw_spans
        else:
            spans = []
    else:
        spans = normalize_source_spans(item.get("source_span")) or normalize_source_spans(item.get("source_spans"))
    if not spans:
        result.error(f"{label} must include source_span or source_spans for script-lock validation.")
        return
    extracted = extract_script_span_text(locked_text, spans, result, label, strict_contract=strict_contract)
    if extracted is None:
        return
    declared = normalize_script_text(item.get(text_key)) if strict_contract else clean_text(item.get(text_key))
    if not declared:
        result.error(f"{label} {text_key} must not be empty.")
        return
    matches = declared == extracted if strict_contract else script_fragment_fingerprint(declared) == script_fragment_fingerprint(extracted)
    if not matches:
        result.error(f"{label} {text_key} is not an exact extract from script_lock.locked_text.")


def validate_script_lock(
    data: dict[str, Any],
    result: ValidationResult,
    *,
    data_path: Path | None = None,
    workspace_root: Path | None = None,
) -> None:
    if not requires_source_lock(data):
        return
    lock = as_dict(data.get("script_lock"))
    if not lock:
        result.error("2.4.1+ data must include top-level script_lock.")
        return
    if clean_text(lock.get("status")) != "locked":
        result.error('script_lock.status must be "locked".')
    strict_contract = is_current_contract(data)
    if strict_contract and not isinstance(lock.get("locked_text"), str):
        result.error("script_lock.locked_text 必须是字符串。")
    locked_text = normalize_script_text(lock.get("locked_text"))
    if not locked_text or (strict_contract and not clean_text(locked_text)):
        result.error("script_lock.locked_text must contain the full human-approved script text.")
        return
    expected_hash = normalized_script_text_hash(locked_text) if strict_contract else legacy_script_text_hash(locked_text)
    supplied_hash = clean_text(lock.get("locked_text_hash"))
    if supplied_hash != expected_hash:
        result.error("script_lock.locked_text_hash does not match locked_text.")
    corrections = lock.get("approved_corrections", [])
    if not isinstance(corrections, list):
        result.error("script_lock.approved_corrections must be an array when present.")
    elif strict_contract:
        for index, correction in enumerate(corrections, start=1):
            if not isinstance(correction, dict):
                result.error(f"approved_corrections #{index} 必须是对象。")
                continue
            source = normalize_script_text(correction.get("from"))
            target = normalize_script_text(correction.get("to"))
            if not source or not target or source == target:
                result.error(f"approved_corrections #{index} from/to 必须非空且不同。")
            elif target not in locked_text:
                result.error(f"approved_corrections #{index} to 未出现在批准剧本。")
    if metadata_version(data) in APPROVED_SCRIPT_PATH_VERSIONS:
        declared_path = lock.get("approved_script_path")
        if not clean_text(declared_path):
            result.error(f"{metadata_version(data)} data must include script_lock.approved_script_path.")
        elif strict_contract and not isinstance(declared_path, str):
            result.error("script_lock.approved_script_path 必须是工作区相对路径字符串。")
        elif strict_contract:
            try:
                approved_path = resolve_approved_script_path(declared_path, data_path, workspace_root)
                approved_text = normalize_script_text(approved_path.read_text(encoding="utf-8-sig"))
            except (OSError, UnicodeError, ValueError) as exc:
                result.error(str(exc))
            else:
                if approved_text != locked_text:
                    result.error("script_lock.locked_text 必须与 approved_script_path 文件全文一致。")
    for beat in as_list(data.get("beats")):
        if isinstance(beat, dict):
            label = clean_text(beat.get("beat_id")) or "beat"
            validate_source_text_against_lock(
                locked_text, beat, "source_text", result, label, strict_contract=strict_contract
            )
    for shot in as_list(data.get("shots")):
        if isinstance(shot, dict):
            label = f"shot {shot.get('shot_no')}"
            validate_source_text_against_lock(
                locked_text, shot, "source_paragraph", result, label, strict_contract=strict_contract
            )


def is_auto_whitelist_warning(data: dict[str, Any], message: str) -> bool:
    base_tokens = ("reference missing", "[reference missing]", "[合理补足]")
    if uses_director_integrity_guard(data):
        return any(token in message for token in base_tokens)
    return any(token in message for token in (*base_tokens, "节奏", "cut/min", "短镜比例"))


def validate_warn_resolutions(
    data: dict[str, Any], result: ValidationResult, *, require_all: bool = True
) -> None:
    if not requires_human_gate(data):
        return
    resolutions = as_list(data.get("warn_resolutions"))
    if not isinstance(data.get("warn_resolutions"), list):
        result.error("2.4.0 数据必须包含 warn_resolutions 数组。")
        resolutions = []
    by_id: dict[str, dict[str, Any]] = {}
    for item in resolutions:
        if not isinstance(item, dict):
            result.error("warn_resolutions 每项必须是对象。")
            continue
        warn_id_value = clean_text(item.get("warn_id"))
        if not warn_id_value:
            result.error("warn_resolutions.warn_id 不能为空。")
            continue
        if warn_id_value in by_id and is_current_contract(data):
            result.error(f"warn_resolutions.warn_id 重复：{warn_id_value}")
        by_id[warn_id_value] = item
        if clean_text(item.get("resolution")) not in {"keep", "revise", "accepted_without_change"}:
            result.error(f"{warn_id_value} resolution 必须为 keep / revise / accepted_without_change。")
        if clean_text(item.get("resolved_by")) not in {"human", "auto_whitelist"}:
            result.error(f"{warn_id_value} resolved_by 必须为 human / auto_whitelist。")
        if not clean_text(item.get("note")):
            result.error(f"{warn_id_value} note 不能为空。")
    for message in result.warnings:
        warn_id_value = warning_id(message)
        item = by_id.get(warn_id_value)
        if item is None:
            if require_all:
                result.error(f"WARN 缺少处置记录：{warn_id_value}。")
            continue
        if is_current_contract(data) and clean_text(item.get("resolution")) == "revise":
            if require_all:
                result.error(f"{warn_id_value} 当前 WARN 仍存在，resolution=revise 不能视为已解决。")
            continue
        resolved_by = clean_text(item.get("resolved_by"))
        if resolved_by == "auto_whitelist" and not is_auto_whitelist_warning(data, message):
            result.error(f"{warn_id_value} 不是白名单 WARN，必须由 human 处置。")


def validate_report_truth(
    data: dict[str, Any],
    result: ValidationResult,
    *,
    expected_errors: list[str],
    expected_warnings: list[str],
    final_signoff: bool,
) -> None:
    report_value = data.get("validation_report")
    if not isinstance(report_value, dict):
        result.error("strict validate 要求 validation_report 为对象。")
        return
    report = report_value
    status = clean_text(report.get("status"))
    if status not in VALID_STATUSES:
        result.error("validation_report.status 只允许 PASS / WARN / FAIL / NOT_RUN。")
        return
    fresh = copy.deepcopy(data)
    update_validation_report(
        fresh,
        ValidationResult(errors=list(expected_errors), warnings=list(expected_warnings)),
        final_signoff=final_signoff,
    )
    expected_report = as_dict(fresh.get("validation_report"))
    if status == "NOT_RUN":
        expected_report["status"] = "NOT_RUN"
        expected_report["source_json_hash"] = ""
    for key in sorted(set(report) | set(expected_report)):
        if report.get(key) != expected_report.get(key):
            result.error(f"validation_report.{key} 与本次真实校验/统计结果不一致。")


def normalize_json_pointer(value: Any) -> str:
    if not isinstance(value, str) or not value:
        raise ValueError("JSON Pointer 必须是非空字符串。")
    if not value.startswith("/"):
        if "/" in value or "." in value:
            raise ValueError(f"旧式继承字段只能是单段顶层名：{value}")
        return "/" + value.replace("~", "~0").replace("/", "~1")
    if re.search(r"~(?![01])", value):
        raise ValueError(f"JSON Pointer 含非法 ~ 转义：{value}")
    tokens = [token.replace("~1", "/").replace("~0", "~") for token in value[1:].split("/")]
    return "/" + "/".join(token.replace("~", "~0").replace("/", "~1") for token in tokens)


def json_pointer_tokens(pointer: Any) -> tuple[str, ...]:
    normalized = normalize_json_pointer(pointer)
    return tuple(token.replace("~1", "/").replace("~0", "~") for token in normalized[1:].split("/"))


def resolve_json_pointer(document: Any, pointer: Any) -> Any:
    current = document
    for token in json_pointer_tokens(pointer):
        if isinstance(current, dict):
            if token not in current:
                raise KeyError(normalize_json_pointer(pointer))
            current = current[token]
            continue
        if isinstance(current, list):
            if not re.fullmatch(r"0|[1-9][0-9]*", token):
                raise KeyError(normalize_json_pointer(pointer))
            try:
                index = int(token)
            except ValueError as exc:
                raise KeyError(normalize_json_pointer(pointer)) from exc
            if index >= len(current):
                raise KeyError(normalize_json_pointer(pointer))
            current = current[index]
            continue
        raise KeyError(normalize_json_pointer(pointer))
    return current


def json_pointers_overlap(left: str, right: str) -> bool:
    left_tokens = json_pointer_tokens(left)
    right_tokens = json_pointer_tokens(right)
    shorter = min(len(left_tokens), len(right_tokens))
    return left_tokens[:shorter] == right_tokens[:shorter]


def json_values_equal(left: Any, right: Any) -> bool:
    """Compare JSON values without Python's bool/int equality collapse."""

    if type(left) is not type(right):
        return False
    if isinstance(left, dict):
        return left.keys() == right.keys() and all(json_values_equal(left[key], right[key]) for key in left)
    if isinstance(left, list):
        return len(left) == len(right) and all(json_values_equal(a, b) for a, b in zip(left, right))
    return bool(left == right)


def normalized_pointer_list(
    value: Any,
    *,
    label: str,
    result: ValidationResult,
) -> list[str]:
    if not isinstance(value, list):
        result.error(f"{label} 必须是原生数组。")
        return []
    pointers: list[str] = []
    for item in value:
        try:
            pointer = normalize_json_pointer(item)
        except ValueError as exc:
            result.error(f"{label} {exc}")
            continue
        if pointer in pointers:
            result.error(f"{label} 规范化后路径重复：{pointer}")
            continue
        pointers.append(pointer)
    return pointers


def validate_sound_sources(scene_log: dict[str, Any], scene: str, result: ValidationResult) -> set[str]:
    raw = scene_log.get("sound_sources")
    if not isinstance(raw, dict):
        result.error(f"{scene} sound_sources 必须是对象映射。")
        return set()
    source_ids: set[str] = set()
    for source_id, source in raw.items():
        if not isinstance(source_id, str) or not re.fullmatch(SOUND_SOURCE_ID_PATTERN, source_id):
            result.error(f"{scene} sound_sources 键必须符合 SS[0-9]{{2,}} 格式：{source_id}")
            continue
        source_ids.add(source_id)
        if not isinstance(source, dict):
            result.error(f"{scene} sound_sources.{source_id} 必须是对象。")
            continue
        for field_name in ("name", "visibility", "position", "state"):
            if not isinstance(source.get(field_name), str) or not clean_text(source.get(field_name)):
                result.error(f"{scene} sound_sources.{source_id}.{field_name} 必须是非空字符串。")
        if clean_text(source.get("visibility")) not in SOUND_SOURCE_VISIBILITIES:
            result.error(
                f"{scene} sound_sources.{source_id}.visibility 必须为 onscreen / offscreen。"
            )
    return source_ids


def sound_source_ids(scene_log: dict[str, Any]) -> set[str]:
    return {
        source_id
        for source_id in as_dict(scene_log.get("sound_sources"))
        if isinstance(source_id, str) and re.fullmatch(SOUND_SOURCE_ID_PATTERN, source_id)
    }


def validate_spatial_axes(scene_log: dict[str, Any], scene: str, result: ValidationResult) -> set[str]:
    raw_axes = scene_log.get("spatial_axes")
    if not isinstance(raw_axes, list):
        result.error(f"{scene} spatial_axes 必须是数组。")
        return set()
    axis_ids: set[str] = set()
    for index, axis in enumerate(raw_axes, start=1):
        label = f"{scene} spatial_axes[{index}]"
        if not isinstance(axis, dict):
            result.error(f"{label} 必须是对象。")
            continue
        axis_id = axis.get("axis_id")
        if not isinstance(axis_id, str) or axis_id != clean_text(axis_id) or not re.fullmatch(
            SPATIAL_AXIS_ID_PATTERN, axis_id
        ):
            result.error(f"{label}.axis_id 必须符合 AX[0-9]{{2,}}。")
            continue
        if axis_id in axis_ids:
            result.error(f"{scene} spatial_axes.axis_id 重复：{axis_id}")
        axis_ids.add(axis_id)
        if clean_text(axis.get("axis_type")) not in SPATIAL_AXIS_TYPES:
            result.error(f"{label}.axis_type 不合法。")
        for field_name in ("endpoint_a", "endpoint_b", "side_a_anchor", "side_b_anchor"):
            value = axis.get(field_name)
            if not isinstance(value, str) or not clean_text(value) or value != clean_text(value):
                result.error(f"{label}.{field_name} 必须是无首尾空白的非空字符串。")
        if clean_text(axis.get("endpoint_a")) == clean_text(axis.get("endpoint_b")):
            result.error(f"{label} 两个轴线端点不得相同。")
    if scene_log.get("first_shot_anchor_type") in {"multi_character", "both"} and not axis_ids:
        result.error(f"{scene} 多人锚定场景至少需要一个 spatial_axes 轴线。")
    return axis_ids


def validate_continuity_logs(data: dict[str, Any], result: ValidationResult) -> dict[str, dict[str, Any]]:
    logs_by_scene: dict[str, dict[str, Any]] = {}
    logs: list[dict[str, Any]] = []
    use_scene_id = requires_human_gate(data)
    raw_logs = data.get("continuity_logs")
    if not isinstance(raw_logs, list):
        result.error("continuity_logs 必须是数组。")
    if is_current_contract(data) and not as_list(raw_logs):
        result.error(f"{metadata_version(data)} continuity_logs 必须是非空数组。")
    for log in as_list(raw_logs):
        if not isinstance(log, dict):
            result.error("continuity_logs 中每一项必须是对象。")
            continue
        if is_current_contract(data):
            raw_scene_id = log.get("scene_id")
            if (
                not isinstance(raw_scene_id, str)
                or raw_scene_id != clean_text(raw_scene_id)
                or not re.fullmatch(SCENE_ID_PATTERN, raw_scene_id)
            ):
                result.error(f"continuity_logs scene_id 必须是 S[0-9]{{2,}} 格式字符串：{raw_scene_id}")
        scene = scene_key_for(data, log)
        if not scene:
            result.error("continuity_logs 每项必须包含 scene_id。" if use_scene_id else "continuity_logs 每项必须包含 scene。")
            continue
        if scene in logs_by_scene:
            result.error(f"continuity_logs 场景键重复：{scene}")
        logs_by_scene[scene] = log
        logs.append(log)
        if use_scene_id and not clean_text(log.get("scene")):
            result.error(f"{scene} continuity_logs 必须包含展示用 scene。")
        anchor_type = log.get("first_shot_anchor_type")
        if anchor_type not in ANCHOR_TYPES:
            result.error(
                f"{scene} first_shot_anchor_type 必须为 "
                "space / multi_character / both / single_continuation / subjective。"
            )
        if is_current_contract(data):
            validate_sound_sources(log, scene, result)
        if uses_director_integrity_guard(data):
            validate_spatial_axes(log, scene, result)
    if requires_structured_audit(data):
        parent_map: dict[str, str] = {}
        for log in logs:
            scene = scene_key_for(data, log)
            raw_parent_scene = log.get("inherits_from")
            if is_current_contract(data) and (
                not isinstance(raw_parent_scene, str) or raw_parent_scene != clean_text(raw_parent_scene)
            ):
                result.error(f"{scene} inherits_from 必须是无首尾空白的原生字符串；无继承时写空字符串。")
            parent_scene = clean_text(raw_parent_scene)
            if is_current_contract(data):
                inherited_pointers = normalized_pointer_list(
                    log.get("inherited_states"), label=f"{scene} inherited_states", result=result
                )
                diverged_pointers = normalized_pointer_list(
                    log.get("diverged_states"), label=f"{scene} diverged_states", result=result
                )
                if not parent_scene:
                    if inherited_pointers or diverged_pointers:
                        result.error(
                            f"{scene} inherits_from 为空时 inherited_states/diverged_states 必须都是空数组。"
                        )
                    continue
            else:
                inherited_pointers = normalized_list(log.get("inherited_states"))
                diverged_pointers = normalized_list(log.get("diverged_states"))
            if not parent_scene:
                continue
            parent_map[scene] = parent_scene
            if parent_scene == scene:
                result.error(f"{scene} inherits_from 不得自指。")
            parent = logs_by_scene.get(parent_scene)
            if parent is None:
                result.error(f"{scene} inherits_from 指向不存在的场景：{parent_scene}")
                continue
            if is_current_contract(data):
                for inherited_pointer in inherited_pointers:
                    for diverged_pointer in diverged_pointers:
                        if json_pointers_overlap(inherited_pointer, diverged_pointer):
                            result.error(
                                f"{scene} inherited/diverged JSON Pointer 祖先或子路径冲突："
                                f"{inherited_pointer} <> {diverged_pointer}"
                            )
                for pointer in inherited_pointers:
                    try:
                        resolve_json_pointer(parent, pointer)
                        resolve_json_pointer(log, pointer)
                    except KeyError:
                        result.error(f"{scene} 继承路径必须同时存在于父场景和当前场景：{pointer}")
                for pointer in diverged_pointers:
                    try:
                        resolve_json_pointer(parent, pointer)
                        resolve_json_pointer(log, pointer)
                    except KeyError:
                        result.error(f"{scene} diverged 路径必须同时存在于父场景和当前场景：{pointer}")
            else:
                inherited = set(inherited_pointers)
                diverged = set(diverged_pointers)
                overlap = inherited & diverged
                if overlap:
                    result.error(f"{scene} inherited_states 与 diverged_states 重叠：{', '.join(sorted(overlap))}")
                for field_name in inherited:
                    if field_name not in parent or field_name not in log:
                        result.error(f"{scene} 继承字段 {field_name} 必须同时存在于父场景和当前场景。")
                        continue
                    if log.get(field_name) != parent.get(field_name):
                        result.error(f"{scene} 继承字段 {field_name} 与 {parent_scene} 不一致。")
        for scene in parent_map:
            seen: set[str] = set()
            current = scene
            while current in parent_map:
                if current in seen:
                    result.error(f"continuity_logs inherits_from 存在循环：{scene}")
                    break
                seen.add(current)
                current = parent_map[current]
    return logs_by_scene


def state_key(entity_type: Any, entity: Any, field_name: Any) -> tuple[str, str, str]:
    return clean_text(entity_type), clean_text(entity), clean_text(field_name)


def register_entity_state(states: dict[tuple[str, str, str], str], entity_type: str, entity: str, state: Any) -> None:
    if not entity:
        return
    if isinstance(state, dict):
        for field_name in STATEFUL_CONTINUITY_FIELDS:
            if field_name in state:
                states[state_key(entity_type, entity, field_name)] = clean_text(state.get(field_name))
        return
    if isinstance(state, str):
        text = clean_text(state)
        for field_name in STATEFUL_CONTINUITY_FIELDS:
            match = re.search(rf"{field_name}[：:]\s*([^，。；;]+)", text)
            if match:
                states[state_key(entity_type, entity, field_name)] = match.group(1).strip()


def initial_scene_states(scene_log: dict[str, Any]) -> dict[tuple[str, str, str], str]:
    states: dict[tuple[str, str, str], str] = {}
    for field_name, entity_type in [("characters", "character"), ("props", "prop"), ("fixed_objects", "fixed_object")]:
        collection = scene_log.get(field_name)
        if isinstance(collection, dict):
            for entity, state in collection.items():
                register_entity_state(states, entity_type, str(entity), state)
        for item in as_list(collection):
            if isinstance(item, dict):
                entity = clean_text(item.get("name") or item.get("entity") or item.get("character") or item.get("prop"))
                register_entity_state(states, entity_type, entity, item)
    for source_id, source in as_dict(scene_log.get("sound_sources")).items():
        if isinstance(source, dict):
            register_entity_state(states, "sound_source", clean_text(source_id), source)
    reality = clean_text(scene_log.get("reality_layer"))
    if reality:
        states[state_key("reality_layer", "", "value")] = reality
    return states


def declared_character_names(scene_log: dict[str, Any]) -> set[str]:
    collection = scene_log.get("characters")
    if isinstance(collection, dict):
        return {clean_text(name) for name in collection if clean_text(name)}
    names: set[str] = set()
    for item in as_list(collection):
        if not isinstance(item, dict):
            continue
        name = clean_text(item.get("name") or item.get("entity") or item.get("character"))
        if name:
            names.add(name)
    return names


def declared_entity_names(scene_log: dict[str, Any], field_name: str) -> set[str]:
    collection = scene_log.get(field_name)
    if isinstance(collection, dict):
        return {clean_text(name) for name in collection if clean_text(name)}
    names: set[str] = set()
    for item in as_list(collection):
        if isinstance(item, str) and clean_text(item):
            names.add(clean_text(item))
            continue
        if not isinstance(item, dict):
            continue
        name = clean_text(
            item.get("name") or item.get("entity") or item.get("character") or item.get("prop")
        )
        if name:
            names.add(name)
    return names


def apply_update_to_scene_document(document: dict[str, Any], update: Any) -> bool:
    """Overlay a validated continuity update onto a scene-log deep copy."""

    if not isinstance(update, dict):
        return False
    entity_type = clean_text(update.get("entity_type"))
    entity = clean_text(update.get("entity"))
    field_name = clean_text(update.get("field"))
    to_value = update.get("to")
    if not field_name:
        return False
    if entity_type == "sound_source":
        source = as_dict(document.get("sound_sources")).get(entity)
        if isinstance(source, dict):
            source[field_name] = copy.deepcopy(to_value)
            return True
        return False
    if entity_type == "reality_layer" and field_name == "value":
        document["reality_layer"] = copy.deepcopy(to_value)
        return True
    collection_name = {
        "character": "characters",
        "prop": "props",
        "fixed_object": "fixed_objects",
    }.get(entity_type)
    if collection_name is None:
        return False
    collection = document.get(collection_name)
    if isinstance(collection, dict):
        target = collection.get(entity)
        if isinstance(target, dict):
            target[field_name] = copy.deepcopy(to_value)
            return True
        return False
    for item in as_list(collection):
        if not isinstance(item, dict):
            continue
        item_entity = clean_text(
            item.get("name") or item.get("entity") or item.get("character") or item.get("prop")
        )
        if item_entity == entity:
            item[field_name] = copy.deepcopy(to_value)
            return True
    return False


def validate_final_inherited_states(
    data: dict[str, Any],
    continuity_logs: dict[str, dict[str, Any]],
    final_scene_documents: dict[str, dict[str, Any]],
    result: ValidationResult,
) -> None:
    if not is_current_contract(data):
        return
    for scene, child in continuity_logs.items():
        parent_scene = clean_text(child.get("inherits_from"))
        if not parent_scene or parent_scene not in continuity_logs:
            continue
        parent_final = final_scene_documents.get(parent_scene, continuity_logs[parent_scene])
        for raw_pointer in as_list(child.get("inherited_states")):
            try:
                pointer = normalize_json_pointer(raw_pointer)
                parent_value = resolve_json_pointer(parent_final, pointer)
                child_value = resolve_json_pointer(child, pointer)
            except (KeyError, ValueError):
                # Shape/path errors were already reported while validating the
                # raw ledgers.  Avoid duplicate diagnostics here.
                continue
            if not json_values_equal(parent_value, child_value):
                result.error(
                    f"{scene} 继承路径 {pointer} 必须等于父场景 {parent_scene} 的镜头推进后终态。"
                )


def collect_facts(
    data: dict[str, Any], result: ValidationResult
) -> tuple[dict[str, str], dict[str, str], set[str], dict[str, str]]:
    beat_to_facts: dict[str, set[str]] = {}
    fact_to_beat: dict[str, str] = {}
    fact_types: dict[str, str] = {}
    seen_beats: set[str] = set()
    seen_facts: set[str] = set()
    beat_scenes: dict[str, str] = {}
    previous_beat_number: int | None = None
    previous_beat_order: Decimal | None = None
    seen_beat_orders: set[Decimal] = set()
    scene_logs = {
        scene_key_for(data, log): log
        for log in as_list(data.get("continuity_logs"))
        if isinstance(log, dict) and scene_key_for(data, log)
    }
    scene_ids = set(scene_logs)
    raw_beats = data.get("beats")
    if not isinstance(raw_beats, list):
        result.error("beats 必须是数组。")
    if is_current_contract(data) and not as_list(raw_beats):
        result.error(f"{metadata_version(data)} beats 必须是非空数组。")
    for beat in as_list(raw_beats):
        if not isinstance(beat, dict):
            result.error("beats 中每一项必须是对象。")
            continue
        beat_id = str(beat.get("beat_id", ""))
        beat_number = numeric_beat_id(beat_id)
        beat_pattern = CURRENT_BEAT_ID_PATTERN if is_current_contract(data) else LEGACY_BEAT_ID_PATTERN
        if beat_number is None or not re.fullmatch(beat_pattern, beat_id):
            result.error(f"Beat ID 不合法：{beat_id}")
        elif requires_human_gate(data) and not is_current_contract(data):
            if previous_beat_number is not None and beat_number <= previous_beat_number:
                result.error(f"2.4.0 Beat ID 必须唯一单调递增且禁止重编号：{beat_id}")
            previous_beat_number = beat_number
        if beat_id in seen_beats:
            result.error(f"Beat ID 重复：{beat_id}")
        seen_beats.add(beat_id)
        if is_current_contract(data):
            try:
                order = parse_beat_order(beat.get("beat_order"))
            except ValueError as exc:
                result.error(f"{beat_id} {exc}")
            else:
                if order in seen_beat_orders:
                    result.error(f"beat_order 重复：{beat.get('beat_order')}")
                if previous_beat_order is not None and order <= previous_beat_order:
                    result.error(
                        f"beats 数组必须按 beat_order 严格递增：{beat.get('beat_order')} 不大于前项。"
                    )
                seen_beat_orders.add(order)
                previous_beat_order = order
        beat_scenes[beat_id] = scene_key_for(data, beat)
        if requires_human_gate(data):
            raw_scene_id = beat.get("scene_id")
            scene_id = clean_text(raw_scene_id)
            if is_current_contract(data) and (
                not isinstance(raw_scene_id, str)
                or raw_scene_id != scene_id
                or not re.fullmatch(SCENE_ID_PATTERN, scene_id)
            ):
                result.error(f"{beat_id} scene_id 必须是 S[0-9]{{2,}} 格式字符串。")
            if not scene_id:
                result.error(f"{beat_id} 缺少 scene_id。")
            elif scene_ids and scene_id not in scene_ids:
                result.error(f"{beat_id} scene_id 未在 continuity_logs 登记：{scene_id}")
            display_scene = clean_text(beat.get("scene"))
            if not display_scene:
                result.error(f"{beat_id} 缺少展示用 scene。")
            elif scene_id in scene_logs and display_scene != clean_text(scene_logs[scene_id].get("scene")):
                result.error(f"{beat_id} scene 与 continuity_logs[{scene_id}].scene 不一致。")
        facts = as_list(beat.get("facts"))
        if not facts:
            if is_current_contract(data):
                result.error(f"{beat_id} facts 必须是非空数组。")
            else:
                result.warn(f"{beat_id} 没有 facts。")
        beat_to_facts.setdefault(beat_id, set())
        for fact in facts:
            if not isinstance(fact, dict):
                result.error(f"{beat_id} facts 中存在非对象项。")
                continue
            fact_id = str(fact.get("fact_id", ""))
            fact_pattern = (
                rf"{re.escape(beat_id)}-{CURRENT_FACT_SUFFIX_PATTERN}"
                if is_current_contract(data)
                else rf"{re.escape(beat_id)}-F[0-9]{{2}}"
            )
            if not re.fullmatch(fact_pattern, fact_id):
                result.error(f"事实 ID 必须绑定所属 Beat：{fact_id}")
            if fact.get("type") not in FACT_TYPES:
                result.error(f"{fact_id} 事实类型不合法：{fact.get('type')}")
            if fact_id in seen_facts:
                result.error(f"事实 ID 重复：{fact_id}")
            seen_facts.add(fact_id)
            if is_current_contract(data) and (
                not isinstance(fact.get("text"), str) or not clean_text(fact.get("text"))
            ):
                result.error(f"{fact_id} fact.text 不能为空。")
            if uses_director_integrity_guard(data) and fact.get("type") == "dialogue":
                scene_characters = declared_character_names(as_dict(scene_logs.get(beat_scenes[beat_id])))
                speaker = fact.get("speaker")
                if (
                    not isinstance(speaker, str)
                    or speaker != clean_text(speaker)
                    or not speaker
                    or speaker not in scene_characters
                ):
                    result.error(f"{fact_id} dialogue speaker 必须精确匹配本场连续性台账角色。")
                addressees = fact.get("addressees")
                if not isinstance(addressees, list):
                    result.error(f"{fact_id} dialogue addressees 必须是数组，可为空。")
                else:
                    canonical_addressees = [item for item in addressees if isinstance(item, str)]
                    if len(canonical_addressees) != len(addressees) or any(
                        not item or item != clean_text(item) or item not in scene_characters
                        for item in canonical_addressees
                    ):
                        result.error(f"{fact_id} dialogue addressees 必须精确匹配本场连续性台账角色。")
                    if len(canonical_addressees) != len(set(canonical_addressees)):
                        result.error(f"{fact_id} dialogue addressees 不得重复。")
                if clean_text(fact.get("delivery")) not in DIALOGUE_DELIVERIES:
                    result.error(f"{fact_id} dialogue delivery 必须为 onscreen / offscreen / vo。")
            if uses_director_integrity_guard(data) and fact.get("type") in {"action", "emotion"}:
                scene_characters = declared_character_names(as_dict(scene_logs.get(beat_scenes[beat_id])))
                raw_performers = fact.get("performers")
                if not isinstance(raw_performers, list):
                    result.error(f"{fact_id} {fact.get('type')} performers 必须是数组。")
                else:
                    performers = [item for item in raw_performers if isinstance(item, str)]
                    if len(performers) != len(raw_performers) or any(
                        not item or item != clean_text(item) or item not in scene_characters
                        for item in performers
                    ):
                        result.error(f"{fact_id} performers 必须精确匹配本场连续性台账角色。")
                    if len(performers) != len(set(performers)):
                        result.error(f"{fact_id} performers 不得重复。")
                    if not performers:
                        result.error(f"{fact_id} {fact.get('type')} performers 不能为空。")
                    named_characters = {
                        character
                        for character in scene_characters
                        if character in clean_text(fact.get("text"))
                    }
                    missing_performers = sorted(named_characters - set(performers))
                    if missing_performers:
                        result.error(
                            f"{fact_id} fact.text 点名角色但 performers 未登记：{', '.join(missing_performers)}。"
                        )
            fact_to_beat[fact_id] = beat_id
            fact_types[fact_id] = str(fact.get("type", ""))
            beat_to_facts[beat_id].add(fact_id)
    return fact_to_beat, fact_types, set(fact_to_beat), beat_scenes


def validate_first_scene_shot(
    shot: dict[str, Any],
    scene_log: dict[str, Any],
    result: ValidationResult,
    fact_types: dict[str, str],
    *,
    current_contract: bool,
) -> None:
    shot_no = shot.get("shot_no")
    camera = clean_text(shot.get("camera_main_image"))
    _angle, shot_size, movement = parse_triad(camera)
    anchor_type = scene_log.get("first_shot_anchor_type")
    if current_contract:
        visible_names = {
            item
            for item in as_list(shot.get("visible_characters"))
            if isinstance(item, str) and item and item == clean_text(item)
        }
    else:
        visible_names = set(normalized_list(shot.get("visible_characters")))
    visible_count = len(visible_names)
    has_first_position = "【场景首镜站位】" in camera
    if not has_first_position:
        result.error(f"镜号{shot_no} 是场景首镜，必须包含【场景首镜站位】。")
    requires_panorama = anchor_type in PANORAMIC_REQUIRED_ANCHORS or (
        not current_contract and visible_count >= 2 and anchor_type != "single_continuation"
    )
    if requires_panorama and not is_panoramic_shot_size(shot_size):
        result.error(
            f"镜号{shot_no} 为 {anchor_type or '未标注'} 锚定场景首镜，"
            f"且可见人物数 {visible_count}，景别必须为大远景/大全景/全景/中全景；当前为 {shot_size}。"
        )
    if anchor_type in {"space", "both"} and "斯坦尼康" in clean_text(movement):
        result.error(f"镜号{shot_no} 为 {anchor_type} 锚定场景首镜，不得使用斯坦尼康建立大空间。")
    first_position_line = next(
        (line for line in camera.split("\n") if line.strip().startswith("【场景首镜站位】")),
        "",
    )
    if current_contract and anchor_type in {"multi_character", "both"}:
        if visible_count < 2:
            result.error(f"镜号{shot_no} 为 {anchor_type} 首镜，visible_characters 至少需要2人。")
        if "position" not in fact_types_for_shot(shot, fact_types):
            result.error(f"镜号{shot_no} 为 {anchor_type} 首镜，必须绑定 position fact 完成基础锚定。")
        declared = declared_character_names(scene_log)
        missing_from_ledger = sorted(visible_names - declared)
        if missing_from_ledger:
            result.error(
                f"镜号{shot_no} visible_characters 含 continuity_logs 未登记或不可识别角色："
                f"{', '.join(missing_from_ledger)}"
            )
        missing_from_anchor = sorted(name for name in visible_names if name not in first_position_line)
        if missing_from_anchor:
            result.error(
                f"镜号{shot_no} 的【场景首镜站位】未逐一写明可见互动角色："
                f"{', '.join(missing_from_anchor)}"
            )
    if visible_count >= 2 and anchor_type != "single_continuation":
        blocking_text = first_position_line + "\n" + clean_text(scene_log.get("spatial_axis"))
        if not any(keyword in blocking_text for keyword in ["左", "右", "轴线", "对视", "朝向", "面向"]):
            result.error(f"镜号{shot_no} 多人场景首镜站位必须写清左右关系、朝向或主要对视轴线。")


def validate_adjacent_shot_design(
    data: dict[str, Any],
    shots: list[Any],
    fact_types: dict[str, str],
    result: ValidationResult,
    *,
    guard_reverse_depth_motion: bool,
    allow_marked_exceptions: bool,
) -> None:
    previous: dict[str, Any] | None = None
    for item in shots:
        if not isinstance(item, dict):
            continue
        item_scene = clean_text(item.get("scene_id")) or clean_text(item.get("scene"))
        previous_scene = clean_text(previous.get("scene_id")) or clean_text(previous.get("scene")) if previous is not None else ""
        if previous is not None and item_scene == previous_scene:
            validate_adjacent_pair(
                data,
                previous,
                item,
                fact_types,
                result,
                guard_reverse_depth_motion=guard_reverse_depth_motion,
                allow_marked_exceptions=allow_marked_exceptions,
            )
        previous = item


def validate_adjacent_pair(
    data: dict[str, Any],
    prev: dict[str, Any],
    curr: dict[str, Any],
    fact_types: dict[str, str],
    result: ValidationResult,
    *,
    guard_reverse_depth_motion: bool,
    allow_marked_exceptions: bool,
) -> None:
    prev_no = prev.get("shot_no")
    curr_no = curr.get("shot_no")
    validate_director_adjacent_pair(data, prev, curr, result)
    prev_angle, prev_size, prev_movement = parse_triad(clean_text(prev.get("camera_main_image")))
    curr_angle, curr_size, curr_movement = parse_triad(clean_text(curr.get("camera_main_image")))
    if guard_reverse_depth_motion:
        prev_depth_direction = depth_motion_direction(prev_movement)
        curr_depth_direction = depth_motion_direction(curr_movement)
        if {prev_depth_direction, curr_depth_direction} == {"push_in", "pull_out"}:
            if allow_marked_exceptions and "[反转动机]" in clean_text(curr.get("notes")):
                if is_current_contract(data) and not approved_review(
                    data,
                    "GATE_B",
                    batch_id=batch_id_for_scene(data, curr.get("scene_id")),
                    approved_item=f"shot:{curr_no}:axis-reversal",
                ):
                    result.error(f"镜号{curr_no} [反转动机] 必须由 Gate B approved_items 精确批准。")
                else:
                    result.warn(
                        f"镜号{prev_no}-{curr_no} 同场相邻镜头使用 [反转动机] 保留纵深推拉方向反转，需 Gate B 复核。"
                    )
            else:
                result.error(
                    f"镜号{prev_no}-{curr_no} 同场相邻镜头出现无动机纵深推拉方向反转："
                    f"{prev_movement} 后接 {curr_movement}。应改为固定、横移、环绕、垂直摇移、"
                    "切换主体或切换景别。"
                )
    prev_rank = shot_size_rank(prev_size)
    curr_rank = shot_size_rank(curr_size)
    same_characters = normalized_visible_characters(prev) == normalized_visible_characters(curr)
    if not uses_director_integrity_guard(data) and same_characters and normalized_visible_characters(prev):
        no_position_change = not has_continuity_position_change(prev) and not has_continuity_position_change(curr)
        close_size = prev_rank is not None and curr_rank is not None and abs(prev_rank - curr_rank) <= 1
        same_angle = angle_family(prev_angle) == angle_family(curr_angle)
        soft_movement = is_static_or_soft_movement(prev_movement) and is_static_or_soft_movement(curr_movement)
        if no_position_change and close_size and same_angle and soft_movement:
            if allow_marked_exceptions and ("[必拆相邻]" in clean_text(prev.get("notes")) or "[必拆相邻]" in clean_text(curr.get("notes"))):
                marked = [
                    shot
                    for shot in (prev, curr)
                    if "[必拆相邻]" in clean_text(shot.get("notes"))
                ]
                approved = any(
                    approved_review(
                        data,
                        "GATE_B",
                        batch_id=batch_id_for_scene(data, shot.get("scene_id")),
                        approved_item=f"shot:{shot.get('shot_no')}:required-adjacent",
                    )
                    for shot in marked
                )
                if is_current_contract(data) and not approved:
                    result.error(
                        f"镜号{prev_no}-{curr_no} [必拆相邻] 必须由 Gate B approved_items 精确批准。"
                    )
                else:
                    result.warn(f"镜号{prev_no}-{curr_no} 使用 [必拆相邻] 豁免低增量相邻镜组，需 Gate B 复核。")
            else:
                result.error(
                    f"镜号{prev_no}-{curr_no} 同场相邻镜头主体、视角、景别和运动过近，"
                    "且无站位/朝向迁移；必须合并或提供明确画面信息增量。"
                )

    prev_dialogue = quoted_dialogue_text(prev)
    curr_dialogue = quoted_dialogue_text(curr)
    if prev_dialogue and curr_dialogue and prev_dialogue.endswith(INCOMPLETE_DIALOGUE_ENDINGS):
        prev_speaker = dialogue_speaker(prev)
        curr_speaker = dialogue_speaker(curr)
        if not prev_speaker or not curr_speaker or prev_speaker == curr_speaker:
            result.error(
                f"镜号{prev_no}-{curr_no} 将同一说话人的未完成台词切成相邻两镜；"
                "应合并为一镜，用表演或机位运动承接。"
            )

    prev_types = fact_types_for_shot(prev, fact_types)
    curr_types = fact_types_for_shot(curr, fact_types)
    if "dialogue" in prev_types and is_expression_only_response(curr, fact_types):
        if same_characters or not normalized_visible_characters(curr):
            result.error(
                f"镜号{prev_no}-{curr_no} 是对白后同一人物短表情反应，且没有新的空间、道具、声音或位置事实；"
                "应合并进同一镜。"
            )
    curr_duration = contract_duration_number(
        curr.get("duration_seconds"), strict_contract=is_current_contract(data)
    )
    if (
        clean_text(curr.get("shot_type")) == "reaction"
        and clean_text(curr.get("insert_priority")) == "none"
        and curr_duration <= 3
        and not (curr_types & NEW_FACT_TYPES_FOR_CUT)
        and not has_continuity_position_change(prev)
        and not has_continuity_position_change(curr)
    ):
        result.warn(f"[可合并] 镜号{curr_no} 是无新增空间/道具/声音/位置/现实层事实的短 reaction，建议与相邻镜头合并。")


def validate_hybrid_fields(
    shot: dict[str, Any],
    result: ValidationResult,
    *,
    required: bool,
    strict_contract: bool = False,
) -> None:
    shot_no = shot.get("shot_no")
    duration = contract_duration_number(shot.get("duration_seconds"), strict_contract=strict_contract)
    if not required:
        return
    shot_type = clean_text(shot.get("shot_type"))
    if not shot_type:
        result.error(f"镜号{shot_no} 2.3.4+ 必须填写 shot_type。")
    elif shot_type not in SHOT_TYPES:
        result.error(f"镜号{shot_no} shot_type 不合法：{shot_type}")
    split_reasons = normalized_list(shot.get("split_reason"))
    if not split_reasons:
        result.error(f"镜号{shot_no} 2.3.4+ 必须填写 split_reason。")
    for reason in split_reasons:
        if reason not in SPLIT_REASONS:
            result.error(f"镜号{shot_no} split_reason 不合法：{reason}")
    insert_priority = clean_text(shot.get("insert_priority"))
    if not insert_priority:
        result.error(f"镜号{shot_no} 2.3.4+ 必须填写 insert_priority。")
    elif insert_priority not in INSERT_PRIORITIES:
        result.error(f"镜号{shot_no} insert_priority 不合法：{insert_priority}")
    supports = normalized_list(shot.get("long_take_support"))
    for support in supports:
        if support not in LONG_TAKE_SUPPORTS:
            result.error(f"镜号{shot_no} long_take_support 不合法：{support}")
    if duration > 10 and len(supports) < 2:
        result.error(f"镜号{shot_no} 超过10秒，2.3.4 要求至少填写两项 long_take_support。")
    if shot_type in INSERT_LIKE_TYPES and duration > 5:
        result.warn(f"镜号{shot_no} 是 {shot_type}，插镜建议 2-5 秒；当前 {duration:g} 秒，请确认是否承担镜内推进。")
    causal_reasons = {"new_vfx_state", "causal_reveal", "prop_state_change"}
    if causal_reasons & set(split_reasons) and shot_type not in {"insert", "vfx_anchor", "transition", "action"}:
        result.warn(f"镜号{shot_no} 承担真相/道具/VFX 因果落点，但 shot_type 为 {shot_type}，建议复核插镜意识。")


def spatial_axis_ids(scene_log: dict[str, Any]) -> set[str]:
    return {
        clean_text(axis.get("axis_id"))
        for axis in as_list(scene_log.get("spatial_axes"))
        if isinstance(axis, dict) and clean_text(axis.get("axis_id"))
    }


def screen_direction_map(geometry: dict[str, Any]) -> dict[tuple[str, str], str]:
    output: dict[tuple[str, str], str] = {}
    for item in as_list(geometry.get("screen_directions")):
        if isinstance(item, dict):
            output[(clean_text(item.get("entity")), clean_text(item.get("kind")))] = clean_text(
                item.get("direction")
            )
    return output


def source_span_signature(shot: dict[str, Any]) -> tuple[tuple[Any, Any], ...]:
    raw = shot.get("source_span") if "source_span" in shot else shot.get("source_spans")
    return tuple(
        (span.get("start"), span.get("end"))
        for span in normalize_source_spans(raw)
    )


def validate_shot_director_integrity(
    data: dict[str, Any],
    shot: dict[str, Any],
    scene_log: dict[str, Any],
    facts_by_id: dict[str, dict[str, Any]],
    result: ValidationResult,
    *,
    is_first_scene_shot: bool,
) -> None:
    if not uses_director_integrity_guard(data):
        return
    shot_no = shot.get("shot_no")
    covered = set(normalized_list(shot.get("covered_fact_ids")))
    visible = set(normalized_list(shot.get("visible_characters")))
    offscreen = set(normalized_list(shot.get("offscreen_characters")))
    visible_props = set(normalized_list(shot.get("visible_props")))
    declared = declared_character_names(scene_log)
    unknown = sorted((visible | offscreen) - declared)
    if unknown:
        result.error(f"镜号{shot_no} 人物映射未命中本场台账：{', '.join(unknown)}。")

    raw_geometry = shot.get("camera_geometry")
    if not isinstance(raw_geometry, dict):
        result.error(f"镜号{shot_no} 2.4.5 必须包含 camera_geometry 对象。")
        geometry: dict[str, Any] = {}
    else:
        geometry = raw_geometry
    axis_id = geometry.get("axis_id")
    if axis_id is not None and (
        not isinstance(axis_id, str)
        or axis_id != clean_text(axis_id)
        or axis_id not in spatial_axis_ids(scene_log)
    ):
        result.error(f"镜号{shot_no} camera_geometry.axis_id 必须为本场已登记 AX 编号或 null。")
    axis_side = clean_text(geometry.get("axis_side"))
    if axis_side not in CAMERA_AXIS_SIDES:
        result.error(f"镜号{shot_no} camera_geometry.axis_side 不合法。")
    if axis_id is None and axis_side != "not_applicable":
        result.error(f"镜号{shot_no} axis_id=null 时 axis_side 必须为 not_applicable。")
    if axis_id is not None and axis_side == "not_applicable":
        result.error(f"镜号{shot_no} 已绑定轴线时必须声明 side_a / side_b / on_axis。")
    position = geometry.get("camera_position")
    if not isinstance(position, str) or not clean_text(position) or position != clean_text(position):
        result.error(f"镜号{shot_no} camera_geometry.camera_position 必须是非空规范字符串。")
    if clean_text(geometry.get("camera_facing")) not in CAMERA_FACINGS:
        result.error(f"镜号{shot_no} camera_geometry.camera_facing 不合法。")
    angle_delta = clean_text(geometry.get("angle_delta_from_previous"))
    if angle_delta not in ANGLE_DELTAS:
        result.error(f"镜号{shot_no} camera_geometry.angle_delta_from_previous 不合法。")
    elif is_first_scene_shot and angle_delta != "scene_start":
        result.error(f"镜号{shot_no} 为本场首镜，angle_delta_from_previous 必须为 scene_start。")
    elif not is_first_scene_shot and angle_delta == "scene_start":
        result.error(f"镜号{shot_no} 非本场首镜，不得使用 scene_start。")

    raw_subjects = geometry.get("primary_subjects")
    subjects = normalized_list(raw_subjects)
    if not isinstance(raw_subjects, list) or not subjects:
        result.error(f"镜号{shot_no} camera_geometry.primary_subjects 必须是非空数组。")
    elif (
        any(not isinstance(subject, str) or not subject or subject != clean_text(subject) for subject in raw_subjects)
        or len(subjects) != len(set(subjects))
    ):
        result.error(f"镜号{shot_no} camera_geometry.primary_subjects 必须是去重规范字符串。")
    else:
        pov_subject = {clean_text(geometry.get("pov_character"))} if geometry.get("pov_character") else set()
        visual_characters = visible | pov_subject
        non_character_subjects = declared_entity_names(scene_log, "props") | declared_entity_names(
            scene_log, "fixed_objects"
        )
        unknown_subjects = sorted(set(subjects) - visual_characters - non_character_subjects)
        if unknown_subjects:
            result.error(
                f"镜号{shot_no} primary_subjects 必须映射到本镜可见/POV人物或本场已登记道具、固定物："
                f"{', '.join(unknown_subjects)}。"
            )
        unpictured_entities = sorted(
            subject
            for subject in set(subjects) & non_character_subjects
            if subject not in visible_props and subject not in clean_text(shot.get("camera_main_image"))
        )
        if unpictured_entities:
            result.error(
                f"镜号{shot_no} 非人物 primary_subjects 未出现在 visible_props 或主画面："
                f"{', '.join(unpictured_entities)}。"
            )
    raw_directions = geometry.get("screen_directions")
    if not isinstance(raw_directions, list):
        result.error(f"镜号{shot_no} camera_geometry.screen_directions 必须是数组。")
    seen_directions: set[tuple[str, str]] = set()
    for index, direction in enumerate(as_list(raw_directions), start=1):
        label = f"镜号{shot_no} screen_directions[{index}]"
        if not isinstance(direction, dict):
            result.error(f"{label} 必须是对象。")
            continue
        entity = clean_text(direction.get("entity"))
        kind = clean_text(direction.get("kind"))
        if entity not in declared:
            result.error(f"{label}.entity 必须精确匹配本场角色台账。")
        if kind not in SCREEN_DIRECTION_KINDS:
            result.error(f"{label}.kind 不合法。")
        if clean_text(direction.get("direction")) not in SCREEN_DIRECTIONS:
            result.error(f"{label}.direction 不合法。")
        if (entity, kind) in seen_directions:
            result.error(f"{label} entity/kind 重复。")
        seen_directions.add((entity, kind))

    camera = clean_text(shot.get("camera_main_image"))
    angle, shot_size, _movement = parse_triad(camera)
    subjective = "主观" in angle or "主观" in shot_size
    pov = geometry.get("pov_character")
    if subjective:
        if not isinstance(pov, str) or pov != clean_text(pov) or pov not in declared:
            result.error(f"镜号{shot_no} 主观镜必须填写本场角色 pov_character。")
    elif pov is not None:
        result.error(f"镜号{shot_no} 非主观镜 pov_character 必须为 null。")

    descriptive = descriptive_text_without_dialogue(camera)
    mapped = visible | offscreen | ({clean_text(pov)} if isinstance(pov, str) else set())
    for character in sorted(declared):
        if character in descriptive and character not in mapped:
            result.error(f"镜号{shot_no} 画面写到角色 {character}，但未登记为可见、画外或 POV。")
        if character in offscreen:
            unsupported_lines = [
                line
                for line in descriptive.splitlines()
                if character in line and not re.search(r"画外声|VO|旁白|声音|只闻其声", line, re.IGNORECASE)
            ]
            if unsupported_lines:
                result.error(
                    f"镜号{shot_no} 将 {character} 登记为 offscreen，却在非声音语境中描写其可见动作或机位关系。"
                )
    for fact_id in covered:
        fact = facts_by_id.get(fact_id)
        if not isinstance(fact, dict):
            continue
        fact_type = clean_text(fact.get("type"))
        if fact_type in {"action", "emotion"}:
            unmapped_performers = sorted(set(normalized_list(fact.get("performers"))) - mapped)
            if unmapped_performers:
                result.error(
                    f"镜号{shot_no} {fact_type} fact {fact_id} 的 performers 未映射到镜头："
                    f"{', '.join(unmapped_performers)}。"
                )
            continue
        if fact_type != "dialogue":
            continue
        speaker = clean_text(fact.get("speaker"))
        delivery = clean_text(fact.get("delivery"))
        if delivery == "onscreen" and speaker not in visible:
            result.error(f"镜号{shot_no} dialogue fact {fact_id} 的 onscreen speaker 必须在 visible_characters。")
        if delivery in {"offscreen", "vo"} and speaker not in offscreen:
            result.error(f"镜号{shot_no} dialogue fact {fact_id} 的画外说话人必须在 offscreen_characters。")
        for addressee in normalized_list(fact.get("addressees")):
            if addressee not in mapped:
                result.error(f"镜号{shot_no} dialogue fact {fact_id} 的受话人 {addressee} 未映射到镜头。")
        if delivery in {"offscreen", "vo"} and not any(
            isinstance(facts_by_id.get(other_id), dict)
            and clean_text(facts_by_id[other_id].get("type")) != "dialogue"
            for other_id in covered
        ):
            result.error(
                f"镜号{shot_no} 的 {delivery} 对白不能成为无画面事实支撑的独立镜；"
                "必须与可见动作、反应、空间或其他非对白事实同镜。"
            )

    raw_cues = shot.get("performance_cues")
    cues = as_list(raw_cues)
    if not isinstance(raw_cues, list):
        result.error(f"镜号{shot_no} performance_cues 必须是数组。")
    if visible and not 1 <= len(cues) <= 3:
        result.error(f"镜号{shot_no} 有可见人物时 performance_cues 必须为 1-3 项。")
    if not visible and cues:
        result.error(f"镜号{shot_no} 无可见人物时 performance_cues 必须为空。")
    for index, cue in enumerate(cues, start=1):
        label = f"镜号{shot_no} performance_cues[{index}]"
        if not isinstance(cue, dict):
            result.error(f"{label} 必须是对象。")
            continue
        if clean_text(cue.get("character")) not in visible:
            result.error(f"{label}.character 必须是本镜可见角色。")
        if clean_text(cue.get("channel")) not in PERFORMANCE_CUE_CHANNELS:
            result.error(f"{label}.channel 不合法。")
        description = clean_text(cue.get("description"))
        if not description or description not in camera:
            result.error(f"{label}.description 必须逐字出现于画面与机位。")
        evidence = normalized_list(cue.get("evidence_fact_ids"))
        if not evidence or not set(evidence).issubset(covered):
            result.error(f"{label}.evidence_fact_ids 必须是本镜已覆盖事实的非空子集。")
        cue_character = clean_text(cue.get("character"))
        for fact_id in evidence:
            fact = facts_by_id.get(fact_id)
            if not isinstance(fact, dict):
                continue
            fact_type = clean_text(fact.get("type"))
            if fact_type not in {"action", "emotion", "dialogue", "position"}:
                result.error(
                    f"{label}.evidence_fact_ids 只能引用 action / emotion / dialogue / position 人物事实。"
                )
            elif fact_type in {"action", "emotion"}:
                if cue_character not in normalized_list(fact.get("performers")):
                    result.error(
                        f"{label}.character 必须包含在证据事实 {fact_id} 的 performers 中。"
                    )
            elif fact_type == "dialogue" and cue_character != clean_text(fact.get("speaker")):
                result.error(f"{label}.character 必须等于 dialogue 证据事实 {fact_id} 的 speaker。")
            elif fact_type == "position" and cue_character not in clean_text(fact.get("text")):
                result.error(f"{label}.character 必须在 position 证据事实 {fact_id} 的 fact.text 中明确点名。")

    cue_coverage = {
        (clean_text(cue.get("character")), fact_id)
        for cue in cues
        if isinstance(cue, dict)
        for fact_id in normalized_list(cue.get("evidence_fact_ids"))
    }
    for fact_id in covered:
        fact = facts_by_id.get(fact_id)
        if not isinstance(fact, dict) or clean_text(fact.get("type")) not in {"action", "emotion"}:
            continue
        for performer in set(normalized_list(fact.get("performers"))) & visible:
            if (performer, fact_id) not in cue_coverage:
                result.error(
                    f"镜号{shot_no} 可见表演者 {performer} 的事实 {fact_id} 缺少对应 performance_cue 证据。"
                )

    raw_progressions = shot.get("action_progressions")
    progressions = as_list(raw_progressions)
    if not isinstance(raw_progressions, list):
        result.error(f"镜号{shot_no} action_progressions 必须是数组。")
    progression_required = bool(as_list(shot.get("continuity_updates"))) or bool(
        set(normalized_list(shot.get("split_reason")))
        & {"prop_state_change", "new_vfx_state", "continuity_migration"}
    )
    if progression_required and not progressions:
        result.error(f"镜号{shot_no} 含状态推进时必须填写 action_progressions。")
    for index, progression in enumerate(progressions, start=1):
        label = f"镜号{shot_no} action_progressions[{index}]"
        if not isinstance(progression, dict):
            result.error(f"{label} 必须是对象。")
            continue
        if not clean_text(progression.get("entity")):
            result.error(f"{label}.entity 不能为空。")
        evidence = normalized_list(progression.get("evidence_fact_ids"))
        if not evidence or not set(evidence).issubset(covered):
            result.error(f"{label}.evidence_fact_ids 必须是本镜已覆盖事实的非空子集。")
        for field_name in ("start", "process", "end"):
            value = clean_text(progression.get(field_name))
            if not value or value not in camera:
                result.error(f"{label}.{field_name} 必须逐字出现于画面与机位。")

    raw_environment = shot.get("environment_cues")
    environment = as_list(raw_environment)
    if not isinstance(raw_environment, list):
        result.error(f"镜号{shot_no} environment_cues 必须是数组。")
    if re.search(r"阳光|白光|灯光|光线|阴影|明暗|冷暖|光源|脚步声|摩擦声|碰撞声|呼唤声|响起|传来", descriptive) and not environment:
        result.error(f"镜号{shot_no} 环境声光承担画面信息时必须填写 environment_cues。")
    for index, cue in enumerate(environment, start=1):
        label = f"镜号{shot_no} environment_cues[{index}]"
        if not isinstance(cue, dict):
            result.error(f"{label} 必须是对象。")
            continue
        if clean_text(cue.get("kind")) not in ENVIRONMENT_CUE_KINDS:
            result.error(f"{label}.kind 不合法。")
        if clean_text(cue.get("basis")) not in ENVIRONMENT_CUE_BASES:
            result.error(f"{label}.basis 不合法。")
        description = clean_text(cue.get("description"))
        if not description or description not in camera:
            result.error(f"{label}.description 必须逐字出现于画面与机位。")
        evidence = normalized_list(cue.get("evidence_fact_ids"))
        if not evidence or not set(evidence).issubset(covered):
            result.error(f"{label}.evidence_fact_ids 必须是本镜已覆盖事实的非空子集。")

    raw_breaks = shot.get("chain_breaks")
    if not isinstance(raw_breaks, list):
        result.error(f"镜号{shot_no} chain_breaks 必须是数组。")
    for index, chain_break in enumerate(as_list(raw_breaks), start=1):
        label = f"镜号{shot_no} chain_breaks[{index}]"
        if not isinstance(chain_break, dict):
            result.error(f"{label} 必须是对象。")
            continue
        if not re.fullmatch(PERFORMANCE_CHAIN_ID_PATTERN, clean_text(chain_break.get("chain_id"))):
            result.error(f"{label}.chain_id 不合法。")
        if not is_json_integer(chain_break.get("after_step"), minimum=1):
            result.error(f"{label}.after_step 必须是大于等于1的 JSON 整数。")
        if clean_text(chain_break.get("reason")) not in PERFORMANCE_CHAIN_BREAK_REASONS:
            result.error(f"{label}.reason 不合法。")


def validate_director_adjacent_pair(
    data: dict[str, Any], prev: dict[str, Any], curr: dict[str, Any], result: ValidationResult
) -> None:
    if not uses_director_integrity_guard(data):
        return
    prev_no, curr_no = prev.get("shot_no"), curr.get("shot_no")
    prev_geometry = as_dict(prev.get("camera_geometry"))
    curr_geometry = as_dict(curr.get("camera_geometry"))
    same_axis = clean_text(prev_geometry.get("axis_id")) and (
        clean_text(prev_geometry.get("axis_id")) == clean_text(curr_geometry.get("axis_id"))
    )
    delta = clean_text(curr_geometry.get("angle_delta_from_previous"))
    batch_id = batch_id_for_scene(data, curr.get("scene_id"))
    overlapping_subjects = set(normalized_list(prev_geometry.get("primary_subjects"))) & set(
        normalized_list(curr_geometry.get("primary_subjects"))
    )
    overlapping_visible = set(normalized_list(prev.get("visible_characters"))) & set(
        normalized_list(curr.get("visible_characters"))
    )
    same_position = clean_text(prev_geometry.get("camera_position")) == clean_text(
        curr_geometry.get("camera_position")
    )
    if delta == "not_comparable" and (overlapping_subjects or overlapping_visible or (same_axis and same_position)):
        overlap = overlapping_subjects | overlapping_visible
        reason = (
            f"可见人物或主体仍有 {', '.join(sorted(overlap))}"
            if overlap
            else "同一轴线且 camera_position 未改变"
        )
        result.error(
            f"镜号{prev_no}-{curr_no} {reason}，"
            "不得用 not_comparable 绕过相邻机位角度变化登记。"
        )
    if delta in {"same", "minor"}:
        token = f"shot:{curr_no}:jump-cut"
        if "[跳切说明]" in clean_text(curr.get("notes")) and approved_review(
            data, "GATE_B", batch_id=batch_id, approved_item=token
        ):
            result.warn(f"镜号{prev_no}-{curr_no} 有意保留同轴小于等价30度变化的跳切，已登记 [跳切说明]。")
        else:
            result.error(
                f"镜号{prev_no}-{curr_no} 同轴相邻机位变化为 {delta or 'missing'}；"
                "仅改变景别不能替代等价30度机位变化，必须合并、改机位或经 Gate B 批准跳切。"
            )
    if delta == "substantial" and clean_text(prev_geometry.get("camera_position")) == clean_text(
        curr_geometry.get("camera_position")
    ):
        result.error(f"镜号{prev_no}-{curr_no} 声称 substantial 机位变化，但 camera_position 未改变。")
    sides = {clean_text(prev_geometry.get("axis_side")), clean_text(curr_geometry.get("axis_side"))}
    is_axis_cross = bool(same_axis and sides == {"side_a", "side_b"})
    if delta == "axis_cross" and not is_axis_cross:
        result.error(f"镜号{curr_no} 声明 axis_cross，但相邻镜并未在同一 AX 轴线两侧切换。")
    if is_axis_cross:
        token = f"shot:{curr_no}:axis-cross"
        if "[越轴说明]" in clean_text(curr.get("notes")) and approved_review(
            data, "GATE_B", batch_id=batch_id, approved_item=token
        ):
            result.warn(f"镜号{prev_no}-{curr_no} 有意越轴，已登记 [越轴说明]。")
        else:
            result.error(f"镜号{prev_no}-{curr_no} 从轴线一侧直接切到另一侧，必须重设轴线或经 Gate B 批准越轴。")

    reversals = {
        ("screen_left", "screen_right"),
        ("screen_right", "screen_left"),
        ("toward_camera", "away_camera"),
        ("away_camera", "toward_camera"),
    }
    previous_directions = screen_direction_map(prev_geometry)
    current_directions = screen_direction_map(curr_geometry)
    updated_entities = {
        clean_text(update.get("entity"))
        for update in as_list(curr.get("continuity_updates"))
        if isinstance(update, dict) and clean_text(update.get("field")) in {"position", "facing"}
    }
    for key in sorted(set(previous_directions) & set(current_directions)):
        if (previous_directions[key], current_directions[key]) in reversals and key[0] not in updated_entities:
            result.error(
                f"镜号{prev_no}-{curr_no} 的 {key[0]} {key[1]} 屏幕方向反转，"
                "但 continuity_updates 未登记位置或朝向变化。"
            )

    prev_signature = source_span_signature(prev)
    curr_signature = source_span_signature(curr)
    if prev_signature and prev_signature == curr_signature:
        reuse = curr.get("source_reuse")
        token = f"shot:{curr_no}:source-reuse"
        valid_reuse = (
            isinstance(reuse, dict)
            and reuse.get("from_shot_no") == prev_no
            and clean_text(reuse.get("reason"))
            in {"simultaneous_must_isolate", "single_sentence_multi_action", "source_line_indivisible"}
            and "[原文复用]" in clean_text(curr.get("notes"))
            and approved_review(data, "GATE_B", batch_id=batch_id, approved_item=token)
        )
        if valid_reuse:
            result.warn(f"镜号{prev_no}-{curr_no} 复用同一原文跨度，已按 [原文复用] 登记并批准。")
        else:
            result.error(
                f"镜号{prev_no}-{curr_no} 使用完全相同 source_span；默认视为重复拆镜，"
                "只有结构性不可合并时才可登记 source_reuse 并经 Gate B 批准。"
            )


def validate_fact_cut_contract(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_hybrid_fields(data):
        return
    seen_groups: set[str] = set()
    structured_required = requires_structured_audit(data)
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        beat_id = str(beat.get("beat_id", ""))
        for fact in as_list(beat.get("facts")):
            if not isinstance(fact, dict):
                continue
            fact_id = str(fact.get("fact_id", ""))
            priority = clean_text(fact.get("cut_priority"))
            if priority not in CUT_PRIORITIES:
                result.error(f"{fact_id} cut_priority 必须为 normal / recommended / must_isolate。")
            if "cut_reasons" not in fact:
                result.error(f"{fact_id} 缺少 cut_reasons。")
                reasons: list[str] = []
            else:
                reasons = normalized_list(fact.get("cut_reasons"))
            for reason in reasons:
                if reason not in SPLIT_REASONS:
                    result.error(f"{fact_id} cut_reasons 不合法：{reason}")
            if priority in {"recommended", "must_isolate"} and not reasons:
                result.error(f"{fact_id} cut_priority 为 {priority} 时 cut_reasons 不能为空。")
            group = clean_text(fact.get("cut_group"))
            if not group:
                result.error(f"{fact_id} 缺少 cut_group。")
            else:
                seen_groups.add(group)
            if structured_required:
                category = clean_text(fact.get("cut_category"))
                moment_id = clean_text(fact.get("cut_moment_id"))
                if category not in CUT_CATEGORIES:
                    result.error(f"{fact_id} cut_category 必须为 {' / '.join(sorted(CUT_CATEGORIES))}。")
                if not moment_id:
                    result.error(f"{fact_id} 缺少 cut_moment_id。")
            reason_set = set(reasons)
            if reason_set & MUST_ISOLATE_REASONS and priority != "must_isolate":
                result.error(f"{fact_id} 包含关键切点 {sorted(reason_set & MUST_ISOLATE_REASONS)}，cut_priority 必须为 must_isolate。")
            if reason_set & CRITICAL_SPLIT_REASONS and priority == "normal":
                result.error(f"{fact_id} 包含关键切点 {sorted(reason_set & CRITICAL_SPLIT_REASONS)}，cut_priority 不得为 normal。")
            if group and not group.startswith(beat_id):
                result.warn(f"{fact_id} cut_group 建议以所属 Beat 开头，当前为 {group}。")
    if not seen_groups:
        result.error("2.3.4 数据必须登记事实级 cut_group。")


def validate_fact_text_preservation(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_structured_audit(data):
        return
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        source = clean_text(beat.get("source_text"))
        if not source:
            continue
        for fact in as_list(beat.get("facts")):
            if not isinstance(fact, dict):
                continue
            missing = unsupported_fact_terms(clean_text(fact.get("text")), source)
            if missing:
                result.warn(
                    f"{fact.get('fact_id')} Fact 文本疑似引入 source_text 未出现词组："
                    f"{', '.join(missing[:5])}。请确认不是编造事实。"
                )
            if is_current_contract(data) and clean_text(fact.get("type")) == "dialogue":
                allowed = dialogue_candidates(source)
                fact_text = normalize_script_text(fact.get("text")).strip()
                candidates = dialogue_candidates(fact_text)
                if fact_text and not quoted_segments(fact_text) and not re.search(r"[：:]", fact_text):
                    candidates.add(fact_text)
                if not candidates or not candidates.issubset(allowed):
                    result.error(f"{fact.get('fact_id')} dialogue fact 必须逐字命中所属 Beat 的原文对白。")


def validate_must_isolate_density(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_structured_audit(data):
        return
    scene_stats: dict[str, dict[str, Any]] = {}
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        scene = clean_text(beat.get("scene"))
        stats = scene_stats.setdefault(scene, {"beats": 0, "must_beats": 0, "facts": []})
        stats["beats"] += 1
        beat_must = False
        for fact in as_list(beat.get("facts")):
            if isinstance(fact, dict) and clean_text(fact.get("cut_priority")) == "must_isolate":
                beat_must = True
                stats["facts"].append(str(fact.get("fact_id")))
        if beat_must:
            stats["must_beats"] += 1
    for scene, stats in scene_stats.items():
        beats = int(stats["beats"])
        if beats <= 0:
            continue
        ratio = int(stats["must_beats"]) / beats
        if ratio > MUST_ISOLATE_DENSITY_THRESHOLD:
            result.warn(
                f"场景 {scene} must_isolate Beat 占比 {ratio:.0%}（{stats['must_beats']}/{beats} Beats），"
                f"可能存在过度标记；请复核 {', '.join(stats['facts'][:8])} 是否应降级为 recommended。"
            )


def validate_overcompression(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_hybrid_fields(data):
        return
    structured_required = requires_structured_audit(data)
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    facts_by_id = fact_lookup(data)
    total_duration: int | float = 0 if is_current_contract(data) else 0.0
    short_3 = 0
    long_6_or_more = 0
    has_multi_beat_critical = False
    lexicon_patterns = project_lexicon_patterns(data)
    for shot in shots:
        shot_no = shot.get("shot_no")
        duration = contract_duration_number(
            shot.get("duration_seconds"), strict_contract=is_current_contract(data)
        )
        total_duration += duration
        short_3 += int(duration <= 3)
        long_6_or_more += int(duration >= 6)
        notes = clean_text(shot.get("notes"))
        shot_type = clean_text(shot.get("shot_type"))
        beat_count = len(as_list(shot.get("beat_ids")))
        facts = shot_fact_objects(shot, facts_by_id)
        critical_facts = [fact for fact in facts if is_fact_critical(fact)]
        must_groups = shot_cut_groups(facts, must_only=True)
        must_facts = [fact for fact in facts if clean_text(fact.get("cut_priority")) == "must_isolate"]
        structural_groups = structural_cut_groups(
            must_facts,
            must_only=False,
            include_category=not requires_human_gate(data),
        )
        cut_points = find_required_cut_point_categories(
            clean_text(shot.get("source_paragraph")) + "\n" + clean_text(shot.get("camera_main_image")),
            lexicon_patterns,
            current_contract=is_current_contract(data),
        )
        if beat_count >= 3 and critical_facts:
            has_multi_beat_critical = True
            result.error(f"镜号{shot_no} 覆盖 {beat_count} 个 Beat 且包含关键切点，属于过度压缩。")
        if beat_count >= 2 and critical_facts:
            has_multi_beat_critical = True
        if structured_required:
            if len(structural_groups) > 1:
                result.error(
                    f"镜号{shot_no} 同时覆盖多个 must_isolate 结构切点："
                    f"{', '.join(structured_key_label(key) for key in sorted(structural_groups))}；"
                    "必须拆镜，不能只靠 cut_group 或 cut_category 合并。"
                )
            if len(must_facts) > 1 and len(structural_groups) == 1 and "[不可拆说明]" not in notes:
                result.error(f"镜号{shot_no} 合并同一结构切点的多个 must_isolate 事实时，备注必须写 [不可拆说明]。")
        elif len(must_groups) > 1:
            result.error(
                f"镜号{shot_no} 同时覆盖多个 must_isolate cut_group：{', '.join(sorted(must_groups))}；"
                "除同一不可拆瞬间外必须拆镜。"
            )
        if shot_type == "master" and critical_facts:
            result.error(f"镜号{shot_no} master 镜不得承载关键切点，应拆为 transition / vfx_anchor / insert / action。")
        if shot_type == "dialogue" and critical_facts:
            if len(must_groups) > 1 or duration >= 8 or len(cut_points) >= 2:
                result.error(f"镜号{shot_no} dialogue 镜承载多个/过长关键切点，必须拆分真相、声源或反应落点。")
        if shot_type in INSERT_LIKE_TYPES and duration > 5:
            supports = normalized_list(shot.get("long_take_support"))
            if "[保留理由]" not in notes or len(supports) < 2:
                result.error(f"镜号{shot_no} {shot_type} 超过5秒，必须写 [保留理由] 且至少两项 long_take_support。")
        if 9 <= duration <= 11:
            if "[长镜头]" not in notes or "[保留理由]" not in notes:
                result.error(f"镜号{shot_no} 9-11秒高风险长镜头必须写 [长镜头] 和 [保留理由]。")
        if duration >= 12:
            missing = [label for label in ["[长镜头]", "[保留理由]", "[不可拆说明]"] if label not in notes]
            if missing:
                result.error(f"镜号{shot_no} 12秒及以上镜头缺少 {'、'.join(missing)}。")
        if duration >= 9 and len(cut_points) >= 2:
            result.error(
                f"镜号{shot_no} {fmt_number(duration)}秒镜头包含多个必拆切点：{', '.join(cut_points)}。"
            )
        source = clean_text(shot.get("source_paragraph"))
        camera = clean_text(shot.get("camera_main_image"))
        visible = set(normalized_list(shot.get("visible_characters")))
        for speaker in extract_source_speakers(source):
            if speaker not in visible and re.search(rf"{re.escape(speaker)}[^。；;\n]{{0,8}}画外声", camera):
                if not source_marks_offscreen_voice(source, speaker):
                    result.error(f"镜号{shot_no} 将原文现场对白角色 {speaker} 改为画外声，疑似绕过多人关系镜头。")
    cut_per_minute = cut_rate_per_minute(len(shots), total_duration)
    short_ratio = round(short_3 / len(shots), 4) if shots else 0
    long_ratio = round(long_6_or_more / len(shots), 4) if shots else 0
    if cut_per_minute and cut_per_minute < 10.5:
        message = f"cut/min {cut_per_minute} 低于 10.5，疑似整体压缩。"
        if has_multi_beat_critical:
            result.error(message + " 且存在多 Beat 关键镜头，必须重拆。")
        else:
            result.warn(message)
    if shots and short_ratio < 0.10:
        message = f"短镜比例 {short_ratio:.2%} 低于 10%，疑似缺少必要插镜。"
        if has_multi_beat_critical:
            result.error(message + " 且存在多 Beat 关键镜头，必须重拆。")
        else:
            result.warn(message)
    if uses_director_integrity_guard(data):
        if cut_per_minute > 12.5:
            result.warn(f"节奏预警：cut/min {cut_per_minute} 高于 12.5，疑似过度拆镜，必须由人工审核处置。")
        if short_ratio > 0.18:
            result.warn(f"节奏预警：短镜比例 {short_ratio:.2%} 高于 18%，疑似碎切，必须由人工审核处置。")
        if shots and long_ratio < 0.40:
            result.warn(f"节奏预警：长镜比例 {long_ratio:.2%} 低于 40%，连续表演承载不足，必须由人工审核处置。")


def validate_performance_chains_and_constraints(data: dict[str, Any], result: ValidationResult) -> None:
    if not uses_director_integrity_guard(data):
        return
    facts = fact_lookup(data)
    fact_to_beat: dict[str, str] = {}
    fact_to_scene: dict[str, str] = {}
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        beat_id = clean_text(beat.get("beat_id"))
        scene_id = clean_text(beat.get("scene_id"))
        for fact in as_list(beat.get("facts")):
            if isinstance(fact, dict):
                fact_id = clean_text(fact.get("fact_id"))
                fact_to_beat[fact_id] = beat_id
                fact_to_scene[fact_id] = scene_id
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    shot_by_no = {shot.get("shot_no"): shot for shot in shots}
    shots_by_fact: dict[str, set[int]] = {}
    for shot in shots:
        if not is_json_integer(shot.get("shot_no"), minimum=1):
            continue
        for fact_id in normalized_list(shot.get("covered_fact_ids")):
            shots_by_fact.setdefault(fact_id, set()).add(int(shot["shot_no"]))

    raw_chains = data.get("performance_chains")
    if not isinstance(raw_chains, list):
        result.error("2.4.5 数据必须包含 performance_chains 数组，可为空。")
    seen_chains: set[str] = set()
    chain_metadata: dict[str, tuple[str, int]] = {}
    for index, chain in enumerate(as_list(raw_chains), start=1):
        label = f"performance_chains[{index}]"
        if not isinstance(chain, dict):
            result.error(f"{label} 必须是对象。")
            continue
        chain_id = clean_text(chain.get("chain_id"))
        if not re.fullmatch(PERFORMANCE_CHAIN_ID_PATTERN, chain_id):
            result.error(f"{label}.chain_id 必须符合 PC[0-9]{{3,}}。")
        if chain_id in seen_chains:
            result.error(f"performance_chain.chain_id 重复：{chain_id}")
        seen_chains.add(chain_id)
        scene_id = clean_text(chain.get("scene_id"))
        scene_log = next(
            (
                log
                for log in as_list(data.get("continuity_logs"))
                if isinstance(log, dict) and clean_text(log.get("scene_id")) == scene_id
            ),
            {},
        )
        if not re.fullmatch(SCENE_ID_PATTERN, scene_id):
            result.error(f"{label}.scene_id 不合法。")
        character = clean_text(chain.get("character"))
        if character not in declared_character_names(as_dict(scene_log)):
            result.error(f"{label}.character 必须精确匹配本场角色台账。")
        steps = as_list(chain.get("steps"))
        if not isinstance(chain.get("steps"), list) or len(steps) < 2:
            result.error(f"{label}.steps 必须至少包含2个动作/反应/台词步骤。")
        step_fact_sets: list[set[str]] = []
        chain_fact_ids: set[str] = set()
        for step_index, step in enumerate(steps, start=1):
            step_label = f"{label}.steps[{step_index}]"
            if not isinstance(step, dict):
                result.error(f"{step_label} 必须是对象。")
                step_fact_sets.append(set())
                continue
            if clean_text(step.get("role")) not in PERFORMANCE_CHAIN_ROLES:
                result.error(f"{step_label}.role 必须为 action / reaction / dialogue。")
            evidence = normalized_list(step.get("fact_ids"))
            if not isinstance(step.get("fact_ids"), list) or not evidence:
                result.error(f"{step_label}.fact_ids 必须是非空数组。")
            if len(evidence) != len(set(evidence)):
                result.error(f"{step_label}.fact_ids 不得重复。")
            repeated_across_steps = chain_fact_ids & set(evidence)
            if repeated_across_steps:
                result.error(f"{step_label}.fact_ids 与前序步骤重复：{', '.join(sorted(repeated_across_steps))}")
            chain_fact_ids.update(evidence)
            for fact_id in evidence:
                if fact_id not in facts or fact_to_scene.get(fact_id) != scene_id:
                    result.error(f"{step_label} 引用了不存在或不属于本场的事实：{fact_id}")
                    continue
                fact = facts[fact_id]
                fact_type = clean_text(fact.get("type"))
                if fact_type == "dialogue" and clean_text(fact.get("speaker")) != character:
                    result.error(f"{step_label} 的 dialogue speaker 与 chain.character 不一致。")
                if fact_type in {"action", "emotion"} and character not in normalized_list(
                    fact.get("performers")
                ):
                    result.error(f"{step_label} 的 {fact_type} performers 未包含 chain.character。")
            step_fact_sets.append(set(evidence))
        if chain_id:
            chain_metadata[chain_id] = (scene_id, len(steps))

        def common_covering_shots(groups: list[set[str]]) -> set[int]:
            fact_ids = set().union(*groups) if groups else set()
            return (
                set.intersection(*(shots_by_fact.get(fact_id, set()) for fact_id in fact_ids))
                if fact_ids
                else set()
            )

        common_shots = common_covering_shots(step_fact_sets)
        if common_shots:
            continue
        approved_break_shots: dict[int, int] = {}
        for shot in shots:
            shot_no = shot.get("shot_no")
            for item in as_list(shot.get("chain_breaks")):
                if not isinstance(item, dict) or clean_text(item.get("chain_id")) != chain_id:
                    continue
                boundary = item.get("after_step")
                token = f"shot:{shot_no}:performance-chain-break"
                if (
                    is_json_integer(boundary, minimum=1)
                    and boundary < len(step_fact_sets)
                    and clean_text(item.get("reason")) in PERFORMANCE_CHAIN_BREAK_REASONS
                    and "[动作链拆分]" in clean_text(shot.get("notes"))
                    and approved_review(
                        data,
                        "GATE_B",
                        batch_id=batch_id_for_scene(data, shot.get("scene_id")),
                        approved_item=token,
                    )
                ):
                    approved_break_shots[int(boundary)] = int(shot_no)

        boundaries = sorted(approved_break_shots)
        segment_edges = [0, *boundaries, len(step_fact_sets)]
        valid_partition = bool(boundaries)
        segment_common_shots: list[set[int]] = []
        for start, end in zip(segment_edges, segment_edges[1:]):
            segment_shots = common_covering_shots(step_fact_sets[start:end])
            segment_common_shots.append(segment_shots)
            if not segment_shots:
                valid_partition = False
        for boundary_index, boundary in enumerate(boundaries, start=1):
            later_shots = segment_common_shots[boundary_index]
            if not later_shots or approved_break_shots[boundary] != min(later_shots):
                valid_partition = False
                break

        if valid_partition:
            for boundary in boundaries:
                result.warn(
                    f"{chain_id} 在步骤 {boundary}/{boundary + 1} 之间拆镜，已登记 [动作链拆分]。"
                )
        else:
            result.error(
                f"{chain_id} 的同一人物动作→反应→台词链未落在同一镜；"
                "必须整链合并，或用经 Gate B 精确批准的 chain_breaks 将整链划分为各自可共同覆盖的连续段。"
            )

    registered_chain_facts = [
        {
            fact_id
            for step in as_list(chain.get("steps"))
            if isinstance(step, dict)
            for fact_id in normalized_list(step.get("fact_ids"))
        }
        for chain in as_list(raw_chains)
        if isinstance(chain, dict)
    ]
    performance_sequence: list[tuple[str, str, str, str] | None] = []
    role_by_type = {"action": "action", "emotion": "reaction", "dialogue": "dialogue"}
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        scene_id = clean_text(beat.get("scene_id"))
        for fact in as_list(beat.get("facts")):
            if not isinstance(fact, dict):
                performance_sequence.append(None)
                continue
            role = role_by_type.get(clean_text(fact.get("type")))
            if role is None:
                if clean_text(fact.get("type")) == "reality":
                    performance_sequence.append(None)
                continue
            if role == "dialogue":
                actor = clean_text(fact.get("speaker"))
            else:
                performers = normalized_list(fact.get("performers"))
                actor = performers[0] if len(performers) == 1 else ""
            if not actor:
                performance_sequence.append(None)
                continue
            performance_sequence.append((scene_id, actor, role, clean_text(fact.get("fact_id"))))
    for previous, current in zip(performance_sequence, performance_sequence[1:]):
        if previous is None or current is None or previous[:2] != current[:2]:
            continue
        fact_pair = {previous[3], current[3]}
        if shots_by_fact.get(previous[3], set()) & shots_by_fact.get(current[3], set()):
            continue
        if any(fact_pair.issubset(chain_facts) for chain_facts in registered_chain_facts):
            continue
        result.error(
            f"检测到 {previous[1]} 的连续 {previous[2]}→{current[2]} 事实 "
            f"{previous[3]} / {current[3]} 被拆到不同镜头；必须同镜，或先登记 performance_chain 再走 Gate B 拆分例外。"
        )

    for shot in shots:
        shot_no = shot.get("shot_no")
        for chain_break in as_list(shot.get("chain_breaks")):
            if not isinstance(chain_break, dict):
                continue
            chain_id = clean_text(chain_break.get("chain_id"))
            metadata = chain_metadata.get(chain_id)
            if metadata is None:
                result.error(f"镜号{shot_no} chain_breaks 引用了未登记 performance_chain：{chain_id}")
                continue
            chain_scene, step_count = metadata
            if clean_text(shot.get("scene_id")) != chain_scene:
                result.error(f"镜号{shot_no} chain_breaks.{chain_id} 与 performance_chain 场景不一致。")
            after_step = chain_break.get("after_step")
            if is_json_integer(after_step, minimum=1) and int(after_step) >= step_count:
                result.error(f"镜号{shot_no} chain_breaks.{chain_id}.after_step 必须小于步骤数 {step_count}。")

        reuse = shot.get("source_reuse")
        if reuse is None:
            continue
        if not isinstance(reuse, dict):
            result.error(f"镜号{shot_no} source_reuse 必须是对象。")
            continue
        from_no = reuse.get("from_shot_no")
        reason = clean_text(reuse.get("reason"))
        if not is_json_integer(from_no, minimum=1) or not is_json_integer(shot_no, minimum=1) or from_no != shot_no - 1:
            result.error(f"镜号{shot_no} source_reuse.from_shot_no 必须精确指向同场上一镜。")
            continue
        previous = shot_by_no.get(from_no)
        if previous is None or clean_text(previous.get("scene_id")) != clean_text(shot.get("scene_id")):
            result.error(f"镜号{shot_no} source_reuse 必须指向存在的同场上一镜。")
        if reason not in {"simultaneous_must_isolate", "single_sentence_multi_action", "source_line_indivisible"}:
            result.error(f"镜号{shot_no} source_reuse.reason 不合法。")
        if previous is not None and source_span_signature(previous) != source_span_signature(shot):
            result.error(f"镜号{shot_no} source_reuse 只能用于与上一镜完全相同的 source_span(s)。")

    raw_constraints = data.get("directorial_constraints")
    if not isinstance(raw_constraints, list):
        result.error("2.4.5 数据必须包含 directorial_constraints 数组，可为空。")
    seen_constraints: set[str] = set()
    for index, constraint in enumerate(as_list(raw_constraints), start=1):
        label = f"directorial_constraints[{index}]"
        if not isinstance(constraint, dict):
            result.error(f"{label} 必须是对象。")
            continue
        constraint_id = clean_text(constraint.get("constraint_id"))
        if not re.fullmatch(DIRECTORIAL_CONSTRAINT_ID_PATTERN, constraint_id):
            result.error(f"{label}.constraint_id 必须符合 DC[0-9]{{3,}}。")
        if constraint_id in seen_constraints:
            result.error(f"directorial_constraint.constraint_id 重复：{constraint_id}")
        seen_constraints.add(constraint_id)
        constraint_type = clean_text(constraint.get("type"))
        if constraint_type not in DIRECTORIAL_CONSTRAINT_TYPES:
            result.error(f"{label}.type 不合法。")
        freeze_gate = clean_text(constraint.get("freeze_gate"))
        if freeze_gate not in {"GATE_A", "GATE_B"}:
            result.error(f"{label}.freeze_gate 必须为 GATE_A / GATE_B。")
        origin = constraint.get("origin")
        if not isinstance(origin, dict) or clean_text(origin.get("gate")) not in {"GATE_A", "GATE_B"} or not is_json_integer(
            origin.get("round"), minimum=1
        ):
            result.error(f"{label}.origin 必须包含 gate=GATE_A/GATE_B 与正整数 round。")
        elif clean_text(origin.get("gate")) != freeze_gate:
            result.error(f"{label}.origin.gate 必须与 freeze_gate 一致。")
        fact_ids = normalized_list(constraint.get("fact_ids"))
        if constraint_type in {"same_shot", "same_beat"}:
            if not isinstance(constraint.get("fact_ids"), list) or len(fact_ids) < 2:
                result.error(f"{label}.fact_ids 至少需要2项。")
            if any(fact_id not in facts for fact_id in fact_ids):
                result.error(f"{label}.fact_ids 含不存在事实。")
        if constraint_type == "same_shot" and fact_ids:
            owners = [shots_by_fact.get(fact_id, set()) for fact_id in fact_ids]
            if not owners or not set.intersection(*owners):
                result.error(f"{constraint_id} 要求的事实未落在同一镜。")
        if constraint_type == "same_beat" and len({fact_to_beat.get(fact_id) for fact_id in fact_ids}) > 1:
            result.error(f"{constraint_id} 要求的事实不属于同一 Beat。")
        if constraint_type == "forbid_literal":
            literal = clean_text(constraint.get("literal"))
            raw_scene_scope = constraint.get("scene_ids", [])
            raw_shot_scope = constraint.get("shot_nos", [])
            scene_scope = set(normalized_list(raw_scene_scope))
            shot_scope = set(as_list(raw_shot_scope))
            if not isinstance(raw_scene_scope, list) or any(
                not isinstance(scene_id, str) or not re.fullmatch(SCENE_ID_PATTERN, scene_id)
                for scene_id in as_list(raw_scene_scope)
            ) or len(scene_scope) != len(as_list(raw_scene_scope)):
                result.error(f"{label}.scene_ids 必须是去重合法场景 ID 数组。")
            if not isinstance(raw_shot_scope, list) or any(
                not is_json_integer(shot_no, minimum=1) for shot_no in as_list(raw_shot_scope)
            ) or len(shot_scope) != len(as_list(raw_shot_scope)):
                result.error(f"{label}.shot_nos 必须是去重正整数数组。")
            if not literal or (not scene_scope and not shot_scope):
                result.error(f"{label} forbid_literal 必须填写 literal 及 scene_ids/shot_nos 范围。")
            for shot in shots:
                if clean_text(shot.get("scene_id")) not in scene_scope and shot.get("shot_no") not in shot_scope:
                    continue
                searchable = "\n".join(
                    [
                        clean_text(shot.get("camera_main_image")),
                        clean_text(shot.get("prompt")),
                        " ".join(normalized_list(shot.get("visible_props"))),
                    ]
                )
                if literal and literal in searchable:
                    result.error(f"{constraint_id} 禁止字面元素“{literal}”，但镜号{shot.get('shot_no')} 出现了该元素。")
        if constraint_type == "direct_handoff":
            from_fact_id = clean_text(constraint.get("from_fact_id"))
            target_scene = clean_text(constraint.get("to_scene_id"))
            forbidden = clean_text(constraint.get("forbidden_literal"))
            if from_fact_id not in facts or not re.fullmatch(SCENE_ID_PATTERN, target_scene) or not forbidden:
                result.error(f"{label} direct_handoff 必须填写有效 from_fact_id、to_scene_id、forbidden_literal。")
            target_shots = [shot for shot in shots if clean_text(shot.get("scene_id")) == target_scene]
            if not target_shots:
                result.error(f"{constraint_id} direct_handoff 目标场景没有首镜可校验。")
            elif forbidden in clean_text(target_shots[0].get("camera_main_image")):
                result.error(f"{constraint_id} 要求直接承接，但目标场首镜再次出现禁用元素“{forbidden}”。")
            from_shots = shots_by_fact.get(from_fact_id, set())
            if not from_shots:
                result.error(f"{constraint_id} direct_handoff 的 from_fact_id 未落入任何镜头。")
            elif target_shots and max(from_shots) >= int(target_shots[0].get("shot_no") or 0):
                result.error(f"{constraint_id} direct_handoff 的来源事实必须早于目标场首镜。")


def validate_rows(
    data: dict[str, Any],
    result: ValidationResult,
    fact_to_beat: dict[str, str],
    fact_types: dict[str, str],
    all_fact_ids: set[str],
    continuity_logs: dict[str, dict[str, Any]],
    beat_scenes: dict[str, str],
) -> None:
    raw_shots = data.get("shots")
    if not isinstance(raw_shots, list):
        result.error("shots 必须是数组。")
    shots = as_list(raw_shots)
    if not shots:
        result.error("shots 不能为空。")
        return
    expected_no = 1
    covered: set[str] = set()
    scenes = set(continuity_logs)
    seen_scenes: set[str] = set()
    has_update = False
    hybrid_required = requires_hybrid_fields(data)
    structured_required = requires_structured_audit(data)
    use_scene_id = requires_human_gate(data)
    beat_positions = beat_order_index(data) if is_current_contract(data) else {}
    facts_by_id = fact_lookup(data)
    scene_states = {
        scene: initial_scene_states(log)
        for scene, log in continuity_logs.items()
    } if structured_required else {}
    final_scene_documents = {
        scene: copy.deepcopy(log)
        for scene, log in continuity_logs.items()
    }
    for shot in shots:
        if not isinstance(shot, dict):
            result.error("shots 中每一项必须是对象。")
            continue
        shot_no = shot.get("shot_no")
        if is_current_contract(data) and not is_json_integer(shot_no, minimum=1):
            result.error(f"镜号必须是从1开始的 JSON 整数：{shot_no}")
        if shot_no != expected_no:
            result.error(f"镜号必须从1连续递增：期望 {expected_no}，实际 {shot_no}")
        expected_no += 1
        for key in ["scene", "source_paragraph", "duration_seconds", "duration_breakdown", "camera_main_image", "notes", "prompt", "beat_ids", "covered_fact_ids", "continuity_updates"]:
            if key not in shot:
                result.error(f"镜号{shot_no} 缺少字段：{key}")
        if "keyframe" in shot:
            result.error(f"镜号{shot_no} 不得包含 keyframe 字段；关键帧列已从主表删除。")
        if shot.get("old_shot_no") is not None:
            result.error(f"镜号{shot_no} 不得保留 old_shot_no；禁止后处理合并痕迹进入交付数据。")
        scene = scene_key_for(data, shot)
        if use_scene_id:
            raw_scene_id = shot.get("scene_id")
            if is_current_contract(data) and (
                not isinstance(raw_scene_id, str)
                or raw_scene_id != clean_text(raw_scene_id)
                or not re.fullmatch(SCENE_ID_PATTERN, raw_scene_id)
            ):
                result.error(f"镜号{shot_no} scene_id 必须是 S[0-9]{{2,}} 格式字符串。")
            if not scene:
                result.error(f"镜号{shot_no} 缺少 scene_id。")
            if not clean_text(shot.get("scene")):
                result.error(f"镜号{shot_no} 缺少展示用 scene。")
            elif scene in continuity_logs and clean_text(shot.get("scene")) != clean_text(
                continuity_logs[scene].get("scene")
            ):
                result.error(f"镜号{shot_no} scene 与 continuity_logs[{scene}].scene 不一致。")
        if scenes and scene not in scenes:
            if is_current_contract(data):
                result.error(f"镜号{shot_no} 场景键未在 continuity_logs 中登记：{scene}")
            else:
                result.warn(f"镜号{shot_no} 场景键未在 continuity_logs 中登记：{scene}")
        is_first_scene_shot = str(scene) not in seen_scenes
        if is_first_scene_shot:
            seen_scenes.add(str(scene))
            validate_first_scene_shot(
                shot,
                as_dict(continuity_logs.get(str(scene))),
                result,
                fact_types,
                current_contract=is_current_contract(data),
            )
        beat_ids = [str(item) for item in as_list(shot.get("beat_ids"))]
        fact_ids = [str(item) for item in as_list(shot.get("covered_fact_ids"))]
        if is_current_contract(data):
            if not isinstance(shot.get("beat_ids"), list):
                result.error(f"镜号{shot_no} beat_ids 必须是数组。")
            if not isinstance(shot.get("covered_fact_ids"), list):
                result.error(f"镜号{shot_no} covered_fact_ids 必须是数组。")
            if len(beat_ids) != len(set(beat_ids)):
                result.error(f"镜号{shot_no} beat_ids 不得重复。")
            if len(fact_ids) != len(set(fact_ids)):
                result.error(f"镜号{shot_no} covered_fact_ids 不得重复。")
        if not beat_ids:
            result.error(f"镜号{shot_no} beat_ids 不能为空。")
        for beat_id in beat_ids:
            if beat_id not in beat_scenes:
                result.error(f"镜号{shot_no} 引用了不存在的 Beat：{beat_id}")
            elif is_current_contract(data) and beat_scenes.get(beat_id) != scene:
                result.error(
                    f"镜号{shot_no} 的 scene_id={scene} 与 {beat_id} 的 scene_id={beat_scenes.get(beat_id)} 不一致。"
                )
        if is_current_contract(data) and len(beat_ids) > 1:
            positions = [beat_positions.get(beat_id) for beat_id in beat_ids]
            if all(position is not None for position in positions):
                numeric_positions = [int(position) for position in positions if position is not None]
                strictly_increasing = all(
                    current > previous for previous, current in zip(numeric_positions, numeric_positions[1:])
                )
                if not strictly_increasing:
                    result.error(f"镜号{shot_no} beat_ids 必须按 beat_order 严格递增。")
                has_gap = strictly_increasing and any(
                    current != previous + 1
                    for previous, current in zip(numeric_positions, numeric_positions[1:])
                )
            else:
                has_gap = False
            if has_gap:
                notes = clean_text(shot.get("notes"))
                reason = re.search(r"\[非连续Beat\]\s*(?:[：:]\s*)?([^\[\]\n]+)", notes)
                if reason is None or not clean_text(reason.group(1)):
                    result.error(f"镜号{shot_no} 覆盖非连续 Beat，备注必须包含 [非连续Beat] 及非空原因。")
        safety_exception = approved_safety_exception(data, shot)
        if not fact_ids and not safety_exception:
            result.error(f"镜号{shot_no} covered_fact_ids 不能为空。")
        for fact_id in fact_ids:
            covered.add(fact_id)
            owner = fact_to_beat.get(fact_id)
            if owner is None:
                result.error(f"镜号{shot_no} 覆盖了不存在的事实 ID：{fact_id}")
            elif owner not in beat_ids:
                result.error(f"镜号{shot_no} 覆盖事实 {fact_id}，但 beat_ids 未包含 {owner}。")
        if use_scene_id and clean_text(shot.get("shot_type")) in {"transition", "safety"} and fact_ids:
            bound_types = fact_types_for_shot(shot, fact_types)
            if not (bound_types & {"space", "sound", "reality"}):
                result.error(f"镜号{shot_no} {shot.get('shot_type')} 镜必须绑定空间、声音或现实层事实。")
        if is_current_contract(data):
            character_values: dict[str, set[str]] = {}
            for field_name in ("visible_characters", "offscreen_characters"):
                raw_characters = shot.get(field_name)
                if not isinstance(raw_characters, list):
                    result.error(f"镜号{shot_no} {field_name} 必须是数组。")
                    character_values[field_name] = set()
                    continue
                if any(
                    not isinstance(item, str) or not item or item != clean_text(item)
                    for item in raw_characters
                ):
                    result.error(
                        f"镜号{shot_no} {field_name} 每项必须是无首尾空白的非空字符串。"
                    )
                canonical_items = [item for item in raw_characters if isinstance(item, str) and item]
                if len(canonical_items) != len(set(canonical_items)):
                    result.error(f"镜号{shot_no} {field_name} 不得包含重复角色。")
                character_values[field_name] = {
                    item for item in canonical_items if item == clean_text(item)
                }
            visible = character_values["visible_characters"]
            offscreen = character_values["offscreen_characters"]
            overlap = visible & offscreen
            if overlap:
                result.error(f"镜号{shot_no} 角色不得同时可见又在画外：{', '.join(sorted(overlap))}")
            allowed_markers = set(NOTE_MARKERS if uses_director_integrity_guard(data) else LEGACY_NOTE_MARKERS)
            for marker in re.findall(r"\[[^\]\n]+\]", clean_text(shot.get("notes"))):
                if marker not in allowed_markers:
                    result.error(f"镜号{shot_no} 备注包含封闭清单外标签：{marker}")
        validate_duration(
            shot,
            result,
            version_24=use_scene_id,
            strict_contract=is_current_contract(data),
        )
        validate_camera(data, shot, result)
        validate_prompt(shot, result)
        validate_dialogue_fidelity(data, shot, result)
        validate_hybrid_fields(
            shot,
            result,
            required=hybrid_required,
            strict_contract=is_current_contract(data),
        )
        validate_shot_director_integrity(
            data,
            shot,
            as_dict(continuity_logs.get(str(scene))),
            facts_by_id,
            result,
            is_first_scene_shot=is_first_scene_shot,
        )
        if is_current_contract(data) and not isinstance(shot.get("continuity_updates"), list):
            result.error(f"镜号{shot_no} continuity_updates 必须是数组。")
        updates = as_list(shot.get("continuity_updates"))
        if updates:
            has_update = True
        has_station_move = "【站位位移】" in clean_text(shot.get("camera_main_image"))
        if has_station_move and not updates:
            result.error(f"镜号{shot_no} 写了【站位位移】，但 continuity_updates 为空。")
        shot_sound_fact_ids = {
            fact_id
            for fact_id in fact_ids
            if clean_text(fact_types.get(fact_id)) == "sound"
        }
        for update in updates:
            current_states = scene_states.get(str(scene)) if structured_required else None
            validate_update(
                shot,
                update,
                result,
                has_station_move,
                current_states=current_states,
                known_sound_sources=sound_source_ids(as_dict(continuity_logs.get(str(scene)))),
                sound_fact_ids=shot_sound_fact_ids,
                current_contract=is_current_contract(data),
            )
            scene_document = final_scene_documents.get(str(scene))
            if scene_document is not None:
                applied = apply_update_to_scene_document(scene_document, update)
                if (
                    is_current_contract(data)
                    and isinstance(update, dict)
                    and clean_text(update.get("field")) in STATEFUL_CONTINUITY_FIELDS
                    and not applied
                ):
                    result.error(
                        f"镜号{shot_no} continuity_update 无法写入结构化连续性台账；"
                        "目标实体及字段必须在当前场景初态中可定位。"
                    )
        if is_current_contract(data):
            covered_facts = [facts_by_id.get(fact_id) for fact_id in fact_ids]
            requires_new_sound_source = "new_sound_source" in normalized_list(shot.get("split_reason")) or any(
                isinstance(fact, dict) and "new_sound_source" in fact_cut_reasons(fact)
                for fact in covered_facts
            )
            if requires_new_sound_source:
                sound_fact_ids = {
                    fact_id
                    for fact_id in fact_ids
                    if isinstance(facts_by_id.get(fact_id), dict)
                    and clean_text(facts_by_id[fact_id].get("type")) == "sound"
                }
                sound_updates = [
                    update
                    for update in updates
                    if isinstance(update, dict) and clean_text(update.get("entity_type")) == "sound_source"
                ]
                if not sound_updates:
                    result.error(f"镜号{shot_no} new_sound_source 必须登记 sound_source continuity_update。")
                for update in sound_updates:
                    evidence = set(normalized_list(update.get("evidence_fact_ids")))
                    if not (evidence & sound_fact_ids):
                        result.error(
                            f"镜号{shot_no} new_sound_source update 必须由当前镜头覆盖的 sound fact 取证。"
                        )
    missing = sorted(all_fact_ids - covered)
    if missing:
        result.error("存在未覆盖事实 ID：" + ", ".join(missing[:20]))
    if any("【站位位移】" in clean_text(shot.get("camera_main_image")) for shot in shots if isinstance(shot, dict)) and not has_update:
        result.error("分镜出现站位位移，但没有任何 continuity_updates。")
    validate_final_inherited_states(data, continuity_logs, final_scene_documents, result)
    validate_adjacent_shot_design(
        data,
        shots,
        fact_types,
        result,
        guard_reverse_depth_motion=requires_adjacent_motion_guard(data),
        allow_marked_exceptions=use_scene_id,
    )


def validate_duration(
    shot: dict[str, Any],
    result: ValidationResult,
    *,
    version_24: bool = False,
    strict_contract: bool = False,
) -> None:
    shot_no = shot.get("shot_no")
    if strict_contract and not is_json_integer(shot.get("duration_seconds"), minimum=1):
        result.error(f"镜号{shot_no} duration_seconds 必须是大于0的 JSON 整数。")
        return
    if strict_contract:
        duration: int | float = int(shot["duration_seconds"])
    else:
        duration_value = safe_finite_number(shot.get("duration_seconds"))
        if duration_value is None:
            result.error(f"镜号{shot_no} 镜头时长必须是有限数字。")
            return
        duration = duration_value
    if duration <= 0:
        result.error(f"镜号{shot_no} 镜头时长必须大于0。")
    breakdown = as_dict(shot.get("duration_breakdown"))
    for field_name in DURATION_FIELDS:
        if field_name not in breakdown:
            result.error(f"镜号{shot_no} duration_breakdown 缺少 {field_name}。")
            return
    values: list[int | float] = []
    for field_name in DURATION_FIELDS:
        value = breakdown.get(field_name)
        if strict_contract:
            if not is_json_integer(value, minimum=0):
                result.error(f"镜号{shot_no} duration_breakdown.{field_name} 必须是非负 JSON 整数。")
                return
            values.append(value)
        else:
            numeric = safe_finite_number(value)
            if numeric is None:
                result.error(f"镜号{shot_no} duration_breakdown 必须全为有限数字。")
                return
            if version_24 and (numeric < 0 or not numeric.is_integer()):
                result.error(f"镜号{shot_no} duration_breakdown.{field_name} 必须是非负整数。")
            values.append(numeric)
    expected_value = max(values[0], values[1]) + values[2] + values[3]
    expected = expected_value if strict_contract else math.ceil(expected_value)
    matches = duration == expected if strict_contract else int(duration) == expected
    if not matches:
        result.error(
            f"镜号{shot_no} 时长不符合 max(同步动作, 同步对白) + 非同步动作 + 情绪留白 公式："
            f"表格 {duration:g} 秒，公式 {expected} 秒。"
        )
    quoted = "".join(re.findall(r"[“\"]([^”\"]+)[”\"]", clean_text(shot.get("camera_main_image"))))
    spoken_chars = len(re.sub(r"[，。！？……、\s]", "", quoted))
    if spoken_chars:
        min_dialogue_seconds = math.ceil(spoken_chars / 4)
        if values[1] < min_dialogue_seconds:
            result.error(f"镜号{shot_no} 对白时长过短：{spoken_chars}字至少约 {min_dialogue_seconds} 秒。")
    notes = clean_text(shot.get("notes"))
    if duration > 6 and "[时长估算]" not in notes:
        result.error(f"镜号{shot_no} 超过6秒，备注必须写 [时长估算]。")
    if duration > 10:
        long_take = as_dict(shot.get("long_take"))
        classification = clean_text(long_take.get("classification"))
        if version_24:
            if classification not in LONG_TAKE_CLASSIFICATIONS or classification == "not_applicable":
                result.error(f"镜号{shot_no} 超过10秒，long_take.classification 必须为有效长镜分类。")
        elif long_take.get("classification") in (None, "not_applicable") and "[长镜头]" not in notes:
            result.error(f"镜号{shot_no} 超过10秒，必须标注长镜头分类或人工复核理由。")


def validate_camera(
    data: dict[str, Any],
    shot: dict[str, Any],
    result: ValidationResult,
) -> None:
    shot_no = shot.get("shot_no")
    camera = clean_text(shot.get("camera_main_image"))
    lines = [line.strip() for line in camera.split("\n") if line.strip()]
    if len(lines) < 2:
        result.error(f"镜号{shot_no} 运镜列至少需要三联首行和【机位逻辑】。")
        return
    if not re.fullmatch(r"\[[^,\]]+,\s*[^,\]]+,\s*[^\]]+\]", lines[0]):
        result.error(f"镜号{shot_no} 运镜首行必须是三要素：[视角/高度, 景别/特殊视角, 运镜方式]。")
    angle, shot_size, movement = parse_triad(camera)
    validate_camera_movement_compatibility(shot_no, angle, shot_size, movement, result, clean_text(shot.get("notes")))
    if not lines[1].startswith("【机位逻辑】"):
        result.error(f"镜号{shot_no} 运镜第二行必须以【机位逻辑】开头。")
    descriptive_camera = descriptive_text_without_dialogue(camera)
    forbidden_words = list(FORBIDDEN_VISUAL_WORDS)
    if uses_emotion_performance_guard(data):
        forbidden_words.extend(EMOTION_PERFORMANCE_FORBIDDEN_WORDS)
    for word in forbidden_words:
        if word in descriptive_camera:
            result.error(f"镜号{shot_no} 运镜列包含禁用表达：{word}")
    if uses_emotion_performance_guard(data):
        precise_match = PRECISE_EXECUTION_VALUE_PATTERN.search(descriptive_camera)
        if precise_match:
            result.error(
                f"镜号{shot_no} 运镜列包含精确执行数值：{precise_match.group(0)}；"
                "改用相对距离、相对角度、冷暖倾向或可见结果。"
            )
    for marker in UNSUPPORTED_MARKERS:
        if marker in camera or marker in clean_text(shot.get("notes")):
            result.error(f"镜号{shot_no} 包含后处理/无依据标记：{marker}")
    if "【镜内变化】" in camera:
        result.error(f"镜号{shot_no} 不得在最终主表出现【镜内变化】。")
    if "运镜修正" in clean_text(shot.get("notes")):
        result.error(f"镜号{shot_no} 备注不得保留运镜修正痕迹。")


def validate_camera_movement_compatibility(
    shot_no: Any,
    angle: str,
    shot_size: str,
    movement: str,
    result: ValidationResult,
    notes: str = "",
) -> None:
    if "斯坦尼康" in clean_text(movement) and has_any_keyword(shot_size, ("大远景", "大全景")):
        result.error(f"镜号{shot_no} 大远景/大全景不得使用斯坦尼康；应改用摇臂、伸缩摇臂或航拍。")
    if "航拍" in clean_text(movement):
        if not has_any_keyword(shot_size, AERIAL_ALLOWED_SHOT_SIZE_KEYWORDS):
            result.error(f"镜号{shot_no} 航拍必须服务于大范围空间，景别应为大远景/大全景/全景。")
        if not has_any_keyword(angle, AERIAL_ALLOWED_ANGLE_KEYWORDS):
            result.error(f"镜号{shot_no} 航拍必须匹配俯拍、高角度、正俯拍或微俯视。")
    size_text = f"{angle} {shot_size}"
    if shot_size_has_span(size_text):
        has_span_movement = has_any_keyword(movement, SPAN_MOVEMENT_KEYWORDS)
        has_span_note = "[景别跨度]" in clean_text(notes)
        if not has_span_movement or not has_span_note:
            result.error(
                f"镜号{shot_no} 景别同时包含远景类和近景/特写类，"
                "必须使用光学变焦、快速推拉、急推/急拉、推轨推进/拉出或伸缩摇臂等跨度运镜，"
                "并在备注写 [景别跨度] 理由。"
            )


def validate_prompt(shot: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    prompt = clean_text(shot.get("prompt"))
    expected = derive_prompt(shot)
    if prompt != expected:
        result.error(f"镜号{shot_no} Prompt 必须由锁定主表派生，且与标准五字段模板一致。")
    if len(prompt) > 800:
        result.warn(f"镜号{shot_no} Prompt 总长超过800字符，需申请回流 Gate B 裁决拆镜或保留。")
    lines = [line.strip() for line in prompt.split("\n") if line.strip()]
    if len(lines) != 5:
        result.error(f"镜号{shot_no} Prompt 必须且只能包含五行。")
    for field_name in PROMPT_FIELDS:
        if not any(line.startswith(field_name) for line in lines):
            result.error(f"镜号{shot_no} Prompt 缺少字段：{field_name}")
    for item in PROMPT_FORBIDDEN:
        if item in prompt:
            result.error(f"镜号{shot_no} Prompt 包含禁止进入派生列的内部/关键帧内容：{item}")
    if not quotes_are_balanced(prompt):
        result.error(f"镜号{shot_no} Prompt 引号不闭合。")
    fields = prompt_fields(prompt)
    composition = fields.get("构图：", "")
    if composition and composition_has_field_leak(composition):
        result.error(f"镜号{shot_no} Prompt 构图字段混入对白、声音或动作链；构图只能写主体/位置/道具关系。")


def quoted_segments(value: Any) -> list[str]:
    text = normalize_script_text(value)
    values: list[str] = []
    for match in re.finditer(r'“([^”]*)”|"([^"]*)"', text):
        segment = match.group(1) if match.group(1) is not None else match.group(2)
        if segment is not None:
            values.append(segment)
    return values


def dialogue_candidates(value: Any) -> set[str]:
    text = normalize_script_text(value)
    candidates = {segment for segment in quoted_segments(text) if segment}
    labels = list(
        re.finditer(
            r"[\u4e00-\u9fa5A-Za-z0-9·]{1,8}(?:（VO）|\(VO\)|VO|画外声)?[：:]",
            text,
        )
    )
    for index, match in enumerate(labels):
        next_label = labels[index + 1].start() if index + 1 < len(labels) else len(text)
        next_line = text.find("\n", match.end())
        end = min(next_label, next_line if next_line >= 0 else len(text))
        candidate = text[match.end():end].strip()
        if candidate:
            candidates.add(candidate)
    return candidates


_DIALOGUE_LABEL_RE = re.compile(
    r"(?P<speaker>[\u4e00-\u9fa5A-Za-z0-9·]{0,8}?)(?P<vo>（VO）|\(VO\)|\bVO\b|画外声)?[：:]",
    re.IGNORECASE,
)
_VO_MARKER_RE = re.compile(r"(?:（VO）|\(VO\)|\bVO\b|画外声)", re.IGNORECASE)


def normalize_dialogue_speaker(value: Any) -> str:
    speaker = clean_text(value)
    speaker = re.sub(r"(?:说道|说|问|喊|答|道)$", "", speaker).strip()
    return "" if speaker in {"画外声", "VO"} else speaker


def dialogue_records(value: Any) -> list[tuple[str, str, bool]]:
    text = normalize_script_text(value)
    records: list[tuple[str, str, bool]] = []
    pending: tuple[str, bool] | None = None
    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        labels = list(_DIALOGUE_LABEL_RE.finditer(line))
        if labels:
            for index, match in enumerate(labels):
                end = labels[index + 1].start() if index + 1 < len(labels) else len(line)
                speaker = normalize_dialogue_speaker(match.group("speaker"))
                is_vo = bool(match.group("vo"))
                declared = line[match.end():end].strip()
                texts = quoted_segments(declared) or ([declared] if declared else [])
                for dialogue in texts:
                    if dialogue:
                        records.append((speaker, dialogue, is_vo))
                pending = (speaker, is_vo) if not declared and index == len(labels) - 1 else None
            continue
        if pending is not None:
            speaker, is_vo = pending
            texts = quoted_segments(line) or [line]
            records.extend((speaker, dialogue, is_vo) for dialogue in texts if dialogue)
            pending = None
    return records


def vo_dialogue_candidates(value: Any) -> set[str]:
    return {dialogue for _speaker, dialogue, is_vo in dialogue_records(value) if is_vo}


def speaker_bound_to_vo(value: str, known_speakers: set[str]) -> str:
    for speaker in sorted((item for item in known_speakers if item), key=len, reverse=True):
        escaped = re.escape(speaker)
        if re.search(rf"{escaped}\s*(?:（VO）|\(VO\)|\bVO\b|画外声)", value, re.IGNORECASE):
            return speaker
        if re.search(rf"(?:（VO）|\(VO\)|\bVO\b|画外声)\s*[-—：:]?\s*{escaped}", value, re.IGNORECASE):
            return speaker
    return ""


def vo_claim_for_quote(
    text: str,
    match: re.Match[str],
    known_speakers: set[str] | None = None,
) -> tuple[bool, str]:
    speakers = known_speakers or set()
    line_start = text.rfind("\n", 0, match.start()) + 1
    line_end_value = text.find("\n", match.end())
    line_end = line_end_value if line_end_value >= 0 else len(text)
    line = text[line_start:line_end]
    prefix = text[line_start:match.start()]
    labels = list(_DIALOGUE_LABEL_RE.finditer(prefix))
    speaker = normalize_dialogue_speaker(labels[-1].group("speaker")) if labels else ""
    marked = bool(_VO_MARKER_RE.search(line))
    if marked:
        return True, speaker_bound_to_vo(line, speakers) or speaker
    previous_end = line_start - 1
    if previous_end <= 0 or prefix.strip():
        return False, speaker
    previous_start = text.rfind("\n", 0, previous_end) + 1
    previous_line = text[previous_start:previous_end]
    if not _VO_MARKER_RE.search(previous_line):
        return False, speaker
    previous_labels = list(_DIALOGUE_LABEL_RE.finditer(previous_line))
    previous_speaker = speaker_bound_to_vo(previous_line, speakers) or (
        normalize_dialogue_speaker(previous_labels[-1].group("speaker")) if previous_labels else ""
    )
    return True, previous_speaker


def validate_dialogue_fidelity(data: dict[str, Any], shot: dict[str, Any], result: ValidationResult) -> None:
    if not is_current_contract(data):
        return
    shot_no = shot.get("shot_no")
    source = normalize_script_text(shot.get("source_paragraph"))
    facts_by_id = fact_lookup(data)
    dialogue_facts = [
        normalize_script_text(fact.get("text"))
        for fact_id in normalized_list(shot.get("covered_fact_ids"))
        for fact in [facts_by_id.get(fact_id)]
        if fact is not None and clean_text(fact.get("type")) == "dialogue"
    ]
    allowed = dialogue_candidates(source)
    source_records = dialogue_records(source)
    known_speakers = {speaker for speaker, _dialogue, _is_vo in source_records if speaker}
    allowed_vo = {dialogue for _speaker, dialogue, is_vo in source_records if is_vo}
    allowed_vo_pairs = {(speaker, dialogue) for speaker, dialogue, is_vo in source_records if is_vo and speaker}
    for fact_text in dialogue_facts:
        allowed.update(dialogue_candidates(fact_text))
        if fact_text and not quoted_segments(fact_text) and not re.search(r"[：:]", fact_text):
            allowed.add(fact_text.strip())
    for field_name in ("camera_main_image", "prompt"):
        field_text = normalize_script_text(shot.get(field_name))
        if re.search(r"(?:\bVO\b|画外声)", field_text, re.IGNORECASE) and not re.search(
            r"(?:\bVO\b|画外声)", source, re.IGNORECASE
        ):
            result.error(f"镜号{shot_no} {field_name} 将现场内容泛化为 VO/画外声，原文无对应标记。")
        for segment in quoted_segments(shot.get(field_name)):
            if not segment:
                result.error(f"镜号{shot_no} {field_name} 不得包含空对白引号。")
                continue
            if segment in allowed:
                pass
            else:
                result.error(
                    f"镜号{shot_no} {field_name} 引号内对白未逐字命中 source_paragraph 或 dialogue fact：{segment}"
                )
        for match in re.finditer(r'“([^”]*)”|"([^"]*)"', field_text):
            segment = match.group(1) if match.group(1) is not None else match.group(2)
            marked_vo, claimed_speaker = vo_claim_for_quote(field_text, match, known_speakers)
            if not marked_vo:
                continue
            if claimed_speaker:
                allowed_claim = (claimed_speaker, segment) in allowed_vo_pairs
            else:
                allowed_claim = segment in allowed_vo
            if not allowed_claim:
                result.error(
                    f"镜号{shot_no} {field_name} 画外对白未按说话人及原文逐字命中 VO/画外声："
                    f"{claimed_speaker or '<unspecified>'}:{segment}"
                )


def validate_update(
    shot: dict[str, Any],
    update: Any,
    result: ValidationResult,
    has_station_move: bool,
    *,
    current_states: dict[tuple[str, str, str], str] | None = None,
    known_sound_sources: set[str] | None = None,
    sound_fact_ids: set[str] | None = None,
    current_contract: bool = False,
) -> None:
    shot_no = shot.get("shot_no")
    if not isinstance(update, dict):
        result.error(f"镜号{shot_no} continuity_updates 每项必须是对象。")
        return
    required = ["entity_type", "entity", "field", "from", "to", "evidence_fact_ids"]
    for key in required:
        if key not in update:
            result.error(f"镜号{shot_no} continuity_update 缺少 {key}。")
    if update.get("from") == update.get("to"):
        result.error(f"镜号{shot_no} continuity_update 的 from/to 不得相同。")
    evidence = [str(item) for item in as_list(update.get("evidence_fact_ids"))]
    covered = {str(item) for item in as_list(shot.get("covered_fact_ids"))}
    if not evidence:
        result.error(f"镜号{shot_no} continuity_update 必须有 evidence_fact_ids。")
    for fact_id in evidence:
        if fact_id not in covered:
            result.error(f"镜号{shot_no} continuity_update 证据 {fact_id} 未被当前镜头覆盖。")
    if update.get("field") in POSITION_UPDATE_FIELDS and update.get("entity_type") == "character" and not has_station_move:
        result.error(f"镜号{shot_no} 登记了位置/朝向迁移，但主画面缺少【站位位移】。")
    if current_contract:
        for field_key in ("entity_type", "entity", "field"):
            raw_value = update.get(field_key)
            if not isinstance(raw_value, str) or raw_value != clean_text(raw_value):
                result.error(
                    f"镜号{shot_no} continuity_update.{field_key} 必须是无首尾空白的 canonical 字符串。"
                )
        entity_type = clean_text(update.get("entity_type"))
        entity = clean_text(update.get("entity"))
        field_name = clean_text(update.get("field"))
        if entity_type not in CONTINUITY_ENTITY_TYPES:
            result.error(
                f"镜号{shot_no} continuity_update.entity_type 必须为 "
                "character / prop / fixed_object / sound_source / reality_layer。"
            )
        if field_name not in STATEFUL_CONTINUITY_FIELDS:
            result.error(
                f"镜号{shot_no} continuity_update.field 必须为 "
                "position / facing / state / owner / value / presence / visibility。"
            )
        if entity_type == "reality_layer":
            if entity:
                result.error(f"镜号{shot_no} reality_layer update.entity 必须为空字符串。")
        elif entity_type in CONTINUITY_ENTITY_TYPES and not entity:
            result.error(f"镜号{shot_no} {entity_type} update.entity 必须是非空字符串。")
        if any(
            not isinstance(value, str) or not value or value != clean_text(value)
            for value in (update.get("from"), update.get("to"))
        ):
            result.error(
                f"镜号{shot_no} continuity_update.from/to 必须是无首尾空白的非空字符串。"
            )
    if current_contract and clean_text(update.get("entity_type")) == "sound_source":
        raw_source_id = update.get("entity")
        source_id = clean_text(raw_source_id)
        field_name = clean_text(update.get("field"))
        if not re.fullmatch(SOUND_SOURCE_ID_PATTERN, source_id) or source_id not in (known_sound_sources or set()):
            result.error(f"镜号{shot_no} continuity_update 引用了未知 sound_source：{source_id}")
        if field_name not in SOUND_SOURCE_UPDATE_FIELDS:
            result.error(
                f"镜号{shot_no} sound_source update.field 必须为 position / state / visibility。"
            )
        if not (set(evidence) & (sound_fact_ids or set())):
            result.error(
                f"镜号{shot_no} sound_source update 必须由当前镜头覆盖的 sound fact 取证。"
            )
        if field_name == "visibility":
            if clean_text(update.get("from")) not in SOUND_SOURCE_VISIBILITIES or clean_text(
                update.get("to")
            ) not in SOUND_SOURCE_VISIBILITIES:
                result.error(
                    f"镜号{shot_no} sound_source visibility from/to 必须为 onscreen / offscreen。"
                )
    if current_states is not None and clean_text(update.get("field")) in STATEFUL_CONTINUITY_FIELDS:
        key = state_key(update.get("entity_type"), update.get("entity"), update.get("field"))
        current = current_states.get(key)
        from_value = clean_text(update.get("from"))
        to_value = clean_text(update.get("to"))
        if current is None:
            message = f"镜号{shot_no} continuity_update 缺少可校验的上一状态：{key[0]}/{key[1]}/{key[2]}。"
            if current_contract:
                result.error(message)
            else:
                result.warn(message)
        elif current != from_value:
            result.error(
                f"镜号{shot_no} continuity_update.from 与上一状态不一致："
                f"{key[0]}/{key[1]}/{key[2]} 当前为 {current}，from 写为 {from_value}。"
            )
        if to_value:
            current_states[key] = to_value


def validate_data(
    data: dict[str, Any],
    *,
    strict_status: bool,
    final_signoff: bool = False,
    data_path: Path | None = None,
    workspace_root: Path | None = None,
) -> ValidationResult:
    result = ValidationResult()
    validate_json_compatible(data, result)
    if result.errors:
        return result
    validate_metadata(data, result, strict_status=strict_status, final_signoff=final_signoff)
    validate_project_lexicon(data, result)
    validate_batch_plan(data, result)
    validate_human_reviews(data, result, final_signoff=final_signoff)
    validate_script_lock(data, result, data_path=data_path, workspace_root=workspace_root)
    continuity_logs = validate_continuity_logs(data, result)
    fact_to_beat, fact_types, all_fact_ids, beat_scenes = collect_facts(data, result)
    validate_fact_cut_contract(data, result)
    validate_fact_text_preservation(data, result)
    validate_rows(data, result, fact_to_beat, fact_types, all_fact_ids, continuity_logs, beat_scenes)
    validate_performance_chains_and_constraints(data, result)
    validate_overcompression(data, result)
    validate_must_isolate_density(data, result)
    validate_warn_resolutions(
        data,
        result,
        require_all=final_signoff or not is_current_contract(data),
    )
    if strict_status and is_current_contract(data):
        expected_errors = list(result.errors)
        expected_warnings = list(result.warnings)
        validate_report_truth(
            data,
            result,
            expected_errors=expected_errors,
            expected_warnings=expected_warnings,
            final_signoff=final_signoff,
        )
    return result


def markdown_text(data: dict[str, Any]) -> str:
    lines = ["# 分镜表", ""]
    lines.append("| " + " | ".join(HEADERS) + " |")
    lines.append("| " + " | ".join(["---"] * len(HEADERS)) + " |")
    for row in expected_rows(data):
        lines.append("| " + " | ".join(escape_cell(item) for item in row) + " |")
    report = as_dict(data.get("validation_report"))
    if report:
        lines.extend(
            [
                "",
                "## 校验摘要",
                "",
                f"- status: {report.get('status', '')}",
                f"- shot_count: {report.get('shot_count', '')}",
                f"- column_contract: {report.get('column_contract', '')}",
            ]
        )
        if requires_structured_audit(data):
            for key, value in report_summary_items(data):
                lines.append(f"- {key}: {value}")
    return "\n".join(lines) + "\n"


def build_markdown(data: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(markdown_text(data), encoding="utf-8")


def build_excel(data: dict[str, Any], path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl 不可用，无法生成 Excel。")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    for row in expected_rows(data):
        sheet.append([excel_cell_value(value) for value in row])
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, width in enumerate([8, 24, 44, 14, 72, 36, 56], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str):
                cell.data_type = "s"
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 36 if row_index == 1 else 108
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if requires_structured_audit(data):
        summary = workbook.create_sheet(SUMMARY_SHEET_NAME)
        summary.append(["字段", "值"])
        for key, value in report_summary_items(data):
            summary.append([canonical_cell_text(key), canonical_cell_text(value)])
        for cell in summary[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        summary.column_dimensions["A"].width = 24
        summary.column_dimensions["B"].width = 96
        for row in summary.iter_rows(min_row=2):
            for cell in row:
                cell.data_type = "s"
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def build_report(data: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    report = as_dict(data.get("validation_report"))
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")


def parse_markdown_table(path: Path) -> list[list[str]]:
    text = path.read_text(encoding="utf-8-sig")
    rows: list[list[str]] = []
    for line in text.splitlines():
        if not line.startswith("|"):
            continue
        cells = [unescape_cell(cell.strip()) for cell in line.strip().strip("|").split("|")]
        if cells and all(re.fullmatch(r":?-{3,}:?", cell) for cell in cells):
            continue
        rows.append(cells)
    return rows


def compare_markdown(data: dict[str, Any], path: Path, result: ValidationResult) -> None:
    if not path.exists():
        result.error(f"Markdown 文件不存在：{path}")
        return
    try:
        rows = parse_markdown_table(path)
    except (OSError, UnicodeError) as exc:
        result.error(f"Markdown 文件损坏或不可读：{exc}")
        return
    if not rows:
        result.error("Markdown 未找到分镜表。")
        return
    if rows[0] != HEADERS:
        result.error("Markdown 表头必须是稳定 7 列，且不得包含关键帧列。")
    expected = [[canonical_cell_text(item) for item in row] for row in expected_rows(data)]
    actual = [[canonical_cell_text(item) for item in row] for row in rows[1:]]
    if actual != expected:
        result.error("Markdown 表格内容与 shot_data 派生结果不一致。")
    if requires_structured_audit(data):
        text = path.read_text(encoding="utf-8-sig")
        for key, value in report_summary_items(data):
            expected_line = f"- {key}: {value}"
            if expected_line not in text:
                result.error(f"Markdown 校验摘要缺少或不一致：{key}")


def compare_excel(data: dict[str, Any], path: Path, result: ValidationResult) -> None:
    if load_workbook is None:
        result.error("openpyxl 不可用，无法校验 Excel。")
        return
    if not path.exists():
        result.error(f"Excel 文件不存在：{path}")
        return
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        result.error(f"Excel 文件损坏或不可读：{exc}")
        return
    try:
        expected_sheets = [SHEET_NAME, SUMMARY_SHEET_NAME] if requires_structured_audit(data) else [SHEET_NAME]
        if workbook.sheetnames != expected_sheets:
            result.error(f"Excel Sheet 必须为：{', '.join(expected_sheets)}。")
            return
        sheet = workbook[SHEET_NAME]
        rows = list(sheet.iter_rows(values_only=True))
        summary_rows = []
        if requires_structured_audit(data):
            summary_rows = list(workbook[SUMMARY_SHEET_NAME].iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        result.error("Excel 分镜表为空。")
        return
    header = [canonical_cell_text(item) for item in rows[0]]
    if header != HEADERS:
        result.error("Excel 表头必须是稳定 7 列，且不得包含关键帧列。")
    expected = [[canonical_cell_text(item) for item in row] for row in expected_rows(data)]
    actual = [[canonical_cell_text(item) for item in row] for row in rows[1:]]
    if actual != expected:
        result.error("Excel 内容与 shot_data 派生结果不一致。")
    if requires_structured_audit(data):
        expected_summary = [("字段", "值"), *report_summary_items(data)]
        actual_summary = [(canonical_cell_text(row[0]), canonical_cell_text(row[1])) for row in summary_rows]
        if actual_summary != [(canonical_cell_text(key), canonical_cell_text(value)) for key, value in expected_summary]:
            result.error("Excel 校验摘要与 shot_data.validation_report 不一致。")


def compare_report(data: dict[str, Any], path: Path, result: ValidationResult) -> None:
    if not path.exists():
        result.error(f"validation_report 文件不存在：{path}")
        return
    try:
        report = json.loads(path.read_text(encoding="utf-8-sig"), parse_constant=reject_json_constant)
    except Exception as exc:
        result.error(f"validation_report 不是合法 JSON：{exc}")
        return
    if report != as_dict(data.get("validation_report")):
        result.error("validation_report 文件与 shot_data.validation_report 不一致。")


def print_result(result: ValidationResult) -> None:
    if result.errors:
        print("ERRORS:", file=sys.stderr)
        for item in result.errors:
            print(f"- {item}", file=sys.stderr)
    if result.warnings:
        print("WARNINGS:")
        for item in result.warnings:
            print(f"- {item}")
    print(f"{result.status}: storyboard validation {'passed' if not result.errors else 'failed'}.")


def _temporary_sibling(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=f".tmp{path.suffix}", dir=path.parent)
    os.close(descriptor)
    return Path(name)


def _restore_bytes(path: Path, payload: bytes) -> None:
    temporary = _temporary_sibling(path)
    try:
        temporary.write_bytes(payload)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def atomic_write_delivery(
    data: dict[str, Any],
    *,
    data_path: Path,
    markdown_path: Path,
    excel_path: Path,
    report_path: Path,
    workspace_root: Path,
    final_signoff: bool,
) -> None:
    destinations = [data_path, markdown_path, excel_path, report_path]
    resolved = [path.resolve(strict=False) for path in destinations]
    if len(set(resolved)) != len(resolved):
        raise ValueError("data / markdown / excel / report 输出路径必须互不相同。")
    temporary_paths = {path: _temporary_sibling(path) for path in destinations}
    try:
        write_json(temporary_paths[data_path], data)
        build_markdown(data, temporary_paths[markdown_path])
        build_excel(data, temporary_paths[excel_path])
        build_report(data, temporary_paths[report_path])

        self_result = ValidationResult()
        try:
            rendered_data = load_json(temporary_paths[data_path])
        except (OSError, UnicodeError, ValueError, json.JSONDecodeError) as exc:
            self_result.error(f"临时 shot_data 无法回读：{exc}")
        else:
            if rendered_data != data:
                self_result.error("临时 shot_data 回读内容不一致。")
            strict_result = validate_data(
                rendered_data,
                strict_status=True,
                final_signoff=final_signoff,
                data_path=data_path,
                workspace_root=workspace_root,
            )
            for error in strict_result.errors:
                self_result.error(f"临时 shot_data 严格校验失败：{error}")
        compare_markdown(data, temporary_paths[markdown_path], self_result)
        compare_excel(data, temporary_paths[excel_path], self_result)
        compare_report(data, temporary_paths[report_path], self_result)
        if self_result.errors:
            raise RuntimeError("交付物临时自校验失败：" + " | ".join(self_result.errors))

        backups: dict[Path, bytes | None] = {
            path: path.read_bytes() if path.exists() else None for path in destinations
        }
        committed: list[Path] = []
        try:
            for path in destinations:
                os.replace(temporary_paths[path], path)
                committed.append(path)
        except Exception:
            for path in reversed(committed):
                original = backups[path]
                if original is None:
                    path.unlink(missing_ok=True)
                else:
                    _restore_bytes(path, original)
            raise
    finally:
        for temporary in temporary_paths.values():
            temporary.unlink(missing_ok=True)


def review_package_text(
    data: dict[str, Any], *, gate: str, batch_id: str | None, result: ValidationResult
) -> str:
    scene_scope = _batch_scene_ids(data, batch_id)
    shots = [
        shot
        for shot in as_list(data.get("shots"))
        if isinstance(shot, dict)
        and (scene_scope is None or clean_text(shot.get("scene_id")) in scene_scope)
    ]
    logs = [
        log
        for log in as_list(data.get("continuity_logs"))
        if isinstance(log, dict)
        and (scene_scope is None or clean_text(log.get("scene_id")) in scene_scope)
    ]
    lines = [
        "# Gate B 导演审核包",
        "",
        f"- version: {metadata_version(data)}",
        f"- rule_revision: {clean_text(as_dict(data.get('metadata')).get('rule_revision'))}",
        f"- gate: {gate}",
        f"- batch_id: {batch_id or ''}",
        f"- reviewed_hash: {gate_review_hash(data, gate, batch_id)}",
        f"- shot_count: {len(shots)}",
        "",
        "## 已冻结导演约束",
        "",
        "| ID | 类型 | 冻结门禁 | 原点 | 事实/范围 |",
        "| --- | --- | --- | --- | --- |",
    ]
    for constraint in as_list(data.get("directorial_constraints")):
        if not isinstance(constraint, dict):
            continue
        origin = as_dict(constraint.get("origin"))
        scope = normalized_list(constraint.get("fact_ids")) or normalized_list(constraint.get("scene_ids"))
        if constraint.get("shot_nos"):
            scope.extend(str(value) for value in as_list(constraint.get("shot_nos")))
        lines.append(
            "| "
            + " | ".join(
                escape_cell(value)
                for value in [
                    constraint.get("constraint_id"),
                    constraint.get("type"),
                    constraint.get("freeze_gate"),
                    f"{origin.get('gate', '')}#{origin.get('round', '')}",
                    ", ".join(scope),
                ]
            )
            + " |"
        )
    lines.extend(["", "## 场景轴线", "", "| 场景 | 轴线 | 类型 | 端点 | 两侧锚点 |", "| --- | --- | --- | --- | --- |"])
    for log in logs:
        for axis in as_list(log.get("spatial_axes")):
            if not isinstance(axis, dict):
                continue
            lines.append(
                "| "
                + " | ".join(
                    escape_cell(value)
                    for value in [
                        log.get("scene_id"),
                        axis.get("axis_id"),
                        axis.get("axis_type"),
                        f"{axis.get('endpoint_a', '')} ↔ {axis.get('endpoint_b', '')}",
                        f"{axis.get('side_a_anchor', '')} / {axis.get('side_b_anchor', '')}",
                    ]
                )
                + " |"
            )
    lines.extend(
        [
            "",
            "## 相邻镜头机位检查",
            "",
            "| 镜号 | 场景 | 轴线/侧 | 机位 | 相对角度 | 屏幕方向 | 原文复用 | 备注 |",
            "| --- | --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for shot in shots:
        geometry = as_dict(shot.get("camera_geometry"))
        directions = ", ".join(
            f"{item.get('entity', '')}:{item.get('kind', '')}={item.get('direction', '')}"
            for item in as_list(geometry.get("screen_directions"))
            if isinstance(item, dict)
        )
        lines.append(
            "| "
            + " | ".join(
                escape_cell(value)
                for value in [
                    shot.get("shot_no"),
                    shot.get("scene_id"),
                    f"{geometry.get('axis_id', '')}/{geometry.get('axis_side', '')}",
                    geometry.get("camera_position"),
                    geometry.get("angle_delta_from_previous"),
                    directions,
                    json.dumps(shot.get("source_reuse"), ensure_ascii=False, sort_keys=True)
                    if isinstance(shot.get("source_reuse"), dict)
                    else "",
                    shot.get("notes"),
                ]
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## 表演链",
            "",
            "| Chain | 场景 | 人物 | 步骤 |",
            "| --- | --- | --- | --- |",
        ]
    )
    for chain in as_list(data.get("performance_chains")):
        if not isinstance(chain, dict) or (scene_scope is not None and clean_text(chain.get("scene_id")) not in scene_scope):
            continue
        steps = "; ".join(
            f"{step.get('role', '')}:{','.join(normalized_list(step.get('fact_ids')))}"
            for step in as_list(chain.get("steps"))
            if isinstance(step, dict)
        )
        lines.append(
            "| "
            + " | ".join(
                escape_cell(value)
                for value in [chain.get("chain_id"), chain.get("scene_id"), chain.get("character"), steps]
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## 事实人物映射",
            "",
            "| 事实 | 类型 | 表演者/说话人 | 受话人 | 发声方式 |",
            "| --- | --- | --- | --- | --- |",
        ]
    )
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict) or (
            scene_scope is not None and clean_text(beat.get("scene_id")) not in scene_scope
        ):
            continue
        for fact in as_list(beat.get("facts")):
            if not isinstance(fact, dict) or clean_text(fact.get("type")) not in {
                "action",
                "emotion",
                "dialogue",
            }:
                continue
            fact_type = clean_text(fact.get("type"))
            mapped_people = (
                clean_text(fact.get("speaker"))
                if fact_type == "dialogue"
                else ", ".join(normalized_list(fact.get("performers")))
            )
            lines.append(
                "| "
                + " | ".join(
                    escape_cell(value)
                    for value in [
                        fact.get("fact_id"),
                        fact_type,
                        mapped_people,
                        ", ".join(normalized_list(fact.get("addressees"))),
                        fact.get("delivery"),
                    ]
                )
                + " |"
            )
    lines.extend(
        [
            "",
            "## 人物映射与表演证据",
            "",
            "| 镜号 | 可见 | 画外 | POV | 表演提示 | 动作推进 | 环境提示 |",
            "| --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for shot in shots:
        geometry = as_dict(shot.get("camera_geometry"))
        lines.append(
            "| "
            + " | ".join(
                escape_cell(value)
                for value in [
                    shot.get("shot_no"),
                    ", ".join(normalized_list(shot.get("visible_characters"))),
                    ", ".join(normalized_list(shot.get("offscreen_characters"))),
                    geometry.get("pov_character"),
                    len(as_list(shot.get("performance_cues"))),
                    len(as_list(shot.get("action_progressions"))),
                    len(as_list(shot.get("environment_cues"))),
                ]
            )
            + " |"
        )
    durations = sum(
        contract_duration_number(shot.get("duration_seconds"), strict_contract=True) for shot in shots
    )
    short_count = sum(
        contract_duration_number(shot.get("duration_seconds"), strict_contract=True) <= 3 for shot in shots
    )
    long_count = sum(
        contract_duration_number(shot.get("duration_seconds"), strict_contract=True) >= 6 for shot in shots
    )
    lines.extend(
        [
            "",
            "## 节奏统计（软目标，越界必须人工处置）",
            "",
            f"- cut_per_minute: {cut_rate_per_minute(len(shots), durations)}",
            f"- short_shot_ratio: {round(short_count / len(shots), 4) if shots else 0}",
            f"- long_shot_ratio: {round(long_count / len(shots), 4) if shots else 0}",
            "",
            "## 校验发现",
            "",
            f"- status: {result.status}",
        ]
    )
    lines.extend(f"- ERROR: {message}" for message in result.errors)
    lines.extend(f"- WARN: {message}" for message in result.warnings)
    return "\n".join(lines) + "\n"


def command_review_package(args: argparse.Namespace) -> int:
    data_path = Path(args.data)
    data = load_json(data_path)
    result = ValidationResult()
    workspace_root = Path(args.workspace_root).resolve(strict=False)
    output_path = Path(args.output).resolve(strict=False)
    if output_path == data_path.resolve(strict=False):
        result.error("review-package --output 不得与输入 shot_data 路径相同。")
    try:
        output_path.relative_to(workspace_root)
    except ValueError:
        result.error("review-package --output 必须位于 --workspace-root 内。")
    if metadata_version(data) != VERSION or not uses_director_integrity_guard(data):
        result.error(f"review-package 只接受当前 {VERSION} 导演完整性合同。")
    if clean_text(args.gate) != "GATE_B":
        result.error("review-package 当前只生成 GATE_B 审核包。")
    batch_id = clean_text(args.batch_id) or None
    batches = {
        clean_text(batch.get("batch_id"))
        for batch in as_list(as_dict(data.get("batch_plan")).get("batches"))
        if isinstance(batch, dict) and clean_text(batch.get("batch_id"))
    }
    if batches and batch_id not in batches:
        result.error("分批任务的 GATE_B review-package 必须绑定已登记 --batch-id。")
    if not batches and batch_id is not None:
        result.error("非分批任务不得填写 --batch-id。")
    if result.errors:
        print_result(result)
        return 1
    audit = validate_data(
        data,
        strict_status=False,
        data_path=data_path,
        workspace_root=workspace_root,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        review_package_text(data, gate="GATE_B", batch_id=batch_id, result=audit),
        encoding="utf-8",
    )
    print(f"WROTE: Gate B review package ({audit.status}) to {output_path}")
    return 0


def command_build(args: argparse.Namespace) -> int:
    data_path = Path(args.data)
    original = load_json(data_path)
    if metadata_version(original) != VERSION:
        result = ValidationResult(errors=[f"build 只允许当前 {VERSION} 合同；历史版本只能 validate。"])
        print_result(result)
        return 1
    if not args.report:
        result = ValidationResult(errors=[f"{VERSION} build 必须提供 --report，确保四文件原子交付。"])
        print_result(result)
        return 1
    if not args.workspace_root:
        result = ValidationResult(errors=[f"{VERSION} build 必须显式提供 --workspace-root。"])
        print_result(result)
        return 1
    data = copy.deepcopy(original)
    derive_prompts(data)
    workspace_root = Path(args.workspace_root)
    result = validate_data(
        data,
        strict_status=False,
        final_signoff=bool(args.final_signoff),
        data_path=data_path,
        workspace_root=workspace_root,
    )
    if result.errors:
        print_result(result)
        return 1
    update_validation_report(data, result, final_signoff=bool(args.final_signoff))
    atomic_write_delivery(
        data,
        data_path=data_path,
        markdown_path=Path(args.markdown),
        excel_path=Path(args.excel),
        report_path=Path(args.report),
        workspace_root=workspace_root,
        final_signoff=bool(args.final_signoff),
    )
    print_result(result)
    return 0


def command_review_hash(args: argparse.Namespace) -> int:
    """Print the deterministic human-review hash without mutating delivery data."""

    data = load_json(Path(args.data))
    result = ValidationResult()
    validate_json_compatible(data, result)
    metadata = as_dict(data.get("metadata"))
    if metadata.get("skill_name") != "su-fenjingskill-zh":
        result.error("metadata.skill_name 必须为 su-fenjingskill-zh。")
    if metadata_version(data) != VERSION:
        result.error(f"review-hash 只接受当前 {VERSION} 合同。")
    if clean_text(metadata.get("rule_revision")) != RULE_REVISION:
        result.error(f"metadata.rule_revision 必须为 {RULE_REVISION}。")
    gate = clean_text(args.gate)
    batch_id = clean_text(args.batch_id) or None
    batches = {
        clean_text(batch.get("batch_id"))
        for batch in as_list(as_dict(data.get("batch_plan")).get("batches"))
        if isinstance(batch, dict) and clean_text(batch.get("batch_id"))
    }
    if gate == "GATE_0":
        if not batches:
            result.error("仅 batch_plan.batches 非空时允许计算 GATE_0 review hash。")
        if batch_id is not None:
            result.error("GATE_0 是全局审核，不得填写 --batch-id。")
    elif gate in {"GATE_A", "GATE_B"}:
        if batches and batch_id not in batches:
            result.error(f"分批任务的 {gate} 必须用 --batch-id 绑定已登记批次。")
        if not batches and batch_id is not None:
            result.error(f"非分批任务的 {gate} 不得填写 --batch-id。")
    elif gate == "GATE_C" and batch_id is not None:
        result.error("GATE_C 是全局审核，不得填写 --batch-id。")
    if result.errors:
        print_result(result)
        return 1
    output = {
        "gate": gate,
        "batch_id": batch_id,
        "reviewed_hash": gate_review_hash(data, gate, batch_id),
    }
    print(json.dumps(output, ensure_ascii=False, sort_keys=True))
    return 0


def command_validate(args: argparse.Namespace) -> int:
    data_path = Path(args.data)
    data = load_json(data_path)
    if is_current_contract(data) and not args.report:
        result = ValidationResult(
            errors=[f"{metadata_version(data)} validate 必须提供 --report，确保四文件一致性。"]
        )
        print_result(result)
        return 1
    if is_current_contract(data) and not args.workspace_root:
        result = ValidationResult(
            errors=[f"{metadata_version(data)} validate 必须显式提供 --workspace-root。"]
        )
        print_result(result)
        return 1
    workspace_root = Path(args.workspace_root) if args.workspace_root else None
    result = validate_data(
        data,
        strict_status=True,
        final_signoff=bool(args.final_signoff),
        data_path=data_path,
        workspace_root=workspace_root,
    )
    compare_markdown(data, Path(args.markdown), result)
    compare_excel(data, Path(args.excel), result)
    if args.report:
        compare_report(data, Path(args.report), result)
    if clean_text(as_dict(data.get("validation_report")).get("status")) == "NOT_RUN" and not result.errors:
        if result.warnings:
            for item in result.warnings:
                print(f"- {item}")
        print("NOT_RUN: final validation was explicitly accepted without execution; this is not PASS.")
    else:
        print_result(result)
    return 1 if result.errors else 0


def make_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    build = subparsers.add_parser("build", help="derive Prompt, update JSON, and write Markdown/Excel")
    build.add_argument("--data", required=True)
    build.add_argument("--markdown", required=True)
    build.add_argument("--excel", required=True)
    build.add_argument("--report", help=f"required for {VERSION} four-file atomic delivery")
    build.add_argument(
        "--workspace-root",
        help=f"required for {VERSION}; resolves approved_script_path within the workspace",
    )
    build.add_argument("--final-signoff", action="store_true", help="require a valid Gate C approval")
    build.set_defaults(func=command_build)
    review_hash = subparsers.add_parser(
        "review-hash",
        help="print a deterministic Gate 0/A/B/C reviewed_hash without writing files",
    )
    review_hash.add_argument("--data", required=True)
    review_hash.add_argument("--gate", required=True, choices=sorted(GATES))
    review_hash.add_argument("--batch-id")
    review_hash.set_defaults(func=command_review_hash)
    review_package = subparsers.add_parser(
        "review-package",
        help="write a deterministic Gate B director review package without mutating shot data",
    )
    review_package.add_argument("--data", required=True)
    review_package.add_argument("--gate", required=True, choices=["GATE_B"])
    review_package.add_argument("--batch-id")
    review_package.add_argument("--workspace-root", required=True)
    review_package.add_argument("--output", required=True)
    review_package.set_defaults(func=command_review_package)
    validate = subparsers.add_parser("validate", help="validate JSON/Markdown/Excel consistency")
    validate.add_argument("--data", required=True)
    validate.add_argument("--markdown", required=True)
    validate.add_argument("--excel", required=True)
    validate.add_argument(
        "--report",
        help="required for strict-contract versions; optional for older legacy validation",
    )
    validate.add_argument(
        "--workspace-root",
        help="required for strict-contract versions; optional for older legacy validation",
    )
    validate.add_argument("--final-signoff", action="store_true", help="require Gate C approved for 2.4.x final deliveries")
    validate.set_defaults(func=command_validate)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = make_parser()
    args = parser.parse_args(argv)
    try:
        return int(args.func(args))
    except Exception as exc:
        print(f"FAIL: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
