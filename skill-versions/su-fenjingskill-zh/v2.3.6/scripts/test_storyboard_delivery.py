#!/usr/bin/env python3
"""Tests for the stable 7-column storyboard delivery contract."""

from __future__ import annotations

import copy
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

import storyboard_delivery as delivery


RULE_REVISION_234 = "2.3.4-overcompression-guard-2026-06-30"
RULE_REVISION_235 = "2.3.5-adjacent-motion-guard-2026-06-30"
RULE_REVISION_236 = "2.3.6-tri-source-audit-guard-2026-06-30"


def valid_data() -> dict:
    return {
        "metadata": {
            "skill_name": "su-fenjingskill-zh",
            "version": "2.3.0",
            "title": "稳定化样片",
            "reference_status": {
                "continuity-shot-data": "loaded",
                "camera-language": "loaded",
                "seedance-prompt-rules": "loaded",
            },
        },
        "continuity_logs": [
            {
                "scene": "1 室内 日 内",
                "first_shot_anchor_type": "single_continuation",
                "spatial_axis": "A在门口，B在桌边。",
                "fixed_objects": [],
                "characters": [],
                "props": [],
                "sound_sources": [],
                "reality_layer": "现实",
            }
        ],
        "beats": [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "A站在门口。",
                "facts": [{"fact_id": "B001-F01", "type": "position", "text": "A站在门口。"}],
            },
            {
                "beat_id": "B002",
                "scene": "1 室内 日 内",
                "source_text": "A走到桌边。",
                "facts": [{"fact_id": "B002-F01", "type": "position", "text": "A走到桌边。"}],
            },
            {
                "beat_id": "B003",
                "scene": "1 室内 日 内",
                "source_text": "A说：到了。",
                "facts": [{"fact_id": "B003-F01", "type": "dialogue", "text": "到了。"}],
            },
        ],
        "shots": [
            {
                "shot_no": 1,
                "scene": "1 室内 日 内",
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01"],
                "source_paragraph": "A站在门口。",
                "duration_seconds": 2,
                "duration_breakdown": {
                    "sync_action_seconds": 1,
                    "sync_dialogue_seconds": 0,
                    "non_sync_action_seconds": 0,
                    "emotional_pause_seconds": 1,
                },
                "camera_main_image": "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机在桌边看向门口。\n【场景首镜站位】（A站在门口，面向桌边。）\nA站在门口，手扶门框。",
                "notes": "建立初始站位。",
                "prompt": "",
                "visible_characters": ["A"],
                "offscreen_characters": [],
                "visible_props": [],
                "continuity_updates": [],
            },
            {
                "shot_no": 2,
                "scene": "1 室内 日 内",
                "beat_ids": ["B002", "B003"],
                "covered_fact_ids": ["B002-F01", "B003-F01"],
                "source_paragraph": "A走到桌边。A说：到了。",
                "duration_seconds": 3,
                "duration_breakdown": {
                    "sync_action_seconds": 2,
                    "sync_dialogue_seconds": 1,
                    "non_sync_action_seconds": 0,
                    "emotional_pause_seconds": 1,
                },
                "camera_main_image": "[侧面平视, 中景, 横移跟拍]\n【机位逻辑】摄影机沿桌边横移，跟住A的脚步。\n【站位位移】A从门口走到桌边，面向B的位置。\nA在桌边停下，说：“到了。”",
                "notes": "A完成位置迁移。",
                "prompt": "",
                "visible_characters": ["A"],
                "offscreen_characters": [],
                "visible_props": [],
                "continuity_updates": [
                    {
                        "entity_type": "character",
                        "entity": "A",
                        "field": "position",
                        "from": "门口",
                        "to": "桌边",
                        "evidence_fact_ids": ["B002-F01"],
                    }
                ],
            },
        ],
        "validation_report": {"status": "PASS", "warnings": [], "errors": []},
    }


def valid_data_234() -> dict:
    data = valid_data()
    data["metadata"]["version"] = "2.3.4"
    data["metadata"]["rule_revision"] = RULE_REVISION_234
    for beat in data["beats"]:
        for fact in beat["facts"]:
            fact["cut_priority"] = "normal"
            fact["cut_reasons"] = []
            fact["cut_group"] = fact["fact_id"]
    for shot in data["shots"]:
        shot["shot_type"] = "master" if shot["shot_no"] == 1 else "action"
        shot["split_reason"] = ["spatial_anchor"] if shot["shot_no"] == 1 else ["performance_continuity", "continuity_migration"]
        shot["insert_priority"] = "none"
        shot["long_take_support"] = []
    return data


def valid_data_235() -> dict:
    data = valid_data_234()
    data["metadata"]["version"] = "2.3.5"
    data["metadata"]["rule_revision"] = RULE_REVISION_235
    return data


def valid_data_236() -> dict:
    data = valid_data_235()
    data["metadata"]["version"] = "2.3.6"
    data["metadata"]["rule_revision"] = RULE_REVISION_236
    data["continuity_logs"][0]["characters"] = [{"name": "A", "position": "门口", "facing": "桌边"}]
    for beat in data["beats"]:
        for fact in beat["facts"]:
            fact_type = fact["type"]
            category = {
                "position": "space",
                "prop": "prop",
                "sound": "sound",
                "reality": "reality",
                "dialogue": "dialogue",
                "emotion": "emotion",
                "character": "character",
            }.get(fact_type, "action")
            fact["cut_category"] = category
            fact["cut_moment_id"] = f"{beat['beat_id']}-{category}-moment"
    return data


class DeliveryContractTests(unittest.TestCase):
    def test_build_writes_7_column_markdown_and_excel(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data_path = root / "sample.shot_data.json"
            md_path = root / "sample.md"
            xlsx_path = root / "sample.xlsx"
            data_path.write_text(json.dumps(valid_data(), ensure_ascii=False), encoding="utf-8")

            rc = delivery.main(["build", "--data", str(data_path), "--markdown", str(md_path), "--excel", str(xlsx_path)])
            self.assertEqual(0, rc)

            built = json.loads(data_path.read_text(encoding="utf-8"))
            self.assertEqual("PASS", built["validation_report"]["status"])
            self.assertNotIn("keyframe", built["shots"][0])
            self.assertEqual(delivery.derive_prompt(built["shots"][0]), built["shots"][0]["prompt"])

            markdown = md_path.read_text(encoding="utf-8")
            self.assertIn("| 镜号 | 场景 | 原剧本段落 | 镜头时长(秒) | 运镜+主画面描述(含台词) | 备注 | Prompt |", markdown)
            self.assertNotIn("关键帧", markdown)
            self.assertNotIn("【镜内变化】", markdown)

            workbook = load_workbook(xlsx_path, read_only=True)
            try:
                self.assertEqual(["分镜表"], workbook.sheetnames)
                sheet = workbook["分镜表"]
                self.assertEqual(3, sheet.max_row)
                self.assertEqual(7, sheet.max_column)
            finally:
                workbook.close()

    def test_validate_wrapper_passes_built_files(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data_path = root / "sample.shot_data.json"
            md_path = root / "sample.md"
            xlsx_path = root / "sample.xlsx"
            data_path.write_text(json.dumps(valid_data(), ensure_ascii=False), encoding="utf-8")
            self.assertEqual(0, delivery.main(["build", "--data", str(data_path), "--markdown", str(md_path), "--excel", str(xlsx_path)]))

            process = subprocess.run(
                [
                    "node",
                    str(SCRIPT_DIR / "validate_storyboard.js"),
                    "--python",
                    sys.executable,
                    "--data",
                    str(data_path),
                    "--markdown",
                    str(md_path),
                    "--excel",
                    str(xlsx_path),
                ],
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertEqual(0, process.returncode, process.stderr + process.stdout)

    def test_keyframe_field_is_rejected(self) -> None:
        data = valid_data()
        data["shots"][0]["keyframe"] = "场景：旧关键帧"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("keyframe" in item for item in result.errors))

    def test_pass_internal_status_is_rejected(self) -> None:
        data = valid_data()
        data["validation_report"]["status"] = "PASS_INTERNAL"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("PASS / WARN / FAIL" in item for item in result.errors))

    def test_station_move_requires_continuity_update(self) -> None:
        data = valid_data()
        data["shots"][1]["continuity_updates"] = []
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("continuity_updates 为空" in item for item in result.errors))

    def test_position_update_requires_station_move(self) -> None:
        data = valid_data()
        data["shots"][1]["camera_main_image"] = data["shots"][1]["camera_main_image"].replace("【站位位移】", "")
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("缺少【站位位移】" in item for item in result.errors))

    def test_prompt_internal_label_is_rejected(self) -> None:
        data = valid_data()
        delivery.derive_prompts(data)
        data["shots"][0]["prompt"] += "\n【镜内变化】污染"
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("Prompt" in item and "内部" in item for item in result.errors))

    def test_dialogue_duration_cannot_be_underestimated(self) -> None:
        data = valid_data()
        data["shots"][1]["camera_main_image"] = "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机拍A站在桌边。\nA说：“这句话很长很长不能压到一秒。”"
        data["shots"][1]["duration_seconds"] = 1
        data["shots"][1]["duration_breakdown"]["sync_action_seconds"] = 1
        data["shots"][1]["duration_breakdown"]["sync_dialogue_seconds"] = 1
        data["shots"][1]["duration_breakdown"]["emotional_pause_seconds"] = 0
        data["shots"][1]["continuity_updates"] = []
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("对白时长过短" in item for item in result.errors))

    def test_outdoor_multi_character_panorama_first_shot_passes(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "both"
        data["continuity_logs"][0]["spatial_axis"] = "A在画面左侧，B在画面右侧，两人沿山路面向远处。"
        data["shots"][0]["camera_main_image"] = (
            "[微俯视, 大全景, 固定镜头]\n"
            "【机位逻辑】摄影机在山路上方俯看两人和远处入口。\n"
            "【场景首镜站位】（A在画面左侧，B在画面右侧，两人相距三米，面向远处入口。）\n"
            "A和B站在山路上，远处入口被雾压住。"
        )
        data["shots"][0]["visible_characters"] = ["A", "B"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_indoor_multi_character_full_shot_first_shot_passes(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "multi_character"
        data["continuity_logs"][0]["spatial_axis"] = "A在画面左侧门口，B在画面右侧桌边，两人隔桌对视。"
        data["shots"][0]["camera_main_image"] = (
            "[平视, 全景, 固定镜头]\n"
            "【机位逻辑】摄影机在房间侧墙拍到门、桌和两人的左右关系。\n"
            "【场景首镜站位】（A在画面左侧门口，面向右侧桌边；B在画面右侧桌边，面向A，两人隔桌对视。）\n"
            "A扶着门框，B站在桌边。"
        )
        data["shots"][0]["visible_characters"] = ["A", "B"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_multi_character_first_shot_medium_close_fails(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "multi_character"
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中近景, 固定镜头]\n"
            "【机位逻辑】摄影机拍两人上半身。\n"
            "【场景首镜站位】（A在左，B在右，两人面向彼此。）\n"
            "A和B看着对方。"
        )
        data["shots"][0]["visible_characters"] = ["A", "B"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("景别必须为大远景/大全景/全景/中全景" in item for item in result.errors))

    def test_single_continuation_medium_first_shot_passes(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "single_continuation"
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 固定镜头]\n"
            "【机位逻辑】摄影机在岩壁前拍A的半身和身后空间边界。\n"
            "【场景首镜站位】（A靠在岩壁前，面向画面右侧微光。）\n"
            "A靠在岩壁上，抬头看向右侧微光。"
        )
        data["shots"][0]["visible_characters"] = ["A"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_extreme_wide_steadicam_fails(self) -> None:
        data = valid_data()
        data["shots"][0]["camera_main_image"] = (
            "[微俯视, 大全景, 斯坦尼康平稳跟随]\n"
            "【机位逻辑】摄影机拍到山路全貌和人物位置。\n"
            "【场景首镜站位】（A站在画面左侧，面向远处。）\n"
            "A站在山路上。"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("大远景/大全景不得使用斯坦尼康" in item for item in result.errors))

    def test_extreme_wide_crane_passes(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "both"
        data["continuity_logs"][0]["spatial_axis"] = "A在画面左侧，B在画面右侧，两人面向远处入口。"
        data["shots"][0]["camera_main_image"] = (
            "[微俯视, 大全景, 伸缩摇臂缓慢下降]\n"
            "【机位逻辑】摄影机从林冠上方下降，拍到山路、入口和两人左右关系。\n"
            "【场景首镜站位】（A在画面左侧，B在画面右侧，两人相距三米，面向远处入口。）\n"
            "A和B站在山路上。"
        )
        data["shots"][0]["visible_characters"] = ["A", "B"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_aerial_large_landscape_passes(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "space"
        data["shots"][0]["camera_main_image"] = (
            "[高角度俯拍, 大远景, 航拍缓慢推进]\n"
            "【机位逻辑】摄影机从高速公路上方推进，拍到道路纵深和远处雾区。\n"
            "【场景首镜站位】（道路由画面左下延伸至右上，A处在远处路肩，面向雾区。）\n"
            "浓雾压过高速公路。"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_steadicam_medium_follow_passes(self) -> None:
        data = valid_data()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 斯坦尼康跟随]\n"
            "【机位逻辑】摄影机在地面跟随A从门口走向桌边。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A从门口向桌边走去。"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_aerial_medium_close_fails(self) -> None:
        data = valid_data()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中近景, 航拍缓慢推进]\n"
            "【机位逻辑】摄影机贴近A的上半身推进。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A看向桌边。"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("航拍必须服务于大范围空间" in item for item in result.errors))

    def test_adjacent_similar_shots_without_position_change_fail(self) -> None:
        data = valid_data()
        data["beats"] = data["beats"][:2]
        data["beats"][1]["source_text"] = "A抬眼看向桌边。"
        data["beats"][1]["facts"] = [{"fact_id": "B002-F01", "type": "action", "text": "A抬眼看向桌边。"}]
        data["shots"][1].update(
            {
                "beat_ids": ["B002"],
                "covered_fact_ids": ["B002-F01"],
                "source_paragraph": "A抬眼看向桌边。",
                "duration_seconds": 2,
                "duration_breakdown": {
                    "sync_action_seconds": 1,
                    "sync_dialogue_seconds": 0,
                    "non_sync_action_seconds": 0,
                    "emotional_pause_seconds": 1,
                },
                "camera_main_image": "[平视, 中近景, 固定镜头]\n【机位逻辑】摄影机继续拍A的上半身。\nA抬眼看向桌边。",
                "notes": "同一主体反应。",
                "continuity_updates": [],
            }
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("主体、视角、景别和运动过近" in item for item in result.errors))

    def test_incomplete_dialogue_split_across_adjacent_shots_fails(self) -> None:
        data = valid_data()
        data["beats"] = data["beats"][:2]
        data["beats"][0]["source_text"] = "A：你听我说，"
        data["beats"][0]["facts"] = [{"fact_id": "B001-F01", "type": "dialogue", "text": "你听我说，"}]
        data["beats"][1]["source_text"] = "A：我不是故意的。"
        data["beats"][1]["facts"] = [{"fact_id": "B002-F01", "type": "dialogue", "text": "我不是故意的。"}]
        data["shots"][0]["source_paragraph"] = "A：你听我说，"
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 固定镜头]\n"
            "【机位逻辑】摄影机在桌边拍A。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A看向桌边，说：“你听我说，”"
        )
        data["shots"][1].update(
            {
                "beat_ids": ["B002"],
                "covered_fact_ids": ["B002-F01"],
                "source_paragraph": "A：我不是故意的。",
                "duration_seconds": 3,
                "duration_breakdown": {
                    "sync_action_seconds": 1,
                    "sync_dialogue_seconds": 2,
                    "non_sync_action_seconds": 0,
                    "emotional_pause_seconds": 1,
                },
                "camera_main_image": "[平视, 中近景, 固定镜头]\n【机位逻辑】摄影机切近A的脸。\nA继续说：“我不是故意的。”",
                "notes": "同一说话人续句。",
                "continuity_updates": [],
            }
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("未完成台词" in item for item in result.errors))

    def test_dialogue_followed_by_short_expression_reaction_fails(self) -> None:
        data = valid_data()
        data["beats"] = data["beats"][:2]
        data["beats"][0]["source_text"] = "A：够了。"
        data["beats"][0]["facts"] = [{"fact_id": "B001-F01", "type": "dialogue", "text": "够了。"}]
        data["beats"][1]["source_text"] = "A笑了一下。"
        data["beats"][1]["facts"] = [{"fact_id": "B002-F01", "type": "emotion", "text": "A笑了一下。"}]
        data["shots"][0]["source_paragraph"] = "A：够了。"
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 固定镜头]\n"
            "【机位逻辑】摄影机在桌边拍A。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A看向桌边，说：“够了。”"
        )
        data["shots"][1].update(
            {
                "beat_ids": ["B002"],
                "covered_fact_ids": ["B002-F01"],
                "source_paragraph": "A笑了一下。",
                "duration_seconds": 2,
                "duration_breakdown": {
                    "sync_action_seconds": 1,
                    "sync_dialogue_seconds": 0,
                    "non_sync_action_seconds": 0,
                    "emotional_pause_seconds": 1,
                },
                "camera_main_image": "[平视, 中近景, 固定镜头]\n【机位逻辑】摄影机切近A的嘴角。\nA嘴角轻轻抬起，笑了一下。",
                "notes": "短表情反应。",
                "continuity_updates": [],
            }
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("短表情反应" in item for item in result.errors))

    def test_wide_to_close_shot_size_without_span_movement_fails(self) -> None:
        data = valid_data()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 全景->特写, 固定镜头]\n"
            "【机位逻辑】摄影机同时描述房间全貌和A的眼睛。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A站在门口，眼神收紧。"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("[景别跨度]" in item for item in result.errors))

    def test_wide_to_close_shot_size_with_span_movement_passes(self) -> None:
        data = valid_data()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 全景->特写, 光学变焦]\n"
            "【机位逻辑】摄影机从房间全貌压到A的眼睛。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A站在门口，眼神收紧。"
        )
        data["shots"][0]["notes"] = "[景别跨度] 从房间关系压入A的关键反应。"
        data["shots"][1]["camera_main_image"] = (
            "[侧面平视, 中景, 横移跟拍]\n"
            "【机位逻辑】摄影机沿桌边横移，跟住A的脚步。\n"
            "【站位位移】A从门口走到桌边，面向B的位置。\n"
            "A在桌边停下，说：“到了。”"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_233_data_does_not_require_hybrid_fields(self) -> None:
        data = valid_data()
        data["metadata"]["version"] = "2.3.3"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(any("shot_type" in item or "split_reason" in item or "insert_priority" in item or "long_take_support" in item for item in result.errors), result.errors)

    def test_234_build_keeps_7_column_outputs_with_hybrid_fields(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data_path = root / "sample.shot_data.json"
            md_path = root / "sample.md"
            xlsx_path = root / "sample.xlsx"
            data_path.write_text(json.dumps(valid_data_234(), ensure_ascii=False), encoding="utf-8")

            rc = delivery.main(["build", "--data", str(data_path), "--markdown", str(md_path), "--excel", str(xlsx_path)])
            self.assertEqual(0, rc)
            built = json.loads(data_path.read_text(encoding="utf-8"))
            self.assertIn("hybrid_audit", built["validation_report"])
            self.assertEqual("master", built["shots"][0]["shot_type"])

            workbook = load_workbook(xlsx_path, read_only=True)
            try:
                sheet = workbook[delivery.SHEET_NAME]
                self.assertEqual(7, sheet.max_column)
            finally:
                workbook.close()

    def test_234_missing_shot_type_fails(self) -> None:
        data = valid_data_234()
        del data["shots"][0]["shot_type"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("shot_type" in item for item in result.errors))

    def test_234_invalid_split_reason_fails(self) -> None:
        data = valid_data_234()
        data["shots"][0]["split_reason"] = ["wrong_reason"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("split_reason" in item and "wrong_reason" in item for item in result.errors))

    def test_234_long_take_requires_two_supports(self) -> None:
        data = valid_data_234()
        data["shots"][1]["duration_seconds"] = 11
        data["shots"][1]["source_paragraph"] = "A移至桌边。A说：到了。"
        data["shots"][1]["duration_breakdown"] = {
            "sync_action_seconds": 11,
            "sync_dialogue_seconds": 0,
            "non_sync_action_seconds": 0,
            "emotional_pause_seconds": 0,
        }
        data["shots"][1]["camera_main_image"] = (
            "[侧面平视, 中景, 横移跟拍]\n"
            "【机位逻辑】摄影机沿桌边横移，跟住A的脚步。\n"
            "【站位位移】A从门口移至桌边，面向B的位置。\n"
            "A在桌边整理账册，随后示意。"
        )
        data["shots"][1]["notes"] = "[时长估算] 同步动作11秒 + 同步对白0秒 + 非同步动作0秒 + 情绪停顿0秒。 [长镜头] [保留理由] 动作长镜承载位移。"
        data["shots"][1]["long_take"] = {"classification": "action_long_take", "reason": "动作长镜"}
        data["shots"][1]["long_take_support"] = ["character_blocking"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("long_take_support" in item for item in result.errors))

        data["shots"][1]["long_take_support"] = ["character_blocking", "spatial_progression"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_234_long_insert_duration_fails_without_support(self) -> None:
        data = valid_data_234()
        data["shots"][0]["shot_type"] = "insert"
        data["shots"][0]["split_reason"] = ["prop_state_change"]
        data["shots"][0]["insert_priority"] = "recommended"
        data["shots"][0]["duration_seconds"] = 6
        data["shots"][0]["duration_breakdown"] = {
            "sync_action_seconds": 5,
            "sync_dialogue_seconds": 0,
            "non_sync_action_seconds": 0,
            "emotional_pause_seconds": 1,
        }
        data["shots"][0]["notes"] = "[时长估算] 同步动作5秒 + 同步对白0秒 + 非同步动作0秒 + 情绪停顿1秒。插镜偏长，仅用于警告测试。"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("超过5秒" in item for item in result.errors))

    def test_234_three_truth_reveals_in_one_dialogue_shot_fail(self) -> None:
        data = valid_data_234()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "沈夜：二十年前它就想找你。",
                "facts": [{"fact_id": "B001-F01", "type": "dialogue", "text": "二十年前它就想找你。", "cut_priority": "must_isolate", "cut_reasons": ["causal_reveal"], "cut_group": "B001-truth"}],
            },
            {
                "beat_id": "B002",
                "scene": "1 室内 日 内",
                "source_text": "顾成：林晓杰从一开始就是替身。",
                "facts": [{"fact_id": "B002-F01", "type": "dialogue", "text": "林晓杰从一开始就是替身。", "cut_priority": "must_isolate", "cut_reasons": ["causal_reveal"], "cut_group": "B002-substitute"}],
            },
            {
                "beat_id": "B003",
                "scene": "1 室内 日 内",
                "source_text": "沈夜：本来你才是最好的宿主。",
                "facts": [{"fact_id": "B003-F01", "type": "dialogue", "text": "本来你才是最好的宿主。", "cut_priority": "must_isolate", "cut_reasons": ["causal_reveal"], "cut_group": "B003-host"}],
            },
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "shot_no": 1,
                "beat_ids": ["B001", "B002", "B003"],
                "covered_fact_ids": ["B001-F01", "B002-F01", "B003-F01"],
                "source_paragraph": "沈夜：二十年前它就想找你。 / 顾成：林晓杰从一开始就是替身。 / 沈夜：本来你才是最好的宿主。",
                "duration_seconds": 13,
                "duration_breakdown": {"sync_action_seconds": 2, "sync_dialogue_seconds": 10, "non_sync_action_seconds": 2, "emotional_pause_seconds": 1},
                "long_take": {"classification": "dialogue_long_take", "reason": "测试长对白"},
                "camera_main_image": "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机拍沈夜和林晓彤。\n【场景首镜站位】沈夜在左，林晓彤在右，两人面向彼此。\n沈夜说：“二十年前它就想找你。”顾成说：“林晓杰从一开始就是替身。”沈夜说：“本来你才是最好的宿主。”",
                "notes": "[时长估算] 同步动作2秒 + 同步台词10秒 + 非同步动作2秒 + 情绪留白1秒。 [长镜头] [保留理由] [不可拆说明] 测试故意压缩。",
                "visible_characters": ["沈夜", "林晓彤", "顾成"],
                "shot_type": "dialogue",
                "split_reason": ["causal_reveal", "performance_continuity"],
                "insert_priority": "must_have",
                "long_take_support": ["character_blocking", "emotional_turn"],
            }
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("过度压缩" in item or "must_isolate cut_group" in item for item in result.errors), result.errors)

    def test_234_first_alien_voice_and_mocking_turn_in_one_long_dialogue_fail(self) -> None:
        data = valid_data_234()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "雾深处传来低沉震动。林晓杰：别催。",
                "facts": [
                    {"fact_id": "B001-F01", "type": "sound", "text": "雾深处传来低沉震动。", "cut_priority": "must_isolate", "cut_reasons": ["new_sound_source"], "cut_group": "B001-sound"},
                    {"fact_id": "B001-F02", "type": "dialogue", "text": "别催。", "cut_priority": "normal", "cut_reasons": [], "cut_group": "B001-dialogue"},
                ],
            },
            {
                "beat_id": "B002",
                "scene": "1 室内 日 内",
                "source_text": "林晓杰笑了一下，转为嘲讽。",
                "facts": [{"fact_id": "B002-F01", "type": "emotion", "text": "林晓杰笑了一下，转为嘲讽。", "cut_priority": "must_isolate", "cut_reasons": ["emotional_turn"], "cut_group": "B002-mock"}],
            },
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "shot_no": 1,
                "beat_ids": ["B001", "B002"],
                "covered_fact_ids": ["B001-F01", "B001-F02", "B002-F01"],
                "source_paragraph": "雾深处传来低沉震动。林晓杰：别催。 / 林晓杰笑了一下，转为嘲讽。",
                "duration_seconds": 13,
                "duration_breakdown": {"sync_action_seconds": 3, "sync_dialogue_seconds": 9, "non_sync_action_seconds": 3, "emotional_pause_seconds": 1},
                "long_take": {"classification": "dialogue_long_take", "reason": "测试长对白"},
                "camera_main_image": "[低角度仰拍, 中近景, 固定镜头]\n【机位逻辑】摄影机固定拍林晓杰。\n【场景首镜站位】林晓杰靠在岩壁前，面向远处黑雾。\n雾深处传来低沉震动，林晓杰说：“别催。”他笑了一下，转为嘲讽。",
                "notes": "[时长估算] 同步动作3秒 + 同步台词9秒 + 非同步动作3秒 + 情绪留白1秒。 [长镜头] [保留理由] [不可拆说明] 测试故意压缩。",
                "visible_characters": ["林晓杰"],
                "shot_type": "dialogue",
                "split_reason": ["new_sound_source", "emotional_turn"],
                "insert_priority": "must_have",
                "long_take_support": ["sound_source_change", "emotional_turn"],
            }
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("must_isolate cut_group" in item or "dialogue 镜" in item for item in result.errors), result.errors)

    def test_234_exile_origin_chain_in_one_transition_shot_fails(self) -> None:
        data = valid_data_234()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "光柱熄灭，沈夜坠落，身上变成凡人衣服。",
                "facts": [
                    {"fact_id": "B001-F01", "type": "prop", "text": "光柱熄灭。", "cut_priority": "must_isolate", "cut_reasons": ["new_vfx_state"], "cut_group": "B001-light-off"},
                    {"fact_id": "B001-F02", "type": "action", "text": "沈夜坠落到人间。", "cut_priority": "must_isolate", "cut_reasons": ["reality_layer_shift"], "cut_group": "B001-fall"},
                    {"fact_id": "B001-F03", "type": "prop", "text": "沈夜身上变成凡人衣服。", "cut_priority": "must_isolate", "cut_reasons": ["causal_reveal"], "cut_group": "B001-human-clothes"},
                ],
            }
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "shot_no": 1,
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01", "B001-F02", "B001-F03"],
                "source_paragraph": "光柱熄灭，沈夜坠落，身上变成凡人衣服。",
                "duration_seconds": 9,
                "duration_breakdown": {"sync_action_seconds": 5, "sync_dialogue_seconds": 0, "non_sync_action_seconds": 3, "emotional_pause_seconds": 1},
                "camera_main_image": "[微俯视, 全景, 伸缩摇臂拉出]\n【机位逻辑】摄影机从神殿拉到人间。\n【场景首镜站位】沈夜跪在大殿中央。\n光柱熄灭，沈夜坠落到人间，身上变成凡人衣服。",
                "notes": "[时长估算] 同步动作5秒 + 同步台词0秒 + 非同步动作3秒 + 情绪留白1秒。 [长镜头] [保留理由] 测试故意压缩。",
                "visible_characters": ["沈夜"],
                "shot_type": "transition",
                "split_reason": ["new_vfx_state", "reality_layer_shift", "causal_reveal"],
                "insert_priority": "must_have",
                "long_take_support": ["vfx_state_change", "spatial_progression"],
            }
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("must_isolate cut_group" in item for item in result.errors), result.errors)

    def test_234_same_cut_group_vfx_moment_can_pass(self) -> None:
        data = valid_data_234()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "恶念钻入眉心，瞳孔全黑。",
                "facts": [
                    {"fact_id": "B001-F01", "type": "action", "text": "恶念钻入眉心。", "cut_priority": "must_isolate", "cut_reasons": ["causal_reveal", "new_vfx_state"], "cut_group": "B001-impact"},
                    {"fact_id": "B001-F02", "type": "prop", "text": "瞳孔全黑。", "cut_priority": "must_isolate", "cut_reasons": ["new_vfx_state"], "cut_group": "B001-impact"},
                ],
            }
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "shot_no": 1,
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01", "B001-F02"],
                "source_paragraph": "恶念钻入眉心，瞳孔全黑。",
                "duration_seconds": 4,
                "duration_breakdown": {"sync_action_seconds": 3, "sync_dialogue_seconds": 0, "non_sync_action_seconds": 0, "emotional_pause_seconds": 1},
                "camera_main_image": "[平视, 特写, 急推]\n【机位逻辑】摄影机压到眉心和眼睛。\n【场景首镜站位】小林晓杰位于画面中央，面向镜头。\n恶念钻入小林晓杰眉心，他的瞳孔瞬间全黑。",
                "notes": "[不可拆说明] 恶念入体和瞳孔变化属于同一撞击瞬间。",
                "visible_characters": ["小林晓杰"],
                "shot_type": "vfx_anchor",
                "split_reason": ["causal_reveal", "new_vfx_state"],
                "insert_priority": "must_have",
                "long_take_support": [],
            }
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_234_plain_short_performance_merge_passes(self) -> None:
        data = valid_data_234()
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_234_local_dialogue_changed_to_offscreen_voice_fails(self) -> None:
        data = valid_data_234()
        data["beats"][2]["facts"][0]["cut_priority"] = "must_isolate"
        data["beats"][2]["facts"][0]["cut_reasons"] = ["causal_reveal"]
        data["beats"][2]["facts"][0]["cut_group"] = "B003-truth"
        data["shots"][1]["source_paragraph"] = "顾成：所以林晓杰从一开始就是替身。"
        data["shots"][1]["beat_ids"] = ["B003"]
        data["shots"][1]["covered_fact_ids"] = ["B003-F01"]
        data["shots"][1]["camera_main_image"] = "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机只拍沈夜和林晓彤。\n顾成画外声说：“所以林晓杰从一开始就是替身。”"
        data["shots"][1]["visible_characters"] = ["沈夜", "林晓彤"]
        data["shots"][1]["shot_type"] = "dialogue"
        data["shots"][1]["split_reason"] = ["causal_reveal"]
        data["shots"][1]["insert_priority"] = "must_have"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("改为画外声" in item for item in result.errors), result.errors)

    def test_235_adjacent_push_then_pull_fails(self) -> None:
        data = valid_data_235()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 缓慢推进]\n"
            "【机位逻辑】摄影机从桌边缓慢推进到A。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A站在门口，手扶门框。"
        )
        data["shots"][1]["camera_main_image"] = (
            "[侧面平视, 中景, 缓慢拉出]\n"
            "【机位逻辑】摄影机从A半身缓慢拉出到桌边关系。\n"
            "【站位位移】A从门口走到桌边，面向B的位置。\n"
            "A在桌边停下，说：“到了。”"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("轴向反转" in item for item in result.errors), result.errors)

    def test_235_adjacent_pull_then_push_fails(self) -> None:
        data = valid_data_235()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 缓慢拉出]\n"
            "【机位逻辑】摄影机从A半身缓慢拉出到门口空间。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A站在门口，手扶门框。"
        )
        data["shots"][1]["camera_main_image"] = (
            "[侧面平视, 中景, 缓慢推进]\n"
            "【机位逻辑】摄影机沿桌边缓慢推进到A。\n"
            "【站位位移】A从门口走到桌边，面向B的位置。\n"
            "A在桌边停下，说：“到了。”"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("轴向反转" in item for item in result.errors), result.errors)

    def test_235_adjacent_push_then_lateral_passes(self) -> None:
        data = valid_data_235()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 缓慢推进]\n"
            "【机位逻辑】摄影机从桌边缓慢推进到A。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A站在门口，手扶门框。"
        )
        data["shots"][1]["camera_main_image"] = (
            "[侧面平视, 中景, 缓慢横移]\n"
            "【机位逻辑】摄影机沿桌边横移，跟住A的脚步。\n"
            "【站位位移】A从门口走到桌边，面向B的位置。\n"
            "A在桌边停下，说：“到了。”"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_235_final_shot_crane_up_pull_away_passes(self) -> None:
        data = valid_data_235()
        data["shots"][1]["camera_main_image"] = (
            "[高角度俯拍, 大全景, 摇臂缓慢上升拉远]\n"
            "【机位逻辑】摄影机从桌边上方升起拉远，让A的身影在空间里变小。\n"
            "【站位位移】A从门口走到桌边，面向B的位置。\n"
            "A在桌边停下，说：“到了。”"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_236_build_writes_report_and_audit_summary(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data_path = root / "sample.shot_data.json"
            md_path = root / "sample.md"
            xlsx_path = root / "sample.xlsx"
            report_path = root / "sample.validation_report.json"
            data_path.write_text(json.dumps(valid_data_236(), ensure_ascii=False), encoding="utf-8")

            rc = delivery.main([
                "build",
                "--data",
                str(data_path),
                "--markdown",
                str(md_path),
                "--excel",
                str(xlsx_path),
                "--report",
                str(report_path),
            ])
            self.assertEqual(0, rc)
            built = json.loads(data_path.read_text(encoding="utf-8"))
            report = json.loads(report_path.read_text(encoding="utf-8"))
            self.assertEqual(built["validation_report"], report)
            self.assertEqual(report["source_json_hash"], delivery.canonical_data_hash(built))
            self.assertIn(report["source_json_hash"], md_path.read_text(encoding="utf-8"))

            workbook = load_workbook(xlsx_path, read_only=True)
            try:
                self.assertEqual([delivery.SHEET_NAME, delivery.SUMMARY_SHEET_NAME], workbook.sheetnames)
            finally:
                workbook.close()

            rc = delivery.main([
                "validate",
                "--data",
                str(data_path),
                "--markdown",
                str(md_path),
                "--excel",
                str(xlsx_path),
                "--report",
                str(report_path),
            ])
            self.assertEqual(0, rc)

    def test_236_same_category_different_moment_must_isolate_fails(self) -> None:
        data = valid_data_236()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "光变黑，雾钻出。",
                "facts": [
                    {
                        "fact_id": "B001-F01",
                        "type": "prop",
                        "text": "光变黑。",
                        "cut_priority": "must_isolate",
                        "cut_reasons": ["new_vfx_state"],
                        "cut_group": "B001-vfx",
                        "cut_category": "vfx",
                        "cut_moment_id": "B001-vfx-light",
                    },
                    {
                        "fact_id": "B001-F02",
                        "type": "prop",
                        "text": "雾钻出。",
                        "cut_priority": "must_isolate",
                        "cut_reasons": ["new_vfx_state"],
                        "cut_group": "B001-vfx",
                        "cut_category": "vfx",
                        "cut_moment_id": "B001-vfx-fog",
                    },
                ],
            }
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01", "B001-F02"],
                "source_paragraph": "光变黑，雾钻出。",
                "duration_seconds": 4,
                "duration_breakdown": {"sync_action_seconds": 3, "sync_dialogue_seconds": 0, "non_sync_action_seconds": 0, "emotional_pause_seconds": 1},
                "camera_main_image": "[平视, 特写, 急推]\n【机位逻辑】摄影机压到光和雾。\n【场景首镜站位】A站在门口，面向光源。\n光变黑，雾钻出。",
                "notes": "[不可拆说明] 测试错误合并。",
                "shot_type": "vfx_anchor",
                "split_reason": ["new_vfx_state"],
                "insert_priority": "must_have",
            }
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("结构切点" in item for item in result.errors), result.errors)

    def test_236_same_category_same_moment_with_note_passes(self) -> None:
        data = valid_data_236()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "恶念钻入眉心，瞳孔全黑。",
                "facts": [
                    {
                        "fact_id": "B001-F01",
                        "type": "action",
                        "text": "恶念钻入眉心。",
                        "cut_priority": "must_isolate",
                        "cut_reasons": ["causal_reveal", "new_vfx_state"],
                        "cut_group": "B001-vfx",
                        "cut_category": "vfx",
                        "cut_moment_id": "B001-vfx-impact",
                    },
                    {
                        "fact_id": "B001-F02",
                        "type": "prop",
                        "text": "瞳孔全黑。",
                        "cut_priority": "must_isolate",
                        "cut_reasons": ["new_vfx_state"],
                        "cut_group": "B001-vfx",
                        "cut_category": "vfx",
                        "cut_moment_id": "B001-vfx-impact",
                    },
                ],
            }
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01", "B001-F02"],
                "source_paragraph": "恶念钻入眉心，瞳孔全黑。",
                "duration_seconds": 4,
                "duration_breakdown": {"sync_action_seconds": 3, "sync_dialogue_seconds": 0, "non_sync_action_seconds": 0, "emotional_pause_seconds": 1},
                "camera_main_image": "[平视, 特写, 急推]\n【机位逻辑】摄影机压到眉心和眼睛。\n【场景首镜站位】A位于画面中央，面向镜头。\n恶念钻入眉心，瞳孔全黑。",
                "notes": "[不可拆说明] 恶念入体和瞳孔变化属于同一撞击瞬间。",
                "shot_type": "vfx_anchor",
                "split_reason": ["causal_reveal", "new_vfx_state"],
                "insert_priority": "must_have",
            }
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_236_must_isolate_density_warns(self) -> None:
        data = valid_data_236()
        data["beats"] = data["beats"][:2]
        data["shots"] = data["shots"][:2]
        data["shots"][0]["shot_type"] = "action"
        data["shots"][1]["beat_ids"] = ["B002"]
        data["shots"][1]["covered_fact_ids"] = ["B002-F01"]
        data["shots"][1]["source_paragraph"] = "A走到桌边。"
        for beat in data["beats"]:
            fact = beat["facts"][0]
            fact["cut_priority"] = "must_isolate"
            fact["cut_reasons"] = ["new_information"]
            fact["cut_category"] = "space"
            fact["cut_moment_id"] = f"{beat['beat_id']}-space"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("must_isolate Beat 占比" in item for item in result.warnings), result.warnings)

    def test_236_fact_text_deviation_warns(self) -> None:
        data = valid_data_236()
        data["beats"][0]["source_text"] = "手环变亮。"
        data["beats"][0]["facts"][0]["type"] = "prop"
        data["beats"][0]["facts"][0]["text"] = "手环同频呼吸变亮。"
        data["beats"][0]["facts"][0]["cut_category"] = "prop"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("Fact 文本疑似引入" in item for item in result.warnings), result.warnings)

    def test_236_inherited_state_mismatch_fails(self) -> None:
        data = valid_data_236()
        data["continuity_logs"].append(
            {
                "scene": "2 室内 日 内",
                "inherits_from": "1 室内 日 内",
                "inherited_states": ["spatial_axis", "fixed_objects"],
                "diverged_states": ["characters", "props"],
                "first_shot_anchor_type": "single_continuation",
                "spatial_axis": "A在窗边，B在桌边。",
                "fixed_objects": [],
                "characters": [],
                "props": [],
                "sound_sources": [],
                "reality_layer": "现实",
            }
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("继承字段 spatial_axis" in item for item in result.errors), result.errors)

    def test_236_continuity_update_from_mismatch_fails(self) -> None:
        data = valid_data_236()
        data["shots"][1]["continuity_updates"][0]["from"] = "窗边"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("from 与上一状态不一致" in item for item in result.errors), result.errors)

    def test_236_prompt_composition_dialogue_leak_fails(self) -> None:
        data = valid_data_236()
        delivery.derive_prompts(data)
        data["shots"][1]["prompt"] = data["shots"][1]["prompt"].replace("构图：可见主体：A。", "构图：A说：“到了。”")
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("构图字段混入" in item for item in result.errors), result.errors)

    def test_236_general_short_reaction_warns(self) -> None:
        data = valid_data_236()
        data["beats"] = data["beats"][:2]
        data["beats"][0]["source_text"] = "A看向门口。"
        data["beats"][0]["facts"] = [
            {
                "fact_id": "B001-F01",
                "type": "action",
                "text": "A看向门口。",
                "cut_priority": "normal",
                "cut_reasons": [],
                "cut_group": "B001-action",
                "cut_category": "action",
                "cut_moment_id": "B001-action",
            }
        ]
        data["beats"][1]["source_text"] = "B皱眉。"
        data["beats"][1]["facts"] = [
            {
                "fact_id": "B002-F01",
                "type": "emotion",
                "text": "B皱眉。",
                "cut_priority": "normal",
                "cut_reasons": [],
                "cut_group": "B002-emotion",
                "cut_category": "emotion",
                "cut_moment_id": "B002-emotion",
            }
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01"],
                "source_paragraph": "A看向门口。",
                "camera_main_image": "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机在桌边拍A。\n【场景首镜站位】A在门口，面向桌边。\nA看向门口。",
                "shot_type": "action",
                "split_reason": ["performance_continuity"],
            },
            {
                **data["shots"][1],
                "beat_ids": ["B002"],
                "covered_fact_ids": ["B002-F01"],
                "source_paragraph": "B皱眉。",
                "duration_seconds": 2,
                "duration_breakdown": {"sync_action_seconds": 1, "sync_dialogue_seconds": 0, "non_sync_action_seconds": 0, "emotional_pause_seconds": 1},
                "camera_main_image": "[侧面平视, 近景, 固定镜头]\n【机位逻辑】摄影机切到B的脸。\nB皱眉。",
                "visible_characters": ["B"],
                "continuity_updates": [],
                "shot_type": "reaction",
                "split_reason": ["performance_continuity"],
                "insert_priority": "none",
            },
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)
        self.assertTrue(any("[可合并]" in item for item in result.warnings), result.warnings)


if __name__ == "__main__":
    unittest.main()
