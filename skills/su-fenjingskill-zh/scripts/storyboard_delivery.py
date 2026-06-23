#!/usr/bin/env python3
"""Build and validate stable 7-column su-fenjingskill-zh deliveries."""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - reported at runtime.
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None

VERSION = "2.3.2"
SHEET_NAME = "分镜表"
HEADERS = [
    "镜号",
    "场景",
    "原剧本段落",
    "镜头时长(秒)",
    "运镜+主画面描述(含台词)",
    "备注",
    "Prompt",
]
PROMPT_FIELDS = ["时间：", "景别：", "构图：", "运镜手法：", "画面内容："]
DURATION_FIELDS = [
    "sync_action_seconds",
    "sync_dialogue_seconds",
    "non_sync_action_seconds",
    "emotional_pause_seconds",
]
FACT_TYPES = {
    "character",
    "action",
    "dialogue",
    "prop",
    "space",
    "position",
    "emotion",
    "sound",
    "reality",
}
VALID_STATUSES = {"PASS", "WARN", "FAIL"}
ANCHOR_TYPES = {"space", "multi_character", "both", "single_continuation", "subjective"}
PANORAMIC_SHOT_SIZE_KEYWORDS = ("大远景", "大全景", "中全景", "全景")
PANORAMIC_REQUIRED_ANCHORS = {"space", "multi_character", "both"}
AERIAL_ALLOWED_SHOT_SIZE_KEYWORDS = ("大远景", "大全景", "全景")
AERIAL_ALLOWED_ANGLE_KEYWORDS = ("俯拍", "高角度", "正俯拍", "微俯视")
INTERNAL_LABELS = [
    "【镜内变化】",
    "【人物关系】",
    "【人物关系补充】",
    "【站位位移】",
    "【机位逻辑】",
    "【场景首镜站位】",
]
PROMPT_FORBIDDEN = [
    "【镜内变化】",
    "【人物关系】",
    "【人物关系补充】",
    "【站位位移】",
    "【机位逻辑】",
    "【场景首镜站位】",
    "运镜修正",
    "关键帧",
    "生图",
    "AI生图提示词",
    "风格：",
    "禁止：",
]
UNSUPPORTED_MARKERS = [
    "承上镜动作",
    "承上镜",
    "补充镜头",
    "新增镜头",
    "新增过渡",
    "无原文依据",
    "非原文",
    "原文未写",
    "自行添加",
    "自行补充",
    "为了过渡",
    "建议保留此镜",
]
FORBIDDEN_VISUAL_WORDS = [
    "令人叹为观止",
    "令人惊叹",
    "令人着迷",
    "精心打造",
    "匠心独运",
    "独具匠心",
    "视觉盛宴",
    "光影交响",
    "完美呈现",
    "极致体验",
    "引人入胜",
    "震撼人心",
    "巧妙融合",
    "仿佛",
    "犹如",
    "宛如",
    "好似",
]


@dataclass
class ValidationResult:
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def status(self) -> str:
        if self.errors:
            return "FAIL"
        if self.warnings:
            return "WARN"
        return "PASS"

    def error(self, message: str) -> None:
        self.errors.append(message)

    def warn(self, message: str) -> None:
        if message not in self.warnings:
            self.warnings.append(message)


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig") as handle:
        value = json.load(handle)
    if not isinstance(value, dict):
        raise ValueError("shot_data 顶层必须是 JSON 对象。")
    return value


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def clean_text(value: Any) -> str:
    return str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()


def one_line(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def fmt_number(value: Any) -> str:
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:g}"


def numeric_beat_id(beat_id: str) -> int | None:
    match = re.fullmatch(r"B(\d{3})", str(beat_id))
    return int(match.group(1)) if match else None


def format_beat_ids(beat_ids: list[Any]) -> str:
    ids = [str(item) for item in beat_ids if item]
    if not ids:
        return "B000"
    numbers = [numeric_beat_id(item) for item in ids]
    if all(number is not None for number in numbers):
        first = numbers[0]
        expected = list(range(first, first + len(numbers)))
        if numbers == expected:
            if len(ids) == 1:
                return ids[0]
            return f"{ids[0]}-{ids[-1]}"
    return "+".join(ids)


def source_with_beats(shot: dict[str, Any]) -> str:
    return f"{format_beat_ids(as_list(shot.get('beat_ids')))}～{one_line(shot.get('source_paragraph'))}"


def escape_cell(value: Any) -> str:
    return clean_text(value).replace("|", "｜").replace("\n", "<br>")


def parse_triad(camera: str) -> tuple[str, str, str]:
    first = clean_text(camera).split("\n", 1)[0].strip()
    match = re.fullmatch(r"\[([^,\]]+),\s*([^,\]]+),\s*([^\]]+)\]", first)
    if not match:
        return "平视", "中景", "固定镜头"
    return tuple(part.strip() for part in match.groups())


def is_panoramic_shot_size(shot_size: str) -> bool:
    value = clean_text(shot_size)
    return any(keyword in value for keyword in PANORAMIC_SHOT_SIZE_KEYWORDS)


def has_any_keyword(value: str, keywords: tuple[str, ...]) -> bool:
    text = clean_text(value)
    return any(keyword in text for keyword in keywords)


def strip_internal_prefix(line: str) -> str:
    value = line.strip()
    for label in INTERNAL_LABELS:
        if value.startswith(label):
            return value[len(label):].lstrip(" ：:").strip()
    return value


def prompt_visual_text(camera: str, source: str) -> str:
    lines = clean_text(camera).split("\n")
    body: list[str] = []
    for index, raw in enumerate(lines):
        line = raw.strip()
        if not line:
            continue
        if index == 0 and line.startswith("["):
            continue
        if line.startswith("【机位逻辑】") or line.startswith("【场景首镜站位】"):
            continue
        if line.startswith("【站位位移】"):
            continue
        stripped = strip_internal_prefix(line)
        if stripped:
            body.append(stripped)
    text = "；".join(body) if body else one_line(source)
    return sanitize_prompt_text(text)


def sanitize_prompt_text(text: str) -> str:
    value = clean_text(text)
    for label in PROMPT_FORBIDDEN:
        value = value.replace(label, "")
    value = re.sub(r"\s*；\s*", "；", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip(" ；")


def first_sentence(text: str) -> str:
    value = sanitize_prompt_text(text)
    if not value:
        return "当前镜头可见主体动作。"
    match = re.search(r"[。！？；]", value)
    if match:
        return value[: match.end()]
    return value[:80]


def derive_prompt(shot: dict[str, Any]) -> str:
    duration = fmt_number(shot.get("duration_seconds", 0))
    _angle, shot_size, movement = parse_triad(clean_text(shot.get("camera_main_image")))
    visual = prompt_visual_text(
        clean_text(shot.get("camera_main_image")),
        clean_text(shot.get("source_paragraph")),
    )
    return "\n".join(
        [
            f"时间：0秒-{duration}秒",
            f"景别：{shot_size}",
            f"构图：{first_sentence(visual)}",
            f"运镜手法：{movement}",
            f"画面内容：{visual}",
        ]
    )


def derive_prompts(data: dict[str, Any]) -> None:
    for shot in as_list(data.get("shots")):
        if isinstance(shot, dict):
            shot["prompt"] = derive_prompt(shot)


def expected_rows(data: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for shot in as_list(data.get("shots")):
        if not isinstance(shot, dict):
            continue
        rows.append(
            [
                shot.get("shot_no", ""),
                shot.get("scene", ""),
                source_with_beats(shot),
                shot.get("duration_seconds", ""),
                clean_text(shot.get("camera_main_image")),
                clean_text(shot.get("notes")),
                clean_text(shot.get("prompt")),
            ]
        )
    return rows


def update_validation_report(data: dict[str, Any], result: ValidationResult) -> None:
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    scene_counts: dict[str, int] = {}
    scene_logs = {
        str(log.get("scene")): log
        for log in as_list(data.get("continuity_logs"))
        if isinstance(log, dict) and log.get("scene")
    }
    first_shot_anchor_audit: list[dict[str, Any]] = []
    seen_scenes: set[str] = set()
    total_duration = 0.0
    over_6 = over_8 = over_10 = 0
    for shot in shots:
        scene = str(shot.get("scene", "")).split(" ")[0] or str(shot.get("scene", ""))
        scene_counts[scene] = scene_counts.get(scene, 0) + 1
        full_scene = str(shot.get("scene", ""))
        if full_scene not in seen_scenes:
            seen_scenes.add(full_scene)
            _angle, shot_size, _movement = parse_triad(clean_text(shot.get("camera_main_image")))
            anchor_type = as_dict(scene_logs.get(full_scene)).get("first_shot_anchor_type", "missing")
            visible_count = len(as_list(shot.get("visible_characters")))
            requires_panorama = anchor_type in PANORAMIC_REQUIRED_ANCHORS or (
                visible_count >= 2 and anchor_type != "single_continuation"
            )
            first_shot_anchor_audit.append(
                {
                    "scene": full_scene,
                    "shot_no": shot.get("shot_no"),
                    "anchor_type": anchor_type,
                    "shot_size": shot_size,
                    "visible_character_count": visible_count,
                    "requires_panoramic_first_shot": requires_panorama,
                    "passed": (not requires_panorama) or is_panoramic_shot_size(shot_size),
                }
            )
        duration = float(shot.get("duration_seconds") or 0)
        total_duration += duration
        over_6 += int(duration > 6)
        over_8 += int(duration > 8)
        over_10 += int(duration > 10)
    data["validation_report"] = {
        "status": result.status,
        "warnings": result.warnings,
        "errors": result.errors,
        "shot_count": len(shots),
        "scene_counts": scene_counts,
        "total_duration_seconds": int(total_duration) if total_duration.is_integer() else total_duration,
        "average_duration_seconds": round(total_duration / len(shots), 2) if shots else 0,
        "shots_over_6_seconds": over_6,
        "shots_over_8_seconds": over_8,
        "shots_over_10_seconds": over_10,
        "column_contract": "7-column stable storyboard table; keyframe column removed",
        "first_shot_anchor_audit": first_shot_anchor_audit,
    }


def validate_metadata(data: dict[str, Any], result: ValidationResult, *, strict_status: bool) -> None:
    metadata = as_dict(data.get("metadata"))
    if metadata.get("skill_name") != "su-fenjingskill-zh":
        result.error("metadata.skill_name 必须为 su-fenjingskill-zh。")
    if "reference_status" not in metadata:
        result.warn("metadata.reference_status 缺失，建议记录 reference loaded/missing 状态。")
    report = as_dict(data.get("validation_report"))
    status = report.get("status")
    if strict_status and status is not None and status not in VALID_STATUSES:
        result.error("validation_report.status 只允许 PASS / WARN / FAIL。")


def validate_continuity_logs(data: dict[str, Any], result: ValidationResult) -> dict[str, dict[str, Any]]:
    logs_by_scene: dict[str, dict[str, Any]] = {}
    for log in as_list(data.get("continuity_logs")):
        if not isinstance(log, dict):
            result.error("continuity_logs 中每一项必须是对象。")
            continue
        scene = str(log.get("scene") or "")
        if not scene:
            result.error("continuity_logs 每项必须包含 scene。")
            continue
        if scene in logs_by_scene:
            result.error(f"continuity_logs 场景重复：{scene}")
        logs_by_scene[scene] = log
        anchor_type = log.get("first_shot_anchor_type")
        if anchor_type not in ANCHOR_TYPES:
            result.error(
                f"{scene} first_shot_anchor_type 必须为 "
                "space / multi_character / both / single_continuation / subjective。"
            )
    return logs_by_scene


def collect_facts(data: dict[str, Any], result: ValidationResult) -> tuple[dict[str, str], set[str]]:
    beat_to_facts: dict[str, set[str]] = {}
    fact_to_beat: dict[str, str] = {}
    seen_beats: set[str] = set()
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            result.error("beats 中每一项必须是对象。")
            continue
        beat_id = str(beat.get("beat_id", ""))
        if not re.fullmatch(r"B\d{3}", beat_id):
            result.error(f"Beat ID 不合法：{beat_id}")
        if beat_id in seen_beats:
            result.error(f"Beat ID 重复：{beat_id}")
        seen_beats.add(beat_id)
        facts = as_list(beat.get("facts"))
        if not facts:
            result.warn(f"{beat_id} 没有 facts。")
        beat_to_facts.setdefault(beat_id, set())
        for fact in facts:
            if not isinstance(fact, dict):
                result.error(f"{beat_id} facts 中存在非对象项。")
                continue
            fact_id = str(fact.get("fact_id", ""))
            if not re.fullmatch(rf"{beat_id}-F\d{{2}}", fact_id):
                result.error(f"事实 ID 必须绑定所属 Beat：{fact_id}")
            if fact.get("type") not in FACT_TYPES:
                result.error(f"{fact_id} 事实类型不合法：{fact.get('type')}")
            fact_to_beat[fact_id] = beat_id
            beat_to_facts[beat_id].add(fact_id)
    return fact_to_beat, set(fact_to_beat)


def validate_first_scene_shot(shot: dict[str, Any], scene_log: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    camera = clean_text(shot.get("camera_main_image"))
    _angle, shot_size, movement = parse_triad(camera)
    anchor_type = scene_log.get("first_shot_anchor_type")
    visible_count = len(as_list(shot.get("visible_characters")))
    has_first_position = "【场景首镜站位】" in camera
    if not has_first_position:
        result.error(f"镜号{shot_no} 是场景首镜，必须包含【场景首镜站位】。")
    requires_panorama = anchor_type in PANORAMIC_REQUIRED_ANCHORS or (
        visible_count >= 2 and anchor_type != "single_continuation"
    )
    if requires_panorama and not is_panoramic_shot_size(shot_size):
        result.error(
            f"镜号{shot_no} 为 {anchor_type or '未标注'} 锚定场景首镜，"
            f"且可见人物数 {visible_count}，景别必须为大远景/大全景/全景/中全景；当前为 {shot_size}。"
        )
    if anchor_type in {"space", "both"} and "斯坦尼康" in clean_text(movement):
        result.error(f"镜号{shot_no} 为 {anchor_type} 锚定场景首镜，不得使用斯坦尼康建立大空间。")
    if visible_count >= 2 and anchor_type != "single_continuation":
        first_position_line = next(
            (line for line in camera.split("\n") if line.strip().startswith("【场景首镜站位】")),
            "",
        )
        blocking_text = first_position_line + "\n" + clean_text(scene_log.get("spatial_axis"))
        if not any(keyword in blocking_text for keyword in ["左", "右", "轴线", "对视", "朝向", "面向"]):
            result.error(f"镜号{shot_no} 多人场景首镜站位必须写清左右关系、朝向或主要对视轴线。")


def validate_rows(
    data: dict[str, Any],
    result: ValidationResult,
    fact_to_beat: dict[str, str],
    all_fact_ids: set[str],
    continuity_logs: dict[str, dict[str, Any]],
) -> None:
    shots = as_list(data.get("shots"))
    if not shots:
        result.error("shots 不能为空。")
        return
    expected_no = 1
    covered: set[str] = set()
    scenes = set(continuity_logs)
    seen_scenes: set[str] = set()
    has_update = False
    for shot in shots:
        if not isinstance(shot, dict):
            result.error("shots 中每一项必须是对象。")
            continue
        shot_no = shot.get("shot_no")
        if shot_no != expected_no:
            result.error(f"镜号必须从1连续递增：期望 {expected_no}，实际 {shot_no}")
        expected_no += 1
        for key in ["scene", "source_paragraph", "duration_seconds", "duration_breakdown", "camera_main_image", "notes", "prompt", "beat_ids", "covered_fact_ids", "continuity_updates"]:
            if key not in shot:
                result.error(f"镜号{shot_no} 缺少字段：{key}")
        if "keyframe" in shot:
            result.error(f"镜号{shot_no} 不得包含 keyframe 字段；关键帧列已从主表删除。")
        if shot.get("old_shot_no") is not None:
            result.error(f"镜号{shot_no} 不得保留 old_shot_no；禁止后处理合并痕迹进入交付数据。")
        scene = shot.get("scene")
        if scenes and scene not in scenes:
            result.warn(f"镜号{shot_no} 场景未在 continuity_logs 中登记：{scene}")
        is_first_scene_shot = str(scene) not in seen_scenes
        if is_first_scene_shot:
            seen_scenes.add(str(scene))
            validate_first_scene_shot(shot, as_dict(continuity_logs.get(str(scene))), result)
        beat_ids = [str(item) for item in as_list(shot.get("beat_ids"))]
        fact_ids = [str(item) for item in as_list(shot.get("covered_fact_ids"))]
        if not beat_ids:
            result.error(f"镜号{shot_no} beat_ids 不能为空。")
        if not fact_ids:
            result.error(f"镜号{shot_no} covered_fact_ids 不能为空。")
        for fact_id in fact_ids:
            covered.add(fact_id)
            owner = fact_to_beat.get(fact_id)
            if owner is None:
                result.error(f"镜号{shot_no} 覆盖了不存在的事实 ID：{fact_id}")
            elif owner not in beat_ids:
                result.error(f"镜号{shot_no} 覆盖事实 {fact_id}，但 beat_ids 未包含 {owner}。")
        validate_duration(shot, result)
        validate_camera(shot, result)
        validate_prompt(shot, result)
        updates = as_list(shot.get("continuity_updates"))
        if updates:
            has_update = True
        has_station_move = "【站位位移】" in clean_text(shot.get("camera_main_image"))
        if has_station_move and not updates:
            result.error(f"镜号{shot_no} 写了【站位位移】，但 continuity_updates 为空。")
        for update in updates:
            validate_update(shot, update, result, has_station_move)
    missing = sorted(all_fact_ids - covered)
    if missing:
        result.error("存在未覆盖事实 ID：" + ", ".join(missing[:20]))
    if any("【站位位移】" in clean_text(shot.get("camera_main_image")) for shot in shots if isinstance(shot, dict)) and not has_update:
        result.error("分镜出现站位位移，但没有任何 continuity_updates。")


def validate_duration(shot: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    try:
        duration = float(shot.get("duration_seconds"))
    except Exception:
        result.error(f"镜号{shot_no} 镜头时长必须是数字。")
        return
    if duration <= 0:
        result.error(f"镜号{shot_no} 镜头时长必须大于0。")
    breakdown = as_dict(shot.get("duration_breakdown"))
    for field_name in DURATION_FIELDS:
        if field_name not in breakdown:
            result.error(f"镜号{shot_no} duration_breakdown 缺少 {field_name}。")
            return
    try:
        expected = math.ceil(
            max(float(breakdown["sync_action_seconds"]), float(breakdown["sync_dialogue_seconds"]))
            + float(breakdown["non_sync_action_seconds"])
            + float(breakdown["emotional_pause_seconds"])
        )
    except Exception:
        result.error(f"镜号{shot_no} duration_breakdown 必须全为数字。")
        return
    if int(duration) != expected:
        result.error(f"镜号{shot_no} 时长不符合公式：表格 {duration:g} 秒，公式 {expected} 秒。")
    quoted = "".join(re.findall(r"[“\"]([^”\"]+)[”\"]", clean_text(shot.get("camera_main_image"))))
    spoken_chars = len(re.sub(r"[，。！？……、\s]", "", quoted))
    if spoken_chars:
        min_dialogue_seconds = math.ceil(spoken_chars / 4)
        if float(breakdown["sync_dialogue_seconds"]) < min_dialogue_seconds:
            result.error(f"镜号{shot_no} 对白时长过短：{spoken_chars}字至少约 {min_dialogue_seconds} 秒。")
    notes = clean_text(shot.get("notes"))
    if duration > 6 and "[时长估算]" not in notes:
        result.error(f"镜号{shot_no} 超过6秒，备注必须写 [时长估算]。")
    if duration > 10:
        long_take = as_dict(shot.get("long_take"))
        if long_take.get("classification") in (None, "not_applicable") and "[长镜头]" not in notes:
            result.error(f"镜号{shot_no} 超过10秒，必须标注长镜头分类或人工复核理由。")


def validate_camera(shot: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    camera = clean_text(shot.get("camera_main_image"))
    lines = [line.strip() for line in camera.split("\n") if line.strip()]
    if len(lines) < 2:
        result.error(f"镜号{shot_no} 运镜列至少需要三联首行和【机位逻辑】。")
        return
    if not re.fullmatch(r"\[[^,\]]+,\s*[^,\]]+,\s*[^\]]+\]", lines[0]):
        result.error(f"镜号{shot_no} 运镜首行必须是三要素：[视角/高度, 景别/特殊视角, 运镜方式]。")
    angle, shot_size, movement = parse_triad(camera)
    validate_camera_movement_compatibility(shot_no, angle, shot_size, movement, result)
    if not lines[1].startswith("【机位逻辑】"):
        result.error(f"镜号{shot_no} 运镜第二行必须以【机位逻辑】开头。")
    for word in FORBIDDEN_VISUAL_WORDS:
        if word in camera:
            result.error(f"镜号{shot_no} 运镜列包含禁用表达：{word}")
    for marker in UNSUPPORTED_MARKERS:
        if marker in camera or marker in clean_text(shot.get("notes")):
            result.error(f"镜号{shot_no} 包含后处理/无依据标记：{marker}")
    if "【镜内变化】" in camera:
        result.error(f"镜号{shot_no} 不得在最终主表出现【镜内变化】。")
    if "运镜修正" in clean_text(shot.get("notes")):
        result.error(f"镜号{shot_no} 备注不得保留运镜修正痕迹。")


def validate_camera_movement_compatibility(
    shot_no: Any,
    angle: str,
    shot_size: str,
    movement: str,
    result: ValidationResult,
) -> None:
    if "斯坦尼康" in clean_text(movement) and has_any_keyword(shot_size, ("大远景", "大全景")):
        result.error(f"镜号{shot_no} 大远景/大全景不得使用斯坦尼康；应改用摇臂、伸缩摇臂或航拍。")
    if "航拍" in clean_text(movement):
        if not has_any_keyword(shot_size, AERIAL_ALLOWED_SHOT_SIZE_KEYWORDS):
            result.error(f"镜号{shot_no} 航拍必须服务于大范围空间，景别应为大远景/大全景/全景。")
        if not has_any_keyword(angle, AERIAL_ALLOWED_ANGLE_KEYWORDS):
            result.error(f"镜号{shot_no} 航拍必须匹配俯拍、高角度、正俯拍或微俯视。")


def validate_prompt(shot: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    prompt = clean_text(shot.get("prompt"))
    expected = derive_prompt(shot)
    if prompt != expected:
        result.error(f"镜号{shot_no} Prompt 必须由锁定主表派生，且与标准五字段模板一致。")
    lines = [line.strip() for line in prompt.split("\n") if line.strip()]
    if len(lines) != 5:
        result.error(f"镜号{shot_no} Prompt 必须且只能包含五行。")
    for field_name in PROMPT_FIELDS:
        if not any(line.startswith(field_name) for line in lines):
            result.error(f"镜号{shot_no} Prompt 缺少字段：{field_name}")
    for item in PROMPT_FORBIDDEN:
        if item in prompt:
            result.error(f"镜号{shot_no} Prompt 包含禁止进入派生列的内部/关键帧内容：{item}")


def validate_update(shot: dict[str, Any], update: Any, result: ValidationResult, has_station_move: bool) -> None:
    shot_no = shot.get("shot_no")
    if not isinstance(update, dict):
        result.error(f"镜号{shot_no} continuity_updates 每项必须是对象。")
        return
    required = ["entity_type", "entity", "field", "from", "to", "evidence_fact_ids"]
    for key in required:
        if key not in update:
            result.error(f"镜号{shot_no} continuity_update 缺少 {key}。")
    if update.get("from") == update.get("to"):
        result.error(f"镜号{shot_no} continuity_update 的 from/to 不得相同。")
    evidence = [str(item) for item in as_list(update.get("evidence_fact_ids"))]
    covered = {str(item) for item in as_list(shot.get("covered_fact_ids"))}
    if not evidence:
        result.error(f"镜号{shot_no} continuity_update 必须有 evidence_fact_ids。")
    for fact_id in evidence:
        if fact_id not in covered:
            result.error(f"镜号{shot_no} continuity_update 证据 {fact_id} 未被当前镜头覆盖。")
    if update.get("field") in {"position", "facing"} and not has_station_move:
        result.error(f"镜号{shot_no} 登记了位置/朝向迁移，但主画面缺少【站位位移】。")


def validate_data(data: dict[str, Any], *, strict_status: bool) -> ValidationResult:
    result = ValidationResult()
    validate_metadata(data, result, strict_status=strict_status)
    continuity_logs = validate_continuity_logs(data, result)
    fact_to_beat, all_fact_ids = collect_facts(data, result)
    validate_rows(data, result, fact_to_beat, all_fact_ids, continuity_logs)
    return result


def markdown_text(data: dict[str, Any]) -> str:
    lines = ["# 分镜表", ""]
    lines.append("| " + " | ".join(HEADERS) + " |")
    lines.append("| " + " | ".join(["---"] * len(HEADERS)) + " |")
    for row in expected_rows(data):
        lines.append("| " + " | ".join(escape_cell(item) for item in row) + " |")
    report = as_dict(data.get("validation_report"))
    if report:
        lines.extend(
            [
                "",
                "## 校验摘要",
                "",
                f"- status: {report.get('status', '')}",
                f"- shot_count: {report.get('shot_count', '')}",
                f"- column_contract: {report.get('column_contract', '')}",
            ]
        )
    return "\n".join(lines) + "\n"


def build_markdown(data: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(markdown_text(data), encoding="utf-8")


def build_excel(data: dict[str, Any], path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl 不可用，无法生成 Excel。")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    for row in expected_rows(data):
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, width in enumerate([8, 24, 44, 14, 72, 36, 56], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 36 if row_index == 1 else 108
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def parse_markdown_table(path: Path) -> list[list[str]]:
    text = path.read_text(encoding="utf-8-sig")
    rows: list[list[str]] = []
    for line in text.splitlines():
        if not line.startswith("|"):
            continue
        cells = [cell.strip().replace("<br>", "\n").replace("｜", "|") for cell in line.strip().strip("|").split("|")]
        if cells and all(re.fullmatch(r":?-{3,}:?", cell) for cell in cells):
            continue
        rows.append(cells)
    return rows


def compare_markdown(data: dict[str, Any], path: Path, result: ValidationResult) -> None:
    if not path.exists():
        result.error(f"Markdown 文件不存在：{path}")
        return
    rows = parse_markdown_table(path)
    if not rows:
        result.error("Markdown 未找到分镜表。")
        return
    if rows[0] != HEADERS:
        result.error("Markdown 表头必须是稳定 7 列，且不得包含关键帧列。")
    expected = [[str(item) for item in row] for row in expected_rows(data)]
    actual = [[str(item) for item in row] for row in rows[1:]]
    if actual != expected:
        result.error("Markdown 表格内容与 shot_data 派生结果不一致。")


def compare_excel(data: dict[str, Any], path: Path, result: ValidationResult) -> None:
    if load_workbook is None:
        result.error("openpyxl 不可用，无法校验 Excel。")
        return
    if not path.exists():
        result.error(f"Excel 文件不存在：{path}")
        return
    workbook = load_workbook(path, read_only=True)
    try:
        if workbook.sheetnames != [SHEET_NAME]:
            result.error("Excel 必须只有一个 Sheet：分镜表。")
            return
        sheet = workbook[SHEET_NAME]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        result.error("Excel 分镜表为空。")
        return
    header = [str(item or "") for item in rows[0]]
    if header != HEADERS:
        result.error("Excel 表头必须是稳定 7 列，且不得包含关键帧列。")
    expected = [[str(item) for item in row] for row in expected_rows(data)]
    actual = [[str(item or "") for item in row] for row in rows[1:]]
    if actual != expected:
        result.error("Excel 内容与 shot_data 派生结果不一致。")


def print_result(result: ValidationResult) -> None:
    if result.errors:
        print("ERRORS:", file=sys.stderr)
        for item in result.errors:
            print(f"- {item}", file=sys.stderr)
    if result.warnings:
        print("WARNINGS:")
        for item in result.warnings:
            print(f"- {item}")
    print(f"{result.status}: storyboard validation {'passed' if not result.errors else 'failed'}.")


def command_build(args: argparse.Namespace) -> int:
    data_path = Path(args.data)
    data = load_json(data_path)
    derive_prompts(data)
    result = validate_data(data, strict_status=False)
    update_validation_report(data, result)
    if result.errors:
        write_json(data_path, data)
        print_result(result)
        return 1
    write_json(data_path, data)
    build_markdown(data, Path(args.markdown))
    build_excel(data, Path(args.excel))
    print_result(result)
    return 0


def command_validate(args: argparse.Namespace) -> int:
    data = load_json(Path(args.data))
    result = validate_data(data, strict_status=True)
    compare_markdown(data, Path(args.markdown), result)
    compare_excel(data, Path(args.excel), result)
    print_result(result)
    return 1 if result.errors else 0


def make_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    build = subparsers.add_parser("build", help="derive Prompt, update JSON, and write Markdown/Excel")
    build.add_argument("--data", required=True)
    build.add_argument("--markdown", required=True)
    build.add_argument("--excel", required=True)
    build.set_defaults(func=command_build)
    validate = subparsers.add_parser("validate", help="validate JSON/Markdown/Excel consistency")
    validate.add_argument("--data", required=True)
    validate.add_argument("--markdown", required=True)
    validate.add_argument("--excel", required=True)
    validate.set_defaults(func=command_validate)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = make_parser()
    args = parser.parse_args(argv)
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
