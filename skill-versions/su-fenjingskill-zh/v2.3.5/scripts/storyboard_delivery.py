#!/usr/bin/env python3
"""Build and validate stable 7-column su-fenjingskill-zh deliveries."""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - reported at runtime.
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None

VERSION = "2.3.5"
RULE_REVISION = "2.3.5-adjacent-motion-guard-2026-06-30"
RULE_REVISIONS = {
    "2.3.4": "2.3.4-overcompression-guard-2026-06-30",
    "2.3.5": RULE_REVISION,
}
SHEET_NAME = "分镜表"
HEADERS = [
    "镜号",
    "场景",
    "原剧本段落",
    "镜头时长(秒)",
    "运镜+主画面描述(含台词)",
    "备注",
    "Prompt",
]
PROMPT_FIELDS = ["时间：", "景别：", "构图：", "运镜手法：", "画面内容："]
DURATION_FIELDS = [
    "sync_action_seconds",
    "sync_dialogue_seconds",
    "non_sync_action_seconds",
    "emotional_pause_seconds",
]
FACT_TYPES = {
    "character",
    "action",
    "dialogue",
    "prop",
    "space",
    "position",
    "emotion",
    "sound",
    "reality",
}
SHOT_TYPES = {
    "master",
    "action",
    "dialogue",
    "reaction",
    "insert",
    "transition",
    "vfx_anchor",
    "safety",
}
SPLIT_REASONS = {
    "spatial_anchor",
    "performance_continuity",
    "new_information",
    "prop_state_change",
    "new_vfx_state",
    "new_sound_source",
    "reality_layer_shift",
    "causal_reveal",
    "emotional_turn",
    "continuity_migration",
}
INSERT_PRIORITIES = {"none", "recommended", "must_have"}
CUT_PRIORITIES = {"normal", "recommended", "must_isolate"}
CRITICAL_SPLIT_REASONS = {
    "causal_reveal",
    "reality_layer_shift",
    "new_vfx_state",
    "new_sound_source",
    "emotional_turn",
    "prop_state_change",
}
MUST_ISOLATE_REASONS = {
    "causal_reveal",
    "reality_layer_shift",
    "new_vfx_state",
    "new_sound_source",
    "emotional_turn",
}
LONG_TAKE_SUPPORTS = {
    "shot_size_change",
    "character_blocking",
    "foreground_background_change",
    "sound_source_change",
    "vfx_state_change",
    "emotional_turn",
    "spatial_progression",
}
INSERT_LIKE_TYPES = {"insert", "reaction", "vfx_anchor", "safety"}
HYBRID_VERSIONS = {"2.3.4", "2.3.5"}
ADJACENT_MOTION_GUARD_VERSION = "2.3.5"
VALID_STATUSES = {"PASS", "WARN", "FAIL"}
ANCHOR_TYPES = {"space", "multi_character", "both", "single_continuation", "subjective"}
PANORAMIC_SHOT_SIZE_KEYWORDS = ("大远景", "大全景", "中全景", "全景")
PANORAMIC_REQUIRED_ANCHORS = {"space", "multi_character", "both"}
AERIAL_ALLOWED_SHOT_SIZE_KEYWORDS = ("大远景", "大全景", "全景")
AERIAL_ALLOWED_ANGLE_KEYWORDS = ("俯拍", "高角度", "正俯拍", "微俯视")
WIDE_SHOT_SIZE_KEYWORDS = ("大远景", "大全景", "远景", "全景", "中全景")
CLOSE_SHOT_SIZE_KEYWORDS = ("中近景", "近景", "中特写", "特写", "微距特写")
SHOT_SIZE_RANKS = [
    (0, ("大远景", "大全景", "远景")),
    (1, ("全景", "中全景")),
    (2, ("中景", "七分身", "半身")),
    (3, ("中近景", "近景", "中特写")),
    (4, ("微距特写", "特写")),
]
STATIC_OR_SOFT_MOVEMENT_KEYWORDS = (
    "固定",
    "轻微",
    "缓慢",
    "微晃",
    "呼吸式",
    "下摇",
    "纵摇",
    "横摇",
)
PUSH_IN_MOVEMENT_KEYWORDS = (
    "推轨推进",
    "缓慢推进",
    "航拍推进",
    "急推",
    "推进",
)
PULL_OUT_MOVEMENT_KEYWORDS = (
    "推轨拉出",
    "缓慢拉出",
    "拉远",
    "拉出",
)
SPAN_MOVEMENT_KEYWORDS = (
    "光学变焦",
    "快速推拉",
    "急推",
    "急拉",
    "推轨推进",
    "推轨拉出",
    "缓慢推进",
    "缓慢拉出",
    "伸缩摇臂",
)
INCOMPLETE_DIALOGUE_ENDINGS = ("，", "、", "：", "；", ",", ":", ";")
EXPRESSION_ONLY_KEYWORDS = (
    "笑",
    "苦笑",
    "嘴角",
    "眼神",
    "抬眼",
    "闭眼",
    "呼吸",
    "停顿",
    "沉默",
    "表情",
    "眉",
    "瞳孔",
    "下颌",
    "嘴唇",
)
NEW_FACT_TYPES_FOR_CUT = {"position", "prop", "space", "sound", "reality"}
INTERNAL_LABELS = [
    "【镜内变化】",
    "【人物关系】",
    "【人物关系补充】",
    "【站位位移】",
    "【机位逻辑】",
    "【场景首镜站位】",
]
PROMPT_FORBIDDEN = [
    "【镜内变化】",
    "【人物关系】",
    "【人物关系补充】",
    "【站位位移】",
    "【机位逻辑】",
    "【场景首镜站位】",
    "运镜修正",
    "关键帧",
    "生图",
    "AI生图提示词",
    "风格：",
    "禁止：",
]
UNSUPPORTED_MARKERS = [
    "承上镜动作",
    "承上镜",
    "补充镜头",
    "新增镜头",
    "新增过渡",
    "无原文依据",
    "非原文",
    "原文未写",
    "自行添加",
    "自行补充",
    "为了过渡",
    "建议保留此镜",
]
FORBIDDEN_VISUAL_WORDS = [
    "令人叹为观止",
    "令人惊叹",
    "令人着迷",
    "精心打造",
    "匠心独运",
    "独具匠心",
    "视觉盛宴",
    "光影交响",
    "完美呈现",
    "极致体验",
    "引人入胜",
    "震撼人心",
    "巧妙融合",
    "仿佛",
    "犹如",
    "宛如",
    "好似",
]

REQUIRED_CUT_POINT_CATEGORIES = [
    (
        "发言权转移",
        [re.compile(r"[\u4e00-\u9fa5A-Za-z0-9·]{1,8}[：:]")],
        True,
    ),
    ("问答关系变化", [re.compile(r"问|回答|回应|反问|是不是|为什么|什么|吗|？|\?")], False),
    (
        "角色明显反应",
        [re.compile(r"一怔|脸色|眼眶|嘴唇|冷笑|大笑|苦笑|痛苦|不可思议|回头|低头|抬头|摇头|点头|踉跄|绷|僵|停住")],
        False,
    ),
    (
        "道具状态变化",
        [re.compile(r"手环|短棍|长棍|磁卡|魂钉|裂痕|项链|绿灯|显示|亮起|暗下|恢复|变长|延伸|变成|熄灭|铜镯")],
        False,
    ),
    (
        "攻击/命中/结果",
        [re.compile(r"抬手|挥|扑|冲|拽|拦|扫|斩|震|炸|击|挡|断裂|碎裂|灰飞烟灭|跳入|钻进|坠落|吞没")],
        False,
    ),
    (
        "空间方向改变",
        [re.compile(r"东侧|西侧|南侧|中央|高台|祭池边|走向|走到|冲向|转身|回到|两侧|不同方向|分头|裂缝|地底|神殿|人间")],
        False,
    ),
    ("层级切换", [re.compile(r"切回|切——|梦境|幻境|现实|闪回|脑海|灵魂|亡魂|残魂|画外声|VO|回忆")], False),
    (
        "阵法/VFX状态变化",
        [re.compile(r"阵眼|破阵|续阵|阵破|符文|祭池|锁魂柱|柱身|黑纹|雾气|黑雾|龙卷|漩涡|回收|膨胀|光柱|封印|赤光|蓝光")],
        False,
    ),
]


@dataclass
class ValidationResult:
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def status(self) -> str:
        if self.errors:
            return "FAIL"
        if self.warnings:
            return "WARN"
        return "PASS"

    def error(self, message: str) -> None:
        self.errors.append(message)

    def warn(self, message: str) -> None:
        if message not in self.warnings:
            self.warnings.append(message)


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig") as handle:
        value = json.load(handle)
    if not isinstance(value, dict):
        raise ValueError("shot_data 顶层必须是 JSON 对象。")
    return value


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def clean_text(value: Any) -> str:
    return str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()


def one_line(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def fmt_number(value: Any) -> str:
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:g}"


def numeric_beat_id(beat_id: str) -> int | None:
    match = re.fullmatch(r"B(\d{3})", str(beat_id))
    return int(match.group(1)) if match else None


def format_beat_ids(beat_ids: list[Any]) -> str:
    ids = [str(item) for item in beat_ids if item]
    if not ids:
        return "B000"
    numbers = [numeric_beat_id(item) for item in ids]
    if all(number is not None for number in numbers):
        first = numbers[0]
        expected = list(range(first, first + len(numbers)))
        if numbers == expected:
            if len(ids) == 1:
                return ids[0]
            return f"{ids[0]}-{ids[-1]}"
    return "+".join(ids)


def source_with_beats(shot: dict[str, Any]) -> str:
    return f"{format_beat_ids(as_list(shot.get('beat_ids')))}～{one_line(shot.get('source_paragraph'))}"


def escape_cell(value: Any) -> str:
    return clean_text(value).replace("|", "｜").replace("\n", "<br>")


def parse_triad(camera: str) -> tuple[str, str, str]:
    first = clean_text(camera).split("\n", 1)[0].strip()
    match = re.fullmatch(r"\[([^,\]]+),\s*([^,\]]+),\s*([^\]]+)\]", first)
    if not match:
        return "平视", "中景", "固定镜头"
    return tuple(part.strip() for part in match.groups())


def is_panoramic_shot_size(shot_size: str) -> bool:
    value = clean_text(shot_size)
    return any(keyword in value for keyword in PANORAMIC_SHOT_SIZE_KEYWORDS)


def shot_size_rank(shot_size: str) -> int | None:
    value = clean_text(shot_size)
    for rank, keywords in SHOT_SIZE_RANKS:
        if any(keyword in value for keyword in keywords):
            return rank
    return None


def shot_size_has_span(shot_size: str) -> bool:
    value = clean_text(shot_size)
    return has_any_keyword(value, WIDE_SHOT_SIZE_KEYWORDS) and has_any_keyword(value, CLOSE_SHOT_SIZE_KEYWORDS)


def angle_family(angle: str) -> str:
    value = clean_text(angle)
    if "主观" in value:
        return "subjective"
    if "过肩" in value:
        return "over_shoulder"
    if "仰" in value:
        return "low"
    if "俯" in value or "高角度" in value:
        return "high"
    if "侧" in value:
        return "side"
    return "level"


def is_static_or_soft_movement(movement: str) -> bool:
    return has_any_keyword(movement, STATIC_OR_SOFT_MOVEMENT_KEYWORDS)


def motion_axis_direction(movement: str) -> str | None:
    value = clean_text(movement)
    if has_any_keyword(value, PUSH_IN_MOVEMENT_KEYWORDS):
        return "push_in"
    if has_any_keyword(value, PULL_OUT_MOVEMENT_KEYWORDS):
        return "pull_out"
    return None


def has_continuity_position_change(shot: dict[str, Any]) -> bool:
    if "【站位位移】" in clean_text(shot.get("camera_main_image")):
        return True
    for update in as_list(shot.get("continuity_updates")):
        if isinstance(update, dict) and update.get("field") in {"position", "facing"}:
            return True
    return False


def normalized_visible_characters(shot: dict[str, Any]) -> tuple[str, ...]:
    return tuple(sorted(str(item) for item in as_list(shot.get("visible_characters")) if str(item)))


def fact_types_for_shot(shot: dict[str, Any], fact_types: dict[str, str]) -> set[str]:
    return {fact_types[fact_id] for fact_id in map(str, as_list(shot.get("covered_fact_ids"))) if fact_id in fact_types}


def quoted_dialogue_text(shot: dict[str, Any]) -> str:
    camera = clean_text(shot.get("camera_main_image"))
    quoted = "".join(re.findall(r"[“\"]([^”\"]+)[”\"]", camera))
    if quoted:
        return quoted.strip()
    source = one_line(shot.get("source_paragraph"))
    match = re.search(r"[：:](.+)$", source)
    return match.group(1).strip() if match else ""


def dialogue_speaker(shot: dict[str, Any]) -> str:
    source = one_line(shot.get("source_paragraph"))
    match = re.match(r"([^：:]{1,20})[：:]", source)
    return match.group(1).strip() if match else ""


def is_expression_only_response(shot: dict[str, Any], fact_types: dict[str, str]) -> bool:
    types = fact_types_for_shot(shot, fact_types)
    if types & NEW_FACT_TYPES_FOR_CUT:
        return False
    if not (types & {"action", "emotion", "character"}):
        return False
    text = one_line(shot.get("source_paragraph")) + " " + one_line(shot.get("camera_main_image"))
    return has_any_keyword(text, EXPRESSION_ONLY_KEYWORDS)


def has_any_keyword(value: str, keywords: tuple[str, ...]) -> bool:
    text = clean_text(value)
    return any(keyword in text for keyword in keywords)


def metadata_version(data: dict[str, Any]) -> str:
    return clean_text(as_dict(data.get("metadata")).get("version"))


def requires_hybrid_fields(data: dict[str, Any]) -> bool:
    return metadata_version(data) in HYBRID_VERSIONS


def requires_adjacent_motion_guard(data: dict[str, Any]) -> bool:
    return metadata_version(data) == ADJACENT_MOTION_GUARD_VERSION


def current_rule_revision(data: dict[str, Any]) -> str:
    return RULE_REVISIONS.get(metadata_version(data), RULE_REVISION)


def normalized_list(value: Any) -> list[str]:
    return [str(item) for item in as_list(value) if str(item)]


def fact_lookup(data: dict[str, Any]) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        for fact in as_list(beat.get("facts")):
            if isinstance(fact, dict) and fact.get("fact_id"):
                lookup[str(fact.get("fact_id"))] = fact
    return lookup


def fact_cut_reasons(fact: dict[str, Any]) -> set[str]:
    return set(normalized_list(fact.get("cut_reasons")))


def is_fact_critical(fact: dict[str, Any]) -> bool:
    return clean_text(fact.get("cut_priority")) == "must_isolate" or bool(fact_cut_reasons(fact) & CRITICAL_SPLIT_REASONS)


def find_required_cut_point_categories(text_value: str) -> list[str]:
    found: list[str] = []
    for name, patterns, count_matches in REQUIRED_CUT_POINT_CATEGORIES:
        if count_matches:
            count = sum(len(pattern.findall(text_value)) for pattern in patterns)
            if count >= 2:
                found.append(name)
            continue
        if any(pattern.search(text_value) for pattern in patterns):
            found.append(name)
    return found


def extract_source_speakers(text_value: str) -> list[str]:
    forbidden_labels = {"时间", "景别", "构图", "运镜手法", "画面内容", "场景", "主体", "画面", "风格", "禁止"}
    speakers: list[str] = []
    for match in re.finditer(r"(?:^|[\n/。；;，,])\s*([\u4e00-\u9fa5A-Za-z0-9·]{1,8})(?:（VO）|\(VO\))?[：:]", clean_text(text_value)):
        speaker = match.group(1).strip()
        if speaker and speaker not in forbidden_labels and speaker not in speakers:
            speakers.append(speaker)
    return speakers


def source_marks_offscreen_voice(text_value: str, speaker: str) -> bool:
    source = clean_text(text_value)
    return bool(re.search(rf"{re.escape(speaker)}\s*(?:（VO）|\(VO\)|VO|画外声)", source))


def shot_fact_objects(shot: dict[str, Any], facts_by_id: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    facts: list[dict[str, Any]] = []
    for fact_id in normalized_list(shot.get("covered_fact_ids")):
        fact = facts_by_id.get(fact_id)
        if fact is not None:
            facts.append(fact)
    return facts


def shot_cut_groups(facts: list[dict[str, Any]], *, must_only: bool) -> dict[str, list[str]]:
    groups: dict[str, list[str]] = {}
    for fact in facts:
        if must_only and clean_text(fact.get("cut_priority")) != "must_isolate":
            continue
        group = clean_text(fact.get("cut_group"))
        if not group:
            continue
        groups.setdefault(group, []).append(str(fact.get("fact_id")))
    return groups


def strip_internal_prefix(line: str) -> str:
    value = line.strip()
    for label in INTERNAL_LABELS:
        if value.startswith(label):
            return value[len(label):].lstrip(" ：:").strip()
    return value


def build_overcompression_audit(data: dict[str, Any]) -> dict[str, Any]:
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    facts_by_id = fact_lookup(data)
    at_risk_shots: list[dict[str, Any]] = []
    total_duration = 0.0
    short_3 = 0
    multi_beat_critical = 0
    for shot in shots:
        duration = float(shot.get("duration_seconds") or 0)
        total_duration += duration
        short_3 += int(duration <= 3)
        facts = shot_fact_objects(shot, facts_by_id)
        critical_facts = [fact for fact in facts if is_fact_critical(fact)]
        must_groups = shot_cut_groups(facts, must_only=True)
        cut_points = find_required_cut_point_categories(
            clean_text(shot.get("source_paragraph")) + "\n" + clean_text(shot.get("camera_main_image"))
        )
        beat_count = len(as_list(shot.get("beat_ids")))
        has_multi_beat_critical = beat_count >= 2 and bool(critical_facts)
        multi_beat_critical += int(has_multi_beat_critical)
        if critical_facts or len(must_groups) > 1 or cut_points or beat_count >= 3 or duration >= 9:
            at_risk_shots.append(
                {
                    "shot_no": shot.get("shot_no"),
                    "duration_seconds": int(duration) if duration.is_integer() else duration,
                    "beat_count": beat_count,
                    "shot_type": shot.get("shot_type"),
                    "critical_fact_count": len(critical_facts),
                    "must_isolate_groups": sorted(must_groups),
                    "required_cut_points": cut_points,
                    "multi_beat_critical": has_multi_beat_critical,
                }
            )
    cut_per_minute = round(((len(shots) - 1) / total_duration) * 60, 2) if total_duration > 0 and len(shots) > 1 else 0
    short_ratio = round(short_3 / len(shots), 4) if shots else 0
    return {
        "rule_revision": current_rule_revision(data),
        "at_risk_shots": at_risk_shots,
        "multi_beat_critical_shots": multi_beat_critical,
        "cut_per_minute": cut_per_minute,
        "short_shot_ratio": short_ratio,
    }


def prompt_visual_text(camera: str, source: str) -> str:
    lines = clean_text(camera).split("\n")
    body: list[str] = []
    for index, raw in enumerate(lines):
        line = raw.strip()
        if not line:
            continue
        if index == 0 and line.startswith("["):
            continue
        if line.startswith("【机位逻辑】") or line.startswith("【场景首镜站位】"):
            continue
        if line.startswith("【站位位移】"):
            continue
        stripped = strip_internal_prefix(line)
        if stripped:
            body.append(stripped)
    text = "；".join(body) if body else one_line(source)
    return sanitize_prompt_text(text)


def sanitize_prompt_text(text: str) -> str:
    value = clean_text(text)
    for label in PROMPT_FORBIDDEN:
        value = value.replace(label, "")
    value = re.sub(r"\s*；\s*", "；", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip(" ；")


def first_sentence(text: str) -> str:
    value = sanitize_prompt_text(text)
    if not value:
        return "当前镜头可见主体动作。"
    match = re.search(r"[。！？；]", value)
    if match:
        return value[: match.end()]
    return value[:80]


def derive_prompt(shot: dict[str, Any]) -> str:
    duration = fmt_number(shot.get("duration_seconds", 0))
    _angle, shot_size, movement = parse_triad(clean_text(shot.get("camera_main_image")))
    visual = prompt_visual_text(
        clean_text(shot.get("camera_main_image")),
        clean_text(shot.get("source_paragraph")),
    )
    return "\n".join(
        [
            f"时间：0秒-{duration}秒",
            f"景别：{shot_size}",
            f"构图：{first_sentence(visual)}",
            f"运镜手法：{movement}",
            f"画面内容：{visual}",
        ]
    )


def derive_prompts(data: dict[str, Any]) -> None:
    for shot in as_list(data.get("shots")):
        if isinstance(shot, dict):
            shot["prompt"] = derive_prompt(shot)


def expected_rows(data: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for shot in as_list(data.get("shots")):
        if not isinstance(shot, dict):
            continue
        rows.append(
            [
                shot.get("shot_no", ""),
                shot.get("scene", ""),
                source_with_beats(shot),
                shot.get("duration_seconds", ""),
                clean_text(shot.get("camera_main_image")),
                clean_text(shot.get("notes")),
                clean_text(shot.get("prompt")),
            ]
        )
    return rows


def update_validation_report(data: dict[str, Any], result: ValidationResult) -> None:
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    scene_counts: dict[str, int] = {}
    scene_logs = {
        str(log.get("scene")): log
        for log in as_list(data.get("continuity_logs"))
        if isinstance(log, dict) and log.get("scene")
    }
    first_shot_anchor_audit: list[dict[str, Any]] = []
    seen_scenes: set[str] = set()
    total_duration = 0.0
    over_6 = over_8 = over_10 = 0
    short_3 = 0
    long_6_or_more = 0
    shot_type_counts: dict[str, int] = {}
    insert_priority_counts: dict[str, int] = {}
    long_take_audit: list[dict[str, Any]] = []
    for shot in shots:
        scene = str(shot.get("scene", "")).split(" ")[0] or str(shot.get("scene", ""))
        scene_counts[scene] = scene_counts.get(scene, 0) + 1
        full_scene = str(shot.get("scene", ""))
        if full_scene not in seen_scenes:
            seen_scenes.add(full_scene)
            _angle, shot_size, _movement = parse_triad(clean_text(shot.get("camera_main_image")))
            anchor_type = as_dict(scene_logs.get(full_scene)).get("first_shot_anchor_type", "missing")
            visible_count = len(as_list(shot.get("visible_characters")))
            requires_panorama = anchor_type in PANORAMIC_REQUIRED_ANCHORS or (
                visible_count >= 2 and anchor_type != "single_continuation"
            )
            first_shot_anchor_audit.append(
                {
                    "scene": full_scene,
                    "shot_no": shot.get("shot_no"),
                    "anchor_type": anchor_type,
                    "shot_size": shot_size,
                    "visible_character_count": visible_count,
                    "requires_panoramic_first_shot": requires_panorama,
                    "passed": (not requires_panorama) or is_panoramic_shot_size(shot_size),
                }
            )
        duration = float(shot.get("duration_seconds") or 0)
        total_duration += duration
        over_6 += int(duration > 6)
        over_8 += int(duration > 8)
        over_10 += int(duration > 10)
        short_3 += int(duration <= 3)
        long_6_or_more += int(duration >= 6)
        shot_type = clean_text(shot.get("shot_type")) or "missing"
        insert_priority = clean_text(shot.get("insert_priority")) or "missing"
        shot_type_counts[shot_type] = shot_type_counts.get(shot_type, 0) + 1
        insert_priority_counts[insert_priority] = insert_priority_counts.get(insert_priority, 0) + 1
        if duration > 10:
            long_take_audit.append(
                {
                    "shot_no": shot.get("shot_no"),
                    "duration_seconds": int(duration) if duration.is_integer() else duration,
                    "long_take_support": normalized_list(shot.get("long_take_support")),
                }
            )
    cut_per_minute = round(((len(shots) - 1) / total_duration) * 60, 2) if total_duration > 0 and len(shots) > 1 else 0
    short_ratio = round(short_3 / len(shots), 4) if shots else 0
    long_ratio = round(long_6_or_more / len(shots), 4) if shots else 0
    hybrid_audit = {
        "short_shots_le_3_seconds": short_3,
        "short_shot_ratio": short_ratio,
        "long_shots_ge_6_seconds": long_6_or_more,
        "long_shot_ratio": long_ratio,
        "cut_per_minute": cut_per_minute,
        "shot_type_counts": shot_type_counts,
        "insert_priority_counts": insert_priority_counts,
        "long_take_audit": long_take_audit,
        "soft_targets": {
            "short_shot_ratio": "0.10-0.18 for full-episode references only",
            "long_shot_ratio": "0.40-0.55 for full-episode references only",
            "cut_per_minute": "10.5-12.5 for full-episode references only",
        },
    }
    data["validation_report"] = {
        "status": result.status,
        "warnings": result.warnings,
        "errors": result.errors,
        "shot_count": len(shots),
        "scene_counts": scene_counts,
        "total_duration_seconds": int(total_duration) if total_duration.is_integer() else total_duration,
        "average_duration_seconds": round(total_duration / len(shots), 2) if shots else 0,
        "shots_over_6_seconds": over_6,
        "shots_over_8_seconds": over_8,
        "shots_over_10_seconds": over_10,
        "column_contract": "7-column stable storyboard table; keyframe column removed",
        "first_shot_anchor_audit": first_shot_anchor_audit,
        "hybrid_audit": hybrid_audit,
        "overcompression_audit": build_overcompression_audit(data),
    }


def validate_metadata(data: dict[str, Any], result: ValidationResult, *, strict_status: bool) -> None:
    metadata = as_dict(data.get("metadata"))
    if metadata.get("skill_name") != "su-fenjingskill-zh":
        result.error("metadata.skill_name 必须为 su-fenjingskill-zh。")
    version = metadata_version(data)
    expected_rule_revision = RULE_REVISIONS.get(version)
    if expected_rule_revision and metadata.get("rule_revision") != expected_rule_revision:
        result.error(f"metadata.rule_revision 必须为 {expected_rule_revision}。")
    if "reference_status" not in metadata:
        result.warn("metadata.reference_status 缺失，建议记录 reference loaded/missing 状态。")
    report = as_dict(data.get("validation_report"))
    status = report.get("status")
    if strict_status and status is not None and status not in VALID_STATUSES:
        result.error("validation_report.status 只允许 PASS / WARN / FAIL。")


def validate_continuity_logs(data: dict[str, Any], result: ValidationResult) -> dict[str, dict[str, Any]]:
    logs_by_scene: dict[str, dict[str, Any]] = {}
    for log in as_list(data.get("continuity_logs")):
        if not isinstance(log, dict):
            result.error("continuity_logs 中每一项必须是对象。")
            continue
        scene = str(log.get("scene") or "")
        if not scene:
            result.error("continuity_logs 每项必须包含 scene。")
            continue
        if scene in logs_by_scene:
            result.error(f"continuity_logs 场景重复：{scene}")
        logs_by_scene[scene] = log
        anchor_type = log.get("first_shot_anchor_type")
        if anchor_type not in ANCHOR_TYPES:
            result.error(
                f"{scene} first_shot_anchor_type 必须为 "
                "space / multi_character / both / single_continuation / subjective。"
            )
    return logs_by_scene


def collect_facts(data: dict[str, Any], result: ValidationResult) -> tuple[dict[str, str], dict[str, str], set[str]]:
    beat_to_facts: dict[str, set[str]] = {}
    fact_to_beat: dict[str, str] = {}
    fact_types: dict[str, str] = {}
    seen_beats: set[str] = set()
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            result.error("beats 中每一项必须是对象。")
            continue
        beat_id = str(beat.get("beat_id", ""))
        if not re.fullmatch(r"B\d{3}", beat_id):
            result.error(f"Beat ID 不合法：{beat_id}")
        if beat_id in seen_beats:
            result.error(f"Beat ID 重复：{beat_id}")
        seen_beats.add(beat_id)
        facts = as_list(beat.get("facts"))
        if not facts:
            result.warn(f"{beat_id} 没有 facts。")
        beat_to_facts.setdefault(beat_id, set())
        for fact in facts:
            if not isinstance(fact, dict):
                result.error(f"{beat_id} facts 中存在非对象项。")
                continue
            fact_id = str(fact.get("fact_id", ""))
            if not re.fullmatch(rf"{beat_id}-F\d{{2}}", fact_id):
                result.error(f"事实 ID 必须绑定所属 Beat：{fact_id}")
            if fact.get("type") not in FACT_TYPES:
                result.error(f"{fact_id} 事实类型不合法：{fact.get('type')}")
            fact_to_beat[fact_id] = beat_id
            fact_types[fact_id] = str(fact.get("type", ""))
            beat_to_facts[beat_id].add(fact_id)
    return fact_to_beat, fact_types, set(fact_to_beat)


def validate_first_scene_shot(shot: dict[str, Any], scene_log: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    camera = clean_text(shot.get("camera_main_image"))
    _angle, shot_size, movement = parse_triad(camera)
    anchor_type = scene_log.get("first_shot_anchor_type")
    visible_count = len(as_list(shot.get("visible_characters")))
    has_first_position = "【场景首镜站位】" in camera
    if not has_first_position:
        result.error(f"镜号{shot_no} 是场景首镜，必须包含【场景首镜站位】。")
    requires_panorama = anchor_type in PANORAMIC_REQUIRED_ANCHORS or (
        visible_count >= 2 and anchor_type != "single_continuation"
    )
    if requires_panorama and not is_panoramic_shot_size(shot_size):
        result.error(
            f"镜号{shot_no} 为 {anchor_type or '未标注'} 锚定场景首镜，"
            f"且可见人物数 {visible_count}，景别必须为大远景/大全景/全景/中全景；当前为 {shot_size}。"
        )
    if anchor_type in {"space", "both"} and "斯坦尼康" in clean_text(movement):
        result.error(f"镜号{shot_no} 为 {anchor_type} 锚定场景首镜，不得使用斯坦尼康建立大空间。")
    if visible_count >= 2 and anchor_type != "single_continuation":
        first_position_line = next(
            (line for line in camera.split("\n") if line.strip().startswith("【场景首镜站位】")),
            "",
        )
        blocking_text = first_position_line + "\n" + clean_text(scene_log.get("spatial_axis"))
        if not any(keyword in blocking_text for keyword in ["左", "右", "轴线", "对视", "朝向", "面向"]):
            result.error(f"镜号{shot_no} 多人场景首镜站位必须写清左右关系、朝向或主要对视轴线。")


def validate_adjacent_shot_design(
    shots: list[Any],
    fact_types: dict[str, str],
    result: ValidationResult,
    *,
    guard_reverse_axis_motion: bool,
) -> None:
    previous: dict[str, Any] | None = None
    for item in shots:
        if not isinstance(item, dict):
            continue
        if previous is not None and item.get("scene") == previous.get("scene"):
            validate_adjacent_pair(
                previous,
                item,
                fact_types,
                result,
                guard_reverse_axis_motion=guard_reverse_axis_motion,
            )
        previous = item


def validate_adjacent_pair(
    prev: dict[str, Any],
    curr: dict[str, Any],
    fact_types: dict[str, str],
    result: ValidationResult,
    *,
    guard_reverse_axis_motion: bool,
) -> None:
    prev_no = prev.get("shot_no")
    curr_no = curr.get("shot_no")
    prev_angle, prev_size, prev_movement = parse_triad(clean_text(prev.get("camera_main_image")))
    curr_angle, curr_size, curr_movement = parse_triad(clean_text(curr.get("camera_main_image")))
    if guard_reverse_axis_motion:
        prev_axis = motion_axis_direction(prev_movement)
        curr_axis = motion_axis_direction(curr_movement)
        if {prev_axis, curr_axis} == {"push_in", "pull_out"}:
            result.error(
                f"镜号{prev_no}-{curr_no} 同场相邻镜头出现无动机轴向反转："
                f"{prev_movement} 后接 {curr_movement}。应改为固定、横移、环绕、垂直摇移、"
                "切换主体或切换景别。"
            )
    prev_rank = shot_size_rank(prev_size)
    curr_rank = shot_size_rank(curr_size)
    same_characters = normalized_visible_characters(prev) == normalized_visible_characters(curr)
    if same_characters and normalized_visible_characters(prev):
        no_position_change = not has_continuity_position_change(prev) and not has_continuity_position_change(curr)
        close_size = prev_rank is not None and curr_rank is not None and abs(prev_rank - curr_rank) <= 1
        same_angle = angle_family(prev_angle) == angle_family(curr_angle)
        soft_movement = is_static_or_soft_movement(prev_movement) and is_static_or_soft_movement(curr_movement)
        if no_position_change and close_size and same_angle and soft_movement:
            result.error(
                f"镜号{prev_no}-{curr_no} 同场相邻镜头主体、视角、景别和运动过近，"
                "且无站位/朝向迁移；必须合并或提供明确画面信息增量。"
            )

    prev_dialogue = quoted_dialogue_text(prev)
    curr_dialogue = quoted_dialogue_text(curr)
    if prev_dialogue and curr_dialogue and prev_dialogue.endswith(INCOMPLETE_DIALOGUE_ENDINGS):
        prev_speaker = dialogue_speaker(prev)
        curr_speaker = dialogue_speaker(curr)
        if not prev_speaker or not curr_speaker or prev_speaker == curr_speaker:
            result.error(
                f"镜号{prev_no}-{curr_no} 将同一说话人的未完成台词切成相邻两镜；"
                "应合并为一镜，用表演或机位运动承接。"
            )

    prev_types = fact_types_for_shot(prev, fact_types)
    if "dialogue" in prev_types and is_expression_only_response(curr, fact_types):
        if same_characters or not normalized_visible_characters(curr):
            result.error(
                f"镜号{prev_no}-{curr_no} 是对白后同一人物短表情反应，且没有新的空间、道具、声音或位置事实；"
                "应合并进同一镜。"
            )


def validate_hybrid_fields(shot: dict[str, Any], result: ValidationResult, *, required: bool) -> None:
    shot_no = shot.get("shot_no")
    duration = float(shot.get("duration_seconds") or 0)
    if not required:
        return
    shot_type = clean_text(shot.get("shot_type"))
    if not shot_type:
        result.error(f"镜号{shot_no} 2.3.4+ 必须填写 shot_type。")
    elif shot_type not in SHOT_TYPES:
        result.error(f"镜号{shot_no} shot_type 不合法：{shot_type}")
    split_reasons = normalized_list(shot.get("split_reason"))
    if not split_reasons:
        result.error(f"镜号{shot_no} 2.3.4+ 必须填写 split_reason。")
    for reason in split_reasons:
        if reason not in SPLIT_REASONS:
            result.error(f"镜号{shot_no} split_reason 不合法：{reason}")
    insert_priority = clean_text(shot.get("insert_priority"))
    if not insert_priority:
        result.error(f"镜号{shot_no} 2.3.4+ 必须填写 insert_priority。")
    elif insert_priority not in INSERT_PRIORITIES:
        result.error(f"镜号{shot_no} insert_priority 不合法：{insert_priority}")
    supports = normalized_list(shot.get("long_take_support"))
    for support in supports:
        if support not in LONG_TAKE_SUPPORTS:
            result.error(f"镜号{shot_no} long_take_support 不合法：{support}")
    if duration > 10 and len(supports) < 2:
        result.error(f"镜号{shot_no} 超过10秒，2.3.4 要求至少填写两项 long_take_support。")
    if shot_type in INSERT_LIKE_TYPES and duration > 5:
        result.warn(f"镜号{shot_no} 是 {shot_type}，插镜建议 2-5 秒；当前 {duration:g} 秒，请确认是否承担镜内推进。")
    causal_reasons = {"new_vfx_state", "causal_reveal", "prop_state_change"}
    if causal_reasons & set(split_reasons) and shot_type not in {"insert", "vfx_anchor", "transition", "action"}:
        result.warn(f"镜号{shot_no} 承担真相/道具/VFX 因果落点，但 shot_type 为 {shot_type}，建议复核插镜意识。")


def validate_fact_cut_contract(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_hybrid_fields(data):
        return
    seen_groups: set[str] = set()
    for beat in as_list(data.get("beats")):
        if not isinstance(beat, dict):
            continue
        beat_id = str(beat.get("beat_id", ""))
        for fact in as_list(beat.get("facts")):
            if not isinstance(fact, dict):
                continue
            fact_id = str(fact.get("fact_id", ""))
            priority = clean_text(fact.get("cut_priority"))
            if priority not in CUT_PRIORITIES:
                result.error(f"{fact_id} cut_priority 必须为 normal / recommended / must_isolate。")
            if "cut_reasons" not in fact:
                result.error(f"{fact_id} 缺少 cut_reasons。")
                reasons: list[str] = []
            else:
                reasons = normalized_list(fact.get("cut_reasons"))
            for reason in reasons:
                if reason not in SPLIT_REASONS:
                    result.error(f"{fact_id} cut_reasons 不合法：{reason}")
            if priority in {"recommended", "must_isolate"} and not reasons:
                result.error(f"{fact_id} cut_priority 为 {priority} 时 cut_reasons 不能为空。")
            group = clean_text(fact.get("cut_group"))
            if not group:
                result.error(f"{fact_id} 缺少 cut_group。")
            else:
                seen_groups.add(group)
            reason_set = set(reasons)
            if reason_set & MUST_ISOLATE_REASONS and priority != "must_isolate":
                result.error(f"{fact_id} 包含关键切点 {sorted(reason_set & MUST_ISOLATE_REASONS)}，cut_priority 必须为 must_isolate。")
            if reason_set & CRITICAL_SPLIT_REASONS and priority == "normal":
                result.error(f"{fact_id} 包含关键切点 {sorted(reason_set & CRITICAL_SPLIT_REASONS)}，cut_priority 不得为 normal。")
            if group and not group.startswith(beat_id):
                result.warn(f"{fact_id} cut_group 建议以所属 Beat 开头，当前为 {group}。")
    if not seen_groups:
        result.error("2.3.4 数据必须登记事实级 cut_group。")


def validate_overcompression(data: dict[str, Any], result: ValidationResult) -> None:
    if not requires_hybrid_fields(data):
        return
    shots = [shot for shot in as_list(data.get("shots")) if isinstance(shot, dict)]
    facts_by_id = fact_lookup(data)
    total_duration = 0.0
    short_3 = 0
    has_multi_beat_critical = False
    for shot in shots:
        shot_no = shot.get("shot_no")
        duration = float(shot.get("duration_seconds") or 0)
        total_duration += duration
        short_3 += int(duration <= 3)
        notes = clean_text(shot.get("notes"))
        shot_type = clean_text(shot.get("shot_type"))
        beat_count = len(as_list(shot.get("beat_ids")))
        facts = shot_fact_objects(shot, facts_by_id)
        critical_facts = [fact for fact in facts if is_fact_critical(fact)]
        must_groups = shot_cut_groups(facts, must_only=True)
        cut_points = find_required_cut_point_categories(
            clean_text(shot.get("source_paragraph")) + "\n" + clean_text(shot.get("camera_main_image"))
        )
        if beat_count >= 3 and critical_facts:
            has_multi_beat_critical = True
            result.error(f"镜号{shot_no} 覆盖 {beat_count} 个 Beat 且包含关键切点，属于过度压缩。")
        if beat_count >= 2 and critical_facts:
            has_multi_beat_critical = True
        if len(must_groups) > 1:
            result.error(
                f"镜号{shot_no} 同时覆盖多个 must_isolate cut_group：{', '.join(sorted(must_groups))}；"
                "除同一不可拆瞬间外必须拆镜。"
            )
        if shot_type == "master" and critical_facts:
            result.error(f"镜号{shot_no} master 镜不得承载关键切点，应拆为 transition / vfx_anchor / insert / action。")
        if shot_type == "dialogue" and critical_facts:
            if len(must_groups) > 1 or duration >= 8 or len(cut_points) >= 2:
                result.error(f"镜号{shot_no} dialogue 镜承载多个/过长关键切点，必须拆分真相、声源或反应落点。")
        if shot_type in INSERT_LIKE_TYPES and duration > 5:
            supports = normalized_list(shot.get("long_take_support"))
            if "[保留理由]" not in notes or len(supports) < 2:
                result.error(f"镜号{shot_no} {shot_type} 超过5秒，必须写 [保留理由] 且至少两项 long_take_support。")
        if 9 <= duration <= 11:
            if "[长镜头]" not in notes or "[保留理由]" not in notes:
                result.error(f"镜号{shot_no} 9-11秒高风险长镜头必须写 [长镜头] 和 [保留理由]。")
        if duration >= 12:
            missing = [label for label in ["[长镜头]", "[保留理由]", "[不可拆说明]"] if label not in notes]
            if missing:
                result.error(f"镜号{shot_no} 12秒及以上镜头缺少 {'、'.join(missing)}。")
        if duration >= 9 and len(cut_points) >= 2:
            result.error(f"镜号{shot_no} {duration:g}秒镜头包含多个必拆切点：{', '.join(cut_points)}。")
        source = clean_text(shot.get("source_paragraph"))
        camera = clean_text(shot.get("camera_main_image"))
        visible = set(normalized_list(shot.get("visible_characters")))
        for speaker in extract_source_speakers(source):
            if speaker not in visible and re.search(rf"{re.escape(speaker)}[^。；;\n]{{0,8}}画外声", camera):
                if not source_marks_offscreen_voice(source, speaker):
                    result.error(f"镜号{shot_no} 将原文现场对白角色 {speaker} 改为画外声，疑似绕过多人关系镜头。")
    cut_per_minute = round(((len(shots) - 1) / total_duration) * 60, 2) if total_duration > 0 and len(shots) > 1 else 0
    short_ratio = round(short_3 / len(shots), 4) if shots else 0
    if cut_per_minute and cut_per_minute < 10.5:
        message = f"cut/min {cut_per_minute} 低于 10.5，疑似整体压缩。"
        if has_multi_beat_critical:
            result.error(message + " 且存在多 Beat 关键镜头，必须重拆。")
        else:
            result.warn(message)
    if shots and short_ratio < 0.10:
        message = f"短镜比例 {short_ratio:.2%} 低于 10%，疑似缺少必要插镜。"
        if has_multi_beat_critical:
            result.error(message + " 且存在多 Beat 关键镜头，必须重拆。")
        else:
            result.warn(message)


def validate_rows(
    data: dict[str, Any],
    result: ValidationResult,
    fact_to_beat: dict[str, str],
    fact_types: dict[str, str],
    all_fact_ids: set[str],
    continuity_logs: dict[str, dict[str, Any]],
) -> None:
    shots = as_list(data.get("shots"))
    if not shots:
        result.error("shots 不能为空。")
        return
    expected_no = 1
    covered: set[str] = set()
    scenes = set(continuity_logs)
    seen_scenes: set[str] = set()
    has_update = False
    hybrid_required = requires_hybrid_fields(data)
    for shot in shots:
        if not isinstance(shot, dict):
            result.error("shots 中每一项必须是对象。")
            continue
        shot_no = shot.get("shot_no")
        if shot_no != expected_no:
            result.error(f"镜号必须从1连续递增：期望 {expected_no}，实际 {shot_no}")
        expected_no += 1
        for key in ["scene", "source_paragraph", "duration_seconds", "duration_breakdown", "camera_main_image", "notes", "prompt", "beat_ids", "covered_fact_ids", "continuity_updates"]:
            if key not in shot:
                result.error(f"镜号{shot_no} 缺少字段：{key}")
        if "keyframe" in shot:
            result.error(f"镜号{shot_no} 不得包含 keyframe 字段；关键帧列已从主表删除。")
        if shot.get("old_shot_no") is not None:
            result.error(f"镜号{shot_no} 不得保留 old_shot_no；禁止后处理合并痕迹进入交付数据。")
        scene = shot.get("scene")
        if scenes and scene not in scenes:
            result.warn(f"镜号{shot_no} 场景未在 continuity_logs 中登记：{scene}")
        is_first_scene_shot = str(scene) not in seen_scenes
        if is_first_scene_shot:
            seen_scenes.add(str(scene))
            validate_first_scene_shot(shot, as_dict(continuity_logs.get(str(scene))), result)
        beat_ids = [str(item) for item in as_list(shot.get("beat_ids"))]
        fact_ids = [str(item) for item in as_list(shot.get("covered_fact_ids"))]
        if not beat_ids:
            result.error(f"镜号{shot_no} beat_ids 不能为空。")
        if not fact_ids:
            result.error(f"镜号{shot_no} covered_fact_ids 不能为空。")
        for fact_id in fact_ids:
            covered.add(fact_id)
            owner = fact_to_beat.get(fact_id)
            if owner is None:
                result.error(f"镜号{shot_no} 覆盖了不存在的事实 ID：{fact_id}")
            elif owner not in beat_ids:
                result.error(f"镜号{shot_no} 覆盖事实 {fact_id}，但 beat_ids 未包含 {owner}。")
        validate_duration(shot, result)
        validate_camera(shot, result)
        validate_prompt(shot, result)
        validate_hybrid_fields(shot, result, required=hybrid_required)
        updates = as_list(shot.get("continuity_updates"))
        if updates:
            has_update = True
        has_station_move = "【站位位移】" in clean_text(shot.get("camera_main_image"))
        if has_station_move and not updates:
            result.error(f"镜号{shot_no} 写了【站位位移】，但 continuity_updates 为空。")
        for update in updates:
            validate_update(shot, update, result, has_station_move)
    missing = sorted(all_fact_ids - covered)
    if missing:
        result.error("存在未覆盖事实 ID：" + ", ".join(missing[:20]))
    if any("【站位位移】" in clean_text(shot.get("camera_main_image")) for shot in shots if isinstance(shot, dict)) and not has_update:
        result.error("分镜出现站位位移，但没有任何 continuity_updates。")
    validate_adjacent_shot_design(
        shots,
        fact_types,
        result,
        guard_reverse_axis_motion=requires_adjacent_motion_guard(data),
    )


def validate_duration(shot: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    try:
        duration = float(shot.get("duration_seconds"))
    except Exception:
        result.error(f"镜号{shot_no} 镜头时长必须是数字。")
        return
    if duration <= 0:
        result.error(f"镜号{shot_no} 镜头时长必须大于0。")
    breakdown = as_dict(shot.get("duration_breakdown"))
    for field_name in DURATION_FIELDS:
        if field_name not in breakdown:
            result.error(f"镜号{shot_no} duration_breakdown 缺少 {field_name}。")
            return
    try:
        expected = math.ceil(
            max(float(breakdown["sync_action_seconds"]), float(breakdown["sync_dialogue_seconds"]))
            + float(breakdown["non_sync_action_seconds"])
            + float(breakdown["emotional_pause_seconds"])
        )
    except Exception:
        result.error(f"镜号{shot_no} duration_breakdown 必须全为数字。")
        return
    if int(duration) != expected:
        result.error(f"镜号{shot_no} 时长不符合公式：表格 {duration:g} 秒，公式 {expected} 秒。")
    quoted = "".join(re.findall(r"[“\"]([^”\"]+)[”\"]", clean_text(shot.get("camera_main_image"))))
    spoken_chars = len(re.sub(r"[，。！？……、\s]", "", quoted))
    if spoken_chars:
        min_dialogue_seconds = math.ceil(spoken_chars / 4)
        if float(breakdown["sync_dialogue_seconds"]) < min_dialogue_seconds:
            result.error(f"镜号{shot_no} 对白时长过短：{spoken_chars}字至少约 {min_dialogue_seconds} 秒。")
    notes = clean_text(shot.get("notes"))
    if duration > 6 and "[时长估算]" not in notes:
        result.error(f"镜号{shot_no} 超过6秒，备注必须写 [时长估算]。")
    if duration > 10:
        long_take = as_dict(shot.get("long_take"))
        if long_take.get("classification") in (None, "not_applicable") and "[长镜头]" not in notes:
            result.error(f"镜号{shot_no} 超过10秒，必须标注长镜头分类或人工复核理由。")


def validate_camera(shot: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    camera = clean_text(shot.get("camera_main_image"))
    lines = [line.strip() for line in camera.split("\n") if line.strip()]
    if len(lines) < 2:
        result.error(f"镜号{shot_no} 运镜列至少需要三联首行和【机位逻辑】。")
        return
    if not re.fullmatch(r"\[[^,\]]+,\s*[^,\]]+,\s*[^\]]+\]", lines[0]):
        result.error(f"镜号{shot_no} 运镜首行必须是三要素：[视角/高度, 景别/特殊视角, 运镜方式]。")
    angle, shot_size, movement = parse_triad(camera)
    validate_camera_movement_compatibility(shot_no, angle, shot_size, movement, result, clean_text(shot.get("notes")))
    if not lines[1].startswith("【机位逻辑】"):
        result.error(f"镜号{shot_no} 运镜第二行必须以【机位逻辑】开头。")
    for word in FORBIDDEN_VISUAL_WORDS:
        if word in camera:
            result.error(f"镜号{shot_no} 运镜列包含禁用表达：{word}")
    for marker in UNSUPPORTED_MARKERS:
        if marker in camera or marker in clean_text(shot.get("notes")):
            result.error(f"镜号{shot_no} 包含后处理/无依据标记：{marker}")
    if "【镜内变化】" in camera:
        result.error(f"镜号{shot_no} 不得在最终主表出现【镜内变化】。")
    if "运镜修正" in clean_text(shot.get("notes")):
        result.error(f"镜号{shot_no} 备注不得保留运镜修正痕迹。")


def validate_camera_movement_compatibility(
    shot_no: Any,
    angle: str,
    shot_size: str,
    movement: str,
    result: ValidationResult,
    notes: str = "",
) -> None:
    if "斯坦尼康" in clean_text(movement) and has_any_keyword(shot_size, ("大远景", "大全景")):
        result.error(f"镜号{shot_no} 大远景/大全景不得使用斯坦尼康；应改用摇臂、伸缩摇臂或航拍。")
    if "航拍" in clean_text(movement):
        if not has_any_keyword(shot_size, AERIAL_ALLOWED_SHOT_SIZE_KEYWORDS):
            result.error(f"镜号{shot_no} 航拍必须服务于大范围空间，景别应为大远景/大全景/全景。")
        if not has_any_keyword(angle, AERIAL_ALLOWED_ANGLE_KEYWORDS):
            result.error(f"镜号{shot_no} 航拍必须匹配俯拍、高角度、正俯拍或微俯视。")
    size_text = f"{angle} {shot_size}"
    if shot_size_has_span(size_text):
        has_span_movement = has_any_keyword(movement, SPAN_MOVEMENT_KEYWORDS)
        has_span_note = "[景别跨度]" in clean_text(notes)
        if not has_span_movement or not has_span_note:
            result.error(
                f"镜号{shot_no} 景别同时包含远景类和近景/特写类，"
                "必须使用光学变焦、快速推拉、急推/急拉、推轨推进/拉出或伸缩摇臂等跨度运镜，"
                "并在备注写 [景别跨度] 理由。"
            )


def validate_prompt(shot: dict[str, Any], result: ValidationResult) -> None:
    shot_no = shot.get("shot_no")
    prompt = clean_text(shot.get("prompt"))
    expected = derive_prompt(shot)
    if prompt != expected:
        result.error(f"镜号{shot_no} Prompt 必须由锁定主表派生，且与标准五字段模板一致。")
    lines = [line.strip() for line in prompt.split("\n") if line.strip()]
    if len(lines) != 5:
        result.error(f"镜号{shot_no} Prompt 必须且只能包含五行。")
    for field_name in PROMPT_FIELDS:
        if not any(line.startswith(field_name) for line in lines):
            result.error(f"镜号{shot_no} Prompt 缺少字段：{field_name}")
    for item in PROMPT_FORBIDDEN:
        if item in prompt:
            result.error(f"镜号{shot_no} Prompt 包含禁止进入派生列的内部/关键帧内容：{item}")


def validate_update(shot: dict[str, Any], update: Any, result: ValidationResult, has_station_move: bool) -> None:
    shot_no = shot.get("shot_no")
    if not isinstance(update, dict):
        result.error(f"镜号{shot_no} continuity_updates 每项必须是对象。")
        return
    required = ["entity_type", "entity", "field", "from", "to", "evidence_fact_ids"]
    for key in required:
        if key not in update:
            result.error(f"镜号{shot_no} continuity_update 缺少 {key}。")
    if update.get("from") == update.get("to"):
        result.error(f"镜号{shot_no} continuity_update 的 from/to 不得相同。")
    evidence = [str(item) for item in as_list(update.get("evidence_fact_ids"))]
    covered = {str(item) for item in as_list(shot.get("covered_fact_ids"))}
    if not evidence:
        result.error(f"镜号{shot_no} continuity_update 必须有 evidence_fact_ids。")
    for fact_id in evidence:
        if fact_id not in covered:
            result.error(f"镜号{shot_no} continuity_update 证据 {fact_id} 未被当前镜头覆盖。")
    if update.get("field") in {"position", "facing"} and not has_station_move:
        result.error(f"镜号{shot_no} 登记了位置/朝向迁移，但主画面缺少【站位位移】。")


def validate_data(data: dict[str, Any], *, strict_status: bool) -> ValidationResult:
    result = ValidationResult()
    validate_metadata(data, result, strict_status=strict_status)
    continuity_logs = validate_continuity_logs(data, result)
    fact_to_beat, fact_types, all_fact_ids = collect_facts(data, result)
    validate_fact_cut_contract(data, result)
    validate_rows(data, result, fact_to_beat, fact_types, all_fact_ids, continuity_logs)
    validate_overcompression(data, result)
    return result


def markdown_text(data: dict[str, Any]) -> str:
    lines = ["# 分镜表", ""]
    lines.append("| " + " | ".join(HEADERS) + " |")
    lines.append("| " + " | ".join(["---"] * len(HEADERS)) + " |")
    for row in expected_rows(data):
        lines.append("| " + " | ".join(escape_cell(item) for item in row) + " |")
    report = as_dict(data.get("validation_report"))
    if report:
        lines.extend(
            [
                "",
                "## 校验摘要",
                "",
                f"- status: {report.get('status', '')}",
                f"- shot_count: {report.get('shot_count', '')}",
                f"- column_contract: {report.get('column_contract', '')}",
            ]
        )
    return "\n".join(lines) + "\n"


def build_markdown(data: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(markdown_text(data), encoding="utf-8")


def build_excel(data: dict[str, Any], path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl 不可用，无法生成 Excel。")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    for row in expected_rows(data):
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, width in enumerate([8, 24, 44, 14, 72, 36, 56], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 36 if row_index == 1 else 108
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def parse_markdown_table(path: Path) -> list[list[str]]:
    text = path.read_text(encoding="utf-8-sig")
    rows: list[list[str]] = []
    for line in text.splitlines():
        if not line.startswith("|"):
            continue
        cells = [cell.strip().replace("<br>", "\n").replace("｜", "|") for cell in line.strip().strip("|").split("|")]
        if cells and all(re.fullmatch(r":?-{3,}:?", cell) for cell in cells):
            continue
        rows.append(cells)
    return rows


def compare_markdown(data: dict[str, Any], path: Path, result: ValidationResult) -> None:
    if not path.exists():
        result.error(f"Markdown 文件不存在：{path}")
        return
    rows = parse_markdown_table(path)
    if not rows:
        result.error("Markdown 未找到分镜表。")
        return
    if rows[0] != HEADERS:
        result.error("Markdown 表头必须是稳定 7 列，且不得包含关键帧列。")
    expected = [[str(item) for item in row] for row in expected_rows(data)]
    actual = [[str(item) for item in row] for row in rows[1:]]
    if actual != expected:
        result.error("Markdown 表格内容与 shot_data 派生结果不一致。")


def compare_excel(data: dict[str, Any], path: Path, result: ValidationResult) -> None:
    if load_workbook is None:
        result.error("openpyxl 不可用，无法校验 Excel。")
        return
    if not path.exists():
        result.error(f"Excel 文件不存在：{path}")
        return
    workbook = load_workbook(path, read_only=True)
    try:
        if workbook.sheetnames != [SHEET_NAME]:
            result.error("Excel 必须只有一个 Sheet：分镜表。")
            return
        sheet = workbook[SHEET_NAME]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        result.error("Excel 分镜表为空。")
        return
    header = [str(item or "") for item in rows[0]]
    if header != HEADERS:
        result.error("Excel 表头必须是稳定 7 列，且不得包含关键帧列。")
    expected = [[str(item) for item in row] for row in expected_rows(data)]
    actual = [[str(item or "") for item in row] for row in rows[1:]]
    if actual != expected:
        result.error("Excel 内容与 shot_data 派生结果不一致。")


def print_result(result: ValidationResult) -> None:
    if result.errors:
        print("ERRORS:", file=sys.stderr)
        for item in result.errors:
            print(f"- {item}", file=sys.stderr)
    if result.warnings:
        print("WARNINGS:")
        for item in result.warnings:
            print(f"- {item}")
    print(f"{result.status}: storyboard validation {'passed' if not result.errors else 'failed'}.")


def command_build(args: argparse.Namespace) -> int:
    data_path = Path(args.data)
    data = load_json(data_path)
    derive_prompts(data)
    result = validate_data(data, strict_status=False)
    update_validation_report(data, result)
    if result.errors:
        write_json(data_path, data)
        print_result(result)
        return 1
    write_json(data_path, data)
    build_markdown(data, Path(args.markdown))
    build_excel(data, Path(args.excel))
    print_result(result)
    return 0


def command_validate(args: argparse.Namespace) -> int:
    data = load_json(Path(args.data))
    result = validate_data(data, strict_status=True)
    compare_markdown(data, Path(args.markdown), result)
    compare_excel(data, Path(args.excel), result)
    print_result(result)
    return 1 if result.errors else 0


def make_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    build = subparsers.add_parser("build", help="derive Prompt, update JSON, and write Markdown/Excel")
    build.add_argument("--data", required=True)
    build.add_argument("--markdown", required=True)
    build.add_argument("--excel", required=True)
    build.set_defaults(func=command_build)
    validate = subparsers.add_parser("validate", help="validate JSON/Markdown/Excel consistency")
    validate.add_argument("--data", required=True)
    validate.add_argument("--markdown", required=True)
    validate.add_argument("--excel", required=True)
    validate.set_defaults(func=command_validate)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = make_parser()
    args = parser.parse_args(argv)
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
