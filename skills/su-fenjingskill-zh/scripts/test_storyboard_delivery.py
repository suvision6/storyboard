#!/usr/bin/env python3
"""Tests for the stable 7-column storyboard delivery contract."""

from __future__ import annotations

import copy
import json
import math
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

import storyboard_delivery as delivery


RULE_REVISION_234 = "2.3.4-overcompression-guard-2026-06-30"
RULE_REVISION_235 = "2.3.5-adjacent-motion-guard-2026-06-30"
RULE_REVISION_236 = "2.3.6-tri-source-audit-guard-2026-06-30"
RULE_REVISION_240 = "2.4.0-human-gate-stability-2026-07-06"
RULE_REVISION_241 = "2.4.1-source-lock-2026-07-07"
RULE_REVISION_242 = "2.4.2-source-lock-entry-guard-2026-07-07"
RULE_REVISION_243 = "2.4.3-contract-integrity-p2-2026-07-12"


def valid_data() -> dict:
    return {
        "metadata": {
            "skill_name": "su-fenjingskill-zh",
            "version": "2.3.2",
            "title": "稳定化样片",
            "reference_status": {
                "continuity-shot-data": "loaded",
                "camera-language": "loaded",
                "seedance-prompt-rules": "loaded",
            },
        },
        "continuity_logs": [
            {
                "scene": "1 室内 日 内",
                "first_shot_anchor_type": "single_continuation",
                "spatial_axis": "A在门口，B在桌边。",
                "fixed_objects": [],
                "characters": [],
                "props": [],
                "sound_sources": [],
                "reality_layer": "现实",
            }
        ],
        "beats": [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "A站在门口。",
                "facts": [{"fact_id": "B001-F01", "type": "position", "text": "A站在门口。"}],
            },
            {
                "beat_id": "B002",
                "scene": "1 室内 日 内",
                "source_text": "A走到桌边。",
                "facts": [{"fact_id": "B002-F01", "type": "position", "text": "A走到桌边。"}],
            },
            {
                "beat_id": "B003",
                "scene": "1 室内 日 内",
                "source_text": "A说：到了。",
                "facts": [{"fact_id": "B003-F01", "type": "dialogue", "text": "到了。"}],
            },
        ],
        "shots": [
            {
                "shot_no": 1,
                "scene": "1 室内 日 内",
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01"],
                "source_paragraph": "A站在门口。",
                "duration_seconds": 2,
                "duration_breakdown": {
                    "sync_action_seconds": 1,
                    "sync_dialogue_seconds": 0,
                    "non_sync_action_seconds": 0,
                    "emotional_pause_seconds": 1,
                },
                "camera_main_image": "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机在桌边看向门口。\n【场景首镜站位】（A站在门口，面向桌边。）\nA站在门口，手扶门框。",
                "notes": "建立初始站位。",
                "prompt": "",
                "visible_characters": ["A"],
                "offscreen_characters": [],
                "visible_props": [],
                "continuity_updates": [],
            },
            {
                "shot_no": 2,
                "scene": "1 室内 日 内",
                "beat_ids": ["B002", "B003"],
                "covered_fact_ids": ["B002-F01", "B003-F01"],
                "source_paragraph": "A走到桌边。A说：到了。",
                "duration_seconds": 3,
                "duration_breakdown": {
                    "sync_action_seconds": 2,
                    "sync_dialogue_seconds": 1,
                    "non_sync_action_seconds": 0,
                    "emotional_pause_seconds": 1,
                },
                "camera_main_image": "[侧面平视, 中景, 横移跟拍]\n【机位逻辑】摄影机沿桌边横移，跟住A的脚步。\n【站位位移】A从门口走到桌边，面向B的位置。\nA在桌边停下，说：“到了。”",
                "notes": "A完成位置迁移。",
                "prompt": "",
                "visible_characters": ["A"],
                "offscreen_characters": [],
                "visible_props": [],
                "continuity_updates": [
                    {
                        "entity_type": "character",
                        "entity": "A",
                        "field": "position",
                        "from": "门口",
                        "to": "桌边",
                        "evidence_fact_ids": ["B002-F01"],
                    }
                ],
            },
        ],
        "validation_report": {"status": "PASS", "warnings": [], "errors": []},
    }


def valid_data_234() -> dict:
    data = valid_data()
    data["metadata"]["version"] = "2.3.4"
    data["metadata"]["rule_revision"] = RULE_REVISION_234
    for beat in data["beats"]:
        for fact in beat["facts"]:
            fact["cut_priority"] = "normal"
            fact["cut_reasons"] = []
            fact["cut_group"] = fact["fact_id"]
    for shot in data["shots"]:
        shot["shot_type"] = "master" if shot["shot_no"] == 1 else "action"
        shot["split_reason"] = ["spatial_anchor"] if shot["shot_no"] == 1 else ["performance_continuity", "continuity_migration"]
        shot["insert_priority"] = "none"
        shot["long_take_support"] = []
    return data


def valid_data_235() -> dict:
    data = valid_data_234()
    data["metadata"]["version"] = "2.3.5"
    data["metadata"]["rule_revision"] = RULE_REVISION_235
    return data


def valid_data_236() -> dict:
    data = valid_data_235()
    data["metadata"]["version"] = "2.3.6"
    data["metadata"]["rule_revision"] = RULE_REVISION_236
    data["continuity_logs"][0]["characters"] = [{"name": "A", "position": "门口", "facing": "桌边"}]
    for beat in data["beats"]:
        for fact in beat["facts"]:
            fact_type = fact["type"]
            category = {
                "position": "space",
                "prop": "prop",
                "sound": "sound",
                "reality": "reality",
                "dialogue": "dialogue",
                "emotion": "emotion",
                "character": "character",
            }.get(fact_type, "action")
            fact["cut_category"] = category
            fact["cut_moment_id"] = f"{beat['beat_id']}-{category}-moment"
    return data


def valid_data_240() -> dict:
    data = valid_data_236()
    data["metadata"]["version"] = "2.4.0"
    data["metadata"]["rule_revision"] = RULE_REVISION_240
    data["metadata"]["reference_status"]["hybrid-shot-audit"] = "loaded"
    data["metadata"]["reference_proof"] = {
        "continuity-shot-data": "# Continuity And Shot Data Contract",
        "hybrid-shot-audit": "# 混合拆镜审计规则",
        "camera-language": "# Camera Language Reference",
        "seedance-prompt-rules": "# Seedance Prompt Rules",
    }
    data["metadata"]["script_status"] = {
        "storyboard_delivery.py": "available",
        "validate_storyboard.js": "available",
    }
    data["metadata"]["revision_log"] = []
    data["batch_plan"] = None
    data["human_reviews"] = [
        {"gate": "GATE_A", "round": 1, "status": "approved", "reviewer": "user", "notes": "GATE_A APPROVED"},
        {"gate": "GATE_B", "round": 1, "status": "approved", "reviewer": "user", "notes": "GATE_B APPROVED"},
    ]
    data["warn_resolutions"] = []
    data["continuity_logs"][0]["scene_id"] = "S01"
    for beat in data["beats"]:
        beat["scene_id"] = "S01"
    for shot in data["shots"]:
        shot["scene_id"] = "S01"
    return data


def valid_data_242() -> dict:
    data = valid_data_240()
    data["metadata"]["version"] = "2.4.2"
    data["metadata"]["rule_revision"] = RULE_REVISION_242
    locked_text = "A站在门口。\nA走到桌边。\nA说：到了。"
    data["script_lock"] = {
        "status": "locked",
        "approved_script_path": "outputs/2026-07-07/docs/sample.approved_script.txt",
        "locked_text": locked_text,
        "locked_text_hash": delivery.script_text_hash(locked_text),
        "approved_corrections": [],
    }
    spans = {
        "B001": {"start": 0, "end": 6},
        "B002": {"start": 7, "end": 13},
        "B003": {"start": 14, "end": 20},
    }
    for beat in data["beats"]:
        beat["source_span"] = spans[beat["beat_id"]]
    data["shots"][0]["source_span"] = spans["B001"]
    data["shots"][1]["source_spans"] = [spans["B002"], spans["B003"]]
    return data


def valid_data_243(*, approved_script_path: str = "approved/sample.approved_script.txt") -> dict:
    data = valid_data_242()
    data["metadata"]["version"] = "2.4.3"
    data["metadata"]["rule_revision"] = RULE_REVISION_243
    data["script_lock"]["approved_script_path"] = approved_script_path
    data["script_lock"]["locked_text_hash"] = delivery.normalized_script_text_hash(
        data["script_lock"]["locked_text"]
    )
    data["continuity_logs"][0]["sound_sources"] = {}
    data["continuity_logs"][0]["inherits_from"] = ""
    data["continuity_logs"][0]["inherited_states"] = []
    data["continuity_logs"][0]["diverged_states"] = []
    for order, beat in enumerate(data["beats"], start=1):
        beat["beat_order"] = str(order)
    data["shots"][1]["source_paragraph"] = "A走到桌边。\nA说：到了。"
    data["human_reviews"] = []
    return data


def append_review(
    data: dict,
    gate: str,
    *,
    round_no: int = 1,
    status: str = "approved",
    batch_id: str | None = None,
    approved_items: list[str] | None = None,
    notes: str | None = None,
) -> dict:
    review = {
        "gate": gate,
        "round": round_no,
        "status": status,
        "reviewer": "user",
        "notes": notes or f"{gate} {status.upper()}",
        "reviewed_hash": delivery.gate_review_hash(data, gate, batch_id=batch_id),
        "approved_items": approved_items or [],
    }
    if batch_id is not None:
        review["batch_id"] = batch_id
    data.setdefault("human_reviews", []).append(review)
    return review


def prepare_243_fixture(
    root: Path,
    *,
    add_gate_c: bool = False,
    approved_file_text: str | None = None,
) -> tuple[dict, Path, Path, Path, Path]:
    data_path = root / "sample.shot_data.json"
    markdown_path = root / "sample.md"
    excel_path = root / "sample.xlsx"
    report_path = root / "sample.validation_report.json"
    approved_path = root / "approved" / "sample.approved_script.txt"
    approved_path.parent.mkdir(parents=True, exist_ok=True)

    data = valid_data_243()
    approved_path.write_text(
        approved_file_text if approved_file_text is not None else data["script_lock"]["locked_text"],
        encoding="utf-8",
    )
    append_review(data, "GATE_A")
    append_review(data, "GATE_B")
    delivery.derive_prompts(data)
    result = delivery.validate_data(
        data,
        strict_status=False,
        data_path=data_path,
        workspace_root=root,
    )
    if result.warnings:
        resolve_warnings(data, result)
        result = delivery.validate_data(
            data,
            strict_status=False,
            data_path=data_path,
            workspace_root=root,
        )
    if result.errors:
        raise AssertionError(f"invalid 2.4.3 fixture: {result.errors}")
    delivery.update_validation_report(data, result, final_signoff=False)

    if add_gate_c:
        append_review(data, "GATE_C")
        result = delivery.validate_data(
            data,
            strict_status=False,
            final_signoff=True,
            data_path=data_path,
            workspace_root=root,
        )
        if result.errors:
            raise AssertionError(f"invalid Gate C fixture: {result.errors}")
        delivery.update_validation_report(data, result, final_signoff=True)

    return data, data_path, markdown_path, excel_path, report_path


def resolve_warnings(data: dict, result: delivery.ValidationResult, *, resolved_by: str = "human") -> None:
    data["warn_resolutions"] = [
        {
            "warn_id": delivery.warning_id(message),
            "resolution": "accepted_without_change",
            "resolved_by": resolved_by,
            "note": "test resolution",
        }
        for message in result.warnings
    ]


class DeliveryContractTests(unittest.TestCase):
    def test_build_writes_7_column_markdown_and_excel(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data, data_path, md_path, xlsx_path, report_path = prepare_243_fixture(root)
            delivery.write_json(data_path, data)
            rc = delivery.main(
                [
                    "build",
                    "--data",
                    str(data_path),
                    "--markdown",
                    str(md_path),
                    "--excel",
                    str(xlsx_path),
                    "--report",
                    str(report_path),
                    "--workspace-root",
                    str(root),
                ]
            )
            self.assertEqual(0, rc)

            built = json.loads(data_path.read_text(encoding="utf-8"))
            self.assertEqual("PASS", built["validation_report"]["status"])
            self.assertNotIn("keyframe", built["shots"][0])
            self.assertEqual(delivery.derive_prompt(built["shots"][0]), built["shots"][0]["prompt"])

            markdown = md_path.read_text(encoding="utf-8")
            self.assertIn("| 镜号 | 场景 | 原剧本段落 | 镜头时长(秒) | 运镜+主画面描述(含台词) | 备注 | Prompt |", markdown)
            self.assertNotIn("关键帧", markdown)
            self.assertNotIn("【镜内变化】", markdown)

            workbook = load_workbook(xlsx_path, read_only=True)
            try:
                self.assertEqual([delivery.SHEET_NAME, delivery.SUMMARY_SHEET_NAME], workbook.sheetnames)
                sheet = workbook["分镜表"]
                self.assertEqual(3, sheet.max_row)
                self.assertEqual(7, sheet.max_column)
            finally:
                workbook.close()

    def test_validate_wrapper_passes_built_files(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data, data_path, md_path, xlsx_path, report_path = prepare_243_fixture(root)
            delivery.write_json(data_path, data)
            common = [
                "--data",
                str(data_path),
                "--markdown",
                str(md_path),
                "--excel",
                str(xlsx_path),
                "--report",
                str(report_path),
                "--workspace-root",
                str(root),
            ]
            self.assertEqual(0, delivery.main(["build", *common]))

            process = subprocess.run(
                [
                    "node",
                    str(SCRIPT_DIR / "validate_storyboard.js"),
                    "--python",
                    sys.executable,
                    *common,
                ],
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertEqual(0, process.returncode, process.stderr + process.stdout)

    def test_keyframe_field_is_rejected(self) -> None:
        data = valid_data()
        data["shots"][0]["keyframe"] = "场景：旧关键帧"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("keyframe" in item for item in result.errors))

    def test_pass_internal_status_is_rejected(self) -> None:
        data = valid_data()
        data["validation_report"]["status"] = "PASS_INTERNAL"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("PASS / WARN / FAIL" in item for item in result.errors))

    def test_station_move_requires_continuity_update(self) -> None:
        data = valid_data()
        data["shots"][1]["continuity_updates"] = []
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("continuity_updates 为空" in item for item in result.errors))

    def test_position_update_requires_station_move(self) -> None:
        data = valid_data()
        data["shots"][1]["camera_main_image"] = data["shots"][1]["camera_main_image"].replace("【站位位移】", "")
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("缺少【站位位移】" in item for item in result.errors))

    def test_prompt_internal_label_is_rejected(self) -> None:
        data = valid_data()
        delivery.derive_prompts(data)
        data["shots"][0]["prompt"] += "\n【镜内变化】污染"
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("Prompt" in item and "内部" in item for item in result.errors))

    def test_dialogue_duration_cannot_be_underestimated(self) -> None:
        data = valid_data()
        data["shots"][1]["camera_main_image"] = "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机拍A站在桌边。\nA说：“这句话很长很长不能压到一秒。”"
        data["shots"][1]["duration_seconds"] = 1
        data["shots"][1]["duration_breakdown"]["sync_action_seconds"] = 1
        data["shots"][1]["duration_breakdown"]["sync_dialogue_seconds"] = 1
        data["shots"][1]["duration_breakdown"]["emotional_pause_seconds"] = 0
        data["shots"][1]["continuity_updates"] = []
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("对白时长过短" in item for item in result.errors))

    def test_outdoor_multi_character_panorama_first_shot_passes(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "both"
        data["continuity_logs"][0]["spatial_axis"] = "A在画面左侧，B在画面右侧，两人沿山路面向远处。"
        data["shots"][0]["camera_main_image"] = (
            "[微俯视, 大全景, 固定镜头]\n"
            "【机位逻辑】摄影机在山路上方俯看两人和远处入口。\n"
            "【场景首镜站位】（A在画面左侧，B在画面右侧，两人相距三米，面向远处入口。）\n"
            "A和B站在山路上，远处入口被雾压住。"
        )
        data["shots"][0]["visible_characters"] = ["A", "B"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_indoor_multi_character_full_shot_first_shot_passes(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "multi_character"
        data["continuity_logs"][0]["spatial_axis"] = "A在画面左侧门口，B在画面右侧桌边，两人隔桌对视。"
        data["shots"][0]["camera_main_image"] = (
            "[平视, 全景, 固定镜头]\n"
            "【机位逻辑】摄影机在房间侧墙拍到门、桌和两人的左右关系。\n"
            "【场景首镜站位】（A在画面左侧门口，面向右侧桌边；B在画面右侧桌边，面向A，两人隔桌对视。）\n"
            "A扶着门框，B站在桌边。"
        )
        data["shots"][0]["visible_characters"] = ["A", "B"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_multi_character_first_shot_medium_close_fails(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "multi_character"
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中近景, 固定镜头]\n"
            "【机位逻辑】摄影机拍两人上半身。\n"
            "【场景首镜站位】（A在左，B在右，两人面向彼此。）\n"
            "A和B看着对方。"
        )
        data["shots"][0]["visible_characters"] = ["A", "B"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("景别必须为大远景/大全景/全景/中全景" in item for item in result.errors))

    def test_single_continuation_medium_first_shot_passes(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "single_continuation"
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 固定镜头]\n"
            "【机位逻辑】摄影机在岩壁前拍A的半身和身后空间边界。\n"
            "【场景首镜站位】（A靠在岩壁前，面向画面右侧微光。）\n"
            "A靠在岩壁上，抬头看向右侧微光。"
        )
        data["shots"][0]["visible_characters"] = ["A"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_extreme_wide_steadicam_fails(self) -> None:
        data = valid_data()
        data["shots"][0]["camera_main_image"] = (
            "[微俯视, 大全景, 斯坦尼康平稳跟随]\n"
            "【机位逻辑】摄影机拍到山路全貌和人物位置。\n"
            "【场景首镜站位】（A站在画面左侧，面向远处。）\n"
            "A站在山路上。"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("大远景/大全景不得使用斯坦尼康" in item for item in result.errors))

    def test_extreme_wide_crane_passes(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "both"
        data["continuity_logs"][0]["spatial_axis"] = "A在画面左侧，B在画面右侧，两人面向远处入口。"
        data["shots"][0]["camera_main_image"] = (
            "[微俯视, 大全景, 伸缩摇臂缓慢下降]\n"
            "【机位逻辑】摄影机从林冠上方下降，拍到山路、入口和两人左右关系。\n"
            "【场景首镜站位】（A在画面左侧，B在画面右侧，两人相距三米，面向远处入口。）\n"
            "A和B站在山路上。"
        )
        data["shots"][0]["visible_characters"] = ["A", "B"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_aerial_large_landscape_passes(self) -> None:
        data = valid_data()
        data["continuity_logs"][0]["first_shot_anchor_type"] = "space"
        data["shots"][0]["camera_main_image"] = (
            "[高角度俯拍, 大远景, 航拍缓慢推进]\n"
            "【机位逻辑】摄影机从高速公路上方推进，拍到道路纵深和远处雾区。\n"
            "【场景首镜站位】（道路由画面左下延伸至右上，A处在远处路肩，面向雾区。）\n"
            "浓雾压过高速公路。"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_steadicam_medium_follow_passes(self) -> None:
        data = valid_data()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 斯坦尼康跟随]\n"
            "【机位逻辑】摄影机在地面跟随A从门口走向桌边。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A从门口向桌边走去。"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_aerial_medium_close_fails(self) -> None:
        data = valid_data()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中近景, 航拍缓慢推进]\n"
            "【机位逻辑】摄影机贴近A的上半身推进。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A看向桌边。"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("航拍必须服务于大范围空间" in item for item in result.errors))

    def test_adjacent_similar_shots_without_position_change_fail(self) -> None:
        data = valid_data()
        data["beats"] = data["beats"][:2]
        data["beats"][1]["source_text"] = "A抬眼看向桌边。"
        data["beats"][1]["facts"] = [{"fact_id": "B002-F01", "type": "action", "text": "A抬眼看向桌边。"}]
        data["shots"][1].update(
            {
                "beat_ids": ["B002"],
                "covered_fact_ids": ["B002-F01"],
                "source_paragraph": "A抬眼看向桌边。",
                "duration_seconds": 2,
                "duration_breakdown": {
                    "sync_action_seconds": 1,
                    "sync_dialogue_seconds": 0,
                    "non_sync_action_seconds": 0,
                    "emotional_pause_seconds": 1,
                },
                "camera_main_image": "[平视, 中近景, 固定镜头]\n【机位逻辑】摄影机继续拍A的上半身。\nA抬眼看向桌边。",
                "notes": "同一主体反应。",
                "continuity_updates": [],
            }
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("主体、视角、景别和运动过近" in item for item in result.errors))

    def test_incomplete_dialogue_split_across_adjacent_shots_fails(self) -> None:
        data = valid_data()
        data["beats"] = data["beats"][:2]
        data["beats"][0]["source_text"] = "A：你听我说，"
        data["beats"][0]["facts"] = [{"fact_id": "B001-F01", "type": "dialogue", "text": "你听我说，"}]
        data["beats"][1]["source_text"] = "A：我不是故意的。"
        data["beats"][1]["facts"] = [{"fact_id": "B002-F01", "type": "dialogue", "text": "我不是故意的。"}]
        data["shots"][0]["source_paragraph"] = "A：你听我说，"
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 固定镜头]\n"
            "【机位逻辑】摄影机在桌边拍A。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A看向桌边，说：“你听我说，”"
        )
        data["shots"][1].update(
            {
                "beat_ids": ["B002"],
                "covered_fact_ids": ["B002-F01"],
                "source_paragraph": "A：我不是故意的。",
                "duration_seconds": 3,
                "duration_breakdown": {
                    "sync_action_seconds": 1,
                    "sync_dialogue_seconds": 2,
                    "non_sync_action_seconds": 0,
                    "emotional_pause_seconds": 1,
                },
                "camera_main_image": "[平视, 中近景, 固定镜头]\n【机位逻辑】摄影机切近A的脸。\nA继续说：“我不是故意的。”",
                "notes": "同一说话人续句。",
                "continuity_updates": [],
            }
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("未完成台词" in item for item in result.errors))

    def test_dialogue_followed_by_short_expression_reaction_fails(self) -> None:
        data = valid_data()
        data["beats"] = data["beats"][:2]
        data["beats"][0]["source_text"] = "A：够了。"
        data["beats"][0]["facts"] = [{"fact_id": "B001-F01", "type": "dialogue", "text": "够了。"}]
        data["beats"][1]["source_text"] = "A笑了一下。"
        data["beats"][1]["facts"] = [{"fact_id": "B002-F01", "type": "emotion", "text": "A笑了一下。"}]
        data["shots"][0]["source_paragraph"] = "A：够了。"
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 固定镜头]\n"
            "【机位逻辑】摄影机在桌边拍A。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A看向桌边，说：“够了。”"
        )
        data["shots"][1].update(
            {
                "beat_ids": ["B002"],
                "covered_fact_ids": ["B002-F01"],
                "source_paragraph": "A笑了一下。",
                "duration_seconds": 2,
                "duration_breakdown": {
                    "sync_action_seconds": 1,
                    "sync_dialogue_seconds": 0,
                    "non_sync_action_seconds": 0,
                    "emotional_pause_seconds": 1,
                },
                "camera_main_image": "[平视, 中近景, 固定镜头]\n【机位逻辑】摄影机切近A的嘴角。\nA嘴角轻轻抬起，笑了一下。",
                "notes": "短表情反应。",
                "continuity_updates": [],
            }
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("短表情反应" in item for item in result.errors))

    def test_wide_to_close_shot_size_without_span_movement_fails(self) -> None:
        data = valid_data()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 全景->特写, 固定镜头]\n"
            "【机位逻辑】摄影机同时描述房间全貌和A的眼睛。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A站在门口，眼神收紧。"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("[景别跨度]" in item for item in result.errors))

    def test_wide_to_close_shot_size_with_span_movement_passes(self) -> None:
        data = valid_data()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 全景->特写, 光学变焦]\n"
            "【机位逻辑】摄影机从房间全貌压到A的眼睛。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A站在门口，眼神收紧。"
        )
        data["shots"][0]["notes"] = "[景别跨度] 从房间关系压入A的关键反应。"
        data["shots"][1]["camera_main_image"] = (
            "[侧面平视, 中景, 横移跟拍]\n"
            "【机位逻辑】摄影机沿桌边横移，跟住A的脚步。\n"
            "【站位位移】A从门口走到桌边，面向B的位置。\n"
            "A在桌边停下，说：“到了。”"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_233_data_does_not_require_hybrid_fields(self) -> None:
        data = valid_data()
        data["metadata"]["version"] = "2.3.3"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(any("shot_type" in item or "split_reason" in item or "insert_priority" in item or "long_take_support" in item for item in result.errors), result.errors)

    def test_234_read_only_artifacts_keep_7_columns_and_hybrid_fields(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data_path = root / "sample.shot_data.json"
            md_path = root / "sample.md"
            xlsx_path = root / "sample.xlsx"
            built = valid_data_234()
            delivery.derive_prompts(built)
            result = delivery.validate_data(built, strict_status=False)
            self.assertFalse(result.errors, result.errors)
            delivery.update_validation_report(built, result)
            delivery.write_json(data_path, built)
            delivery.build_markdown(built, md_path)
            delivery.build_excel(built, xlsx_path)
            self.assertIn("hybrid_audit", built["validation_report"])
            self.assertEqual("master", built["shots"][0]["shot_type"])

            workbook = load_workbook(xlsx_path, read_only=True)
            try:
                sheet = workbook[delivery.SHEET_NAME]
                self.assertEqual(7, sheet.max_column)
            finally:
                workbook.close()

    def test_234_missing_shot_type_fails(self) -> None:
        data = valid_data_234()
        del data["shots"][0]["shot_type"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("shot_type" in item for item in result.errors))

    def test_234_invalid_split_reason_fails(self) -> None:
        data = valid_data_234()
        data["shots"][0]["split_reason"] = ["wrong_reason"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("split_reason" in item and "wrong_reason" in item for item in result.errors))

    def test_234_long_take_requires_two_supports(self) -> None:
        data = valid_data_234()
        data["shots"][1]["duration_seconds"] = 11
        data["shots"][1]["source_paragraph"] = "A移至桌边。A说：到了。"
        data["shots"][1]["duration_breakdown"] = {
            "sync_action_seconds": 11,
            "sync_dialogue_seconds": 0,
            "non_sync_action_seconds": 0,
            "emotional_pause_seconds": 0,
        }
        data["shots"][1]["camera_main_image"] = (
            "[侧面平视, 中景, 横移跟拍]\n"
            "【机位逻辑】摄影机沿桌边横移，跟住A的脚步。\n"
            "【站位位移】A从门口移至桌边，面向B的位置。\n"
            "A在桌边整理账册，随后示意。"
        )
        data["shots"][1]["notes"] = "[时长估算] 同步动作11秒 + 同步对白0秒 + 非同步动作0秒 + 情绪停顿0秒。 [长镜头] [保留理由] 动作长镜承载位移。"
        data["shots"][1]["long_take"] = {"classification": "action_long_take", "reason": "动作长镜"}
        data["shots"][1]["long_take_support"] = ["character_blocking"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("long_take_support" in item for item in result.errors))

        data["shots"][1]["long_take_support"] = ["character_blocking", "spatial_progression"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_234_long_insert_duration_fails_without_support(self) -> None:
        data = valid_data_234()
        data["shots"][0]["shot_type"] = "insert"
        data["shots"][0]["split_reason"] = ["prop_state_change"]
        data["shots"][0]["insert_priority"] = "recommended"
        data["shots"][0]["duration_seconds"] = 6
        data["shots"][0]["duration_breakdown"] = {
            "sync_action_seconds": 5,
            "sync_dialogue_seconds": 0,
            "non_sync_action_seconds": 0,
            "emotional_pause_seconds": 1,
        }
        data["shots"][0]["notes"] = "[时长估算] 同步动作5秒 + 同步对白0秒 + 非同步动作0秒 + 情绪停顿1秒。插镜偏长，仅用于警告测试。"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("超过5秒" in item for item in result.errors))

    def test_234_three_truth_reveals_in_one_dialogue_shot_fail(self) -> None:
        data = valid_data_234()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "沈夜：二十年前它就想找你。",
                "facts": [{"fact_id": "B001-F01", "type": "dialogue", "text": "二十年前它就想找你。", "cut_priority": "must_isolate", "cut_reasons": ["causal_reveal"], "cut_group": "B001-truth"}],
            },
            {
                "beat_id": "B002",
                "scene": "1 室内 日 内",
                "source_text": "顾成：林晓杰从一开始就是替身。",
                "facts": [{"fact_id": "B002-F01", "type": "dialogue", "text": "林晓杰从一开始就是替身。", "cut_priority": "must_isolate", "cut_reasons": ["causal_reveal"], "cut_group": "B002-substitute"}],
            },
            {
                "beat_id": "B003",
                "scene": "1 室内 日 内",
                "source_text": "沈夜：本来你才是最好的宿主。",
                "facts": [{"fact_id": "B003-F01", "type": "dialogue", "text": "本来你才是最好的宿主。", "cut_priority": "must_isolate", "cut_reasons": ["causal_reveal"], "cut_group": "B003-host"}],
            },
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "shot_no": 1,
                "beat_ids": ["B001", "B002", "B003"],
                "covered_fact_ids": ["B001-F01", "B002-F01", "B003-F01"],
                "source_paragraph": "沈夜：二十年前它就想找你。 / 顾成：林晓杰从一开始就是替身。 / 沈夜：本来你才是最好的宿主。",
                "duration_seconds": 13,
                "duration_breakdown": {"sync_action_seconds": 2, "sync_dialogue_seconds": 10, "non_sync_action_seconds": 2, "emotional_pause_seconds": 1},
                "long_take": {"classification": "dialogue_long_take", "reason": "测试长对白"},
                "camera_main_image": "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机拍沈夜和林晓彤。\n【场景首镜站位】沈夜在左，林晓彤在右，两人面向彼此。\n沈夜说：“二十年前它就想找你。”顾成说：“林晓杰从一开始就是替身。”沈夜说：“本来你才是最好的宿主。”",
                "notes": "[时长估算] 同步动作2秒 + 同步台词10秒 + 非同步动作2秒 + 情绪留白1秒。 [长镜头] [保留理由] [不可拆说明] 测试故意压缩。",
                "visible_characters": ["沈夜", "林晓彤", "顾成"],
                "shot_type": "dialogue",
                "split_reason": ["causal_reveal", "performance_continuity"],
                "insert_priority": "must_have",
                "long_take_support": ["character_blocking", "emotional_turn"],
            }
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("过度压缩" in item or "must_isolate cut_group" in item for item in result.errors), result.errors)

    def test_234_first_alien_voice_and_mocking_turn_in_one_long_dialogue_fail(self) -> None:
        data = valid_data_234()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "雾深处传来低沉震动。林晓杰：别催。",
                "facts": [
                    {"fact_id": "B001-F01", "type": "sound", "text": "雾深处传来低沉震动。", "cut_priority": "must_isolate", "cut_reasons": ["new_sound_source"], "cut_group": "B001-sound"},
                    {"fact_id": "B001-F02", "type": "dialogue", "text": "别催。", "cut_priority": "normal", "cut_reasons": [], "cut_group": "B001-dialogue"},
                ],
            },
            {
                "beat_id": "B002",
                "scene": "1 室内 日 内",
                "source_text": "林晓杰笑了一下，转为嘲讽。",
                "facts": [{"fact_id": "B002-F01", "type": "emotion", "text": "林晓杰笑了一下，转为嘲讽。", "cut_priority": "must_isolate", "cut_reasons": ["emotional_turn"], "cut_group": "B002-mock"}],
            },
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "shot_no": 1,
                "beat_ids": ["B001", "B002"],
                "covered_fact_ids": ["B001-F01", "B001-F02", "B002-F01"],
                "source_paragraph": "雾深处传来低沉震动。林晓杰：别催。 / 林晓杰笑了一下，转为嘲讽。",
                "duration_seconds": 13,
                "duration_breakdown": {"sync_action_seconds": 3, "sync_dialogue_seconds": 9, "non_sync_action_seconds": 3, "emotional_pause_seconds": 1},
                "long_take": {"classification": "dialogue_long_take", "reason": "测试长对白"},
                "camera_main_image": "[低角度仰拍, 中近景, 固定镜头]\n【机位逻辑】摄影机固定拍林晓杰。\n【场景首镜站位】林晓杰靠在岩壁前，面向远处黑雾。\n雾深处传来低沉震动，林晓杰说：“别催。”他笑了一下，转为嘲讽。",
                "notes": "[时长估算] 同步动作3秒 + 同步台词9秒 + 非同步动作3秒 + 情绪留白1秒。 [长镜头] [保留理由] [不可拆说明] 测试故意压缩。",
                "visible_characters": ["林晓杰"],
                "shot_type": "dialogue",
                "split_reason": ["new_sound_source", "emotional_turn"],
                "insert_priority": "must_have",
                "long_take_support": ["sound_source_change", "emotional_turn"],
            }
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("must_isolate cut_group" in item or "dialogue 镜" in item for item in result.errors), result.errors)

    def test_234_exile_origin_chain_in_one_transition_shot_fails(self) -> None:
        data = valid_data_234()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "光柱熄灭，沈夜坠落，身上变成凡人衣服。",
                "facts": [
                    {"fact_id": "B001-F01", "type": "prop", "text": "光柱熄灭。", "cut_priority": "must_isolate", "cut_reasons": ["new_vfx_state"], "cut_group": "B001-light-off"},
                    {"fact_id": "B001-F02", "type": "action", "text": "沈夜坠落到人间。", "cut_priority": "must_isolate", "cut_reasons": ["reality_layer_shift"], "cut_group": "B001-fall"},
                    {"fact_id": "B001-F03", "type": "prop", "text": "沈夜身上变成凡人衣服。", "cut_priority": "must_isolate", "cut_reasons": ["causal_reveal"], "cut_group": "B001-human-clothes"},
                ],
            }
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "shot_no": 1,
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01", "B001-F02", "B001-F03"],
                "source_paragraph": "光柱熄灭，沈夜坠落，身上变成凡人衣服。",
                "duration_seconds": 9,
                "duration_breakdown": {"sync_action_seconds": 5, "sync_dialogue_seconds": 0, "non_sync_action_seconds": 3, "emotional_pause_seconds": 1},
                "camera_main_image": "[微俯视, 全景, 伸缩摇臂拉出]\n【机位逻辑】摄影机从神殿拉到人间。\n【场景首镜站位】沈夜跪在大殿中央。\n光柱熄灭，沈夜坠落到人间，身上变成凡人衣服。",
                "notes": "[时长估算] 同步动作5秒 + 同步台词0秒 + 非同步动作3秒 + 情绪留白1秒。 [长镜头] [保留理由] 测试故意压缩。",
                "visible_characters": ["沈夜"],
                "shot_type": "transition",
                "split_reason": ["new_vfx_state", "reality_layer_shift", "causal_reveal"],
                "insert_priority": "must_have",
                "long_take_support": ["vfx_state_change", "spatial_progression"],
            }
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("must_isolate cut_group" in item for item in result.errors), result.errors)

    def test_234_same_cut_group_vfx_moment_can_pass(self) -> None:
        data = valid_data_234()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "恶念钻入眉心，瞳孔全黑。",
                "facts": [
                    {"fact_id": "B001-F01", "type": "action", "text": "恶念钻入眉心。", "cut_priority": "must_isolate", "cut_reasons": ["causal_reveal", "new_vfx_state"], "cut_group": "B001-impact"},
                    {"fact_id": "B001-F02", "type": "prop", "text": "瞳孔全黑。", "cut_priority": "must_isolate", "cut_reasons": ["new_vfx_state"], "cut_group": "B001-impact"},
                ],
            }
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "shot_no": 1,
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01", "B001-F02"],
                "source_paragraph": "恶念钻入眉心，瞳孔全黑。",
                "duration_seconds": 4,
                "duration_breakdown": {"sync_action_seconds": 3, "sync_dialogue_seconds": 0, "non_sync_action_seconds": 0, "emotional_pause_seconds": 1},
                "camera_main_image": "[平视, 特写, 急推]\n【机位逻辑】摄影机压到眉心和眼睛。\n【场景首镜站位】小林晓杰位于画面中央，面向镜头。\n恶念钻入小林晓杰眉心，他的瞳孔瞬间全黑。",
                "notes": "[不可拆说明] 恶念入体和瞳孔变化属于同一撞击瞬间。",
                "visible_characters": ["小林晓杰"],
                "shot_type": "vfx_anchor",
                "split_reason": ["causal_reveal", "new_vfx_state"],
                "insert_priority": "must_have",
                "long_take_support": [],
            }
        ]
        delivery.derive_prompts(data)
        first = delivery.validate_data(data, strict_status=True)
        resolve_warnings(data, first)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_234_plain_short_performance_merge_passes(self) -> None:
        data = valid_data_234()
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_234_local_dialogue_changed_to_offscreen_voice_fails(self) -> None:
        data = valid_data_234()
        data["beats"][2]["facts"][0]["cut_priority"] = "must_isolate"
        data["beats"][2]["facts"][0]["cut_reasons"] = ["causal_reveal"]
        data["beats"][2]["facts"][0]["cut_group"] = "B003-truth"
        data["shots"][1]["source_paragraph"] = "顾成：所以林晓杰从一开始就是替身。"
        data["shots"][1]["beat_ids"] = ["B003"]
        data["shots"][1]["covered_fact_ids"] = ["B003-F01"]
        data["shots"][1]["camera_main_image"] = "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机只拍沈夜和林晓彤。\n顾成画外声说：“所以林晓杰从一开始就是替身。”"
        data["shots"][1]["visible_characters"] = ["沈夜", "林晓彤"]
        data["shots"][1]["shot_type"] = "dialogue"
        data["shots"][1]["split_reason"] = ["causal_reveal"]
        data["shots"][1]["insert_priority"] = "must_have"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("改为画外声" in item for item in result.errors), result.errors)

    def test_235_adjacent_push_then_pull_fails(self) -> None:
        data = valid_data_235()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 缓慢推进]\n"
            "【机位逻辑】摄影机从桌边缓慢推进到A。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A站在门口，手扶门框。"
        )
        data["shots"][1]["camera_main_image"] = (
            "[侧面平视, 中景, 缓慢拉出]\n"
            "【机位逻辑】摄影机从A半身缓慢拉出到桌边关系。\n"
            "【站位位移】A从门口走到桌边，面向B的位置。\n"
            "A在桌边停下，说：“到了。”"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("纵深推拉方向反转" in item for item in result.errors), result.errors)

    def test_235_adjacent_pull_then_push_fails(self) -> None:
        data = valid_data_235()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 缓慢拉出]\n"
            "【机位逻辑】摄影机从A半身缓慢拉出到门口空间。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A站在门口，手扶门框。"
        )
        data["shots"][1]["camera_main_image"] = (
            "[侧面平视, 中景, 缓慢推进]\n"
            "【机位逻辑】摄影机沿桌边缓慢推进到A。\n"
            "【站位位移】A从门口走到桌边，面向B的位置。\n"
            "A在桌边停下，说：“到了。”"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("纵深推拉方向反转" in item for item in result.errors), result.errors)

    def test_235_adjacent_push_then_lateral_passes(self) -> None:
        data = valid_data_235()
        data["shots"][0]["camera_main_image"] = (
            "[平视, 中景, 缓慢推进]\n"
            "【机位逻辑】摄影机从桌边缓慢推进到A。\n"
            "【场景首镜站位】（A站在门口，面向桌边。）\n"
            "A站在门口，手扶门框。"
        )
        data["shots"][1]["camera_main_image"] = (
            "[侧面平视, 中景, 缓慢横移]\n"
            "【机位逻辑】摄影机沿桌边横移，跟住A的脚步。\n"
            "【站位位移】A从门口走到桌边，面向B的位置。\n"
            "A在桌边停下，说：“到了。”"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_235_final_shot_crane_up_pull_away_passes(self) -> None:
        data = valid_data_235()
        data["shots"][1]["camera_main_image"] = (
            "[高角度俯拍, 大全景, 摇臂缓慢上升拉远]\n"
            "【机位逻辑】摄影机从桌边上方升起拉远，让A的身影在空间里变小。\n"
            "【站位位移】A从门口走到桌边，面向B的位置。\n"
            "A在桌边停下，说：“到了。”"
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_236_read_only_artifacts_keep_report_and_audit_summary(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data_path = root / "sample.shot_data.json"
            md_path = root / "sample.md"
            xlsx_path = root / "sample.xlsx"
            report_path = root / "sample.validation_report.json"
            built = valid_data_236()
            delivery.derive_prompts(built)
            result = delivery.validate_data(built, strict_status=False)
            self.assertFalse(result.errors, result.errors)
            delivery.update_validation_report(built, result)
            delivery.write_json(data_path, built)
            delivery.build_markdown(built, md_path)
            delivery.build_excel(built, xlsx_path)
            delivery.build_report(built, report_path)
            report = json.loads(report_path.read_text(encoding="utf-8"))
            self.assertEqual(built["validation_report"], report)
            self.assertEqual(report["source_json_hash"], delivery.canonical_data_hash(built))
            self.assertIn(report["source_json_hash"], md_path.read_text(encoding="utf-8"))

            workbook = load_workbook(xlsx_path, read_only=True)
            try:
                self.assertEqual([delivery.SHEET_NAME, delivery.SUMMARY_SHEET_NAME], workbook.sheetnames)
            finally:
                workbook.close()

            rc = delivery.main([
                "validate",
                "--data",
                str(data_path),
                "--markdown",
                str(md_path),
                "--excel",
                str(xlsx_path),
                "--report",
                str(report_path),
            ])
            self.assertEqual(0, rc)

    def test_236_same_category_different_moment_must_isolate_fails(self) -> None:
        data = valid_data_236()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "光变黑，雾钻出。",
                "facts": [
                    {
                        "fact_id": "B001-F01",
                        "type": "prop",
                        "text": "光变黑。",
                        "cut_priority": "must_isolate",
                        "cut_reasons": ["new_vfx_state"],
                        "cut_group": "B001-vfx",
                        "cut_category": "vfx",
                        "cut_moment_id": "B001-vfx-light",
                    },
                    {
                        "fact_id": "B001-F02",
                        "type": "prop",
                        "text": "雾钻出。",
                        "cut_priority": "must_isolate",
                        "cut_reasons": ["new_vfx_state"],
                        "cut_group": "B001-vfx",
                        "cut_category": "vfx",
                        "cut_moment_id": "B001-vfx-fog",
                    },
                ],
            }
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01", "B001-F02"],
                "source_paragraph": "光变黑，雾钻出。",
                "duration_seconds": 4,
                "duration_breakdown": {"sync_action_seconds": 3, "sync_dialogue_seconds": 0, "non_sync_action_seconds": 0, "emotional_pause_seconds": 1},
                "camera_main_image": "[平视, 特写, 急推]\n【机位逻辑】摄影机压到光和雾。\n【场景首镜站位】A站在门口，面向光源。\n光变黑，雾钻出。",
                "notes": "[不可拆说明] 测试错误合并。",
                "shot_type": "vfx_anchor",
                "split_reason": ["new_vfx_state"],
                "insert_priority": "must_have",
            }
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("结构切点" in item for item in result.errors), result.errors)

    def test_236_same_category_same_moment_with_note_passes(self) -> None:
        data = valid_data_236()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene": "1 室内 日 内",
                "source_text": "恶念钻入眉心，瞳孔全黑。",
                "facts": [
                    {
                        "fact_id": "B001-F01",
                        "type": "action",
                        "text": "恶念钻入眉心。",
                        "cut_priority": "must_isolate",
                        "cut_reasons": ["causal_reveal", "new_vfx_state"],
                        "cut_group": "B001-vfx",
                        "cut_category": "vfx",
                        "cut_moment_id": "B001-vfx-impact",
                    },
                    {
                        "fact_id": "B001-F02",
                        "type": "prop",
                        "text": "瞳孔全黑。",
                        "cut_priority": "must_isolate",
                        "cut_reasons": ["new_vfx_state"],
                        "cut_group": "B001-vfx",
                        "cut_category": "vfx",
                        "cut_moment_id": "B001-vfx-impact",
                    },
                ],
            }
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01", "B001-F02"],
                "source_paragraph": "恶念钻入眉心，瞳孔全黑。",
                "duration_seconds": 4,
                "duration_breakdown": {"sync_action_seconds": 3, "sync_dialogue_seconds": 0, "non_sync_action_seconds": 0, "emotional_pause_seconds": 1},
                "camera_main_image": "[平视, 特写, 急推]\n【机位逻辑】摄影机压到眉心和眼睛。\n【场景首镜站位】A位于画面中央，面向镜头。\n恶念钻入眉心，瞳孔全黑。",
                "notes": "[不可拆说明] 恶念入体和瞳孔变化属于同一撞击瞬间。",
                "shot_type": "vfx_anchor",
                "split_reason": ["causal_reveal", "new_vfx_state"],
                "insert_priority": "must_have",
            }
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_236_must_isolate_density_warns(self) -> None:
        data = valid_data_236()
        data["beats"] = data["beats"][:2]
        data["shots"] = data["shots"][:2]
        data["shots"][0]["shot_type"] = "action"
        data["shots"][1]["beat_ids"] = ["B002"]
        data["shots"][1]["covered_fact_ids"] = ["B002-F01"]
        data["shots"][1]["source_paragraph"] = "A走到桌边。"
        for beat in data["beats"]:
            fact = beat["facts"][0]
            fact["cut_priority"] = "must_isolate"
            fact["cut_reasons"] = ["new_information"]
            fact["cut_category"] = "space"
            fact["cut_moment_id"] = f"{beat['beat_id']}-space"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("must_isolate Beat 占比" in item for item in result.warnings), result.warnings)

    def test_236_fact_text_deviation_warns(self) -> None:
        data = valid_data_236()
        data["beats"][0]["source_text"] = "手环变亮。"
        data["beats"][0]["facts"][0]["type"] = "prop"
        data["beats"][0]["facts"][0]["text"] = "手环同频呼吸变亮。"
        data["beats"][0]["facts"][0]["cut_category"] = "prop"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("Fact 文本疑似引入" in item for item in result.warnings), result.warnings)

    def test_236_inherited_state_mismatch_fails(self) -> None:
        data = valid_data_236()
        data["continuity_logs"].append(
            {
                "scene": "2 室内 日 内",
                "inherits_from": "1 室内 日 内",
                "inherited_states": ["spatial_axis", "fixed_objects"],
                "diverged_states": ["characters", "props"],
                "first_shot_anchor_type": "single_continuation",
                "spatial_axis": "A在窗边，B在桌边。",
                "fixed_objects": [],
                "characters": [],
                "props": [],
                "sound_sources": [],
                "reality_layer": "现实",
            }
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("继承字段 spatial_axis" in item for item in result.errors), result.errors)

    def test_236_continuity_update_from_mismatch_fails(self) -> None:
        data = valid_data_236()
        data["shots"][1]["continuity_updates"][0]["from"] = "窗边"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("from 与上一状态不一致" in item for item in result.errors), result.errors)

    def test_236_prompt_composition_dialogue_leak_fails(self) -> None:
        data = valid_data_236()
        delivery.derive_prompts(data)
        data["shots"][1]["prompt"] = data["shots"][1]["prompt"].replace("构图：可见主体：A。", "构图：A说：“到了。”")
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("构图字段混入" in item for item in result.errors), result.errors)

    def test_236_general_short_reaction_warns(self) -> None:
        data = valid_data_236()
        data["beats"] = data["beats"][:2]
        data["beats"][0]["source_text"] = "A看向门口。"
        data["beats"][0]["facts"] = [
            {
                "fact_id": "B001-F01",
                "type": "action",
                "text": "A看向门口。",
                "cut_priority": "normal",
                "cut_reasons": [],
                "cut_group": "B001-action",
                "cut_category": "action",
                "cut_moment_id": "B001-action",
            }
        ]
        data["beats"][1]["source_text"] = "B皱眉。"
        data["beats"][1]["facts"] = [
            {
                "fact_id": "B002-F01",
                "type": "emotion",
                "text": "B皱眉。",
                "cut_priority": "normal",
                "cut_reasons": [],
                "cut_group": "B002-emotion",
                "cut_category": "emotion",
                "cut_moment_id": "B002-emotion",
            }
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01"],
                "source_paragraph": "A看向门口。",
                "camera_main_image": "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机在桌边拍A。\n【场景首镜站位】A在门口，面向桌边。\nA看向门口。",
                "shot_type": "action",
                "split_reason": ["performance_continuity"],
            },
            {
                **data["shots"][1],
                "beat_ids": ["B002"],
                "covered_fact_ids": ["B002-F01"],
                "source_paragraph": "B皱眉。",
                "duration_seconds": 2,
                "duration_breakdown": {"sync_action_seconds": 1, "sync_dialogue_seconds": 0, "non_sync_action_seconds": 0, "emotional_pause_seconds": 1},
                "camera_main_image": "[侧面平视, 近景, 固定镜头]\n【机位逻辑】摄影机切到B的脸。\nB皱眉。",
                "visible_characters": ["B"],
                "continuity_updates": [],
                "shot_type": "reaction",
                "split_reason": ["performance_continuity"],
                "insert_priority": "none",
            },
        ]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)
        self.assertTrue(any("[可合并]" in item for item in result.warnings), result.warnings)

    def test_240_valid_data_passes_pre_signoff_without_gate_c(self) -> None:
        data = valid_data_240()
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_240_final_signoff_requires_gate_c(self) -> None:
        data = valid_data_240()
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True, final_signoff=True)
        self.assertTrue(any("GATE_C" in item for item in result.errors), result.errors)

    def test_240_missing_gate_a_or_b_fails(self) -> None:
        data = valid_data_240()
        data["human_reviews"] = data["human_reviews"][:1]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("GATE_B" in item for item in result.errors), result.errors)

    def test_240_scene_id_is_required_and_checked(self) -> None:
        data = valid_data_240()
        data["shots"][0]["scene_id"] = "S99"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("S99" in item for item in result.warnings), result.warnings)

    def test_240_duplicate_scene_id_fails(self) -> None:
        data = valid_data_240()
        data["continuity_logs"].append({**data["continuity_logs"][0], "scene": "2 室内 日 内"})
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("场景键重复" in item for item in result.errors), result.errors)

    def test_240_beat_gap_is_allowed(self) -> None:
        data = valid_data_240()
        data["beats"] = [data["beats"][0], data["beats"][2]]
        data["shots"][1]["beat_ids"] = ["B003"]
        data["shots"][1]["covered_fact_ids"] = ["B003-F01"]
        data["shots"][1]["source_paragraph"] = "A说：到了。"
        data["shots"][1]["camera_main_image"] = "[低角度仰拍, 特写, 急推]\n【机位逻辑】摄影机压到A的脸。\nA停在门口，说：“到了。”"
        data["shots"][1]["duration_seconds"] = 2
        data["shots"][1]["duration_breakdown"] = {"sync_action_seconds": 1, "sync_dialogue_seconds": 1, "non_sync_action_seconds": 0, "emotional_pause_seconds": 1}
        data["shots"][1]["continuity_updates"] = []
        delivery.derive_prompts(data)
        first = delivery.validate_data(data, strict_status=True)
        resolve_warnings(data, first)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)
        self.assertEqual("B001+B003", delivery.format_beat_ids(["B001", "B003"]))

    def test_240_beat_order_must_be_monotonic(self) -> None:
        data = valid_data_240()
        data["beats"] = [data["beats"][1], data["beats"][0]]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("单调递增" in item for item in result.errors), result.errors)

    def test_240_same_moment_different_category_must_isolate_passes(self) -> None:
        data = valid_data_240()
        data["beats"] = [
            {
                "beat_id": "B001",
                "scene_id": "S01",
                "scene": "1 室内 日 内",
                "source_text": "恶念钻入眉心，瞳孔全黑。",
                "facts": [
                    {
                        "fact_id": "B001-F01",
                        "type": "action",
                        "text": "恶念钻入眉心。",
                        "cut_priority": "must_isolate",
                        "cut_reasons": ["causal_reveal"],
                        "cut_group": "B001-causal",
                        "cut_category": "action",
                        "cut_moment_id": "B001-impact",
                    },
                    {
                        "fact_id": "B001-F02",
                        "type": "prop",
                        "text": "瞳孔全黑。",
                        "cut_priority": "must_isolate",
                        "cut_reasons": ["new_vfx_state"],
                        "cut_group": "B001-vfx",
                        "cut_category": "vfx",
                        "cut_moment_id": "B001-impact",
                    },
                ],
            }
        ]
        data["shots"] = [
            {
                **data["shots"][0],
                "beat_ids": ["B001"],
                "covered_fact_ids": ["B001-F01", "B001-F02"],
                "source_paragraph": "恶念钻入眉心，瞳孔全黑。",
                "duration_seconds": 4,
                "duration_breakdown": {"sync_action_seconds": 3, "sync_dialogue_seconds": 0, "non_sync_action_seconds": 0, "emotional_pause_seconds": 1},
                "camera_main_image": "[平视, 特写, 急推]\n【机位逻辑】摄影机压到眉心和眼睛。\n【场景首镜站位】A位于画面中央，面向镜头。\n恶念钻入眉心，瞳孔全黑。",
                "notes": "[不可拆说明] 同一撞击瞬间。",
                "shot_type": "vfx_anchor",
                "split_reason": ["causal_reveal", "new_vfx_state"],
                "insert_priority": "must_have",
            }
        ]
        delivery.derive_prompts(data)
        first = delivery.validate_data(data, strict_status=True)
        resolve_warnings(data, first)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_240_different_moment_must_isolate_fails(self) -> None:
        data = valid_data_240()
        data["beats"][0]["facts"].append(
            {
                "fact_id": "B001-F02",
                "type": "prop",
                "text": "门裂开。",
                "cut_priority": "must_isolate",
                "cut_reasons": ["new_vfx_state"],
                "cut_group": "B001-vfx",
                "cut_category": "vfx",
                "cut_moment_id": "B001-other",
            }
        )
        data["beats"][0]["facts"][0]["cut_priority"] = "must_isolate"
        data["beats"][0]["facts"][0]["cut_reasons"] = ["causal_reveal"]
        data["beats"][0]["facts"][0]["cut_moment_id"] = "B001-first"
        data["shots"][0]["covered_fact_ids"] = ["B001-F01", "B001-F02"]
        data["shots"][0]["notes"] = "[不可拆说明] 测试错误合并。"
        data["shots"][0]["shot_type"] = "vfx_anchor"
        data["shots"][0]["split_reason"] = ["causal_reveal", "new_vfx_state"]
        data["shots"][0]["insert_priority"] = "must_have"
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("结构切点" in item for item in result.errors), result.errors)

    def test_240_safety_without_gate_b_approval_fails(self) -> None:
        data = valid_data_240()
        data["shots"].append({**data["shots"][1], "shot_no": 3, "covered_fact_ids": [], "shot_type": "safety", "notes": "[安全镜][人工批准] 测试。", "continuity_updates": []})
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("covered_fact_ids" in item for item in result.errors), result.errors)

    def test_240_safety_with_gate_b_approval_passes(self) -> None:
        data = valid_data_240()
        data["human_reviews"][1]["notes"] = "GATE_B APPROVED 安全镜"
        data["shots"].append(
            {
                **data["shots"][1],
                "shot_no": 3,
                "covered_fact_ids": [],
                "shot_type": "safety",
                "notes": "[安全镜][人工批准] Gate B批准无事实安全镜。",
                "camera_main_image": "[平视, 空镜全景, 固定镜头]\n【机位逻辑】摄影机留在室内空桌和门框之间。\n空房间维持静止。",
                "visible_characters": [],
                "continuity_updates": [],
            }
        )
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_240_required_adjacent_marker_warns_with_resolution(self) -> None:
        data = valid_data_240()
        data["shots"][1]["camera_main_image"] = "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机仍在桌边看向门口。\nA站在门口，手扶门框。"
        data["shots"][1]["visible_characters"] = ["A"]
        data["shots"][1]["continuity_updates"] = []
        data["shots"][1]["covered_fact_ids"] = ["B002-F01", "B003-F01"]
        data["shots"][1]["beat_ids"] = ["B002", "B003"]
        data["shots"][1]["notes"] = "[必拆相邻] must_isolate 独立落点。"
        delivery.derive_prompts(data)
        first = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("[必拆相邻]" in item for item in first.warnings), first.warnings)
        resolve_warnings(data, first)
        second = delivery.validate_data(data, strict_status=True)
        self.assertFalse(second.errors, second.errors)

    def test_240_reverse_axis_marker_warns_with_resolution(self) -> None:
        data = valid_data_240()
        data["shots"][0]["camera_main_image"] = "[平视, 中景, 缓慢推进]\n【机位逻辑】摄影机向A推进。\n【场景首镜站位】A站在门口，面向桌边。\nA站在门口。"
        data["shots"][1]["camera_main_image"] = "[平视, 中景, 缓慢拉出]\n【机位逻辑】摄影机从A身前拉出，暴露空房间。\n【站位位移】A从门口走到桌边，面向B的位置。\nA在桌边停下，说：“到了。”"
        data["shots"][1]["notes"] = "[反转动机] 揭示后拉出暴露空间代价。"
        delivery.derive_prompts(data)
        first = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("[反转动机]" in item for item in first.warnings), first.warnings)
        resolve_warnings(data, first)
        second = delivery.validate_data(data, strict_status=True)
        self.assertFalse(second.errors, second.errors)

    def test_240_prompt_overlength_warn_requires_resolution(self) -> None:
        data = valid_data_240()
        data["shots"][0]["camera_main_image"] += "\n" + "A反复确认门口、桌边、墙面和手部动作。" * 80
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("超过800字符" in item for item in result.warnings), result.warnings)
        self.assertTrue(any("WARN 缺少处置记录" in item for item in result.errors), result.errors)

    def test_240_not_run_hash_and_gate_c_contract(self) -> None:
        data = valid_data_240()
        data["validation_report"] = {"status": "NOT_RUN", "warnings": [], "errors": [], "source_json_hash": "fake"}
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("source_json_hash 必须为空" in item for item in result.errors), result.errors)
        data["validation_report"]["source_json_hash"] = ""
        data["human_reviews"].append(
            {"gate": "GATE_C", "round": 1, "status": "approved", "reviewer": "user", "notes": "accepted_without_validation"}
        )
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_242_valid_source_lock_and_approved_script_path_passes(self) -> None:
        data = valid_data_242()
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertFalse(result.errors, result.errors)

    def test_242_missing_approved_script_path_fails(self) -> None:
        data = valid_data_242()
        del data["script_lock"]["approved_script_path"]
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=True)
        self.assertTrue(any("approved_script_path" in item for item in result.errors), result.errors)


class ContractIntegrity243Tests(unittest.TestCase):
    def test_243_version_profile_is_fail_closed_and_build_only(self) -> None:
        self.assertEqual("2.4.3", delivery.VERSION)
        self.assertEqual(RULE_REVISION_243, delivery.RULE_REVISION)
        self.assertIn("2.4.2", delivery.VERSION_PROFILES)

        unknown = valid_data_243()
        unknown["metadata"]["version"] = "9.9.9"
        delivery.derive_prompts(unknown)
        result = delivery.validate_data(unknown, strict_status=False)
        self.assertTrue(any("version" in item for item in result.errors), result.errors)

        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data_path = root / "legacy.shot_data.json"
            markdown_path = root / "legacy.md"
            excel_path = root / "legacy.xlsx"
            report_path = root / "legacy.validation_report.json"
            original = json.dumps(valid_data_242(), ensure_ascii=False, indent=2) + "\n"
            data_path.write_text(original, encoding="utf-8")
            rc = delivery.main(
                [
                    "build",
                    "--data",
                    str(data_path),
                    "--markdown",
                    str(markdown_path),
                    "--excel",
                    str(excel_path),
                    "--report",
                    str(report_path),
                    "--workspace-root",
                    str(root),
                ]
            )
            self.assertEqual(1, rc)
            self.assertEqual(original, data_path.read_text(encoding="utf-8"))
            self.assertFalse(markdown_path.exists())
            self.assertFalse(excel_path.exists())

    def test_243_version_and_rule_revision_markers_are_consistent_everywhere(self) -> None:
        skill_root = SCRIPT_DIR.parent
        expected_version = "2.4.3"
        expected_revision = "2.4.3-contract-integrity-p2-2026-07-12"
        self.assertEqual(expected_version, (skill_root / "VERSION").read_text(encoding="utf-8-sig").strip())
        self.assertEqual(expected_version, delivery.VERSION)
        self.assertEqual(expected_revision, delivery.RULE_REVISION)

        wrapper = (SCRIPT_DIR / "validate_storyboard.js").read_text(encoding="utf-8-sig")
        self.assertIn(f'const VERSION = "{expected_version}";', wrapper)
        self.assertIn(f'const RULE_REVISION = "{expected_revision}";', wrapper)
        self.assertIn("2.4.3 validation requires both", wrapper)

        contract_files = [
            skill_root / "SKILL.md",
            skill_root / "agents" / "openai.yaml",
            skill_root / "references" / "continuity-shot-data.md",
            skill_root / "references" / "hybrid-shot-audit.md",
            skill_root / "references" / "camera-language.md",
            skill_root / "references" / "seedance-prompt-rules.md",
            skill_root / "references" / "project-notes.md",
        ]
        for path in contract_files:
            with self.subTest(path=path.name):
                text = path.read_text(encoding="utf-8-sig")
                self.assertIn(expected_version, text)
                self.assertIn(expected_revision, text)

    def test_243_wrong_rule_revision_fails(self) -> None:
        data = valid_data_243()
        data["metadata"]["rule_revision"] = RULE_REVISION_242
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=False)
        self.assertTrue(any("rule_revision" in item for item in result.errors), result.errors)

    def test_243_script_status_must_match_actual_builder_and_wrapper_presence(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            for script_name in ("storyboard_delivery.py", "validate_storyboard.js"):
                with self.subTest(script_name=script_name):
                    data = valid_data_243()
                    approved.write_text(data["script_lock"]["locked_text"], encoding="utf-8")
                    data["metadata"]["script_status"][script_name] = "missing"
                    append_review(data, "GATE_A")
                    append_review(data, "GATE_B")
                    delivery.derive_prompts(data)
                    result = delivery.validate_data(
                        data,
                        strict_status=False,
                        data_path=root / "sample.shot_data.json",
                        workspace_root=root,
                    )
                    self.assertTrue(
                        any(script_name in item and "available" in item for item in result.errors),
                        result.errors,
                    )

    def test_243_real_approved_script_path_and_content_pass(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data, data_path, _, _, _ = prepare_243_fixture(root)
            result = delivery.validate_data(
                data,
                strict_status=True,
                data_path=data_path,
                workspace_root=root,
            )
            self.assertFalse(result.errors, result.errors)

    def test_243_script_path_missing_mismatch_and_escape_fail(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data_path = root / "sample.shot_data.json"
            cases: list[tuple[str, dict, str]] = []

            missing = valid_data_243(approved_script_path="approved/missing.txt")
            cases.append(("missing", missing, "approved_script_path"))

            mismatch = valid_data_243()
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            approved.write_text("不是锁定剧本。", encoding="utf-8")
            cases.append(("mismatch", mismatch, "locked_text"))

            escaped = valid_data_243(approved_script_path="../outside.txt")
            cases.append(("escape", escaped, "越出工作区"))

            invalid_utf8 = valid_data_243(approved_script_path="approved/invalid-utf8.txt")
            (root / "approved" / "invalid-utf8.txt").write_bytes(b"\xff\xfe\x00")
            cases.append(("invalid utf8", invalid_utf8, "codec"))

            for name, data, expected in cases:
                with self.subTest(name=name):
                    append_review(data, "GATE_A")
                    append_review(data, "GATE_B")
                    delivery.derive_prompts(data)
                    result = delivery.validate_data(
                        data,
                        strict_status=False,
                        data_path=data_path,
                        workspace_root=root,
                    )
                    self.assertTrue(
                        any(expected.lower() in item.lower() for item in result.errors),
                        result.errors,
                    )

    def test_243_approved_script_path_requires_a_raw_json_string(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            for value in (123, True, ["approved/sample.approved_script.txt"]):
                with self.subTest(value=repr(value)):
                    data = valid_data_243()
                    data["script_lock"]["approved_script_path"] = value
                    append_review(data, "GATE_A")
                    append_review(data, "GATE_B")
                    delivery.derive_prompts(data)
                    result = delivery.validate_data(
                        data,
                        strict_status=False,
                        data_path=root / "sample.shot_data.json",
                        workspace_root=root,
                    )
                    self.assertTrue(
                        any("approved_script_path" in item and "字符串" in item for item in result.errors),
                        result.errors,
                    )

    def test_243_crlf_bom_and_unicode_spans_pass(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data_path = root / "unicode.shot_data.json"
            approved_path = root / "approved" / "sample.approved_script.txt"
            approved_path.parent.mkdir(parents=True, exist_ok=True)

            data = valid_data_243()
            normalized = "A🙂站在门口。\nA走到桌边。\nA说：到了。"
            approved_path.write_bytes(b"\xef\xbb\xbf" + normalized.replace("\n", "\r\n").encode("utf-8"))
            data["script_lock"]["locked_text"] = normalized
            data["script_lock"]["locked_text_hash"] = delivery.normalized_script_text_hash(normalized)
            spans = {
                "B001": {"start": 0, "end": 7},
                "B002": {"start": 8, "end": 14},
                "B003": {"start": 15, "end": 21},
            }
            data["beats"][0]["source_text"] = "A🙂站在门口。"
            data["beats"][0]["facts"][0]["text"] = "A🙂站在门口。"
            for beat in data["beats"]:
                beat["source_span"] = spans[beat["beat_id"]]
            data["shots"][0]["source_paragraph"] = "A🙂站在门口。"
            data["shots"][0]["source_span"] = spans["B001"]
            data["shots"][0]["camera_main_image"] = data["shots"][0]["camera_main_image"].replace(
                "A站在门口", "A🙂站在门口"
            )
            data["shots"][1]["source_spans"] = [spans["B002"], spans["B003"]]
            append_review(data, "GATE_A")
            append_review(data, "GATE_B")
            delivery.derive_prompts(data)
            result = delivery.validate_data(
                data,
                strict_status=False,
                data_path=data_path,
                workspace_root=root,
            )
            if result.warnings:
                resolve_warnings(data, result)
                result = delivery.validate_data(
                    data,
                    strict_status=False,
                    data_path=data_path,
                    workspace_root=root,
                )
            self.assertFalse(result.errors, result.errors)
            self.assertEqual(normalized, delivery.normalize_script_text(approved_path.read_text(encoding="utf-8-sig")))

    def test_243_span_types_order_overlap_duplicate_and_conflict_fail(self) -> None:
        mutators = {
            "string": lambda data: data["beats"][0].update(source_span={"start": "0", "end": 6}),
            "bool": lambda data: data["beats"][0].update(source_span={"start": False, "end": 6}),
            "float": lambda data: data["beats"][0].update(source_span={"start": 0.0, "end": 6}),
            "order": lambda data: data["shots"][1].update(
                source_spans=[{"start": 14, "end": 20}, {"start": 7, "end": 13}]
            ),
            "overlap": lambda data: data["shots"][1].update(
                source_spans=[{"start": 7, "end": 13}, {"start": 12, "end": 20}]
            ),
            "duplicate": lambda data: data["shots"][1].update(
                source_spans=[{"start": 7, "end": 13}, {"start": 7, "end": 13}]
            ),
            "conflict": lambda data: data["shots"][0].update(source_spans=[{"start": 0, "end": 6}]),
            "conflict null source_spans": lambda data: data["shots"][0].update(source_spans=None),
            "conflict null source_span": lambda data: data["shots"][1].update(source_span=None),
        }
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            approved.write_text(valid_data_243()["script_lock"]["locked_text"], encoding="utf-8")
            for name, mutate in mutators.items():
                with self.subTest(name=name):
                    data = valid_data_243()
                    mutate(data)
                    append_review(data, "GATE_A")
                    append_review(data, "GATE_B")
                    delivery.derive_prompts(data)
                    result = delivery.validate_data(
                        data,
                        strict_status=False,
                        data_path=root / "sample.shot_data.json",
                        workspace_root=root,
                    )
                    self.assertTrue(
                        any("source_span" in item or "span" in item.lower() for item in result.errors),
                        result.errors,
                    )

    def test_243_camera_and_prompt_dialogue_changes_fail(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            approved.write_text(valid_data_243()["script_lock"]["locked_text"], encoding="utf-8")

            camera_changed = valid_data_243()
            camera_changed["shots"][1]["camera_main_image"] = camera_changed["shots"][1]["camera_main_image"].replace(
                "到了。", "到啦。"
            )
            append_review(camera_changed, "GATE_A")
            append_review(camera_changed, "GATE_B")
            delivery.derive_prompts(camera_changed)
            result = delivery.validate_data(
                camera_changed,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(any("对白" in item or "dialogue" in item.lower() for item in result.errors), result.errors)

            prompt_changed = valid_data_243()
            append_review(prompt_changed, "GATE_A")
            append_review(prompt_changed, "GATE_B")
            delivery.derive_prompts(prompt_changed)
            prompt_changed["shots"][1]["prompt"] = prompt_changed["shots"][1]["prompt"].replace("到了。", "到啦。")
            result = delivery.validate_data(
                prompt_changed,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(any("Prompt" in item or "对白" in item for item in result.errors), result.errors)

    def test_243_dialogue_fact_camera_and_prompt_cannot_change_together(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            data = valid_data_243()
            approved.write_text(data["script_lock"]["locked_text"], encoding="utf-8")
            data["beats"][2]["facts"][0]["text"] = "到啦。"
            data["shots"][1]["camera_main_image"] = data["shots"][1]["camera_main_image"].replace("到了。", "到啦。")
            append_review(data, "GATE_A")
            append_review(data, "GATE_B")
            delivery.derive_prompts(data)
            result = delivery.validate_data(
                data,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(any("dialogue fact" in item.lower() or "对白事实" in item for item in result.errors), result.errors)

            truncated = valid_data_243()
            truncated["shots"][1]["camera_main_image"] = truncated["shots"][1]["camera_main_image"].replace(
                "到了。", "到"
            )
            append_review(truncated, "GATE_A")
            append_review(truncated, "GATE_B")
            delivery.derive_prompts(truncated)
            result = delivery.validate_data(
                truncated,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(any("对白" in item or "dialogue" in item.lower() for item in result.errors), result.errors)

    def test_243_dialogue_label_stops_at_newline_before_action(self) -> None:
        data = {
            "metadata": {"version": "2.4.3"},
            "beats": [
                {
                    "beat_id": "B001",
                    "facts": [{"fact_id": "B001-F01", "type": "dialogue", "text": "你好"}],
                }
            ],
        }
        shot = {
            "shot_no": 1,
            "covered_fact_ids": ["B001-F01"],
            "source_paragraph": "张三：你好\n张三起身",
            "camera_main_image": "张三说：“你好”\n张三起身。",
            "prompt": "画面内容：张三说：“你好”，随后起身。",
        }
        result = delivery.ValidationResult()
        delivery.validate_dialogue_fidelity(data, shot, result)
        self.assertFalse(result.errors, result.errors)

    def test_243_one_speakers_vo_does_not_authorize_another_speakers_live_dialogue(self) -> None:
        data = {
            "metadata": {"version": "2.4.3"},
            "beats": [
                {
                    "beat_id": "B001",
                    "facts": [
                        {"fact_id": "B001-F01", "type": "dialogue", "text": "别怕。"},
                        {"fact_id": "B001-F02", "type": "dialogue", "text": "别走。"},
                    ],
                }
            ],
        }
        shot = {
            "shot_no": 1,
            "covered_fact_ids": ["B001-F01", "B001-F02"],
            "source_paragraph": "甲（VO）：别怕。\n乙：别走。",
            "camera_main_image": "甲画外声：“别怕。”\n乙画外声：“别走。”",
            "prompt": "画面内容：甲画外声：“别怕。”，乙画外声：“别走。”",
        }
        result = delivery.ValidationResult()
        delivery.validate_dialogue_fidelity(data, shot, result)
        self.assertTrue(
            any(("乙" in item or "别走" in item) and ("VO" in item or "画外声" in item) for item in result.errors),
            result.errors,
        )

    def test_243_vo_markers_after_quote_before_newline_and_cross_speaker_all_fail(self) -> None:
        cases = [
            (
                "marker after quote",
                "乙：“现场。”",
                "乙：“现场。”（画外声）",
                [{"fact_id": "B001-F01", "type": "dialogue", "text": "现场。"}],
                ["B001-F01"],
                "现场",
            ),
            (
                "marker on preceding label line",
                "乙：“现场。”",
                "乙画外声：\n“现场。”",
                [{"fact_id": "B001-F01", "type": "dialogue", "text": "现场。"}],
                ["B001-F01"],
                "现场",
            ),
            (
                "one speaker vo cannot authorize another",
                "甲（VO）：“别怕。”\n乙：“真相。”",
                "甲画外声：“别怕。”\n乙画外声：“真相。”",
                [
                    {"fact_id": "B001-F01", "type": "dialogue", "text": "别怕。"},
                    {"fact_id": "B001-F02", "type": "dialogue", "text": "真相。"},
                ],
                ["B001-F01", "B001-F02"],
                "真相",
            ),
            (
                "speaker after quoted vo cannot reuse another speaker text",
                "甲（VO）：“甲台词。”\n乙：“乙台词。”",
                "“甲台词。”——乙画外声",
                [
                    {"fact_id": "B001-F01", "type": "dialogue", "text": "甲台词。"},
                    {"fact_id": "B001-F02", "type": "dialogue", "text": "乙台词。"},
                ],
                ["B001-F01", "B001-F02"],
                "甲台词",
            ),
            (
                "label without colon on preceding line cannot reuse another speaker text",
                "甲（VO）：“甲台词。”\n乙：“乙台词。”",
                "乙画外声\n“甲台词。”",
                [
                    {"fact_id": "B001-F01", "type": "dialogue", "text": "甲台词。"},
                    {"fact_id": "B001-F02", "type": "dialogue", "text": "乙台词。"},
                ],
                ["B001-F01", "B001-F02"],
                "甲台词",
            ),
            (
                "trailing known speaker cannot reuse vo text",
                "甲（VO）：“别怕。”\n乙：“真相。”",
                "“别怕。”——乙画外声",
                [
                    {"fact_id": "B001-F01", "type": "dialogue", "text": "别怕。"},
                    {"fact_id": "B001-F02", "type": "dialogue", "text": "真相。"},
                ],
                ["B001-F01", "B001-F02"],
                "别怕",
            ),
            (
                "preceding known speaker without colon cannot reuse vo text",
                "甲（VO）：“别怕。”\n乙：“真相。”",
                "乙画外声\n“别怕。”",
                [
                    {"fact_id": "B001-F01", "type": "dialogue", "text": "别怕。"},
                    {"fact_id": "B001-F02", "type": "dialogue", "text": "真相。"},
                ],
                ["B001-F01", "B001-F02"],
                "别怕",
            ),
        ]
        for name, source, camera, facts, fact_ids, expected_text in cases:
            with self.subTest(name=name):
                data = {
                    "metadata": {"version": "2.4.3"},
                    "beats": [{"beat_id": "B001", "facts": facts}],
                }
                shot = {
                    "shot_no": 1,
                    "covered_fact_ids": fact_ids,
                    "source_paragraph": source,
                    "camera_main_image": camera,
                    "prompt": "画面内容：" + camera,
                }
                result = delivery.ValidationResult()
                delivery.validate_dialogue_fidelity(data, shot, result)
                self.assertTrue(
                    any(expected_text in item and ("VO" in item or "画外声" in item) for item in result.errors),
                    result.errors,
                )

    def test_243_generic_vo_marker_without_speaker_can_bind_by_unique_vo_text(self) -> None:
        data = {
            "metadata": {"version": "2.4.3"},
            "beats": [
                {
                    "beat_id": "B001",
                    "facts": [
                        {"fact_id": "B001-F01", "type": "dialogue", "text": "别怕。"},
                        {"fact_id": "B001-F02", "type": "dialogue", "text": "真相。"},
                    ],
                }
            ],
        }
        shot = {
            "shot_no": 1,
            "covered_fact_ids": ["B001-F01", "B001-F02"],
            "source_paragraph": "甲（VO）：“别怕。”\n乙：“真相。”",
            "camera_main_image": "“别怕。”（画外声）\n乙说：“真相。”",
            "prompt": "",
        }
        result = delivery.ValidationResult()
        delivery.validate_dialogue_fidelity(data, shot, result)
        self.assertFalse(result.errors, result.errors)

    def test_243_latest_review_rejection_duplicate_round_and_stale_hash_fail(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            base, data_path, _, _, _ = prepare_243_fixture(root)

            rejected = copy.deepcopy(base)
            append_review(rejected, "GATE_B", round_no=2, status="rejected")
            result = delivery.validate_data(
                rejected,
                strict_status=False,
                data_path=data_path,
                workspace_root=root,
            )
            self.assertTrue(any("GATE_B" in item for item in result.errors), result.errors)
            self.assertEqual("rejected", delivery.latest_review(rejected, "GATE_B")["status"])

            duplicate = copy.deepcopy(base)
            append_review(duplicate, "GATE_B", round_no=1)
            result = delivery.validate_data(
                duplicate,
                strict_status=False,
                data_path=data_path,
                workspace_root=root,
            )
            self.assertTrue(any("round" in item for item in result.errors), result.errors)

            stale = copy.deepcopy(base)
            stale["shots"][0]["notes"] += " 审核后改动。"
            delivery.derive_prompts(stale)
            result = delivery.validate_data(
                stale,
                strict_status=False,
                data_path=data_path,
                workspace_root=root,
            )
            self.assertTrue(any("reviewed_hash" in item or "GATE_B" in item for item in result.errors), result.errors)

    def test_243_final_signoff_requires_current_gate_c_hash(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data, data_path, _, _, _ = prepare_243_fixture(root)
            missing = delivery.validate_data(
                data,
                strict_status=False,
                final_signoff=True,
                data_path=data_path,
                workspace_root=root,
            )
            self.assertTrue(any("GATE_C" in item for item in missing.errors), missing.errors)

            append_review(data, "GATE_C")
            result = delivery.validate_data(
                data,
                strict_status=False,
                final_signoff=True,
                data_path=data_path,
                workspace_root=root,
            )
            self.assertFalse(result.errors, result.errors)
            delivery.update_validation_report(data, result, final_signoff=True)
            strict = delivery.validate_data(
                data,
                strict_status=True,
                final_signoff=True,
                data_path=data_path,
                workspace_root=root,
            )
            self.assertFalse(strict.errors, strict.errors)

            data["shots"][0]["notes"] += " Gate C 后改动。"
            delivery.derive_prompts(data)
            stale = delivery.validate_data(
                data,
                strict_status=False,
                final_signoff=True,
                data_path=data_path,
                workspace_root=root,
            )
            self.assertTrue(any("GATE_C" in item or "reviewed_hash" in item for item in stale.errors), stale.errors)

    def test_243_batch_plan_gate_binding_and_cycle_checks(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            approved.write_text(valid_data_243()["script_lock"]["locked_text"], encoding="utf-8")

            data = valid_data_243()
            data["batch_plan"] = {
                "batches": [
                    {
                        "batch_id": "BT01",
                        "scene_ids": ["S01"],
                        "expected_shot_count": 2,
                        "depends_on": [],
                    }
                ]
            }
            append_review(data, "GATE_0")
            append_review(data, "GATE_A", batch_id="BT01")
            append_review(data, "GATE_B", batch_id="BT01")
            delivery.derive_prompts(data)
            result = delivery.validate_data(
                data,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            if result.warnings:
                resolve_warnings(data, result)
                result = delivery.validate_data(
                    data,
                    strict_status=False,
                    data_path=root / "sample.shot_data.json",
                    workspace_root=root,
                )
            self.assertFalse(result.errors, result.errors)

            missing_batch_review = copy.deepcopy(data)
            missing_batch_review["human_reviews"] = [
                review for review in missing_batch_review["human_reviews"] if review["gate"] != "GATE_B"
            ]
            result = delivery.validate_data(
                missing_batch_review,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(any("GATE_B" in item for item in result.errors), result.errors)

            cyclic = valid_data_243()
            cyclic["batch_plan"] = {
                "batches": [
                    {
                        "batch_id": "BT01",
                        "scene_ids": ["S01"],
                        "expected_shot_count": 2,
                        "depends_on": ["BT01"],
                    }
                ]
            }
            append_review(cyclic, "GATE_0")
            append_review(cyclic, "GATE_A", batch_id="BT01")
            append_review(cyclic, "GATE_B", batch_id="BT01")
            delivery.derive_prompts(cyclic)
            result = delivery.validate_data(
                cyclic,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(any("自指" in item or "循环" in item for item in result.errors), result.errors)

            empty_plan = valid_data_243()
            empty_plan["batch_plan"] = {}
            append_review(empty_plan, "GATE_A")
            append_review(empty_plan, "GATE_B")
            delivery.derive_prompts(empty_plan)
            result = delivery.validate_data(
                empty_plan,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(any("batch_plan.batches" in item for item in result.errors), result.errors)

            gate_order = valid_data_243()
            gate_order["batch_plan"] = copy.deepcopy(data["batch_plan"])
            append_review(gate_order, "GATE_A", batch_id="BT01")
            append_review(gate_order, "GATE_0")
            append_review(gate_order, "GATE_B", batch_id="BT01")
            delivery.derive_prompts(gate_order)
            result = delivery.validate_data(
                gate_order,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(any("GATE_0" in item and "GATE_A" in item for item in result.errors), result.errors)

            unused_scene = valid_data_243()
            unused_scene["batch_plan"] = {
                "batches": [
                    {
                        "batch_id": "BT01",
                        "scene_ids": ["S01", "S99"],
                        "expected_shot_count": 2,
                        "depends_on": [],
                    }
                ]
            }
            append_review(unused_scene, "GATE_0")
            append_review(unused_scene, "GATE_A", batch_id="BT01")
            append_review(unused_scene, "GATE_B", batch_id="BT01")
            delivery.derive_prompts(unused_scene)
            result = delivery.validate_data(
                unused_scene,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(any("S99" in item for item in result.errors), result.errors)

    def test_243_safety_exception_must_bind_exact_shot(self) -> None:
        def safety_data(approved_item: str) -> dict:
            data = valid_data_243()
            safety = copy.deepcopy(data["shots"][0])
            safety.update(
                {
                    "shot_no": 3,
                    "covered_fact_ids": [],
                    "shot_type": "safety",
                    "notes": "[安全镜][人工批准] 无事实安全镜。",
                    "camera_main_image": "[平视, 空镜全景, 固定镜头]\n【机位逻辑】摄影机看向门框。\n空房间维持静止。",
                    "visible_characters": [],
                    "continuity_updates": [],
                }
            )
            data["shots"].append(safety)
            append_review(data, "GATE_A")
            append_review(data, "GATE_B", approved_items=[approved_item])
            delivery.derive_prompts(data)
            return data

        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            approved.write_text(valid_data_243()["script_lock"]["locked_text"], encoding="utf-8")

            wrong = safety_data("shot:2:safety")
            result = delivery.validate_data(
                wrong,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(any("covered_fact_ids" in item for item in result.errors), result.errors)

            right = safety_data("shot:3:safety")
            result = delivery.validate_data(
                right,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertFalse(any("covered_fact_ids" in item for item in result.errors), result.errors)

    def test_243_axis_and_required_adjacent_exceptions_bind_exact_shots(self) -> None:
        def source_ready_root(root: Path) -> None:
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            approved.write_text(valid_data_243()["script_lock"]["locked_text"], encoding="utf-8")

        def axis_data(token: str) -> dict:
            data = valid_data_243()
            data["shots"][0]["camera_main_image"] = (
                "[平视, 中景, 缓慢推进]\n【机位逻辑】摄影机向A推进。\n"
                "【场景首镜站位】A站在门口，面向桌边。\nA站在门口。"
            )
            data["shots"][1]["camera_main_image"] = (
                "[平视, 中景, 缓慢拉出]\n【机位逻辑】摄影机从A身前拉出。\n"
                "【站位位移】A从门口走到桌边，面向B的位置。\nA在桌边停下，说：“到了。”"
            )
            data["shots"][1]["notes"] = "[反转动机] 揭示后拉出暴露空间。"
            append_review(data, "GATE_A")
            append_review(data, "GATE_B", approved_items=[token])
            delivery.derive_prompts(data)
            return data

        def adjacent_data(token: str) -> dict:
            data = valid_data_243()
            data["shots"][1]["camera_main_image"] = (
                "[平视, 中景, 固定镜头]\n【机位逻辑】摄影机仍在桌边看向门口。\n"
                "A站在门口，手扶门框。"
            )
            data["shots"][1]["continuity_updates"] = []
            data["shots"][1]["notes"] = "[必拆相邻] 独立事实落点。"
            append_review(data, "GATE_A")
            append_review(data, "GATE_B", approved_items=[token])
            delivery.derive_prompts(data)
            return data

        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            source_ready_root(root)
            data_path = root / "sample.shot_data.json"
            cases = [
                ("axis wrong", axis_data("shot:1:axis-reversal"), "axis-reversal", True),
                ("axis right", axis_data("shot:2:axis-reversal"), "axis-reversal", False),
                ("adjacent wrong", adjacent_data("shot:1:required-adjacent"), "required-adjacent", True),
                ("adjacent right", adjacent_data("shot:2:required-adjacent"), "required-adjacent", False),
            ]
            for name, data, token, should_error in cases:
                with self.subTest(name=name):
                    result = delivery.validate_data(
                        data,
                        strict_status=False,
                        data_path=data_path,
                        workspace_root=root,
                    )
                    binding_errors = [item for item in result.errors if "approved_items 精确批准" in item]
                    self.assertEqual(should_error, bool(binding_errors), result.errors)

    def test_243_revise_resolution_cannot_close_a_warning_that_still_exists(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            data = valid_data_243()
            approved.write_text(data["script_lock"]["locked_text"], encoding="utf-8")
            data["shots"][0]["camera_main_image"] += "\n" + "A反复确认门口、桌边和墙面。" * 80
            append_review(data, "GATE_A")
            append_review(data, "GATE_B")
            delivery.derive_prompts(data)
            first = delivery.validate_data(
                data,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(first.warnings, first.errors)
            data["warn_resolutions"] = [
                {
                    "warn_id": delivery.warning_id(message),
                    "resolution": "revise",
                    "resolved_by": "human",
                    "note": "记录已修改，但当前数据仍保留原警告。",
                }
                for message in first.warnings
            ]
            result = delivery.validate_data(
                data,
                strict_status=False,
                final_signoff=True,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(any("revise" in item and "仍存在" in item for item in result.errors), result.errors)

    def test_243_nonempty_unique_references_visibility_and_note_contract(self) -> None:
        mutators = {
            "empty continuity_logs": (lambda data: data.update(continuity_logs=[]), "continuity_logs"),
            "empty beats": (lambda data: data.update(beats=[]), "beats"),
            "empty shots": (lambda data: data.update(shots=[]), "shots"),
            "empty fact text": (lambda data: data["beats"][0]["facts"][0].update(text=""), "fact.text"),
            "duplicate fact": (
                lambda data: data["beats"][0]["facts"].append(copy.deepcopy(data["beats"][0]["facts"][0])),
                "事实 ID 重复",
            ),
            "unknown fact": (lambda data: data["shots"][0].update(covered_fact_ids=["B001-F99"]), "不存在的事实"),
            "unknown beat": (lambda data: data["shots"][0].update(beat_ids=["B999"]), "beat"),
            "unknown scene": (lambda data: data["shots"][0].update(scene_id="S99"), "scene_id"),
            "beat display scene mismatch": (
                lambda data: data["beats"][0].update(scene="2 室外 夜 外"),
                "continuity_logs[S01].scene 不一致",
            ),
            "shot display scene mismatch": (
                lambda data: data["shots"][0].update(scene="2 室外 夜 外"),
                "continuity_logs[S01].scene 不一致",
            ),
            "visible offscreen overlap": (
                lambda data: data["shots"][0].update(offscreen_characters=["A"]),
                "同时可见",
            ),
            "unknown note marker": (lambda data: data["shots"][0].update(notes="[未知标签] 测试"), "标签"),
            "non-contiguous beat": (
                lambda data: data["shots"][1].update(beat_ids=["B001", "B003"]),
                "[非连续Beat]",
            ),
            "non-contiguous beat without reason": (
                lambda data: data["shots"][1].update(beat_ids=["B001", "B003"], notes="[非连续Beat]"),
                "原因",
            ),
            "foreign evidence fact": (
                lambda data: data["shots"][1]["continuity_updates"][0].update(evidence_fact_ids=["B001-F01"]),
                "证据",
            ),
        }
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            approved.write_text(valid_data_243()["script_lock"]["locked_text"], encoding="utf-8")
            for name, (mutate, expected) in mutators.items():
                with self.subTest(name=name):
                    data = valid_data_243()
                    mutate(data)
                    append_review(data, "GATE_A")
                    append_review(data, "GATE_B")
                    delivery.derive_prompts(data)
                    result = delivery.validate_data(
                        data,
                        strict_status=False,
                        data_path=root / "sample.shot_data.json",
                        workspace_root=root,
                    )
                    self.assertTrue(
                        any(expected.lower() in item.lower() for item in result.errors),
                        result.errors,
                    )

    def test_243_duration_fields_reject_non_integer_non_finite_and_bool(self) -> None:
        values = [2.5, "2", True, math.nan, math.inf]
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            approved.write_text(valid_data_243()["script_lock"]["locked_text"], encoding="utf-8")
            for field in ["duration_seconds", "sync_action_seconds"]:
                for value in values:
                    with self.subTest(field=field, value=repr(value)):
                        data = valid_data_243()
                        if field == "duration_seconds":
                            data["shots"][0][field] = value
                        else:
                            data["shots"][0]["duration_breakdown"][field] = value
                        append_review(data, "GATE_A")
                        append_review(data, "GATE_B")
                        delivery.derive_prompts(data)
                        result = delivery.validate_data(
                            data,
                            strict_status=False,
                            data_path=root / "sample.shot_data.json",
                            workspace_root=root,
                        )
                        self.assertTrue(any("整数" in item for item in result.errors), result.errors)

    def test_243_strict_report_hash_status_warnings_and_errors_are_recomputed(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            base, data_path, _, _, _ = prepare_243_fixture(root)
            cases = {
                "missing hash": lambda data: data["validation_report"].pop("source_json_hash", None),
                "fake hash": lambda data: data["validation_report"].update(source_json_hash="0" * 64),
                "fake status": lambda data: data["validation_report"].update(status="FAIL"),
                "fake warnings": lambda data: data["validation_report"].update(warnings=[{"warn_id": "W-fake", "message": "fake"}]),
                "fake errors": lambda data: data["validation_report"].update(errors=["fake"]),
                "fake scene counts": lambda data: data["validation_report"].update(scene_counts={"S99": 99}),
                "fake total duration": lambda data: data["validation_report"].update(total_duration_seconds=999),
                "fake duration thresholds": lambda data: data["validation_report"].update(
                    shots_over_6_seconds=99,
                    shots_over_8_seconds=99,
                    shots_over_10_seconds=99,
                ),
            }
            for name, mutate in cases.items():
                with self.subTest(name=name):
                    data = copy.deepcopy(base)
                    mutate(data)
                    result = delivery.validate_data(
                        data,
                        strict_status=True,
                        data_path=data_path,
                        workspace_root=root,
                    )
                    self.assertTrue(any("validation_report" in item for item in result.errors), result.errors)

    def test_legacy_242_is_validate_only_but_remains_readable(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data_path = root / "legacy.shot_data.json"
            markdown_path = root / "legacy.md"
            excel_path = root / "legacy.xlsx"
            report_path = root / "legacy.validation_report.json"
            data = valid_data_242()
            delivery.derive_prompts(data)
            result = delivery.validate_data(data, strict_status=False)
            if result.warnings:
                resolve_warnings(data, result)
                result = delivery.validate_data(data, strict_status=False)
            self.assertFalse(result.errors, result.errors)
            delivery.update_validation_report(data, result)
            delivery.write_json(data_path, data)
            delivery.build_markdown(data, markdown_path)
            delivery.build_excel(data, excel_path)
            delivery.build_report(data, report_path)
            rc = delivery.main(
                [
                    "validate",
                    "--data",
                    str(data_path),
                    "--markdown",
                    str(markdown_path),
                    "--excel",
                    str(excel_path),
                    "--report",
                    str(report_path),
                ]
            )
            self.assertEqual(0, rc)

    def test_legacy_242_duration_compatibility_does_not_inherit_integer_rule(self) -> None:
        data = valid_data_242()
        data["shots"][0]["duration_seconds"] = 2.5
        data["shots"][0]["duration_breakdown"] = {
            "sync_action_seconds": 1.5,
            "sync_dialogue_seconds": 0,
            "non_sync_action_seconds": 0,
            "emotional_pause_seconds": 1,
        }
        delivery.derive_prompts(data)
        result = delivery.validate_data(data, strict_status=False)
        self.assertFalse(any("JSON 整数" in item for item in result.errors), result.errors)

    def test_243_full_build_gate_c_validate_and_node_wrapper_flow(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data, data_path, markdown_path, excel_path, report_path = prepare_243_fixture(root)
            delivery.write_json(data_path, data)
            common = [
                "--data",
                str(data_path),
                "--markdown",
                str(markdown_path),
                "--excel",
                str(excel_path),
                "--report",
                str(report_path),
                "--workspace-root",
                str(root),
            ]
            self.assertEqual(0, delivery.main(["build", *common]))

            built = delivery.load_json(data_path)
            append_review(built, "GATE_C")
            delivery.write_json(data_path, built)
            self.assertEqual(0, delivery.main(["build", *common, "--final-signoff"]))
            self.assertEqual(0, delivery.main(["validate", *common, "--final-signoff"]))

            process = subprocess.run(
                [
                    "node",
                    str(SCRIPT_DIR / "validate_storyboard.js"),
                    "--python",
                    sys.executable,
                    *common,
                    "--final-signoff",
                ],
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertEqual(0, process.returncode, process.stderr + process.stdout)
            final = delivery.load_json(data_path)
            self.assertTrue(all("keyframe" not in shot for shot in final["shots"]))
            workbook = load_workbook(excel_path, read_only=True)
            try:
                self.assertEqual(7, workbook[delivery.SHEET_NAME].max_column)
            finally:
                workbook.close()

    def test_243_build_requires_report_without_mutating_input(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data, data_path, markdown_path, excel_path, _ = prepare_243_fixture(root)
            original = json.dumps(data, ensure_ascii=False, indent=2) + "\n"
            data_path.write_text(original, encoding="utf-8")
            rc = delivery.main(
                [
                    "build",
                    "--data",
                    str(data_path),
                    "--markdown",
                    str(markdown_path),
                    "--excel",
                    str(excel_path),
                    "--workspace-root",
                    str(root),
                ]
            )
            self.assertEqual(1, rc)
            self.assertEqual(original, data_path.read_text(encoding="utf-8"))
            self.assertFalse(markdown_path.exists())
            self.assertFalse(excel_path.exists())

    def test_243_cli_and_node_require_explicit_workspace_root_without_mutation(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data, data_path, markdown_path, excel_path, report_path = prepare_243_fixture(root)
            delivery.write_json(data_path, data)
            original = data_path.read_bytes()
            without_root = [
                "--data",
                str(data_path),
                "--markdown",
                str(markdown_path),
                "--excel",
                str(excel_path),
                "--report",
                str(report_path),
            ]
            self.assertEqual(1, delivery.main(["build", *without_root]))
            self.assertEqual(original, data_path.read_bytes())
            self.assertFalse(markdown_path.exists())
            self.assertFalse(excel_path.exists())

            process = subprocess.run(
                [
                    "node",
                    str(SCRIPT_DIR / "validate_storyboard.js"),
                    "--python",
                    sys.executable,
                    *without_root,
                ],
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertNotEqual(0, process.returncode, process.stderr + process.stdout)
            self.assertEqual(original, data_path.read_bytes())

    def test_243_pre_signoff_warn_can_be_kept_then_finally_signed(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            data = valid_data_243()
            approved.write_text(data["script_lock"]["locked_text"], encoding="utf-8")
            data["shots"][0]["camera_main_image"] += "\n" + "A反复确认门口、桌边和墙面。" * 80
            append_review(data, "GATE_A")
            append_review(data, "GATE_B")
            data_path = root / "sample.shot_data.json"
            markdown_path = root / "sample.md"
            excel_path = root / "sample.xlsx"
            report_path = root / "sample.validation_report.json"
            delivery.write_json(data_path, data)
            common = [
                "--data",
                str(data_path),
                "--markdown",
                str(markdown_path),
                "--excel",
                str(excel_path),
                "--report",
                str(report_path),
                "--workspace-root",
                str(root),
            ]
            self.assertEqual(0, delivery.main(["build", *common]))
            built = delivery.load_json(data_path)
            self.assertEqual("WARN", built["validation_report"]["status"])
            built["warn_resolutions"] = [
                {
                    "warn_id": delivery.warning_id(message),
                    "resolution": "keep",
                    "resolved_by": "human",
                    "note": "用户确认保留当前节奏和镜头信息。",
                }
                for message in built["validation_report"]["warnings"]
            ]
            append_review(built, "GATE_C")
            delivery.write_json(data_path, built)
            self.assertEqual(0, delivery.main(["build", *common, "--final-signoff"]))
            self.assertEqual(0, delivery.main(["validate", *common, "--final-signoff"]))

    def test_243_markdown_excel_special_text_and_numeric_round_trip(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data = valid_data_243()
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            approved.write_text(data["script_lock"]["locked_text"], encoding="utf-8")
            data["shots"][0]["notes"] = "=字面量<br>、竖线 | 与\n换行"
            append_review(data, "GATE_A")
            append_review(data, "GATE_B")
            delivery.derive_prompts(data)
            result = delivery.validate_data(
                data,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            if result.warnings:
                resolve_warnings(data, result)
                result = delivery.validate_data(
                    data,
                    strict_status=False,
                    data_path=root / "sample.shot_data.json",
                    workspace_root=root,
                )
            self.assertFalse(result.errors, result.errors)
            delivery.update_validation_report(data, result)
            data_path = root / "sample.shot_data.json"
            markdown_path = root / "sample.md"
            excel_path = root / "sample.xlsx"
            report_path = root / "sample.validation_report.json"
            delivery.write_json(data_path, data)
            self.assertEqual(
                0,
                delivery.main(
                    [
                        "build",
                        "--data",
                        str(data_path),
                        "--markdown",
                        str(markdown_path),
                        "--excel",
                        str(excel_path),
                        "--report",
                        str(report_path),
                        "--workspace-root",
                        str(root),
                    ]
                ),
            )
            self.assertEqual(
                0,
                delivery.main(
                    [
                        "validate",
                        "--data",
                        str(data_path),
                        "--markdown",
                        str(markdown_path),
                        "--excel",
                        str(excel_path),
                        "--report",
                        str(report_path),
                        "--workspace-root",
                        str(root),
                    ]
                ),
            )
            workbook = load_workbook(excel_path, read_only=False, data_only=False)
            try:
                note_cell = workbook[delivery.SHEET_NAME]["F2"]
                self.assertEqual("s", note_cell.data_type)
                self.assertEqual(data["shots"][0]["notes"], note_cell.value)
                self.assertEqual(2, workbook[delivery.SHEET_NAME]["D2"].value)
            finally:
                workbook.close()

    def test_243_corrupt_excel_markdown_and_json_fail_without_traceback(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data, data_path, markdown_path, excel_path, report_path = prepare_243_fixture(root)
            delivery.write_json(data_path, data)
            common = [
                "--data",
                str(data_path),
                "--markdown",
                str(markdown_path),
                "--excel",
                str(excel_path),
                "--report",
                str(report_path),
                "--workspace-root",
                str(root),
            ]
            self.assertEqual(0, delivery.main(["build", *common]))
            excel_path.write_bytes(b"not-an-xlsx")
            self.assertEqual(1, delivery.main(["validate", *common]))

            self.assertEqual(0, delivery.main(["build", *common]))
            markdown_path.write_bytes(b"\xff\xfe\x00")
            self.assertEqual(1, delivery.main(["validate", *common]))

            self.assertEqual(0, delivery.main(["build", *common]))
            data_path.write_bytes(b"{")
            self.assertEqual(1, delivery.main(["validate", *common]))

    def test_243_atomic_build_failure_preserves_all_formal_files(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data, data_path, markdown_path, excel_path, report_path = prepare_243_fixture(root)
            delivery.write_json(data_path, data)
            markdown_path.write_text("old-markdown", encoding="utf-8")
            excel_path.write_bytes(b"old-excel")
            report_path.write_text("old-report", encoding="utf-8")
            before = {
                data_path: data_path.read_bytes(),
                markdown_path: markdown_path.read_bytes(),
                excel_path: excel_path.read_bytes(),
                report_path: report_path.read_bytes(),
            }
            with mock.patch.object(delivery, "build_excel", side_effect=RuntimeError("simulated Excel failure")):
                rc = delivery.main(
                    [
                        "build",
                        "--data",
                        str(data_path),
                        "--markdown",
                        str(markdown_path),
                        "--excel",
                        str(excel_path),
                        "--report",
                        str(report_path),
                        "--workspace-root",
                        str(root),
                    ]
                )
            self.assertEqual(1, rc)
            for path, content in before.items():
                self.assertEqual(content, path.read_bytes(), path)

            original_compare_markdown = delivery.compare_markdown

            def fail_temporary_self_check(data: dict, path: Path, result: delivery.ValidationResult) -> None:
                original_compare_markdown(data, path, result)
                result.error("simulated temporary self-check failure")

            with mock.patch.object(delivery, "compare_markdown", side_effect=fail_temporary_self_check):
                rc = delivery.main(
                    [
                        "build",
                        "--data",
                        str(data_path),
                        "--markdown",
                        str(markdown_path),
                        "--excel",
                        str(excel_path),
                        "--report",
                        str(report_path),
                        "--workspace-root",
                        str(root),
                    ]
                )
            self.assertEqual(1, rc)
            for path, content in before.items():
                self.assertEqual(content, path.read_bytes(), path)

            real_replace = delivery.os.replace
            formal_paths = {path.resolve(strict=False) for path in before}
            commit_count = 0

            def fail_second_commit(source: str | Path, destination: str | Path) -> None:
                nonlocal commit_count
                if Path(destination).resolve(strict=False) in formal_paths:
                    commit_count += 1
                    if commit_count == 2:
                        raise PermissionError("simulated second commit failure")
                real_replace(source, destination)

            with mock.patch.object(delivery.os, "replace", side_effect=fail_second_commit):
                rc = delivery.main(
                    [
                        "build",
                        "--data",
                        str(data_path),
                        "--markdown",
                        str(markdown_path),
                        "--excel",
                        str(excel_path),
                        "--report",
                        str(report_path),
                        "--workspace-root",
                        str(root),
                    ]
                )
            self.assertEqual(1, rc)
            self.assertGreaterEqual(commit_count, 2)
            for path, content in before.items():
                self.assertEqual(content, path.read_bytes(), path)

    def test_excel_formula_prefixes_are_always_written_as_text(self) -> None:
        for prefix in ("=", "+", "-", "@"):
            with self.subTest(prefix=prefix), tempfile.TemporaryDirectory() as root_name:
                data = valid_data_243()
                data["shots"][0]["notes"] = prefix + "普通文本"
                delivery.derive_prompts(data)
                path = Path(root_name) / "prefix.xlsx"
                delivery.build_excel(data, path)
                workbook = load_workbook(path, read_only=False, data_only=False)
                try:
                    cell = workbook[delivery.SHEET_NAME]["F2"]
                    self.assertEqual("s", cell.data_type)
                    self.assertEqual(prefix + "普通文本", cell.value)
                finally:
                    workbook.close()


class P2ContractTests(unittest.TestCase):
    def validate_current(self, data: dict, root: Path) -> delivery.ValidationResult:
        approved = root / "approved" / "sample.approved_script.txt"
        approved.parent.mkdir(parents=True, exist_ok=True)
        approved.write_text(data["script_lock"]["locked_text"], encoding="utf-8")
        data["human_reviews"] = []
        append_review(data, "GATE_A")
        append_review(data, "GATE_B")
        delivery.derive_prompts(data)
        return delivery.validate_data(
            data,
            strict_status=False,
            data_path=root / "sample.shot_data.json",
            workspace_root=root,
        )

    def test_p2_duration_formula_uses_max_for_synchronized_action_and_dialogue(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            valid = valid_data_243()
            valid["shots"][1]["duration_seconds"] = 3
            valid["shots"][1]["duration_breakdown"] = {
                "sync_action_seconds": 2,
                "sync_dialogue_seconds": 2,
                "non_sync_action_seconds": 0,
                "emotional_pause_seconds": 1,
            }
            result = self.validate_current(valid, root)
            self.assertFalse(any("duration" in item.lower() or "时长" in item for item in result.errors), result.errors)

            summed = valid_data_243()
            summed["shots"][1]["duration_seconds"] = 5
            summed["shots"][1]["duration_breakdown"] = {
                "sync_action_seconds": 2,
                "sync_dialogue_seconds": 2,
                "non_sync_action_seconds": 0,
                "emotional_pause_seconds": 1,
            }
            result = self.validate_current(summed, root)
            self.assertTrue(any("max(" in item or "时长公式" in item for item in result.errors), result.errors)

    def test_p2_beat_order_inserts_without_renumbering_old_ids(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data = valid_data_243()
            by_id = {beat["beat_id"]: beat for beat in data["beats"]}
            by_id["B001"]["beat_order"] = "1"
            by_id["B003"]["beat_order"] = "1.5"
            by_id["B002"]["beat_order"] = "2"
            data["beats"] = [by_id["B001"], by_id["B003"], by_id["B002"]]
            data["shots"][1]["beat_ids"] = ["B003", "B002"]
            data["shots"][1]["covered_fact_ids"] = ["B003-F01", "B002-F01"]
            result = self.validate_current(data, root)
            self.assertFalse(any("beat_order" in item or "非连续 Beat" in item for item in result.errors), result.errors)
            self.assertEqual(["B001", "B003", "B002"], [beat["beat_id"] for beat in data["beats"]])
            positions = delivery.beat_order_index(data)
            self.assertEqual("B003+B002", delivery.format_beat_ids(["B003", "B002"], positions))

    def test_p2_beat_order_rejects_duplicates_noncanonical_values_and_wrong_order(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            cases: list[tuple[str, object]] = [
                ("leading zero", "01"),
                ("trailing zero", "1.0"),
                ("exponent", "1e1"),
                ("number", 1),
                ("zero", "0"),
            ]
            for name, value in cases:
                with self.subTest(name=name):
                    data = valid_data_243()
                    data["beats"][0]["beat_order"] = value
                    result = self.validate_current(data, root)
                    self.assertTrue(any("beat_order" in item for item in result.errors), result.errors)

            duplicate = valid_data_243()
            duplicate["beats"][1]["beat_order"] = duplicate["beats"][0]["beat_order"]
            result = self.validate_current(duplicate, root)
            self.assertTrue(any("beat_order 重复" in item for item in result.errors), result.errors)

            array_out_of_order = valid_data_243()
            array_out_of_order["beats"][0], array_out_of_order["beats"][1] = (
                array_out_of_order["beats"][1],
                array_out_of_order["beats"][0],
            )
            result = self.validate_current(array_out_of_order, root)
            self.assertTrue(any("beat_order 严格递增" in item for item in result.errors), result.errors)

            shot_out_of_order = valid_data_243()
            shot_out_of_order["shots"][1]["beat_ids"] = ["B003", "B002"]
            shot_out_of_order["shots"][1]["covered_fact_ids"] = ["B003-F01", "B002-F01"]
            result = self.validate_current(shot_out_of_order, root)
            self.assertTrue(any("beat_ids" in item and "beat_order" in item for item in result.errors), result.errors)

    def test_p2_b1000_ids_and_range_rendering_follow_narrative_order(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data = valid_data_243()
            last = data["beats"][2]
            last["beat_id"] = "B1000"
            last["facts"][0]["fact_id"] = "B1000-F01"
            data["shots"][1]["beat_ids"] = ["B002", "B1000"]
            data["shots"][1]["covered_fact_ids"] = ["B002-F01", "B1000-F01"]
            result = self.validate_current(data, root)
            self.assertFalse(any("Beat ID 不合法" in item or "事实 ID" in item for item in result.errors), result.errors)

        positions = {"B999": 0, "B1000": 1, "B1001": 2, "B1003": 3}
        self.assertEqual("B999-B1000", delivery.format_beat_ids(["B999", "B1000"], positions))
        self.assertEqual("B1000+B999", delivery.format_beat_ids(["B1000", "B999"], {"B1000": 0, "B999": 1}))
        self.assertEqual("B1001+B1003", delivery.format_beat_ids(["B1001", "B1003"], positions))

    def test_p2_json_pointer_inheritance_supports_nested_dicts_and_lists(self) -> None:
        parent = copy.deepcopy(valid_data_243()["continuity_logs"][0])
        parent["nested"] = {"camera": {"lens": "35mm"}}
        child = copy.deepcopy(parent)
        child.update(
            {
                "scene_id": "S02",
                "scene": "2 室内 日 内",
                "inherits_from": "S01",
                "inherited_states": ["/characters/0/position", "/nested/camera/lens"],
                "diverged_states": ["/characters/0/facing"],
            }
        )
        child["characters"][0]["facing"] = "窗边"
        data = {"metadata": {"version": "2.4.3"}, "continuity_logs": [parent, child]}
        result = delivery.ValidationResult()
        logs = delivery.validate_continuity_logs(data, result)
        self.assertFalse(result.errors, result.errors)
        self.assertEqual("/characters", delivery.normalize_json_pointer("characters"))
        self.assertEqual("门口", delivery.resolve_json_pointer(logs["S02"], "/characters/0/position"))
        self.assertEqual("35mm", delivery.resolve_json_pointer(logs["S02"], "/nested/camera/lens"))
        escaped = {"a/b": {"~key": ["ok"]}}
        self.assertEqual("ok", delivery.resolve_json_pointer(escaped, "/a~1b/~0key/0"))
        with self.assertRaises(ValueError):
            delivery.normalize_json_pointer("characters.0")

    def test_p2_json_pointer_alias_overlap_missing_path_and_value_mismatch_fail(self) -> None:
        def logs() -> tuple[dict, dict]:
            parent = copy.deepcopy(valid_data_243()["continuity_logs"][0])
            parent["nested"] = {"camera": {"lens": "35mm"}}
            child = copy.deepcopy(parent)
            child.update(
                {
                    "scene_id": "S02",
                    "scene": "2 室内 日 内",
                    "inherits_from": "S01",
                    "inherited_states": [],
                    "diverged_states": [],
                }
            )
            return parent, child

        parent, child = logs()
        child["inherited_states"] = ["characters"]
        child["diverged_states"] = ["/characters/0/position"]
        result = delivery.ValidationResult()
        delivery.validate_continuity_logs(
            {"metadata": {"version": "2.4.3"}, "continuity_logs": [parent, child]}, result
        )
        self.assertTrue(any("祖先或子路径冲突" in item for item in result.errors), result.errors)

        parent, child = logs()
        child["inherited_states"] = ["/characters/9/position"]
        result = delivery.ValidationResult()
        delivery.validate_continuity_logs(
            {"metadata": {"version": "2.4.3"}, "continuity_logs": [parent, child]}, result
        )
        self.assertTrue(any("继承路径" in item and "同时存在" in item for item in result.errors), result.errors)

        parent, child = logs()
        child["inherited_states"] = ["/nested/camera/lens"]
        child["nested"]["camera"]["lens"] = "50mm"
        result = delivery.ValidationResult()
        pointer_data = {"metadata": {"version": "2.4.3"}, "continuity_logs": [parent, child]}
        logs_by_scene = delivery.validate_continuity_logs(pointer_data, result)
        delivery.validate_final_inherited_states(
            pointer_data,
            logs_by_scene,
            {scene: copy.deepcopy(log) for scene, log in logs_by_scene.items()},
            result,
        )
        self.assertTrue(any("/nested/camera/lens" in item and "终态" in item for item in result.errors), result.errors)

        parent, child = logs()
        child["inherited_states"] = "characters"
        result = delivery.ValidationResult()
        delivery.validate_continuity_logs(
            {"metadata": {"version": "2.4.3"}, "continuity_logs": [parent, child]}, result
        )
        self.assertTrue(any("inherited_states" in item and "数组" in item for item in result.errors), result.errors)

        parent, child = logs()
        child["inherited_states"] = ["characters", "/characters"]
        result = delivery.ValidationResult()
        delivery.validate_continuity_logs(
            {"metadata": {"version": "2.4.3"}, "continuity_logs": [parent, child]}, result
        )
        self.assertTrue(any("规范化后路径重复" in item for item in result.errors), result.errors)

        parent, _child = logs()
        parent["inherited_states"] = ["/characters"]
        result = delivery.ValidationResult()
        delivery.validate_continuity_logs(
            {"metadata": {"version": "2.4.3"}, "continuity_logs": [parent]}, result
        )
        self.assertTrue(any("inherits_from 为空" in item for item in result.errors), result.errors)

    def test_p2_multi_character_first_shot_requires_position_fact_in_that_shot(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data = valid_data_243()
            log = data["continuity_logs"][0]
            log["first_shot_anchor_type"] = "multi_character"
            log["spatial_axis"] = "A在画面左侧，B在画面右侧，两人面向桌边。"
            log["characters"] = [
                {"name": "A", "position": "画面左侧", "facing": "桌边"},
                {"name": "B", "position": "画面右侧", "facing": "桌边"},
            ]
            first_fact = data["beats"][0]["facts"][0]
            first_fact["type"] = "action"
            first_fact["cut_category"] = "action"
            data["shots"][0]["visible_characters"] = ["A", "B"]
            data["shots"][0]["camera_main_image"] = (
                "[平视, 全景, 固定镜头]\n【机位逻辑】摄影机正对室内两人。\n"
                "【场景首镜站位】A在画面左侧，B在画面右侧，两人面向桌边。\n"
                "A站在门口，B站在桌边。"
            )
            # B002-F01 remains a position fact, but it is not covered until shot 2.
            self.assertEqual("position", data["beats"][1]["facts"][0]["type"])
            result = self.validate_current(data, root)
            self.assertTrue(
                any("首镜" in item and "position" in item for item in result.errors),
                result.errors,
            )

    def test_p2_generic_cut_regex_has_no_project_terms_but_project_lexicon_can_enhance_it(self) -> None:
        old_project_terms = "魂钉祭池锁魂柱黑雾铜镯"
        self.assertEqual([], delivery.find_required_cut_point_categories(old_project_terms))

        data = {
            "metadata": {
                "version": "2.4.3",
                "project_lexicon": {
                    "prop_terms": ["魂钉"],
                    "space_terms": ["祭池"],
                    "vfx_terms": ["黑雾"],
                    "reality_terms": ["幽界"],
                    "sound_terms": ["魂铃声"],
                },
            }
        }
        result = delivery.ValidationResult()
        delivery.validate_project_lexicon(data, result)
        self.assertFalse(result.errors, result.errors)
        patterns = delivery.project_lexicon_patterns(data)
        categories = delivery.find_required_cut_point_categories(
            old_project_terms, patterns, current_contract=True
        )
        self.assertIn("道具状态变化", categories)
        self.assertIn("空间方向改变", categories)
        self.assertIn("视觉/VFX状态变化", categories)
        layer_and_sound_categories = delivery.find_required_cut_point_categories(
            "幽界魂铃声", patterns, current_contract=True
        )
        self.assertEqual(["层级/声音来源变化"], layer_and_sound_categories)

        legacy = valid_data_242()
        legacy["metadata"]["project_lexicon"] = {"sound_terms": ["魂铃声"]}
        self.assertEqual({}, delivery.project_lexicon_patterns(legacy))
        self.assertEqual(
            ["层级切换"],
            delivery.find_required_cut_point_categories("梦境"),
        )
        legacy["shots"][0]["camera_main_image"] += "\n梦境"
        legacy_audit = delivery.build_overcompression_audit(legacy)
        legacy_shot = next(item for item in legacy_audit["at_risk_shots"] if item["shot_no"] == 1)
        self.assertIn("层级切换", legacy_shot["required_cut_points"])
        self.assertNotIn("层级/声音来源变化", legacy_shot["required_cut_points"])

    def test_p2_project_lexicon_rejects_bad_shapes_and_escapes_regex_terms(self) -> None:
        invalid_values = [
            [],
            {},
            {"unknown_terms": ["词"]},
            {"prop_terms": "魂钉"},
            {"prop_terms": []},
            {"prop_terms": [1]},
            {"prop_terms": ["魂钉", "魂钉"]},
            {"prop_terms": [" 魂钉 "]},
            {"prop_terms": ["魂\n钉"]},
        ]
        for value in invalid_values:
            with self.subTest(value=repr(value)):
                data = {"metadata": {"version": "2.4.3", "project_lexicon": value}}
                result = delivery.ValidationResult()
                delivery.validate_project_lexicon(data, result)
                self.assertTrue(result.errors)

        injected = {
            "metadata": {
                "version": "2.4.3",
                "project_lexicon": {"vfx_terms": [".*"]},
            }
        }
        result = delivery.ValidationResult()
        delivery.validate_project_lexicon(injected, result)
        self.assertFalse(result.errors, result.errors)
        patterns = delivery.project_lexicon_patterns(injected)
        self.assertEqual([], delivery.find_required_cut_point_categories("普通文本", patterns))
        self.assertIn("视觉/VFX状态变化", delivery.find_required_cut_point_categories("出现.*字面词", patterns))

    def test_p2_sound_source_schema_and_state_update_chain(self) -> None:
        valid_log = copy.deepcopy(valid_data_243()["continuity_logs"][0])
        valid_log["sound_sources"] = {
            "SS01": {
                "name": "门铃",
                "visibility": "offscreen",
                "position": "门外",
                "state": "silent",
            }
        }
        result = delivery.ValidationResult()
        self.assertEqual({"SS01"}, delivery.validate_sound_sources(valid_log, "S01", result))
        self.assertFalse(result.errors, result.errors)

        invalid_sources = [
            [],
            {"door": {"name": "门铃", "visibility": "offscreen", "position": "门外", "state": "silent"}},
            {"SS01": []},
            {"SS01": {"name": "门铃", "visibility": "hidden", "position": "门外", "state": "silent"}},
            {"SS01": {"name": "门铃", "visibility": "offscreen", "position": "", "state": "silent"}},
        ]
        for value in invalid_sources:
            with self.subTest(schema=repr(value)):
                log = copy.deepcopy(valid_log)
                log["sound_sources"] = value
                result = delivery.ValidationResult()
                delivery.validate_sound_sources(log, "S01", result)
                self.assertTrue(result.errors)

        def sound_data(*, include_update: bool, entity: str = "SS01", evidence: str = "B001-F02", from_state: str = "silent") -> dict:
            data = valid_data_243()
            data["continuity_logs"][0]["sound_sources"] = copy.deepcopy(valid_log["sound_sources"])
            data["beats"][0]["facts"].append(
                {
                    "fact_id": "B001-F02",
                    "type": "sound",
                    "text": "A站在门口。",
                    "cut_priority": "must_isolate",
                    "cut_reasons": ["new_sound_source"],
                    "cut_group": "B001-sound",
                    "cut_category": "sound",
                    "cut_moment_id": "B001-sound-moment",
                }
            )
            data["shots"][0]["covered_fact_ids"].append("B001-F02")
            data["shots"][0]["shot_type"] = "action"
            data["shots"][0]["split_reason"].append("new_sound_source")
            if include_update:
                data["shots"][0]["continuity_updates"].append(
                    {
                        "entity_type": "sound_source",
                        "entity": entity,
                        "field": "state",
                        "from": from_state,
                        "to": "ringing",
                        "evidence_fact_ids": [evidence],
                    }
                )
            return data

        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            missing = self.validate_current(sound_data(include_update=False), root)
            self.assertTrue(any("new_sound_source" in item and "continuity_update" in item for item in missing.errors), missing.errors)

            unknown = self.validate_current(sound_data(include_update=True, entity="SS99"), root)
            self.assertTrue(any("未知 sound_source" in item for item in unknown.errors), unknown.errors)

            wrong_fact = self.validate_current(sound_data(include_update=True, evidence="B001-F01"), root)
            self.assertTrue(any("sound fact" in item for item in wrong_fact.errors), wrong_fact.errors)

            wrong_chain = self.validate_current(sound_data(include_update=True, from_state="ringing"), root)
            self.assertTrue(any("sound_source/SS01/state" in item and "from" in item for item in wrong_chain.errors), wrong_chain.errors)

            valid = self.validate_current(sound_data(include_update=True), root)
            sound_errors = [
                item
                for item in valid.errors
                if "sound_source" in item or "new_sound_source" in item or "sound fact" in item
            ]
            self.assertFalse(sound_errors, valid.errors)

    def test_p2_gate_latest_rounds_cannot_be_retroactively_reordered(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data = valid_data_243()
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            approved.write_text(data["script_lock"]["locked_text"], encoding="utf-8")
            delivery.derive_prompts(data)
            data["human_reviews"] = []
            append_review(data, "GATE_A", round_no=1)
            append_review(data, "GATE_B", round_no=1)
            append_review(data, "GATE_B", round_no=2, status="rejected")
            append_review(data, "GATE_C", round_no=1)
            append_review(data, "GATE_B", round_no=3)
            result = delivery.validate_data(
                data,
                strict_status=False,
                final_signoff=True,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertTrue(
                any("GATE_C" in item and ("GATE_B" in item or "之前" in item) for item in result.errors),
                result.errors,
            )

    def test_p2_parent_final_state_drives_cross_scene_inheritance(self) -> None:
        parent = copy.deepcopy(valid_data_243()["continuity_logs"][0])
        parent["sound_sources"] = {
            "SS01": {"name": "门铃", "visibility": "offscreen", "position": "门外", "state": "silent"}
        }
        child = copy.deepcopy(parent)
        child.update(
            {
                "scene_id": "S02",
                "scene": "2 室内 日 内",
                "inherits_from": "S01",
                "inherited_states": ["/sound_sources/SS01/state"],
                "diverged_states": [],
            }
        )
        data = {"metadata": {"version": "2.4.3"}, "continuity_logs": [parent, child]}
        logs = {"S01": parent, "S02": child}
        final_documents = {scene: copy.deepcopy(log) for scene, log in logs.items()}
        delivery.apply_update_to_scene_document(
            final_documents["S01"],
            {"entity_type": "sound_source", "entity": "SS01", "field": "state", "to": "ringing"},
        )
        result = delivery.ValidationResult()
        delivery.validate_final_inherited_states(data, logs, final_documents, result)
        self.assertTrue(any("父场景 S01" in item and "终态" in item for item in result.errors), result.errors)
        child["sound_sources"]["SS01"]["state"] = "ringing"
        result = delivery.ValidationResult()
        delivery.validate_final_inherited_states(data, logs, final_documents, result)
        self.assertFalse(result.errors, result.errors)

        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            full = valid_data_243()
            full_parent = full["continuity_logs"][0]
            full_parent["sound_sources"] = copy.deepcopy(parent["sound_sources"])
            full_child = copy.deepcopy(full_parent)
            full_child.update(
                {
                    "scene_id": "S02",
                    "scene": "2 室内 日 内",
                    "inherits_from": "S01",
                    "inherited_states": ["/sound_sources/SS01/state"],
                    "diverged_states": [],
                }
            )
            full["continuity_logs"].append(full_child)
            full["beats"][0]["facts"].append(
                {
                    "fact_id": "B001-F02",
                    "type": "sound",
                    "text": "A站在门口。",
                    "cut_priority": "normal",
                    "cut_reasons": [],
                    "cut_group": "B001-sound",
                    "cut_category": "sound",
                    "cut_moment_id": "B001-sound-moment",
                }
            )
            full["shots"][0]["covered_fact_ids"].append("B001-F02")
            full["shots"][0]["continuity_updates"].append(
                {
                    "entity_type": "sound_source",
                    "entity": "SS01",
                    "field": "state",
                    "from": "silent",
                    "to": "ringing",
                    "evidence_fact_ids": ["B001-F02"],
                }
            )
            integration = self.validate_current(full, root)
            self.assertTrue(any("父场景 S01" in item and "终态" in item for item in integration.errors), integration.errors)
            full_child["sound_sources"]["SS01"]["state"] = "ringing"
            integration = self.validate_current(full, root)
            self.assertFalse(any("父场景 S01" in item and "终态" in item for item in integration.errors), integration.errors)

            unstructured = valid_data_243()
            unstructured["continuity_logs"][0]["characters"] = ["A在门口，面向桌边。"]
            integration = self.validate_current(unstructured, root)
            self.assertTrue(
                any("结构化连续性台账" in item or "缺少可校验的上一状态" in item for item in integration.errors),
                integration.errors,
            )

    def test_p2_inherited_values_are_type_sensitive_and_parent_id_is_string(self) -> None:
        parent = copy.deepcopy(valid_data_243()["continuity_logs"][0])
        parent["nested"] = {"flag": 1}
        child = copy.deepcopy(parent)
        child.update(
            {
                "scene_id": "S02",
                "scene": "2 室内 日 内",
                "inherits_from": "S01",
                "inherited_states": ["/nested/flag"],
                "diverged_states": [],
            }
        )
        child["nested"]["flag"] = True
        data = {"metadata": {"version": "2.4.3"}, "continuity_logs": [parent, child]}
        result = delivery.ValidationResult()
        logs = delivery.validate_continuity_logs(data, result)
        delivery.validate_final_inherited_states(
            data,
            logs,
            {scene: copy.deepcopy(log) for scene, log in logs.items()},
            result,
        )
        self.assertTrue(any("终态" in item for item in result.errors), result.errors)

        parent["inherits_from"] = None
        result = delivery.ValidationResult()
        delivery.validate_continuity_logs(
            {"metadata": {"version": "2.4.3"}, "continuity_logs": [parent]}, result
        )
        self.assertTrue(any("inherits_from" in item and "字符串" in item for item in result.errors), result.errors)

    def test_p2_new_multi_character_checks_do_not_change_legacy_validation(self) -> None:
        data = valid_data_242()
        shot = data["shots"][0]
        shot["camera_main_image"] = (
            "[平视, 全景, 固定镜头]\n【机位逻辑】摄影机正对室内。\n"
            "【场景首镜站位】A站在门口，面向桌边。\nA站在门口。"
        )
        shot["visible_characters"] = ["A"]
        log = data["continuity_logs"][0]
        log["first_shot_anchor_type"] = "multi_character"
        result = delivery.ValidationResult()
        delivery.validate_first_scene_shot(
            shot,
            log,
            result,
            {"B001-F01": "action"},
            current_contract=False,
        )
        self.assertFalse(any("至少需要2人" in item or "position fact" in item for item in result.errors), result.errors)

        current = valid_data_243()
        current_shot = current["shots"][0]
        current_shot["visible_characters"] = ["A", "B"]
        current_shot["camera_main_image"] = (
            "[平视, 中近景, 固定镜头]\n【机位逻辑】摄影机面对主观场域。\n"
            "【场景首镜站位】A在左侧，B在右侧，两人面向彼此。\nA与B对视。"
        )
        current_log = current["continuity_logs"][0]
        current_log["first_shot_anchor_type"] = "subjective"
        result = delivery.ValidationResult()
        delivery.validate_first_scene_shot(
            current_shot,
            current_log,
            result,
            {"B001-F01": "position"},
            current_contract=True,
        )
        self.assertFalse(any("景别必须" in item for item in result.errors), result.errors)

    def test_p2_duration_integer_math_is_exact_and_never_uses_float(self) -> None:
        large = 2**53
        shot = copy.deepcopy(valid_data_243()["shots"][0])
        shot["duration_seconds"] = large
        shot["duration_breakdown"] = {
            "sync_action_seconds": large + 1,
            "sync_dialogue_seconds": 0,
            "non_sync_action_seconds": 0,
            "emotional_pause_seconds": 0,
        }
        result = delivery.ValidationResult()
        delivery.validate_duration(shot, result, version_24=True, strict_contract=True)
        self.assertTrue(any("max(" in item for item in result.errors), result.errors)

        exact = large + 1
        self.assertEqual(str(exact), delivery.fmt_number(exact))
        self.assertEqual(str(exact), delivery.canonical_cell_text(exact))
        self.assertEqual(str(exact), delivery.excel_cell_value(exact))
        prompt_shot = copy.deepcopy(valid_data_243()["shots"][0])
        prompt_shot["duration_seconds"] = exact
        self.assertTrue(delivery.derive_prompt(prompt_shot).startswith(f"时间：0秒-{exact}秒"))
        with tempfile.TemporaryDirectory() as root_name:
            data = valid_data_243()
            data["shots"][0]["duration_seconds"] = exact
            delivery.derive_prompts(data)
            excel_path = Path(root_name) / "large-int.xlsx"
            delivery.build_excel(data, excel_path)
            workbook = load_workbook(excel_path, read_only=False, data_only=False)
            try:
                cell = workbook[delivery.SHEET_NAME]["D2"]
                self.assertEqual("s", cell.data_type)
                self.assertEqual(str(exact), cell.value)
            finally:
                workbook.close()

        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data = valid_data_243()
            huge = 10**1000
            data["shots"][0]["duration_seconds"] = huge
            data["shots"][0]["duration_breakdown"] = {
                "sync_action_seconds": huge,
                "sync_dialogue_seconds": 0,
                "non_sync_action_seconds": 0,
                "emotional_pause_seconds": 0,
            }
            data["shots"][0]["notes"] = "[时长估算] 同步动作已按原生整数登记。"
            data["shots"][0]["long_take"] = {"classification": "action_long", "reason": "极端整数反例"}
            data["shots"][0]["long_take_support"] = []
            result = self.validate_current(data, root)
            self.assertTrue(any("long_take_support" in item for item in result.errors), result.errors)
            self.assertTrue(any("12秒及以上" in item for item in result.errors), result.errors)

        shot["duration_seconds"] = 1
        shot["duration_breakdown"]["sync_action_seconds"] = 10**1000
        result = delivery.ValidationResult()
        delivery.validate_duration(shot, result, version_24=True, strict_contract=True)
        self.assertTrue(any("max(" in item for item in result.errors), result.errors)

    def test_p2_sound_updates_require_nonempty_strings_and_sound_fact_evidence(self) -> None:
        shot = {"shot_no": 1, "covered_fact_ids": ["B001-F01"]}
        base_update = {
            "entity_type": "sound_source",
            "entity": "SS01",
            "field": "state",
            "from": "silent",
            "to": "ringing",
            "evidence_fact_ids": ["B001-F01"],
        }
        for value in ("", 1):
            with self.subTest(to=value):
                update = copy.deepcopy(base_update)
                update["to"] = value
                result = delivery.ValidationResult()
                delivery.validate_update(
                    shot,
                    update,
                    result,
                    False,
                    known_sound_sources={"SS01"},
                    sound_fact_ids={"B001-F01"},
                    current_contract=True,
                )
                self.assertTrue(any("from/to" in item for item in result.errors), result.errors)

        result = delivery.ValidationResult()
        delivery.validate_update(
            shot,
            base_update,
            result,
            False,
            known_sound_sources={"SS01"},
            sound_fact_ids={"B001-F02"},
            current_contract=True,
        )
        self.assertTrue(any("sound fact" in item for item in result.errors), result.errors)

        spaced = copy.deepcopy(base_update)
        spaced["entity"] = " SS01 "
        result = delivery.ValidationResult()
        delivery.validate_update(
            shot,
            spaced,
            result,
            False,
            known_sound_sources={"SS01"},
            sound_fact_ids={"B001-F01"},
            current_contract=True,
        )
        self.assertTrue(any("canonical 字符串" in item for item in result.errors), result.errors)

        bogus = {
            "entity_type": "bogus",
            "entity": "",
            "field": "bogus",
            "from": "a",
            "to": "b",
            "evidence_fact_ids": ["B001-F01"],
        }
        result = delivery.ValidationResult()
        delivery.validate_update(shot, bogus, result, False, current_contract=True)
        self.assertTrue(any("entity_type 必须" in item for item in result.errors), result.errors)
        self.assertTrue(any("field 必须" in item for item in result.errors), result.errors)

    def test_p2_character_arrays_and_multi_anchor_reject_phantom_roles(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            for visible in ([None, False], ["A", " A"]):
                with self.subTest(visible=repr(visible)):
                    data = valid_data_243()
                    data["continuity_logs"][0]["first_shot_anchor_type"] = "multi_character"
                    data["shots"][0]["visible_characters"] = visible
                    result = self.validate_current(data, root)
                    self.assertTrue(
                        any("visible_characters" in item and "字符串" in item for item in result.errors),
                        result.errors,
                    )

            data = valid_data_243()
            log = data["continuity_logs"][0]
            log["first_shot_anchor_type"] = "multi_character"
            log["characters"] = [
                {"name": "A", "position": "左侧", "facing": "右侧"},
                {"name": "B", "position": "右侧", "facing": "左侧"},
            ]
            data["shots"][0]["visible_characters"] = ["A", "X"]
            data["shots"][0]["camera_main_image"] = (
                "[平视, 全景, 固定镜头]\n【机位逻辑】摄影机正对室内。\n"
                "【场景首镜站位】A在左侧，X在右侧，两人面向彼此。\nA站在门口。"
            )
            result = self.validate_current(data, root)
            self.assertTrue(any("未登记或不可识别角色" in item and "X" in item for item in result.errors), result.errors)

    def test_p2_ids_are_ascii_canonical_and_pathological_numbers_fail_stably(self) -> None:
        for beat_id in ("B000", "B0001"):
            data = valid_data_243()
            data["beats"][0]["beat_id"] = beat_id
            result = delivery.ValidationResult()
            delivery.collect_facts(data, result)
            self.assertTrue(any("Beat ID 不合法" in item for item in result.errors), result.errors)
        self.assertIsNone(delivery.numeric_beat_id("B" + "9" * 5000))
        with self.assertRaises(KeyError):
            delivery.resolve_json_pointer([], "/" + "9" * 5000)

        for source_id in ("SS٠١", " SS01 "):
            log = copy.deepcopy(valid_data_243()["continuity_logs"][0])
            log["sound_sources"] = {
                source_id: {"name": "声源", "visibility": "offscreen", "position": "门外", "state": "silent"}
            }
            result = delivery.ValidationResult()
            delivery.validate_sound_sources(log, "S01", result)
            self.assertTrue(any("SS[0-9]{2,}" in item for item in result.errors), result.errors)

        for scene_id in ("S٠١", " S01 "):
            log = copy.deepcopy(valid_data_243()["continuity_logs"][0])
            log["scene_id"] = scene_id
            result = delivery.ValidationResult()
            delivery.validate_continuity_logs(
                {"metadata": {"version": "2.4.3"}, "continuity_logs": [log]}, result
            )
            self.assertTrue(any("S[0-9]{2,}" in item for item in result.errors), result.errors)

        for batch_id in ("BT٠١", " BT01 "):
            result = delivery.ValidationResult()
            delivery.validate_batch_plan(
                {
                    "metadata": {"version": "2.4.3"},
                    "batch_plan": {
                        "batches": [
                            {"batch_id": batch_id, "scene_ids": ["S01"], "expected_shot_count": 1, "depends_on": []}
                        ]
                    },
                },
                result,
            )
            self.assertTrue(any("BT[0-9]{2,}" in item for item in result.errors), result.errors)

    def test_p2_project_lexicon_rejects_unicode_control_and_format_characters(self) -> None:
        for term in ("ab\u0085cd", "ab\u202ecd"):
            result = delivery.ValidationResult()
            delivery.validate_project_lexicon(
                {"metadata": {"version": "2.4.3", "project_lexicon": {"prop_terms": [term]}}},
                result,
            )
            self.assertTrue(any("Unicode 控制/格式" in item for item in result.errors), result.errors)

    def test_p2_review_hash_is_public_deterministic_and_read_only(self) -> None:
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            path = root / "review.shot_data.json"
            data = valid_data_243()
            delivery.write_json(path, data)
            before = path.read_bytes()
            args = delivery.make_parser().parse_args(
                ["review-hash", "--data", str(path), "--gate", "GATE_A"]
            )
            with mock.patch("builtins.print") as printer:
                self.assertEqual(0, args.func(args))
            output = json.loads(printer.call_args.args[0])
            self.assertEqual(delivery.gate_review_hash(data, "GATE_A"), output["reviewed_hash"])
            self.assertEqual(before, path.read_bytes())

    def test_p2_source_column_preserves_internal_whitespace_and_line_breaks(self) -> None:
        data = valid_data_243()
        source = "第一行  双空格\n第二行<br>与竖线|"
        data["shots"][0]["source_paragraph"] = source
        expected = f"B001～{source}"
        self.assertEqual(expected, delivery.expected_rows(data)[0][2])
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            markdown_path = root / "source.md"
            excel_path = root / "source.xlsx"
            delivery.build_markdown(data, markdown_path)
            delivery.build_excel(data, excel_path)
            self.assertEqual(expected, delivery.parse_markdown_table(markdown_path)[1][2])
            workbook = load_workbook(excel_path, read_only=True, data_only=False)
            try:
                self.assertEqual(expected, workbook[delivery.SHEET_NAME]["C2"].value)
            finally:
                workbook.close()
        legacy = valid_data_242()
        legacy["shots"][0]["source_paragraph"] = source
        self.assertEqual("B001～第一行 双空格 第二行<br>与竖线|", delivery.expected_rows(legacy)[0][2])

    def test_p2_depth_motion_terminology_keeps_legacy_approval_tokens(self) -> None:
        self.assertEqual("push_in", delivery.depth_motion_direction("缓慢推进"))
        self.assertEqual("pull_out", delivery.depth_motion_direction("缓慢拉出"))
        with tempfile.TemporaryDirectory() as root_name:
            root = Path(root_name)
            data = valid_data_243()
            approved = root / "approved" / "sample.approved_script.txt"
            approved.parent.mkdir(parents=True, exist_ok=True)
            approved.write_text(data["script_lock"]["locked_text"], encoding="utf-8")
            data["shots"][0]["camera_main_image"] = (
                "[平视, 中景, 缓慢推进]\n【机位逻辑】摄影机向A推进。\n"
                "【场景首镜站位】A站在门口，面向桌边。\nA站在门口。"
            )
            data["shots"][1]["camera_main_image"] = (
                "[平视, 中景, 缓慢拉出]\n【机位逻辑】摄影机从A身前拉出。\n"
                "【站位位移】A从门口走到桌边，面向B的位置。\nA在桌边停下，说：“到了。”"
            )
            data["shots"][1]["notes"] = "[反转动机] 揭示后拉出暴露空间。"
            data["human_reviews"] = []
            append_review(data, "GATE_A")
            append_review(data, "GATE_B", approved_items=["shot:2:axis-reversal"])
            delivery.derive_prompts(data)
            result = delivery.validate_data(
                data,
                strict_status=False,
                data_path=root / "sample.shot_data.json",
                workspace_root=root,
            )
            self.assertFalse(any("approved_items 精确批准" in item for item in result.errors), result.errors)
            self.assertTrue(any("纵深推拉方向反转" in item for item in result.warnings), result.warnings)


if __name__ == "__main__":
    unittest.main()
